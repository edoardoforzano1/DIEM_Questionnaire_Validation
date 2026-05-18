import openpyxl
p = r"C:\Temp\questionnaire_validation\kobo_output\report_kobo_ar_PSE_R2_20260518.xlsx"
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb['Choice Changes']
hdr=None; hrow=None
for r in range(1,80):
    vals=[ws.cell(r,c).value for c in range(1,40)]
    if any((str(v).strip()=='Issue type') for v in vals if v is not None):
        hdr=vals; hrow=r; break
cols={str(v).strip():i+1 for i,v in enumerate(hdr) if v is not None}
rows=[]
for r in range(hrow+1, ws.max_row+1):
    it=ws.cell(r, cols['Issue type']).value
    if it is None or str(it).strip()=='' or str(it).strip().lower().startswith('no issues'):
        continue
    d={k:ws.cell(r,c).value for k,c in cols.items()}; rows.append(d)
focus=[('hh_agricactivity','agricactivity'),('hh_wealth_water','water'),('ls_main','lsmain'),('ls_salesmain','salesmain')]
for q,l in focus:
    print('\n===',q,'|',l,'===')
    for d in rows:
        if str(d.get('Q Name') or '')==q and str(d.get('list_name') or '')==l:
            print(str(d.get('Issue type')),'|',str(d.get('Field')),'| CUR=',str(d.get('Current value'))[:110].replace('\n',' / '),'| REF=',str(d.get('Reference / rule'))[:110].replace('\n',' / '),'| SEV=',d.get('Severity'))
