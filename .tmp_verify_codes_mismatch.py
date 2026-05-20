from collections import Counter
from openpyxl import load_workbook
p=r'c:/git/DIEM_Questionnaire_Validation/tmp_report_out5/geopoll_output/report_geopoll_fr_COD_R11_20260520.xlsx'
wb=load_workbook(p)
ws=wb['Option Changes']
h=[ws.cell(2,c).value for c in range(1,20)]
idx={str(v):c for c,v in enumerate(h,1) if v}
issue=idx['Issue type']; q=idx['Q Name']; det=idx['Detail']; sev=idx['Severity']; scope=idx.get('Language scope')
rows=[]
for r in range(3, ws.max_row+1):
    it=ws.cell(r,issue).value
    if it:
        rows.append((str(it), str(ws.cell(r,q).value or ''), str(ws.cell(r,det).value or ''), str(ws.cell(r,sev).value or ''), str(ws.cell(r,scope).value or '')))
ctr=Counter([x[0] for x in rows])
print('counts', {k:v for k,v in ctr.items() if k.startswith('codes_')})
for rec in rows:
    if rec[0]=='codes_option_count_mismatch':
        print('sample',rec)
        break
