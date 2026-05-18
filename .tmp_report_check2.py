import openpyxl, collections
p = r"C:\Temp\questionnaire_validation\kobo_output\report_kobo_ar_PSE_R2_20260518.xlsx"
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb['Choice Changes']
# header
hdr=None; hrow=None
for r in range(1,80):
    vals=[ws.cell(r,c).value for c in range(1,40)]
    if any((str(v).strip()=='Issue type') for v in vals if v is not None):
        hdr=vals; hrow=r; break
cols={str(v).strip():i+1 for i,v in enumerate(hdr) if v is not None}
rows=[]
for r in range(hrow+1, ws.max_row+1):
    it=ws.cell(r, cols['Issue type']).value
    if it is None or str(it).strip()=='' or str(it).strip().lower().startswith('no issues'):
        continue
    d={k:ws.cell(r,c).value for k,c in cols.items()}
    rows.append(d)
by_key=collections.defaultdict(list)
for d in rows:
    key=(str(d.get('Q Name') or ''), str(d.get('list_name') or ''))
    by_key[key].append(d)
print('LABEL_MISMATCH_BY_KEY')
for (q,l),items in sorted(by_key.items()):
    lm=sum(1 for x in items if str(x.get('Issue type'))=='choice_label_mismatch')
    if lm==0:
        continue
    issues=sorted(set(str(x.get('Issue type')) for x in items))
    print(f'{q} | {l} | label_mismatch={lm} | issues={";".join(issues)}')
print('REMOVED_ONLY_OR_ADDED_ONLY_KEYS')
for (q,l),items in sorted(by_key.items()):
    issues=set(str(x.get('Issue type')) for x in items)
    if 'removed_choice' in issues or 'added_choice' in issues:
        print(f'{q} | {l} | issues={";".join(sorted(issues))}')
