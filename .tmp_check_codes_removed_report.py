from collections import Counter
from openpyxl import load_workbook
wb=load_workbook(r'c:/git/DIEM_Questionnaire_Validation/tmp_report_out4/geopoll_output/report_geopoll_fr_COD_R11_20260520.xlsx')
ws=wb['Option Changes']
h=[ws.cell(2,c).value for c in range(1,20)]
idx={str(v):c for c,v in enumerate(h,start=1) if v}
issue=idx['Issue type']; qcol=idx['Q Name']; dcol=idx['Detail']; ccol=idx['Current value']; rcol=idx['Reference / rule']; scol=idx.get('Language scope')
rows=[]
for r in range(3, ws.max_row+1):
    it=ws.cell(r,issue).value
    if it:
        rows.append((r,str(it),str(ws.cell(r,qcol).value or ''),str(ws.cell(r,dcol).value or ''),str(ws.cell(r,ccol).value or ''),str(ws.cell(r,rcol).value or ''),str(ws.cell(r,scol).value or '')))
ctr=Counter(x[1] for x in rows)
print('codes counts:', {k:v for k,v in ctr.items() if k.startswith('codes_')})
for rec in rows:
    if rec[1].startswith('codes_') and rec[2]=='assistance_provider':
        print('row',rec[0],rec[1],'scope=',rec[6],'detail=',rec[3],'cur=',rec[4][:80],'ref=',rec[5][:80])
