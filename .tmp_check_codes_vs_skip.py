import openpyxl
from pathlib import Path
qpath = Path(r'C:/Temp/DRC_R11 HH questionnaire_Geopoll_TEST.xlsx')
wb = openpyxl.load_workbook(qpath, data_only=True, read_only=True)
print('sheets:', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
# find header row containing Q Name
header_row = None
idx = {}
for rr in range(1, 15):
    vals = [ws.cell(rr, c).value for c in range(1, 80)]
    if any(str(v).strip() == 'Q Name' for v in vals if v is not None):
        header_row = rr
        idx = {str(v).strip(): c for c, v in enumerate(vals, start=1) if v is not None}
        break
print('header_row:', header_row)
qcol = idx.get('Q Name')
ccol = idx.get('Codes')
skcol = idx.get('Skip Pattern')
print('columns:', qcol, ccol, skcol)
for r in range((header_row or 3) + 1, ws.max_row + 1):
    q = ws.cell(r, qcol).value if qcol else None
    if q in ('income_sec_control', 'income_sec_comp', 'income_third'):
        cv = ws.cell(r, ccol).value if ccol else None
        sv = ws.cell(r, skcol).value if skcol else None
        print(r, q, '| Codes=', cv, '| Skip=', sv)
