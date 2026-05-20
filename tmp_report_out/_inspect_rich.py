from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
p = r"tmp_report_out/geopoll_output/report_geopoll_fr_COD_R11_20260520.xlsx"
wb = load_workbook(p)
ws = wb['Option Changes']
# row 11 from prior sample, detail col=4
cell = ws.cell(row=11, column=4)
print(type(cell.value).__name__)
print(cell.value)
if isinstance(cell.value, CellRichText):
    for i, part in enumerate(cell.value):
        txt = getattr(part, 'text', str(part))
        font = getattr(part, 'font', None)
        b = getattr(font, 'b', None) if font is not None else None
        print(i, repr(txt), 'bold=', b)
