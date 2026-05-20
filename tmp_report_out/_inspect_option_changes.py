from collections import Counter
from openpyxl import load_workbook
p = r"tmp_report_out/geopoll_output/report_geopoll_fr_COD_R11_20260520.xlsx"
wb = load_workbook(p)
ws = wb['Option Changes']
header = [ws.cell(row=2, column=i).value for i in range(1, 20)]
idx = {str(v): i + 1 for i, v in enumerate(header) if v}
print('HEADER', header)
print('IDX', idx)
issue_col = idx.get('Issue type')
q_col = idx.get('Q Name')
detail_col = idx.get('Detail')
cur_col = idx.get('Current value')
ref_col = idx.get('Reference / rule')
rows = []
for r in range(3, ws.max_row + 1):
    issue = ws.cell(r, issue_col).value if issue_col else None
    if issue:
        rows.append((
            r,
            issue,
            ws.cell(r, q_col).value if q_col else None,
            ws.cell(r, detail_col).value if detail_col else None,
            ws.cell(r, cur_col).value if cur_col else None,
            ws.cell(r, ref_col).value if ref_col else None,
        ))
counts = Counter([x[1] for x in rows])
print('COUNTS', dict(counts))
for t in ['option_changes (added/removed)', 'added_option', 'removed_option', 'removed_option_cascading_drift', 'option_position_drift']:
    print('TYPE', t, '->', counts.get(t, 0))
print('--- SAMPLE ROWS ---')
shown = 0
for rec in rows:
    if rec[1] in ('option_changes (added/removed)', 'added_option', 'removed_option'):
        print('ROW', rec[0], rec[1], rec[2])
        print('DETAIL', rec[3])
        print('CURRENT_HEAD', str(rec[4]).split('\n')[:6])
        print('REFERENCE_HEAD', str(rec[5]).split('\n')[:6])
        print('---')
        shown += 1
        if shown >= 3:
            break
