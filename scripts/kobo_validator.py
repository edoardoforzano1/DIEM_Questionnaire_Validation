# ---------------------------------------------------------------------------
# Terminal output helpers (injected during conversion from notebook)
# ---------------------------------------------------------------------------
import sys as _sys, time as _time, pathlib, warnings as _warnings, atexit as _atexit

_warnings.filterwarnings("ignore")

_STEP_N     = 0
_STEP_TOTAL = 0
_T_START    = 0.0
_BANNER_W   = 66
_REAL_STDOUT = _sys.stdout


class _NullOutput:
    """Swallows intermediate notebook prints."""
    def write(self, *a, **kw): pass
    def flush(self, *a, **kw): pass
    def fileno(self): raise OSError("not a real file")


_atexit.register(lambda: setattr(_sys, "stdout", _REAL_STDOUT))


def _banner(tool: str, language: str, iso3: str, mode: str) -> None:
    global _T_START
    _T_START = _time.perf_counter()
    _sys.stdout = _NullOutput()
    print("", file=_REAL_STDOUT)
    print("=" * _BANNER_W, file=_REAL_STDOUT)
    print("  DIEM Questionnaire Validation v2", file=_REAL_STDOUT)
    print(f"  Tool: {tool.upper()}  |  Language: {language.upper()}  |  Country: {iso3.upper()}", file=_REAL_STDOUT)
    print(f"  Reference mode: {mode}", file=_REAL_STDOUT)
    print("=" * _BANNER_W, file=_REAL_STDOUT)


def _info(label: str, value: str) -> None:
    print(f"  {label:<12} {value}", file=_REAL_STDOUT)


def _step(label: str, detail: str = "") -> None:
    global _STEP_N
    _STEP_N += 1
    tag  = f"[{_STEP_N}/{_STEP_TOTAL}]"
    pad  = "." * max(1, 44 - len(label))
    tail = f"  {detail}" if detail else ""
    print(f"  {tag} {label} {pad}{tail}", file=_REAL_STDOUT, flush=True)


def _summary(all_issues=None, report_path=None, extra_paths=None) -> None:
    _sys.stdout = _REAL_STDOUT
    print()
    print("-" * _BANNER_W)
    print("  VALIDATION SUMMARY")
    print("-" * _BANNER_W)
    try:
        import polars as _pl
        if all_issues is not None and all_issues.height > 0:
            sev = {r["severity"]: r["n"] for r in
                   all_issues.group_by("severity").agg(_pl.len().alias("n")).to_dicts()}
            high   = sev.get("high",   0)
            medium = sev.get("medium", 0)
            info   = sev.get("info",   0)
            print(f"  {'HIGH':<10} {high:>4}   {'Must fix before deployment' if high else 'None'}")
            print(f"  {'MEDIUM':<10} {medium:>4}   {'Review suggested' if medium else 'None'}")
            print(f"  {'INFO':<10} {info:>4}   Informational")
            print(f"  {'PASS':<10} {'yes' if high == 0 and medium == 0 else 'no':>4}")
        else:
            print("  No issues found — questionnaire passed all checks.")
    except Exception as _e:
        print(f"  (summary unavailable: {_e})")
    print()
    if report_path:
        print(f"  Report  : {report_path}")
    if extra_paths:
        for _lbl, _p in extra_paths:
            print(f"  {_lbl:<8}: {_p}")
    print(f"  Time    : {_time.perf_counter() - _T_START:.1f}s")
    print("=" * _BANNER_W)
    print()


# ---------------------------------------------------------------------------

# ======================================================================
_STEP_TOTAL = 13
# SECTION: ## KoBO Questionnaire Validator
# ======================================================================
# --- CODE CELL 1 ---
import re
import polars as pl
import openpyxl
from pathlib import Path
from dataclasses import dataclass

# ======================================================================
# SECTION: ## Step 1 -- Configuration
_step("Loading configuration")
# ======================================================================
# --- CODE CELL 2 ---
import yaml
from datetime import datetime


def _load_effective_config(cfg_path: Path) -> tuple[dict, Path | None, Path]:
    base_cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
    if not isinstance(base_cfg, dict):
        raise TypeError(f"Config must be a mapping in {cfg_path}")

    profile_name = str(base_cfg.get("config_profile") or "").strip()
    profile_path: Path | None = None
    effective_cfg = dict(base_cfg)

    if profile_name:
        profiles_dir = Path(base_cfg.get("config_profiles_dir") or (Path(base_cfg.get("output_dir") or base_cfg.get("working_dir") or "C:/Temp/questionnaire_validation") / "config_profiles"))
        candidate = Path(profile_name)
        if not candidate.is_absolute():
            candidate = profiles_dir / candidate

        if candidate.suffix.lower() not in {".yaml", ".yml"}:
            if candidate.with_suffix(".yaml").exists():
                candidate = candidate.with_suffix(".yaml")
            elif candidate.with_suffix(".yml").exists():
                candidate = candidate.with_suffix(".yml")

        if not candidate.exists():
            raise FileNotFoundError(
                f"Config profile not found: {candidate}\n"
                f"Set 'config_profile' to an existing file or leave it empty."
            )

        profile_cfg = yaml.safe_load(candidate.read_text(encoding="utf-8")) or {}
        if not isinstance(profile_cfg, dict):
            raise TypeError(f"Config profile must be a mapping in {candidate}")

        effective_cfg.update(profile_cfg)
        profile_path = candidate.resolve()

    active_cfg_path = profile_path or cfg_path.resolve()
    return effective_cfg, profile_path, active_cfg_path


# -- Load centralized config ---------------------------------------------------
# Edit validation_config.yaml -- never touch notebook code.
_cfg_base_path = Path(__file__).parent.parent / "configuration" / "validation_config.yaml"
cfg, _cfg_profile_path, _cfg_active_path = _load_effective_config(_cfg_base_path)

# Backward-compatible path resolution:
# - preferred: working_dir + questionnaire_file
# - fallback : questionnaire_path (legacy)
_working_dir = Path(cfg.get("working_dir") or "C:/Temp")
if cfg.get("questionnaire_file"):
    _q_path = _working_dir / str(cfg["questionnaire_file"])
elif cfg.get("questionnaire_path"):
    _q_path = Path(cfg["questionnaire_path"])
else:
    raise ValueError("Provide either 'questionnaire_file' or 'questionnaire_path' in validation_config.yaml")

cfg["working_dir"] = str(_working_dir)
cfg["questionnaire_path"] = str(_q_path)
_output_base = Path(cfg.get("output_dir") or str(_working_dir))
cfg["output_dir"] = str(_output_base / "kobo_output")
cfg["output_base_dir"] = str(_output_base)
Path(cfg["output_dir"]).mkdir(parents=True, exist_ok=True)

_banner("kobo", cfg.get("language","?"), cfg.get("iso3","?"), cfg.get("reference_mode","?"))
_info("Config", str(_cfg_active_path))
if _cfg_profile_path:
    _info("Base cfg", str(_cfg_base_path.resolve()))
_info("File", Path(cfg['questionnaire_path']).name)
_info("Output", cfg.get('output_dir', '(same as questionnaire)'))


# ======================================================================
# SECTION: ## Step 2 -- Reference file resolution
_step("Resolving reference file")
# ======================================================================
# --- CODE CELL 3 ---
# Maps language code -> label column name in KoBO XLSForms
LANG_LABEL_COL = {
    "en": "label::English (en)",
    "fr": "label::French (fr)",
    "ar": "label::Arabic (ar)",
    "es": "label::Spanish (es)",
}

def _write_config_snapshot_kobo(
    cfg: dict,
    run: dict,
    cfg_base_path: Path,
    cfg_active_path: Path,
) -> Path:
    logs_root = Path(cfg.get("output_base_dir") or cfg.get("output_dir") or Path(run["questionnaire_path"]).parent)
    logs_dir = logs_root / "configuration"
    logs_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = logs_dir / f"config_kobo_{str(cfg.get('language','')).upper()}_{cfg.get('iso3','')}_{ts}.yaml"

    payload = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "tool": "kobo",
        "active_config_file": str(cfg_active_path),
        "base_config_file": str(cfg_base_path.resolve()),
        "config_profile": str(cfg.get("config_profile") or ""),
        "resolved_config": dict(cfg),
        "resolved_run_paths": {k: str(v) for k, v in run.items()},
    }

    with out_file.open("w", encoding="utf-8") as f:
        yaml.safe_dump(payload, f, sort_keys=False, allow_unicode=True)
    return out_file


def find_kobo_template(language: str, templates_dir: Path) -> Path:
    lang_upper = language.upper()
    candidates = sorted(
        templates_dir.glob(f"*kobo*{lang_upper}*template*.xlsx"),
        key=lambda p: p.stat().st_mtime, reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"No KoBO template for language {language!r} in {templates_dir}")
    return candidates[0]

TEMPLATES_DIR = Path(cfg["templates_dir"])
working_dir   = Path(cfg.get("working_dir") or Path(cfg["questionnaire_path"]).parent)
template_path = find_kobo_template(cfg["language"], TEMPLATES_DIR)

if cfg["reference_mode"] == "latest_template":
    ref_path = template_path
elif cfg["reference_mode"] == "previous_round":
    if cfg.get("previous_round_file"):
        ref_path = working_dir / str(cfg["previous_round_file"])
    elif cfg.get("previous_round_path"):
        ref_path = Path(cfg["previous_round_path"])
    else:
        raise ValueError("reference_mode='previous_round' requires previous_round_file (or previous_round_path)")
else:
    raise ValueError(f"Unknown reference_mode: {cfg['reference_mode']!r}")

run = {
    "questionnaire_path": cfg["questionnaire_path"],
    "reference_path"    : str(ref_path),
    "template_path"     : str(template_path),
    "reference_mode"    : cfg["reference_mode"],
    "language"          : cfg["language"],
    "iso3"              : cfg["iso3"],
    "working_dir"       : str(working_dir),
    "output_dir"        : cfg.get("output_dir") or str(working_dir),
}

print(f"Reference : {Path(run['reference_path']).name}")
if run["reference_mode"] == "previous_round":
    print(f"Template  : {Path(run['template_path']).name}  (placeholder map only)")
print(f"Country   : {Path(run['questionnaire_path']).name}")
print(f"Language  : {run['language']}   ISO3: {run['iso3']}")


def _reference_scope_label() -> str:
    mode = str(run.get("reference_mode", "") or "").strip().lower()
    if mode == "previous_round":
        return "previous round"
    return "latest template"


def _extract_round_token_kobo(name: str) -> str:
    txt = str(name or "")
    patterns = [
        r"(?i)(?:^|[_\-\s])R(\d{1,3})(?=$|[_\-\s\.])",
        r"(?i)(?:^|[_\-\s])ROUND[_\-\s]*(\d{1,3})(?=$|[_\-\s\.])",
        # Fallback for suffix forms like "...copyr10.xlsx"
        r"(?i)R(?:OUND)?[_\-\s]*([0-9]{1,3})(?=(?:\.[A-Za-z0-9]+)?$)",
    ]
    for pat in patterns:
        m = re.search(pat, txt)
        if m:
            return f"R{m.group(1)}"
    return ""


def _reference_descriptor_kobo() -> str:
    """Human-readable reference descriptor for KoBo Summary header."""
    mode = _reference_scope_label()
    ref_name = Path(run.get("reference_path", "")).name if run.get("reference_path") else ""
    iso = str(run.get("iso3", "") or "").upper()
    rnd = _extract_round_token_kobo(ref_name) or "R?"
    lang = str(run.get("language", "") or "").upper()
    if mode == "previous round":
        return f"{mode} | {iso} {rnd} | {ref_name}"
    return f"{mode} | {lang} template | {ref_name}"


def _language_scope_descriptor_kobo() -> str:
    lang = str(run.get("language", "") or "").lower()
    label_col = LANG_LABEL_COL.get(lang, f"label::{lang}")
    return f"{lang.upper()} labels ({label_col})"


def _not_in_reference_text() -> str:
    return f"(not in {_reference_scope_label()})"

try:
    _config_snapshot_file = _write_config_snapshot_kobo(
        cfg=cfg,
        run=run,
        cfg_base_path=_cfg_base_path,
        cfg_active_path=_cfg_active_path,
    )
    print(f"Config snapshot: {_config_snapshot_file}")
except Exception as _cfg_log_err:
    print(f"Warning: could not write config snapshot: {_cfg_log_err}")


# ======================================================================
# SECTION: ## Step 3 -- Readers
_step("Readers")
# ======================================================================
# --- CODE CELL 4 ---
_METADATA_TYPES = {
    "start", "end", "today", "deviceid", "phonenumber",
    "username", "simserial", "subscriberid", "background-audio",
}
_SELECT_TYPES = {"select_one", "select_multiple"}
_DATA_TYPES = {
    "select_one", "select_multiple", "text", "integer", "decimal",
    "geopoint", "geotrace", "geoshape", "date", "time", "datetime",
    "file", "image", "audio", "video", "barcode", "range", "calculate",
}

def _find_label_col(headers: list, language: str) -> str | None:
    target = LANG_LABEL_COL.get(language.lower(), f"label::{language}")
    if target in headers:
        return target
    return next((h for h in headers if h.startswith("label::")), None)


def _normalize_mandatory(val) -> str:
    v = str(val or "").strip().lower()
    if v in ("yes", "true", "mandatory"):
        return "mandatory"
    if v in ("yes - panel", "yes-panel", "mandatory-panel"):
        return "mandatory-panel"
    if v in ("no", "false"):
        return "non-mandatory"
    return ""


def read_kobo_survey(path: str, language: str = "en", _wb=None) -> pl.DataFrame:
    EMPTY = {
        "Q Name": pl.Utf8, "Q Type": pl.Utf8, "list_name": pl.Utf8,
        "label": pl.Utf8, "required": pl.Utf8, "mandatory_category": pl.Utf8,
        "relevant": pl.Utf8, "constraint": pl.Utf8, "choice_filter": pl.Utf8,
        "appearance": pl.Utf8, "calculation": pl.Utf8, "hint": pl.Utf8,
        "excel_row": pl.Int64, "source_file": pl.Utf8,
    }
    owns_wb = _wb is None
    wb = _wb or openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws   = wb["survey"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        if owns_wb:
            wb.close()

    if not rows:
        return pl.DataFrame(schema=EMPTY)

    headers   = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_idx   = {h: i for i, h in enumerate(headers)}
    label_col = _find_label_col(headers, language)
    mand_col  = next((h for h in headers if h.strip() == "Mandatory"), None)

    def get(row, col):
        i = col_idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    records = []
    for excel_row, row in enumerate(rows[1:], start=2):
        q_type_raw = str(get(row, "type") or "").strip()
        q_name     = str(get(row, "name") or "").strip()
        if not q_name or not q_type_raw or q_type_raw in _METADATA_TYPES:
            continue

        parts     = q_type_raw.split(None, 1)
        q_type    = parts[0]
        list_name = parts[1].strip() if len(parts) > 1 and q_type in _SELECT_TYPES else None

        # Support legacy KoBo encoding where list name is embedded with underscore,
        # e.g. "select_one_debt" / "select_multiple_debt".
        if list_name is None:
            if q_type_raw.startswith("select_one_"):
                q_type = "select_one"
                list_name = q_type_raw[len("select_one_"):].strip() or None
            elif q_type_raw.startswith("select_multiple_"):
                q_type = "select_multiple"
                list_name = q_type_raw[len("select_multiple_"):].strip() or None

        mand_cat = _normalize_mandatory(get(row, mand_col))
        if not mand_cat and q_type in _DATA_TYPES:
            mand_cat = "non-mandatory"
        if q_name.startswith("o_"):
            mand_cat = "optional"

        records.append({
            "Q Name"            : q_name,
            "Q Type"            : q_type,
            "list_name"         : list_name,
            "label"             : str(get(row, label_col) or "") if label_col else "",
            "required"          : str(get(row, "required")    or "").strip().lower(),
            "mandatory_category": mand_cat,
            "relevant"          : str(get(row, "relevant")    or "").strip(),
            "constraint"        : str(get(row, "constraint")  or "").strip(),
            "choice_filter"     : str(get(row, "choice_filter") or "").strip(),
            "appearance"        : str(get(row, "appearance")  or "").strip(),
            "calculation"       : str(get(row, "calculation") or "").strip(),
            "hint"             : str(get(row, "hint")        or "").strip(),
            "excel_row"         : excel_row,
            "source_file"       : str(path),
        })
    return pl.DataFrame(records) if records else pl.DataFrame(schema=EMPTY)


def read_kobo_choices(path: str, language: str = "en", _wb=None) -> pl.DataFrame:
    EMPTY = {"list_name": pl.Utf8, "option_name": pl.Utf8, "option_label": pl.Utf8}
    owns_wb = _wb is None
    wb = _wb or openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws   = wb["choices"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        if owns_wb:
            wb.close()
    if not rows:
        return pl.DataFrame(schema=EMPTY)

    headers   = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_idx   = {h: i for i, h in enumerate(headers)}
    label_col = _find_label_col(headers, language)

    def get(row, col):
        i = col_idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    records = []
    for row in rows[1:]:
        list_name = get(row, "list_name")
        name_raw  = get(row, "name")
        if not list_name or name_raw is None:
            continue
        records.append({
            "list_name"   : str(list_name).strip(),
            "option_name" : str(name_raw).strip(),
            "option_label": str(get(row, label_col) or "").strip() if label_col else "",
        })
    return pl.DataFrame(records) if records else pl.DataFrame(schema=EMPTY)


def build_question_options(survey_df: pl.DataFrame, choices_df: pl.DataFrame) -> pl.DataFrame:
    return (
        survey_df
        .filter(pl.col("Q Type").is_in(list(_SELECT_TYPES)))
        .filter(pl.col("list_name").is_not_null() & (pl.col("list_name") != ""))
        .select(["Q Name", "Q Type", "list_name", "source_file"])
        .join(choices_df, on="list_name", how="left")
        .select(["Q Name", "Q Type", "list_name", "option_name", "option_label", "source_file"])
    )



# ======================================================================
# SECTION: ## Step 4 -- Load files
_step("Loading files")
# ======================================================================
# --- CODE CELL 5 ---
current_wb   = openpyxl.load_workbook(run["questionnaire_path"], data_only=True, read_only=True)
reference_wb = openpyxl.load_workbook(run["reference_path"],     data_only=True, read_only=True)

current_survey    = read_kobo_survey(run["questionnaire_path"],  run["language"], _wb=current_wb)
current_choices   = read_kobo_choices(run["questionnaire_path"], run["language"], _wb=current_wb)
current_wb.close()

reference_survey  = read_kobo_survey(run["reference_path"],  run["language"], _wb=reference_wb)
reference_choices = read_kobo_choices(run["reference_path"], run["language"], _wb=reference_wb)
reference_wb.close()

# Template is needed as placeholder map in previous_round mode.
if run.get("reference_mode") == "previous_round":
    template_wb = openpyxl.load_workbook(run["template_path"], data_only=True, read_only=True)
    template_survey  = read_kobo_survey(run["template_path"],  run["language"], _wb=template_wb)
    template_choices = read_kobo_choices(run["template_path"], run["language"], _wb=template_wb)
    template_wb.close()
else:
    template_survey  = reference_survey
    template_choices = reference_choices

current_options   = build_question_options(current_survey,   current_choices)
reference_options = build_question_options(reference_survey, reference_choices)
template_options  = build_question_options(template_survey,  template_choices)

print(f"Current  : {current_survey.height} questions, {current_choices.height} choice rows")
print(f"Reference: {reference_survey.height} questions, {reference_choices.height} choice rows")
if run.get("reference_mode") == "previous_round":
    print(f"Template : {template_survey.height} questions, {template_choices.height} choice rows")
print()
print(current_survey.select(["Q Name", "Q Type", "mandatory_category", "relevant"]).head(12))


# ======================================================================
# SECTION: ## Step 4b -- Placeholder normalization
_step("Placeholder normalisation")
# ======================================================================
# --- CODE CELL 6 ---
_PLACEHOLDER_RE = re.compile(r'#[^#]+#')


def read_additional_info(country_path: str, language: str = "en") -> dict[str, str]:
    """
    Read the country questionnaire's Additional information sheet.
    Layout: row 1 empty, row 2 = headers (Original | Replacement), rows 3+ = data.
    Returns {#placeholder#: actual_value} for every filled-in entry.
    """
    try:
        wb = openpyxl.load_workbook(country_path, data_only=True, read_only=True)
        if "Additional information" not in wb.sheetnames:
            wb.close()
            return {}
        ws   = wb["Additional information"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        print(f"Warning: could not read Additional information sheet: {e}")
        return {}

    if len(rows) < 2:
        return {}
    headers = [str(h).strip() if h is not None else "" for h in rows[1]]

    orig_idx = next((i for i, h in enumerate(headers) if h.lower().startswith("original")), None)
    if language == "fr":
        repl_idx = next(
            (i for i, h in enumerate(headers) if "replacement" in h.lower() and "fr" in h.lower()), None
        )
    elif language == "ar":
        repl_idx = next(
            (i for i, h in enumerate(headers) if "replacement" in h.lower() and "ar" in h.lower()), None
        )
    else:
        repl_idx = next((i for i, h in enumerate(headers) if h.lower() == "replacement"), None)
        if repl_idx is None:
            repl_idx = next((i for i, h in enumerate(headers) if "replacement" in h.lower()), None)

    if orig_idx is None or repl_idx is None:
        print("Warning: Original/Replacement columns not found in Additional information sheet.")
        return {}

    def _placeholder_aliases(s: str) -> set[str]:
        """Generate tolerant alias tokens (case + simple singular/plural) for replacements."""
        s = (s or "").strip()
        if not s:
            return set()
        aliases = {s, s.lower(), s.upper()}
        low = s.lower()
        # Simple singular/plural tolerance (e.g., unit <-> units)
        if low.endswith("s") and len(s) > 1:
            stem = s[:-1]
            aliases |= {stem, stem.lower(), stem.upper()}
        else:
            plus = s + "s"
            aliases |= {plus, plus.lower(), plus.upper()}
        return aliases

    pairs: dict[str, str] = {}
    for row in rows[2:]:
        orig = row[orig_idx] if orig_idx < len(row) else None
        repl = row[repl_idx] if repl_idx < len(row) else None
        if not orig or str(orig).strip() == "":
            continue
        repl_str = str(repl).strip() if repl is not None else ""
        if repl_str in ("", "nan", "None"):
            continue
        for _alias in _placeholder_aliases(str(orig)):
            pairs[f"#{_alias}#"] = repl_str
    return pairs


def restore_to_vanilla(
    current_df  : pl.DataFrame,
    reference_df: pl.DataFrame,
    cols        : list[str],
    key_cols    : list[str] | None = None,
) -> pl.DataFrame:
    """
    For every cell where the reference contains a #placeholder# token, overwrite
    the current questionnaire's value with the reference (vanilla) value.

    This prevents placeholder-filled country values from being flagged as deviations.
    key_cols: join keys -- defaults to ["Q Name"]; use ["Q Name", "option_name"] for options.
    cols    : columns to restore (e.g. ["label", "constraint"] or ["option_label"]).
    """
    if key_cols is None:
        key_cols = ["Q Name"]
    result = current_df
    for col in cols:
        if col not in reference_df.columns or col not in current_df.columns:
            continue
        ref_ph = (
            reference_df
            .filter(pl.col(col).str.contains(r'#[^#]+#'))
            .select(key_cols + [col])
            .rename({col: "__ref_vanilla__"})
        )
        if ref_ph.height == 0:
            continue
        result = (
            result
            .join(ref_ph, on=key_cols, how="left")
            .with_columns(
                pl.when(pl.col("__ref_vanilla__").is_not_null())
                .then(pl.col("__ref_vanilla__"))
                .otherwise(pl.col(col))
                .alias(col)
            )
            .drop("__ref_vanilla__")
        )
    return result


def apply_replacements(
    df   : pl.DataFrame,
    pairs: dict[str, str],
    cols : list[str],
) -> pl.DataFrame:
    """Apply #placeholder# -> actual_value replacements to specified columns (re-personalise after comparison)."""
    if not pairs:
        return df
    result = df
    for col in cols:
        if col not in result.columns:
            continue
        expr = pl.col(col)
        for key, val in sorted(pairs.items(), key=lambda kv: len(kv[0]), reverse=True):
            expr = expr.str.replace_all(key, val, literal=True)
        result = result.with_columns(expr.alias(col))
    return result


_CROP_SPECIALS_COMPARE = {
    "crop" : [("777", "No crop production"), ("888", "Don't know"), ("999", "Refused")],
    "crop2": [("666", "No other crop"),       ("888", "Don't know"), ("999", "Refused")],
    "crop3": [("666", "No other crop"),       ("888", "Don't know"), ("999", "Refused")],
}


def read_crop_choice_rows_for_compare(country_path: str, language: str) -> dict[str, list[tuple[str, str]]]:
    """Read current questionnaire Crop list and build rows for crop/crop2/crop3 replacement."""
    try:
        wb = openpyxl.load_workbook(country_path, data_only=True, read_only=True)
        if "Crop list" not in wb.sheetnames:
            wb.close()
            return {}
        rows = list(wb["Crop list"].iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        print(f"Warning: could not read Crop list sheet: {e}")
        return {}

    if len(rows) < 3:
        return {}

    headers  = [str(h).strip() if h is not None else "" for h in rows[2]]
    sel_idx  = next((i for i, h in enumerate(headers) if "Select top" in h), None)
    code_idx = next((i for i, h in enumerate(headers) if "Dataset code" in h), None)
    if language == "fr":
        lbl_idx = next((i for i, h in enumerate(headers) if "Label" in h and "FR" in h), None)
    else:
        lbl_idx = next((i for i, h in enumerate(headers) if "Label" in h and "EN" in h), None)

    if code_idx is None or lbl_idx is None:
        return {}

    selected, unselected = [], []
    for row in rows[3:]:
        code = row[code_idx] if code_idx < len(row) else None
        lbl  = row[lbl_idx]  if lbl_idx  < len(row) else None
        sel  = row[sel_idx]  if sel_idx is not None and sel_idx < len(row) else None
        if code is None or str(code).strip() in ("", "nan"):
            continue
        entry = (str(code).strip(), str(lbl or "").strip())
        (selected if sel is not None and str(sel).strip() else unselected).append(entry)

    selected.sort(key=lambda x: x[0])
    unselected.sort(key=lambda x: x[0])
    combined = selected + unselected
    return {name: combined + specials for name, specials in _CROP_SPECIALS_COMPARE.items()}


def fetch_admin_choice_rows_for_compare(iso3: str) -> dict[str, list[tuple]]:
    """Fetch admin1/admin2 rows from AGOL for comparison preprocessing."""
    try:
        import urllib.request as _urlreq
        import json as _json

        _BASE = ("https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services"
                 "/Administrative_Boundaries_Reference_(view_layer)/FeatureServer")

        def _get(layer, fields):
            url = (f"{_BASE}/{layer}/query"
                   f"?where=adm0_ISO3+%3D+%27{iso3}%27"
                   f"&outFields={fields}&returnGeometry=false&outSR=4326&f=json")
            with _urlreq.urlopen(url, timeout=30) as r:
                return _json.loads(r.read().decode())["features"]

        adm1 = sorted(
            [(f["attributes"]["adm1_pcode"], f["attributes"]["adm1_name"])
             for f in _get(1, "adm1_name,adm1_pcode")],
            key=lambda x: x[0],
        )
        adm2 = sorted(
            [(f["attributes"]["adm2_pcode"], f["attributes"]["adm2_name"], f["attributes"]["adm1_pcode"])
             for f in _get(0, "adm2_name,adm2_pcode,adm1_pcode")],
            key=lambda x: x[0],
        )
        print(f"AGOL compare rows loaded: admin1={len(adm1)} admin2={len(adm2)}")
        return {"admin1": adm1, "admin2": adm2}
    except Exception as e:
        print(f"Warning: AGOL fetch failed during compare preprocessing: {e}")
        return {}


def replace_choice_lists_for_compare(choices_df: pl.DataFrame, country_rows: dict[str, list[tuple]]) -> pl.DataFrame:
    """Replace selected list_name blocks in choices_df with provided rows."""
    if not country_rows:
        return choices_df

    skip = list(country_rows.keys())
    kept = choices_df.filter(~pl.col("list_name").is_in(skip)) if choices_df.height > 0 else choices_df

    recs = []
    for list_name, entries in country_rows.items():
        for entry in entries:
            recs.append({
                "list_name": str(list_name),
                "option_name": str(entry[0]),
                "option_label": str(entry[1] if len(entry) > 1 else ""),
            })

    new_df = (
        pl.DataFrame(recs)
        if recs else
        pl.DataFrame(schema={"list_name": pl.Utf8, "option_name": pl.Utf8, "option_label": pl.Utf8})
    )

    if kept.height == 0:
        return new_df
    if new_df.height == 0:
        return kept
    return pl.concat([kept, new_df], how="vertical")


# --- CODE CELL 7 ---
_VANILLA_COLS_SURVEY  = ["label", "constraint", "hint"]
_VANILLA_COLS_OPTIONS = ["option_label"]

# Read placeholder replacements from current questionnaire Additional information sheet
replacement_pairs = read_additional_info(run["questionnaire_path"], run["language"])
print(f"Loaded {len(replacement_pairs)} placeholder pair(s) from Additional information sheet:")
for k, v in replacement_pairs.items():
    print(f"  {k!r}  ->  {v!r}")

# Template is always the placeholder map for restoration in previous_round mode.
_vanilla_ref_survey  = template_survey
_vanilla_ref_options = template_options

# Restore: where template has #placeholder#, copy template value -> current
current_survey_vanilla  = restore_to_vanilla(
    current_survey,  _vanilla_ref_survey,  _VANILLA_COLS_SURVEY
)
current_options_vanilla = restore_to_vanilla(
    current_options, _vanilla_ref_options, _VANILLA_COLS_OPTIONS,
    key_cols=["Q Name", "option_name"],
)

_n_survey  = sum(
    _vanilla_ref_survey.filter(pl.col(c).str.contains(r'#[^#]+#')).height
    for c in _VANILLA_COLS_SURVEY if c in _vanilla_ref_survey.columns
)
_n_options = (
    _vanilla_ref_options.filter(pl.col("option_label").str.contains(r'#[^#]+#')).height
    if "option_label" in _vanilla_ref_options.columns else 0
)
print(f"\nVanilla restoration map (from template): {_n_survey} survey cell(s), {_n_options} option cell(s)")

# -----------------------------------------------------------------------------
# Comparison datasets
# - latest_template: current behaviour (vanilla compare)
# - previous_round : apply current round replacements BEFORE compare,
#                    including crop + AGOL list refresh for current only.
# -----------------------------------------------------------------------------
current_survey_cmp         = current_survey
current_survey_vanilla_cmp = current_survey_vanilla
current_options_vanilla_cmp = current_options_vanilla

_round_param_qnames = set()
_round_param_option_qnames = set()
_round_param_country_lists = set(cfg.get("kobo_options", {}).get("country_specific_list_names", []))

# Placeholder-affected question names from template map.
for _c in _VANILLA_COLS_SURVEY:
    if _c in _vanilla_ref_survey.columns:
        _round_param_qnames |= set(
            _vanilla_ref_survey
            .filter(pl.col(_c).str.contains(r'#[^#]+#'))["Q Name"]
            .to_list()
        )
if "option_label" in _vanilla_ref_options.columns:
    _round_param_option_qnames |= set(
        _vanilla_ref_options
        .filter(pl.col("option_label").str.contains(r'#[^#]+#'))["Q Name"]
        .to_list()
    )

if run.get("reference_mode") == "previous_round":
    print("\nMode previous_round: preprocessing current questionnaire before comparison")

    # 1) Apply Additional information replacements to restored current values.
    current_survey_cmp = apply_replacements(current_survey_vanilla, replacement_pairs, _VANILLA_COLS_SURVEY)

    # 2) Rebuild current choices for country-specific lists from CURRENT round inputs.
    _crop_rows  = read_crop_choice_rows_for_compare(run["questionnaire_path"], run["language"])
    _admin_rows = fetch_admin_choice_rows_for_compare(run["iso3"])
    _country_rows = {**_crop_rows, **_admin_rows}

    current_choices_cmp = replace_choice_lists_for_compare(current_choices, _country_rows)
    current_options_cmp = build_question_options(current_survey_cmp, current_choices_cmp)

    # Restore+replace option labels using template placeholder map.
    current_options_cmp = restore_to_vanilla(
        current_options_cmp, _vanilla_ref_options, _VANILLA_COLS_OPTIONS,
        key_cols=["Q Name", "option_name"],
    )
    current_options_cmp = apply_replacements(current_options_cmp, replacement_pairs, _VANILLA_COLS_OPTIONS)

    current_survey_vanilla_cmp = current_survey_cmp
    current_options_vanilla_cmp = current_options_cmp

    # Country-specific option questions (crop/admin etc) are round-parameter-change candidates.
    if _round_param_country_lists:
        _country_qs = set(
            current_survey_cmp
            .filter(pl.col("list_name").is_in(list(_round_param_country_lists)))["Q Name"]
            .to_list()
        )
        _round_param_option_qnames |= _country_qs

    print(f"Current compare options rebuilt: {current_options_vanilla_cmp.height} rows")
else:
    print("\nMode latest_template: template-style comparison (no pre-replacement before compare)")

_round_param_qnames = sorted(_round_param_qnames)
_round_param_option_qnames = sorted(_round_param_option_qnames)


# ======================================================================
# SECTION: ## Step 5 -- Normalisation helpers
_step("Normalisation helpers")
# ======================================================================
# --- CODE CELL 8 ---
def normalize_text_expr(col_name: str) -> pl.Expr:
    return (
        pl.col(col_name).cast(pl.Utf8).fill_null("")
        .str.to_lowercase()
        .str.replace_all(r"[^\w\s]", "")
        .str.replace_all(r"\s+", " ")
        .str.strip_chars()
    )


def normalize_simple_expr(col_name: str) -> pl.Expr:
    return (
        pl.col(col_name).cast(pl.Utf8).fill_null("")
        .str.to_lowercase()
        .str.replace_all(r"\s+", " ")
        .str.strip_chars()
    )


def normalize_logic_expr(col_name: str) -> pl.Expr:
    return (
        pl.col(col_name).cast(pl.Utf8).fill_null("")
        .str.to_lowercase()
        .str.replace_all(r"\s+", " ")
        .str.replace_all(r"\s*([=<>!+\-*/(),])\s*", "$1")
        .str.strip_chars()
    )


# ======================================================================
# SECTION: ## Step 6 -- Comparison functions
_step("Comparison functions")
# ======================================================================
# --- CODE CELL 9 ---
def compare_question_presence(current, reference):
    """Returns (added, removed). Each row has is_optional flag."""
    key = "Q Name"
    added = (
        current.select([key])
        .join(reference.select([key]), on=key, how="anti")
        .with_columns(pl.col(key).str.starts_with("o_").alias("is_optional"))
    )
    removed = (
        reference.select([key])
        .join(current.select([key]), on=key, how="anti")
        .with_columns(pl.col(key).str.starts_with("o_").alias("is_optional"))
    )
    return added, removed


def compare_mandatory_kobo(current, reference):
    """Compare mandatory_category between current and reference."""
    EMPTY = {
        "issue_type": pl.Utf8, "Q Name": pl.Utf8, "field": pl.Utf8,
        "Mandatory": pl.Utf8, "Mandatory_ref": pl.Utf8, "excel_row": pl.Int64,
    }
    if "mandatory_category" not in current.columns or "mandatory_category" not in reference.columns:
        return pl.DataFrame(schema=EMPTY)

    meaningful = {"mandatory", "mandatory-panel", "optional", "non-mandatory"}
    return (
        current.select(["Q Name", "mandatory_category", "excel_row"])
        .join(reference.select(["Q Name", "mandatory_category"]), on="Q Name", how="inner", suffix="_ref")
        .filter(
            pl.col("mandatory_category").is_in(list(meaningful)) |
            pl.col("mandatory_category_ref").is_in(list(meaningful))
        )
        .filter(pl.col("mandatory_category") != pl.col("mandatory_category_ref"))
        .with_columns([
            pl.lit("mandatory_column_mismatch").alias("issue_type"),
            pl.lit("mandatory_category").alias("field"),
        ])
        .rename({"mandatory_category": "Mandatory", "mandatory_category_ref": "Mandatory_ref"})
        .select(["issue_type", "Q Name", "field", "Mandatory", "Mandatory_ref", "excel_row"])
    )


def compare_list_name_changes(current, reference):
    """Flag questions where the choices list (list_name from type) changed."""
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    cur = current.filter(pl.col("list_name").is_not_null()).select(["Q Name", "list_name", "excel_row"])
    ref = reference.filter(pl.col("list_name").is_not_null()).select(["Q Name", "list_name"])
    result = (
        cur.join(ref, on="Q Name", how="inner", suffix="_ref")
        .filter(pl.col("list_name") != pl.col("list_name_ref"))
        .with_columns([
            pl.lit("choices_list_changed").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("type (list_name)").alias("field"),
            pl.col("list_name").alias("current"),
            pl.col("list_name_ref").alias("reference"),
            pl.lit("medium").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )
    return result if result.height > 0 else pl.DataFrame(schema=EMPTY)


def compare_question_labels(current_vanilla, reference, current_orig=None):
    """
    Compare question label text.
    current_vanilla: current survey with #placeholder# tokens restored (no false positives).
    current_orig   : original current survey -- used so the report shows the actual filled-in label.
    """
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if "label" not in current_vanilla.columns or "label" not in reference.columns:
        return pl.DataFrame(schema=EMPTY)

    diff = (
        current_vanilla.select(["Q Name", "label", "excel_row"])
        .join(reference.select(["Q Name", "label"]), on="Q Name", how="inner", suffix="_ref")
        .with_columns([
            normalize_text_expr("label").alias("_norm"),
            normalize_text_expr("label_ref").alias("_norm_ref"),
        ])
        .filter(pl.col("_norm") != pl.col("_norm_ref"))
        .filter(pl.col("label_ref") != "")
        .drop(["_norm", "_norm_ref"])
    )
    if diff.height == 0:
        return pl.DataFrame(schema=EMPTY)

    # Replace vanilla label with the actual filled-in label for display in the report
    if current_orig is not None and "label" in current_orig.columns:
        orig_labels = current_orig.select(["Q Name", "label"]).rename({"label": "_actual"})
        diff = (
            diff.join(orig_labels, on="Q Name", how="left")
            .with_columns(
                pl.when(pl.col("_actual").is_not_null())
                .then(pl.col("_actual"))
                .otherwise(pl.col("label"))
                .alias("label")
            )
            .drop("_actual")
        )

    return (
        diff
        .with_columns([
            pl.lit("label_mismatch").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("label").alias("field"),
            pl.col("label").alias("current"),
            pl.col("label_ref").alias("reference"),
            pl.lit("medium").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )


def compare_constraint_changes(current_vanilla, reference, current_orig=None):
    """
    Compare constraint expressions.
    Uses vanilla current so #placeholder#-containing constraints are not flagged.
    """
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if "constraint" not in current_vanilla.columns or "constraint" not in reference.columns:
        return pl.DataFrame(schema=EMPTY)

    diff = (
        current_vanilla
        .select(["Q Name", "constraint", "excel_row"])
        .join(
            reference.select(["Q Name", "constraint"]),
            on="Q Name", how="inner", suffix="_ref",
        )
        .with_columns([
            normalize_logic_expr("constraint").alias("_norm"),
            normalize_logic_expr("constraint_ref").alias("_norm_ref"),
        ])
        .filter((pl.col("_norm") != pl.col("_norm_ref")) & ((pl.col("_norm") != "") | (pl.col("_norm_ref") != "")))
        .drop(["_norm", "_norm_ref"])
    )
    if diff.height == 0:
        return pl.DataFrame(schema=EMPTY)

    if current_orig is not None and "constraint" in current_orig.columns:
        orig_c = current_orig.select(["Q Name", "constraint"]).rename({"constraint": "_actual"})
        diff = (
            diff.join(orig_c, on="Q Name", how="left")
            .with_columns(
                pl.when(pl.col("_actual").is_not_null())
                .then(pl.col("_actual"))
                .otherwise(pl.col("constraint"))
                .alias("constraint")
            )
            .drop("_actual")
        )

    return (
        diff
        .with_columns([
            pl.lit("constraint_modified").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("constraint").alias("field"),
            pl.col("constraint").alias("current"),
            pl.col("constraint_ref").alias("reference"),
            pl.lit("medium").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )


def compare_required_changes(current, reference):
    """Compare required column with whitespace/case-tolerant normalization."""
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if "required" not in current.columns or "required" not in reference.columns:
        return pl.DataFrame(schema=EMPTY)

    diff = (
        current.select(["Q Name", "required", "excel_row"])
        .join(reference.select(["Q Name", "required"]), on="Q Name", how="inner", suffix="_ref")
        .with_columns([
            normalize_simple_expr("required").alias("_norm"),
            normalize_simple_expr("required_ref").alias("_norm_ref"),
        ])
        .filter((pl.col("_norm") != pl.col("_norm_ref")) & ((pl.col("_norm") != "") | (pl.col("_norm_ref") != "")))
        .drop(["_norm", "_norm_ref"])
    )
    if diff.height == 0:
        return pl.DataFrame(schema=EMPTY)

    return (
        diff
        .with_columns([
            pl.lit("required_modified").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("required").alias("field"),
            pl.col("required").alias("current"),
            pl.col("required_ref").alias("reference"),
            pl.lit("medium").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )


def compare_choice_filter_changes(current, reference):
    """Compare choice_filter with expression normalization to avoid cosmetic false positives."""
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if "choice_filter" not in current.columns or "choice_filter" not in reference.columns:
        return pl.DataFrame(schema=EMPTY)

    diff = (
        current.select(["Q Name", "choice_filter", "excel_row"])
        .join(reference.select(["Q Name", "choice_filter"]), on="Q Name", how="inner", suffix="_ref")
        .with_columns([
            normalize_logic_expr("choice_filter").alias("_norm"),
            normalize_logic_expr("choice_filter_ref").alias("_norm_ref"),
        ])
        .filter((pl.col("_norm") != pl.col("_norm_ref")) & ((pl.col("_norm") != "") | (pl.col("_norm_ref") != "")))
        .drop(["_norm", "_norm_ref"])
    )
    if diff.height == 0:
        return pl.DataFrame(schema=EMPTY)

    return (
        diff
        .with_columns([
            pl.lit("choice_filter_modified").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("choice_filter").alias("field"),
            pl.col("choice_filter").alias("current"),
            pl.col("choice_filter_ref").alias("reference"),
            pl.lit("medium").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )


def compare_appearance_changes(current, reference):
    """Compare appearance with expression normalization that ignores tiny spacing differences."""
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if "appearance" not in current.columns or "appearance" not in reference.columns:
        return pl.DataFrame(schema=EMPTY)

    diff = (
        current.select(["Q Name", "appearance", "excel_row"])
        .join(reference.select(["Q Name", "appearance"]), on="Q Name", how="inner", suffix="_ref")
        .with_columns([
            normalize_logic_expr("appearance").alias("_norm"),
            normalize_logic_expr("appearance_ref").alias("_norm_ref"),
        ])
        .filter((pl.col("_norm") != pl.col("_norm_ref")) & ((pl.col("_norm") != "") | (pl.col("_norm_ref") != "")))
        .drop(["_norm", "_norm_ref"])
    )
    if diff.height == 0:
        return pl.DataFrame(schema=EMPTY)

    return (
        diff
        .with_columns([
            pl.lit("appearance_modified").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("appearance").alias("field"),
            pl.col("appearance").alias("current"),
            pl.col("appearance_ref").alias("reference"),
            pl.lit("medium").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )


def compare_calculation_changes(current, reference):
    """Compare calculation with expression normalization that ignores tiny spacing differences."""
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if "calculation" not in current.columns or "calculation" not in reference.columns:
        return pl.DataFrame(schema=EMPTY)

    diff = (
        current.select(["Q Name", "calculation", "excel_row"])
        .join(reference.select(["Q Name", "calculation"]), on="Q Name", how="inner", suffix="_ref")
        .with_columns([
            normalize_logic_expr("calculation").alias("_norm"),
            normalize_logic_expr("calculation_ref").alias("_norm_ref"),
        ])
        .filter((pl.col("_norm") != pl.col("_norm_ref")) & ((pl.col("_norm") != "") | (pl.col("_norm_ref") != "")))
        .drop(["_norm", "_norm_ref"])
    )
    if diff.height == 0:
        return pl.DataFrame(schema=EMPTY)

    return (
        diff
        .with_columns([
            pl.lit("calculation_modified").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("calculation").alias("field"),
            pl.col("calculation").alias("current"),
            pl.col("calculation_ref").alias("reference"),
            pl.lit("medium").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )



def _normalize_qtype_value_kobo(value: str) -> str:
    """Canonicalize KoBo Q Type text so cosmetic formatting does not trigger false positives."""
    s = str(value or "").strip().lower()
    if not s:
        return ""
    s = re.sub(r"[\\/_-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    if s in {"select one", "single choice"}:
        return "single choice"
    if s in {"select multiple", "select all that apply"}:
        return "select all that apply"
    return s


def _qtype_mode_kobo(value: str) -> str:
    """Classify Q Type into compatibility families."""
    t = _normalize_qtype_value_kobo(value)
    if not t:
        return ""
    if t == "select all that apply":
        return "multi_select"
    if t == "single choice":
        return "single_select"
    if "select one" in t:
        return "single_select"
    if "select multiple" in t:
        return "multi_select"
    return "non_option"


_KOBO_EXTRA_KNOWN_QTYPES = {
    "note", "rank", "email", "acknowledge", "hidden",
    "begin group", "end group", "begin repeat", "end repeat",
    "begin kobo matrix", "end kobo matrix",
}


_KOBO_KNOWN_QTYPES_NORMALIZED = {
    _normalize_qtype_value_kobo(v)
    for v in (
        set(_SELECT_TYPES)
        | set(_DATA_TYPES)
        | set(_METADATA_TYPES)
        | _KOBO_EXTRA_KNOWN_QTYPES
        | {"single choice", "select all that apply", "select one", "select multiple"}
    )
}
_KOBO_KNOWN_QTYPES_NORMALIZED.discard("")


def _is_known_qtype_value_kobo(value: str) -> bool:
    t = _normalize_qtype_value_kobo(value)
    if not t:
        return False
    if t in _KOBO_KNOWN_QTYPES_NORMALIZED:
        return True
    if t.startswith("select one") or t.startswith("select multiple"):
        return True
    return False


def _qtype_change_severity_kobo(
    curr_mode: str,
    ref_mode: str,
    curr_option_count: int,
    ref_option_count: int,
    curr_type_norm: str,
    ref_type_norm: str,
    curr_type_known: bool,
    ref_type_known: bool,
) -> str:
    """
    Severity policy (independent of mandatory status):
    - HIGH for incompatible/invalid transitions
    - MEDIUM for compatible type changes
    """
    cm = str(curr_mode or "")
    rm = str(ref_mode or "")
    modes = {cm, rm}
    curr_t = str(curr_type_norm or "").strip()
    ref_t = str(ref_type_norm or "").strip()

    # Invalid: current is blank while reference has a defined type.
    if (not curr_t) and ref_t:
        return "high"

    # Invalid: unknown/unlisted Q Type token on either side.
    if curr_t and (not bool(curr_type_known)):
        return "high"
    if ref_t and (not bool(ref_type_known)):
        return "high"

    # Incompatible family change: single <-> multi.
    if "single_select" in modes and "multi_select" in modes:
        return "high"

    # Incompatible option-bearing <-> non-option mode.
    if "non_option" in modes and (("single_select" in modes) or ("multi_select" in modes)):
        return "high"

    # Invalid option-count correctness checks:
    # option-bearing questions should normally have at least 2 options.
    curr_option_like = cm in {"single_select", "multi_select"}
    ref_option_like = rm in {"single_select", "multi_select"}
    if curr_option_like and int(curr_option_count or 0) <= 1:
        return "high"
    if ref_option_like and int(ref_option_count or 0) <= 1:
        return "high"
    if (cm == "non_option") and int(curr_option_count or 0) > 0:
        return "high"
    if (rm == "non_option") and int(ref_option_count or 0) > 0:
        return "high"

    return "medium"


def compare_type_changes(
    current,
    reference,
    current_options: pl.DataFrame | None = None,
    reference_options: pl.DataFrame | None = None,
):
    """Risk-based Q Type comparison with normalization and integrity checks."""
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if "Q Type" not in current.columns or "Q Type" not in reference.columns:
        return pl.DataFrame(schema=EMPTY)

    if current_options is not None and current_options.height > 0 and "Q Name" in current_options.columns:
        _cur_opt_counts = current_options.group_by("Q Name").agg(pl.len().alias("_curr_opt_n"))
    else:
        _cur_opt_counts = pl.DataFrame(schema={"Q Name": pl.Utf8, "_curr_opt_n": pl.Int64})

    if reference_options is not None and reference_options.height > 0 and "Q Name" in reference_options.columns:
        _ref_opt_counts = reference_options.group_by("Q Name").agg(pl.len().alias("_ref_opt_n"))
    else:
        _ref_opt_counts = pl.DataFrame(schema={"Q Name": pl.Utf8, "_ref_opt_n": pl.Int64})

    result = (
        current.select(["Q Name", "Q Type", "excel_row"])
        .join(reference.select(["Q Name", "Q Type"]), on="Q Name", how="inner", suffix="_ref")
        .with_columns([
            pl.col("Q Type").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Q Type"),
            pl.col("Q Type_ref").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Q Type_ref"),
        ])
        .with_columns([
            pl.col("Q Type").map_elements(_normalize_qtype_value_kobo, return_dtype=pl.Utf8).alias("_type_norm"),
            pl.col("Q Type_ref").map_elements(_normalize_qtype_value_kobo, return_dtype=pl.Utf8).alias("_type_norm_ref"),
            pl.col("Q Type").map_elements(_qtype_mode_kobo, return_dtype=pl.Utf8).alias("_type_mode"),
            pl.col("Q Type_ref").map_elements(_qtype_mode_kobo, return_dtype=pl.Utf8).alias("_type_mode_ref"),
            pl.col("Q Type").map_elements(_is_known_qtype_value_kobo, return_dtype=pl.Boolean).alias("_type_known"),
            pl.col("Q Type_ref").map_elements(_is_known_qtype_value_kobo, return_dtype=pl.Boolean).alias("_type_known_ref"),
        ])
        .join(_cur_opt_counts, on="Q Name", how="left")
        .join(_ref_opt_counts, on="Q Name", how="left")
        .with_columns([
            pl.col("_curr_opt_n").fill_null(0).cast(pl.Int64),
            pl.col("_ref_opt_n").fill_null(0).cast(pl.Int64),
        ])
        .filter(pl.col("_type_norm") != pl.col("_type_norm_ref"))
        .with_columns(
            pl.struct([
                "_type_mode", "_type_mode_ref", "_curr_opt_n", "_ref_opt_n",
                "_type_norm", "_type_norm_ref", "_type_known", "_type_known_ref",
            ])
            .map_elements(
                lambda r: _qtype_change_severity_kobo(
                    r.get("_type_mode"),
                    r.get("_type_mode_ref"),
                    r.get("_curr_opt_n"),
                    r.get("_ref_opt_n"),
                    r.get("_type_norm"),
                    r.get("_type_norm_ref"),
                    bool(r.get("_type_known")),
                    bool(r.get("_type_known_ref")),
                ),
                return_dtype=pl.Utf8,
            )
            .alias("severity")
        )
        .with_columns([
            pl.lit("type_changed").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("type").alias("field"),
            pl.col("Q Type").alias("current"),
            pl.col("Q Type_ref").alias("reference"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )
    return result if result.height > 0 else pl.DataFrame(schema=EMPTY)

def _canonical_option_key_expr(col_name: str) -> pl.Expr:
    """Canonical option key: strip spaces and treat integer-like decimals as same key (1 == 1.0)."""
    return (
        pl.col(col_name)
        .cast(pl.Utf8)
        .str.strip_chars()
        .str.replace(r"^([+-]?\d+)\.0+$", "$1")
    )


def compare_option_labels_single(current_opts, reference_opts, lang_scope="EN"):
    """Option label text changes for matching canonical Q Name + option_name."""
    cur = current_opts.with_columns([
        pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.to_lowercase().alias("_qk"),
        _canonical_option_key_expr("option_name").alias("_ok"),
    ]).filter(
        pl.col("_qk").is_not_null() & (pl.col("_qk") != "") &
        pl.col("_ok").is_not_null() & (pl.col("_ok") != "")
    )
    ref = reference_opts.with_columns([
        pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.to_lowercase().alias("_qk"),
        _canonical_option_key_expr("option_name").alias("_ok"),
    ]).filter(
        pl.col("_qk").is_not_null() & (pl.col("_qk") != "") &
        pl.col("_ok").is_not_null() & (pl.col("_ok") != "")
    )
    return (
        cur
        .join(ref, on=["_qk", "_ok"], how="inner", suffix="_ref")
        .with_columns([
            normalize_text_expr("option_label").alias("_norm"),
            normalize_text_expr("option_label_ref").alias("_norm_ref"),
        ])
        .filter(pl.col("_norm") != pl.col("_norm_ref"))
        .drop(["_norm", "_norm_ref", "_qk", "_ok"])
        .with_columns([
            pl.lit("choice_label_mismatch").alias("issue_type"),
            pl.concat_str([pl.lit("option_"), pl.col("option_name")]).alias("field"),
            pl.lit("medium").alias("severity"),
            pl.lit(lang_scope).alias("lang_scope"),
            pl.lit(None).cast(pl.Int64).alias("excel_row"),
        ])
        .select(["issue_type", "Q Name", "list_name", "option_name", "field", "option_label", "option_label_ref",
                 "severity", "excel_row", "lang_scope"])
    )


def compare_option_presence_single(current_opts, reference_opts, lang_scope="EN"):
    """Added / removed options for matching canonical Q Name + option_name."""
    cur = current_opts.with_columns([
        pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.to_lowercase().alias("_qk"),
        _canonical_option_key_expr("option_name").alias("_ok"),
    ]).filter(
        pl.col("_qk").is_not_null() & (pl.col("_qk") != "") &
        pl.col("_ok").is_not_null() & (pl.col("_ok") != "")
    )
    ref = reference_opts.with_columns([
        pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.to_lowercase().alias("_qk"),
        _canonical_option_key_expr("option_name").alias("_ok"),
    ]).filter(
        pl.col("_qk").is_not_null() & (pl.col("_qk") != "") &
        pl.col("_ok").is_not_null() & (pl.col("_ok") != "")
    )
    key_cols = ["_qk", "_ok"]
    removed = (
        ref.select(key_cols + ["Q Name", "list_name", "option_name", "option_label"])
        .join(cur.select(key_cols), on=key_cols, how="anti")
        .drop(["_qk", "_ok"])
        .with_columns([
            pl.lit("removed_choice").alias("issue_type"),
            pl.concat_str([pl.lit("option_"), pl.col("option_name")]).alias("field"),
            pl.lit("high").alias("severity"),
            pl.lit(lang_scope).alias("lang_scope"),
        ])
        .select(["issue_type", "Q Name", "list_name", "option_name", "field", "option_label", "severity", "lang_scope"])
    )
    added = (
        cur.select(key_cols + ["Q Name", "list_name", "option_name", "option_label"])
        .join(ref.select(key_cols), on=key_cols, how="anti")
        .drop(["_qk", "_ok"])
        .with_columns([
            pl.lit("added_choice").alias("issue_type"),
            pl.concat_str([pl.lit("option_"), pl.col("option_name")]).alias("field"),
            pl.lit("medium").alias("severity"),
            pl.lit(lang_scope).alias("lang_scope"),
        ])
        .select(["issue_type", "Q Name", "list_name", "option_name", "field", "option_label", "severity", "lang_scope"])
    )
    return added, removed

def compare_hint_changes(current: pl.DataFrame, reference: pl.DataFrame) -> pl.DataFrame:
    """Flag questions where the hint text differs from the template (medium severity)."""
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    ref_hints = reference.filter(pl.col("hint") != "").select(["Q Name", "hint"])
    if ref_hints.is_empty():
        return pl.DataFrame(schema=EMPTY)
    joined = (
        current.select(["Q Name", "hint", "excel_row"])
        .join(ref_hints, on="Q Name", how="inner", suffix="_ref")
        .filter(pl.col("hint") != pl.col("hint_ref"))
    )
    if joined.is_empty():
        return pl.DataFrame(schema=EMPTY)
    return (
        joined.with_columns([
            pl.lit("hint_changed").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.col("hint").alias("current"),
            pl.col("hint_ref").alias("reference"),
            pl.lit("medium").alias("severity"),
            pl.lit("hint").alias("field"),
        ])
        .select(list(EMPTY.keys()))
    )





# ======================================================================
# SECTION: ## Step 7 -- Relevant validation (KoBO skip logic)
_step("Relevant (skip logic) validation")
# ======================================================================
# --- CODE CELL 10 ---
_KOBO_REF_RE = re.compile(r"\$\{([^}]+)\}")
_HASH_TOKEN_RE = re.compile(r"#([^#]+)#")
_LOOSE_DOLLAR_RE = re.compile(r"(?<!\{)\$[A-Za-z_][A-Za-z0-9_]*")


def validate_relevant(
    current_survey : pl.DataFrame,
    reference_survey: pl.DataFrame,
) -> pl.DataFrame:
    """
    Layer 1a - broken_relevant_reference (high): ${var} points to a name not in the survey at all
    Layer 1b - relevant_inexact_reference (high): exact name missing but o_<name> exists in survey
                (reference is broken in KoBO; note helps debugging)
    Layer 2  - relevant_modified (medium): expression changed from template
    """
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }

    curr_qnames = set(current_survey["Q Name"].to_list())

    # Build lookup: stripped core name -> o_name for all optional-prefixed questions
    _optional_by_core: dict[str, str] = {
        qn[2:]: qn for qn in curr_qnames if qn.startswith("o_")
    }

    cur_rows = {r["Q Name"]: r for r in
                current_survey.select(["Q Name", "relevant", "excel_row"]).iter_rows(named=True)}
    ref_rows = {r["Q Name"]: r for r in
                reference_survey.select(["Q Name", "relevant"]).iter_rows(named=True)}

    issues = []
    for qname, cr in cur_rows.items():
        relevant  = str(cr.get("relevant") or "").strip()
        excel_row = cr.get("excel_row")

        refs = _KOBO_REF_RE.findall(relevant) if relevant else []
        broken = [v for v in refs if v not in curr_qnames]

        has_high_relevant_issue = False
        if broken:
            truly_missing = [v for v in broken if v not in _optional_by_core]
            has_similar   = [v for v in broken if v in _optional_by_core]

            if truly_missing:
                has_high_relevant_issue = True
                issues.append({
                    "issue_type": "broken_relevant_reference",
                    "set_name": "", "Q Name": qname, "field": "relevant",
                    "current"  : relevant[:220],
                    "reference": f"missing variables: {truly_missing}",
                    "severity" : "high", "excel_row": excel_row,
                })

            for v in has_similar:
                has_high_relevant_issue = True
                similar = _optional_by_core[v]
                issues.append({
                    "issue_type": "relevant_inexact_reference",
                    "set_name": "", "Q Name": qname, "field": "relevant",
                    "current"  : relevant[:220],
                    "reference": f"${{{v}}} not found; {similar!r} exists with similar name",
                    "severity" : "high", "excel_row": excel_row,
                })

        # Layer 2: changed from template (only when no high relevant integrity issue).
        # Apply only when question exists in BOTH files.
        # (Added questions should be checked structurally, not compared as template drift.)
        ref_relevant = str(ref_rows.get(qname, {}).get("relevant") or "").strip()
        if (not has_high_relevant_issue) and (qname in ref_rows) and (ref_relevant != relevant):
            issues.append({
                "issue_type": "relevant_modified",
                "set_name": "", "Q Name": qname, "field": "relevant",
                "current"  : relevant[:220],
                "reference": ref_relevant[:220],
                "severity" : "medium", "excel_row": excel_row,
            })

    if not issues:
        return pl.DataFrame(schema=EMPTY)
    return pl.DataFrame(issues).unique(
        subset=["issue_type", "Q Name", "field", "current", "reference"], keep="first"
    )


def validate_questionnaire_structure(
    current_survey: pl.DataFrame,
    current_choices: pl.DataFrame | None = None,
    template_survey: pl.DataFrame | None = None,
    replacement_pairs: dict[str, str] | None = None,
) -> pl.DataFrame:
    """
    Structural checks for token/placeholder integrity in survey text/formula fields.
    - duplicate Q Name in survey
    - duplicate (list_name, name) in choices/options
    - #token# unresolved against template placeholders + Additional information
    - #token# that looks like a survey variable (likely ${var} expected)
    - loose $var syntax (likely ${var} expected)
    - ${var} missing from survey (outside relevant column checks)
    """
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    _fields = [c for c in ["label", "hint", "constraint", "calculation"] if c in current_survey.columns]
    has_text_fields = bool(_fields)

    curr_qnames = set(current_survey["Q Name"].to_list())
    issues = []

    # 0) Structural duplicate checks
    if current_survey.height > 0 and "Q Name" in current_survey.columns:
        dup_q = (
            current_survey
            .select(["Q Name", "excel_row"])
            .with_columns(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().alias("_q"))
            .filter(pl.col("_q") != "")
            .group_by("_q")
            .agg([
                pl.len().alias("_n"),
                pl.col("excel_row").min().alias("_row"),
            ])
            .filter(pl.col("_n") > 1)
        )
        for rr in dup_q.iter_rows(named=True):
            issues.append({
                "issue_type": "duplicate_qname",
                "set_name": "questionnaire_structure",
                "Q Name": rr["_q"],
                "field": "Q Name",
                "current": rr["_q"],
                "reference": f"Duplicate question name appears {rr['_n']} times",
                "severity": "high",
                "excel_row": rr["_row"],
            })

    if (
        current_choices is not None
        and current_choices.height > 0
        and {"list_name", "option_name"}.issubset(set(current_choices.columns))
    ):
        dup_opts = (
            current_choices
            .select(["list_name", "option_name"])
            .with_columns([
                pl.col("list_name").cast(pl.Utf8).str.strip_chars().alias("_list"),
                pl.col("option_name").cast(pl.Utf8).str.strip_chars().alias("_name"),
            ])
            .filter((pl.col("_list") != "") & (pl.col("_name") != ""))
            .group_by(["_list", "_name"])
            .agg([
                pl.len().alias("_n"),
            ])
            .filter(pl.col("_n") > 1)
        )
        for rr in dup_opts.iter_rows(named=True):
            issues.append({
                "issue_type": "duplicate_choice_name",
                "set_name": "questionnaire_structure",
                "Q Name": "",
                "field": "choices",
                "current": f"{rr['_list']} / {rr['_name']}",
                "reference": f"Duplicate (list_name, name) appears {rr['_n']} times",
                "severity": "high",
                "excel_row": None,
            })

    if not has_text_fields:
        return pl.DataFrame(issues) if issues else pl.DataFrame(schema=EMPTY)

    # Tokens considered resolvable: known in template OR in Additional information pairs.
    known_tokens = set()
    if template_survey is not None and template_survey.height > 0:
        tpl_fields = [c for c in _fields if c in template_survey.columns]
        if tpl_fields:
            for row in template_survey.select(tpl_fields).iter_rows(named=True):
                for f in tpl_fields:
                    txt = str(row.get(f) or "")
                    for tok in _HASH_TOKEN_RE.findall(txt):
                        known_tokens.add(f"#{tok}#")
    if replacement_pairs:
        known_tokens |= set(replacement_pairs.keys())
    known_tokens_l = {t.lower() for t in known_tokens}

    iter_cols = ["Q Name", "excel_row"] + _fields
    for row in current_survey.select(iter_cols).iter_rows(named=True):
        qname = row.get("Q Name", "")
        excel_row = row.get("excel_row")
        for field in _fields:
            txt = str(row.get(field) or "").strip()
            if not txt:
                continue

            # 1) #token# checks
            for tok in _HASH_TOKEN_RE.findall(txt):
                full = f"#{tok}#"
                tok_core = str(tok).strip()

                if tok_core in curr_qnames:
                    issues.append({
                        "issue_type": "placeholder_should_use_kobo_ref",
                        "set_name": "questionnaire_structure",
                        "Q Name": qname,
                        "field": field,
                        "current": txt[:220],
                        "reference": f"{full} matches survey variable '{tok_core}' - use ${{{tok_core}}}",
                        "severity": "high",
                        "excel_row": excel_row,
                    })
                elif full.lower() not in known_tokens_l:
                    issues.append({
                        "issue_type": "placeholder_not_found",
                        "set_name": "questionnaire_structure",
                        "Q Name": qname,
                        "field": field,
                        "current": txt[:220],
                        "reference": f"{full} not in template placeholders or Additional information",
                        "severity": "high",
                        "excel_row": excel_row,
                    })

            # 2) $var (missing braces) checks
            for m in _LOOSE_DOLLAR_RE.findall(txt):
                var = m[1:]
                if var in curr_qnames:
                    ref = f"{m} uses loose syntax; use ${{{var}}}"
                    sev = "high"
                else:
                    ref = f"{m} uses loose syntax; expected ${{var}} format"
                    sev = "medium"
                issues.append({
                    "issue_type": "kobo_ref_loose_syntax",
                    "set_name": "questionnaire_structure",
                    "Q Name": qname,
                    "field": field,
                    "current": txt[:220],
                    "reference": ref,
                    "severity": sev,
                    "excel_row": excel_row,
                })

            # 3) ${var} checks (outside relevant field)
            for v in _KOBO_REF_RE.findall(txt):
                if v not in curr_qnames:
                    issues.append({
                        "issue_type": "kobo_ref_missing_variable",
                        "set_name": "questionnaire_structure",
                        "Q Name": qname,
                        "field": field,
                        "current": txt[:220],
                        "reference": f"${{{v}}} not found in survey Q Name",
                        "severity": "high",
                        "excel_row": excel_row,
                    })

    if not issues:
        return pl.DataFrame(schema=EMPTY)
    return pl.DataFrame(issues).unique(
        subset=["issue_type", "Q Name", "field", "current", "reference"], keep="first"
    )



# ======================================================================
# SECTION: ## Step 8 -- Critical sets
_step("Critical sets")
# ======================================================================
# --- CODE CELL 11 ---
import yaml
from pathlib import Path as _Path

_CRIT_YAML = _Path(
    "c:/Users/edoar/WORK/FAO/repo/questionnaire_validation_revision/scripts/critical_sets.yaml"
)
rules = {"exact_sets": {}, "min_count_sets": {}, "crop_harvest": {}}
if _CRIT_YAML.exists():
    with open(_CRIT_YAML, encoding="utf-8") as _f:
        rules = yaml.safe_load(_f) or rules
    print(f"Loaded critical_sets.yaml  exact_sets: {len(rules.get('exact_sets', {}))}")
else:
    print("No critical_sets.yaml found -- critical-set checks will be skipped.")


def validate_critical_sets(questions_df, exact_sets):
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if not exact_sets:
        return pl.DataFrame(schema=EMPTY)

    # KoBO uses mandatory_category (already normalised); add Mandatory alias for compat
    if "Mandatory" not in questions_df.columns and "mandatory_category" in questions_df.columns:
        questions_df = questions_df.with_columns(
            pl.col("mandatory_category").alias("Mandatory")
        )
    # Mandatory_norm: map to yes/no for comparison with yaml expected values
    questions_df = questions_df.with_columns(
        pl.when(pl.col("Mandatory").is_in(["mandatory", "yes"])).then(pl.lit("yes"))
        .when(pl.col("Mandatory").is_in(["mandatory-panel"])).then(pl.lit("yes"))
        .when(pl.col("Mandatory").is_in(["optional", "no"])).then(pl.lit("no"))
        .otherwise(pl.lit("")).alias("Mandatory_norm")
    )

    present_qnames = set(questions_df["Q Name"].to_list())

    def _wealth_optional_variant(q_name: str) -> str:
        q = str(q_name or "").strip()
        if q.startswith("hh_wealth_"):
            return "o_hh_wealth_" + q[len("hh_wealth_"):]
        return ""

    issues = []
    for set_name, set_rules in exact_sets.items():
        for rule in set_rules:
            q_name   = rule["q_name"]
            expected = rule.get("expected_mandatory", "")
            required = rule.get("required", True)
            alt_q_name = _wealth_optional_variant(q_name)
            matched_q_name = q_name if q_name in present_qnames else (alt_q_name if alt_q_name in present_qnames else "")
            if not matched_q_name:
                issues.append({
                    "issue_type": "missing_critical_question" if required else "advisory_question",
                    "set_name": set_name, "Q Name": q_name,
                    "field": "Q Name", "current": "",
                    "reference": (f"present (or optional variant: {alt_q_name})" if alt_q_name else "present"),
                    "severity": "high" if required else "medium", "excel_row": None,
                })
                continue
            if not expected:
                continue
            # If a wealth optional variant is accepted as substitute, treat it as pass
            # and avoid forcing mandatory parity on the alternate form.
            if matched_q_name != q_name:
                continue
            row = questions_df.filter(pl.col("Q Name") == matched_q_name).select(
                ["Q Name", "Mandatory", "Mandatory_norm", "excel_row"]
            ).to_dicts()
            if not row:
                continue
            row = row[0]
            if row["Mandatory_norm"] != expected:
                issues.append({
                    "issue_type": "critical_mandatory_mismatch", "set_name": set_name,
                    "Q Name": q_name, "field": "Mandatory",
                    "current": row["Mandatory"], "reference": expected,
                    "severity": "high", "excel_row": row.get("excel_row"),
                })

    if not issues:
        return pl.DataFrame(schema=EMPTY)
    return pl.DataFrame(issues)


def validate_prefix_counts(questions_df, min_count_sets):
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if not min_count_sets:
        return pl.DataFrame(schema=EMPTY)
    all_qnames = questions_df["Q Name"].to_list()
    issues = []
    for set_name, rule in min_count_sets.items():
        prefix    = rule.get("prefix", "")
        min_count = rule.get("min_count", 1)
        desc      = rule.get("description", f"At least {min_count} '{prefix}*' questions required")
        prefix_alt = "o_hh_wealth_" if str(prefix) == "hh_wealth_" or str(set_name).upper() == "WEALTH" else ""
        matched = sorted(
            q for q in all_qnames
            if (prefix and q.startswith(prefix)) or (prefix_alt and q.startswith(prefix_alt))
        )
        if len(matched) < min_count:
            found_str = f"{len(matched)} found" + (f": {', '.join(matched)}" if matched else " (none)")
            issues.append({
                "issue_type": "missing_critical_question", "set_name": set_name, "Q Name": "",
                "field": "count", "current": found_str, "reference": desc,
                "severity": "high", "excel_row": None,
            })
    if not issues:
        return pl.DataFrame(schema=EMPTY)
    return pl.DataFrame(issues)


def validate_crop_harvest(survey: pl.DataFrame, crop_harvest_cfg: dict) -> pl.DataFrame:
    """
    PASS if only the minimal set is present (no extra full-set questions), OR
    if the full set is completely present (extra questions beyond full are fine).
    FAIL for any partial overlap.
    Returns a DataFrame with ISSUE_SCHEMA columns.
    """
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if not crop_harvest_cfg:
        return pl.DataFrame(schema=EMPTY)

    minimal = set(crop_harvest_cfg.get("minimal", []))
    full    = set(crop_harvest_cfg.get("full", []))
    present = set(survey["Q Name"].to_list())

    if full.issubset(present):
        return pl.DataFrame(schema=EMPTY)

    extra_full = (full - minimal) & present
    if (minimal & present) == minimal and not extra_full:
        return pl.DataFrame(schema=EMPTY)

    missing_from_full = full - present
    partial_found     = (full - minimal) & present
    detail_parts = []
    if missing_from_full:
        detail_parts.append(f"missing from full set: {sorted(missing_from_full)}")
    if partial_found:
        detail_parts.append(f"partial full-set Qs present: {sorted(partial_found)}")
    detail = "; ".join(detail_parts) or "partial crop-harvest set"
    ref    = f"minimal={sorted(minimal)}, full={sorted(full)}"

    return pl.DataFrame([{
        "issue_type": "crop_harvest_violation",
        "set_name":   "CRP_HARV",
        "Q Name":     "crp_harv_*",
        "field":      "",
        "current":    detail,
        "reference":  ref,
        "severity":   "high",
        "excel_row":  -1,
    }], schema=EMPTY)

# ======================================================================
# SECTION: ## Step 9 -- Issue unifiers
_step("Issue unifiers")
# ======================================================================
# --- CODE CELL 12 ---
def make_presence_issues(added, removed, reference_survey=None):
    added_base = added

    # Build optional counterpart lookup from added questions: core -> o_<core>
    # Example: removed "confl_yn" + added "o_confl_yn" => mandatory_to_optional
    added_opt_lookup = (
        added
        .with_columns([
            pl.col("Q Name").cast(pl.Utf8).str.strip_chars().alias("_q"),
            pl.when(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.starts_with("o_"))
            .then(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.slice(2))
            .otherwise(pl.col("Q Name").cast(pl.Utf8).str.strip_chars())
            .str.to_lowercase()
            .alias("_core"),
        ])
        .filter(pl.col("_q").str.starts_with("o_"))
        .select(["_core", pl.col("Q Name").alias("_optional_qname")])
        .unique(subset=["_core"], keep="first")
    )

    # Severity: mandatory / mandatory-panel -> high; non-mandatory / optional -> info
    if reference_survey is not None and "mandatory_category" in reference_survey.columns:
        removed_aug = (
            removed.join(
                reference_survey.select(["Q Name", "mandatory_category"]),
                on="Q Name", how="left"
            )
            .with_columns(
                pl.when(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.starts_with("o_"))
                .then(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.slice(2))
                .otherwise(pl.col("Q Name").cast(pl.Utf8).str.strip_chars())
                .str.to_lowercase()
                .alias("_core")
            )
        )
        severity_expr = (
            pl.when(pl.col("mandatory_category").is_in(["mandatory", "mandatory-panel"]))
            .then(pl.lit("high"))
            .otherwise(pl.lit("info"))
            .alias("severity")
        )

        moved_to_optional = (
            removed_aug
            .join(added_opt_lookup, on="_core", how="inner")
            .filter(pl.col("mandatory_category").is_in(["mandatory", "mandatory-panel"]))
        )
    else:
        removed_aug = removed.with_columns([
            pl.lit("").alias("mandatory_category"),
            pl.when(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.starts_with("o_"))
            .then(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.slice(2))
            .otherwise(pl.col("Q Name").cast(pl.Utf8).str.strip_chars())
            .str.to_lowercase()
            .alias("_core"),
        ])
        severity_expr = (
            pl.when(pl.col("is_optional")).then(pl.lit("info")).otherwise(pl.lit("high"))
            .alias("severity")
        )
        moved_to_optional = pl.DataFrame(schema={
            "Q Name": pl.Utf8, "is_optional": pl.Boolean, "mandatory_category": pl.Utf8,
            "_core": pl.Utf8, "_optional_qname": pl.Utf8,
        })

    moved_i = (
        moved_to_optional.with_columns([
            pl.lit("mandatory_to_optional").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("Q Name").alias("field"),
            pl.col("_optional_qname").alias("current"),
            pl.lit("present (mandatory in reference)").alias("reference"),
            pl.lit("high").alias("severity"),
            pl.lit(None).cast(pl.Int64).alias("excel_row"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )

    added_base = (
        added.join(moved_to_optional.select(pl.col("_optional_qname").alias("Q Name")), on="Q Name", how="anti")
        if moved_to_optional.height > 0 else added
    )
    added_i = added_base.with_columns([
        pl.lit("added_question").alias("issue_type"),
        pl.lit("").alias("set_name"),
        pl.lit("Q Name").alias("field"),
        pl.lit("present").alias("current"),
        pl.lit("missing_in_reference").alias("reference"),
        pl.lit("info").alias("severity"),
        pl.lit(None).cast(pl.Int64).alias("excel_row"),
    ]).select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])

    removed_base = (
        removed_aug.join(moved_to_optional.select(["Q Name"]), on="Q Name", how="anti")
        if moved_to_optional.height > 0 else removed_aug
    )

    removed_i = (
        removed_base.with_columns([
            pl.lit("removed_question").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("Q Name").alias("field"),
            pl.lit("missing_in_current").alias("current"),
            pl.lit("present").alias("reference"),
            severity_expr,
            pl.lit(None).cast(pl.Int64).alias("excel_row"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )

    return pl.concat([added_i, removed_i, moved_i], how="vertical")


def make_mandatory_issues(mandatory_diff):
    return (
        mandatory_diff
        .with_columns([
            pl.lit("").alias("set_name"),
            pl.col("Mandatory").alias("current"),
            pl.col("Mandatory_ref").alias("reference"),
            pl.lit("high").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )


def make_option_issues(option_diff):
    return (
        option_diff
        .with_columns([
            pl.lit("").alias("set_name"),
            pl.col("option_label").alias("current"),
            pl.col("option_label_ref").alias("reference"),
            pl.lit("medium").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "list_name", "option_name", "field", "current", "reference", "severity", "excel_row"])
    )


def make_option_presence_issues(added_opts, removed_opts):
    added_i = added_opts.with_columns([
        pl.lit("").alias("set_name"),
        pl.col("option_label").alias("current"),
        pl.lit(_not_in_reference_text()).alias("reference"),
        pl.lit("medium").alias("severity"),
        pl.lit(None).cast(pl.Int64).alias("excel_row"),
    ]).select(["issue_type", "set_name", "Q Name", "list_name", "option_name", "field", "current", "reference", "severity", "excel_row"])

    removed_i = removed_opts.with_columns([
        pl.lit("").alias("set_name"),
        pl.lit("(removed)").alias("current"),
        pl.col("option_label").alias("reference"),
        pl.lit("high").alias("severity"),
        pl.lit(None).cast(pl.Int64).alias("excel_row"),
    ]).select(["issue_type", "set_name", "Q Name", "list_name", "option_name", "field", "current", "reference", "severity", "excel_row"])

    return pl.concat([added_i, removed_i], how="vertical")


def collapse_cluster_ea_option_changes(
    option_label_issues: pl.DataFrame,
    option_presence_issues: pl.DataFrame,
) -> tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame]:
    """
    Suppress noisy per-option diffs for KoBo cluster/ea lists and replace them
    with a single informational summary row.
    """
    issue_schema = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "list_name": pl.Utf8, "option_name": pl.Utf8, "field": pl.Utf8,
        "current": pl.Utf8, "reference": pl.Utf8, "severity": pl.Utf8,
        "excel_row": pl.Int64,
    }

    q_expr = pl.col("Q Name").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_lowercase()
    list_expr = pl.col("list_name").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_lowercase()
    ea_cluster_mask = (q_expr == "cluster") & (list_expr == "ea")

    modified_n = option_label_issues.filter(ea_cluster_mask).height if option_label_issues.height else 0
    added_n = option_presence_issues.filter(ea_cluster_mask & (pl.col("issue_type") == "added_choice")).height if option_presence_issues.height else 0
    removed_n = option_presence_issues.filter(ea_cluster_mask & (pl.col("issue_type") == "removed_choice")).height if option_presence_issues.height else 0

    option_label_clean = option_label_issues.filter(~ea_cluster_mask) if option_label_issues.height else option_label_issues
    option_presence_clean = option_presence_issues.filter(~ea_cluster_mask) if option_presence_issues.height else option_presence_issues

    total = modified_n + added_n + removed_n
    if total == 0:
        return option_label_clean, option_presence_clean, pl.DataFrame(schema=issue_schema)

    summary_row = pl.DataFrame([{
        "issue_type": "cluster_ea_choice_changes_summary",
        "set_name": "",
        "Q Name": "cluster",
        "list_name": "ea",
        "option_name": "",
        "field": "choices",
        "current": f"added={added_n}; removed={removed_n}; modified={modified_n}",
        "reference": "Detailed cluster/ea per-option differences suppressed (expected round-to-round rotation)",
        "severity": "info",
        "excel_row": None,
    }], schema=issue_schema)

    return option_label_clean, option_presence_clean, summary_row



# ======================================================================
# SECTION: ## Step 10 -- Run pipeline
_step("Running pipeline")
# ======================================================================
# --- CODE CELL 13 ---
import time
_t0 = time.perf_counter()

def _dedupe_questions_for_compare(df: pl.DataFrame) -> pl.DataFrame:
    """
    Keep one deterministic row per Q Name for pairwise comparisons.
    Duplicate Q Names are handled separately by structure checks.
    """
    if df.is_empty() or "Q Name" not in df.columns:
        return df
    order_cols = [c for c in ["Q Name", "excel_row"] if c in df.columns]
    base = df.sort(order_cols) if order_cols else df
    return base.unique(subset=["Q Name"], keep="first")

current_cmp_q   = _dedupe_questions_for_compare(current_survey_cmp)
reference_cmp_q = _dedupe_questions_for_compare(reference_survey)
current_van_q   = _dedupe_questions_for_compare(current_survey_vanilla_cmp)
added, removed = compare_question_presence(current_cmp_q, reference_cmp_q)
print(f"Questions added: {added.height}  removed: {removed.height}")

mandatory_diff = compare_mandatory_kobo(current_cmp_q, reference_cmp_q)
print(f"Mandatory mismatches: {mandatory_diff.height}")

list_name_issues = compare_list_name_changes(current_cmp_q, reference_cmp_q)
print(f"Choices-list changes: {list_name_issues.height}")

# Field-level comparisons -- vanilla versions for label/constraint
label_issues       = compare_question_labels(current_van_q, reference_cmp_q, current_cmp_q)
constraint_issues  = compare_constraint_changes(current_van_q, reference_cmp_q, current_cmp_q)
type_issues        = compare_type_changes(
    current_cmp_q,
    reference_cmp_q,
    current_options=current_options_vanilla_cmp,
    reference_options=reference_options,
)
required_issues    = compare_required_changes(current_cmp_q, reference_cmp_q)
choice_filter_issues = compare_choice_filter_changes(current_cmp_q, reference_cmp_q)
appearance_issues  = compare_appearance_changes(current_cmp_q, reference_cmp_q)
calculation_issues = compare_calculation_changes(current_cmp_q, reference_cmp_q)
hint_issues        = compare_hint_changes(current_cmp_q, reference_cmp_q)
print(f"Label: {label_issues.height}  Constraint: {constraint_issues.height}  Type: {type_issues.height}")
print(f"Required: {required_issues.height}  Choice filter: {choice_filter_issues.height}  Appearance: {appearance_issues.height}  Calculation: {calculation_issues.height}  Hint: {hint_issues.height}")

# Options -- in previous_round mode, include country-specific lists so round updates are visible
_country_lists  = set(cfg.get("kobo_options", {}).get("country_specific_list_names", []))
if run.get("reference_mode") == "previous_round":
    _country_lists = set()

_list_changed_q = set(list_name_issues["Q Name"].to_list())

cur_opts_cmp = (
    current_options_vanilla_cmp
    .filter(~pl.col("list_name").is_in(list(_country_lists)))
    .filter(~pl.col("Q Name").is_in(list(_list_changed_q)))
)
ref_opts_cmp = (
    reference_options
    .filter(~pl.col("list_name").is_in(list(_country_lists)))
    .filter(~pl.col("Q Name").is_in(list(_list_changed_q)))
)

option_diff              = compare_option_labels_single(cur_opts_cmp, ref_opts_cmp)
added_opts, removed_opts = compare_option_presence_single(cur_opts_cmp, ref_opts_cmp)
print(f"Option label changes: {option_diff.height}  added: {added_opts.height}  removed: {removed_opts.height}")

relevant_issues = validate_relevant(current_cmp_q, reference_cmp_q)
print(f"Relevant issues: {relevant_issues.height}")

structure_issues = validate_questionnaire_structure(
    current_survey_cmp,
    current_choices=current_choices,
    template_survey=template_survey,
    replacement_pairs=replacement_pairs,
)
print(f"Questionnaire structure issues: {structure_issues.height}")

critical_issues = validate_critical_sets(current_cmp_q, rules.get("exact_sets", {}))
count_issues    = validate_prefix_counts(current_cmp_q, rules.get("min_count_sets", {}))
harvest_issues  = validate_crop_harvest(current_cmp_q, rules.get("crop_harvest", {}))

print(f"Critical set issues: {critical_issues.height}  Count issues: {count_issues.height}")

# -- Assemble all issues -------------------------------------------------------
presence_issues     = make_presence_issues(added, removed, reference_survey)
mandatory_issues    = make_mandatory_issues(mandatory_diff)
option_label_issues = make_option_issues(option_diff)
option_pres_issues  = make_option_presence_issues(added_opts, removed_opts)

# Exclude option issues for questions not present in both files
_excluded = set(added["Q Name"].to_list()) | set(removed["Q Name"].to_list())
if _excluded:
    option_label_issues = option_label_issues.filter(~pl.col("Q Name").is_in(list(_excluded)))
    option_pres_issues  = option_pres_issues.filter(~pl.col("Q Name").is_in(list(_excluded)))

# Config-driven exclusions for operational fields (case/space-insensitive)
_cfg_excluded = {
    str(q).strip().lower()
    for q in cfg.get("kobo_options", {}).get("exclude_qnames_from_option_compare", [])
    if q is not None and str(q).strip()
}
if _cfg_excluded:
    _q_norm = pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.to_lowercase()
    option_label_issues = option_label_issues.filter(~_q_norm.is_in(list(_cfg_excluded)))
    option_pres_issues  = option_pres_issues.filter(~_q_norm.is_in(list(_cfg_excluded)))

# Suppress noisy country-rotation diffs for cluster/ea and keep one info summary row.
option_label_issues, option_pres_issues, cluster_ea_summary_issues = (
    collapse_cluster_ea_option_changes(option_label_issues, option_pres_issues)
)
if cluster_ea_summary_issues.height > 0:
    print(
        "Cluster/ea option diffs collapsed: "
        + str(cluster_ea_summary_issues.select("current").item())
    )

def _with_choice_cols(df: pl.DataFrame) -> pl.DataFrame:
    out = df
    if "list_name" not in out.columns:
        out = out.with_columns(pl.lit("").alias("list_name"))
    if "option_name" not in out.columns:
        out = out.with_columns(pl.lit("").alias("option_name"))
    return out

all_issues = pl.concat(
    [_with_choice_cols(presence_issues),
     _with_choice_cols(mandatory_issues),
     _with_choice_cols(list_name_issues),
     _with_choice_cols(label_issues),
     _with_choice_cols(constraint_issues),
     _with_choice_cols(type_issues),
     _with_choice_cols(required_issues),
     _with_choice_cols(choice_filter_issues),
     _with_choice_cols(appearance_issues),
     _with_choice_cols(calculation_issues),
     _with_choice_cols(hint_issues),
     _with_choice_cols(option_label_issues),
     _with_choice_cols(option_pres_issues),
     _with_choice_cols(cluster_ea_summary_issues),
     _with_choice_cols(critical_issues),
     _with_choice_cols(count_issues),
     _with_choice_cols(harvest_issues),
     _with_choice_cols(relevant_issues),
     _with_choice_cols(structure_issues)],
    how="diagonal",
)

# -- previous_round special remap ---------------------------------------------
# Differences affecting replacement-driven cells/lists are tracked as
# additional_information_replacement_change (info) instead of hard errors.
if run.get("reference_mode") == "previous_round":
    _rp_types_q = [
        "label_mismatch", "constraint_modified", "hint_changed", "choices_list_changed",
    ]
    _rp_types_opt = ["choice_label_mismatch", "added_choice", "removed_choice"]
    _rp_qnames = sorted(set(_round_param_qnames) | set(_round_param_option_qnames))
    if _rp_qnames:
        _rp_mask_q = (
            pl.col("issue_type").is_in(_rp_types_q)
            & pl.col("Q Name").is_in(_rp_qnames)
        )
        _rp_mask_opt = (
            pl.col("issue_type").is_in(_rp_types_opt)
            & pl.col("Q Name").is_in(_rp_qnames)
        )
        all_issues = all_issues.with_columns(
            pl.col("issue_type").alias("_issue_type_raw")
        ).with_columns([
            pl.when(_rp_mask_q | _rp_mask_opt)
            .then(
                pl.concat_str([
                    pl.lit("additional_information_replacement_change ("),
                    pl.col("_issue_type_raw"),
                    pl.lit(")"),
                ])
            )
            .otherwise(pl.col("issue_type"))
            .alias("issue_type"),
            pl.when(_rp_mask_q | _rp_mask_opt).then(pl.lit("info")).otherwise(pl.col("severity")).alias("severity"),
        ]).drop("_issue_type_raw")
        print(f"Round-parameter remap applied on {_rp_qnames.__len__()} question(s)")

# -- CS downgrade: removed questions in passing prefix-count sets -> info ------
_passing_prefixes = []
for _set_name, _rule in rules.get("min_count_sets", {}).items():
    if count_issues.filter(pl.col("set_name") == _set_name).height == 0:
        _prefix = _rule.get("prefix", "")
        if _prefix:
            _passing_prefixes.append(_prefix)

if _passing_prefixes:
    _exprs    = [pl.col("Q Name").cast(pl.Utf8).str.starts_with(p) for p in _passing_prefixes]
    _q_matches = pl.any_horizontal(_exprs) if len(_exprs) > 1 else _exprs[0]
    all_issues = all_issues.with_columns(
        pl.when((pl.col("issue_type") == "removed_question") & _q_matches)
        .then(pl.lit("info"))
        .otherwise(pl.col("severity"))
        .alias("severity")
    )

# -- Attach mandatory_cat -----------------------------------------------------
_mand_lookup = (
    pl.concat([
        reference_survey.select(["Q Name", "mandatory_category"]).with_columns(pl.lit(0).alias("_p")),
        current_survey_cmp.select(["Q Name", "mandatory_category"]).with_columns(pl.lit(1).alias("_p")),
    ], how="vertical")
    .sort(["Q Name", "_p"])
    .group_by("Q Name")
    .agg(pl.col("mandatory_category").last().alias("mandatory_cat"))
)
all_issues = (
    all_issues
    .join(_mand_lookup, on="Q Name", how="left")
    .with_columns(pl.col("mandatory_cat").fill_null(""))
)

# -- Final sort: high -> medium -> info (AFTER cs-downgrade + mandatory_cat) --
all_issues = (
    all_issues
    .with_columns(
        pl.when(pl.col("severity") == "high").then(pl.lit(0))
        .when(pl.col("severity") == "medium").then(pl.lit(1))
        .otherwise(pl.lit(2))
        .alias("_s")
    )
    .sort(["_s", "issue_type", "Q Name"])
    .drop("_s")
)

# -- Found info for Critical Sets ---------------------------------------------
_found_info = {}
_cur_qnames = current_cmp_q["Q Name"].to_list()

# Exact sets: track which expected questions are present (used for summary fallback).
for _set_name, _set_rules in rules.get("exact_sets", {}).items():
    _expected = [str(r.get("q_name", "")).strip() for r in _set_rules]
    _matched = []
    for q in _expected:
        if not q:
            continue
        if q in _cur_qnames:
            _matched.append(q)
            continue
        if q.startswith("hh_wealth_"):
            _alt = "o_hh_wealth_" + q[len("hh_wealth_"):]
            if _alt in _cur_qnames:
                _matched.append(f"{q} (matched via {_alt})")
    _found_info[_set_name] = _matched

# Prefix-count sets: keep matched names for display + count details.
for _set_name, _rule in rules.get("min_count_sets", {}).items():
    _prefix  = str(_rule.get("prefix", "")).strip()
    _prefix_alt = "o_hh_wealth_" if _prefix == "hh_wealth_" or _set_name.upper() == "WEALTH" else ""
    _matched = sorted(
        q for q in _cur_qnames
        if (_prefix and q.startswith(_prefix)) or (_prefix_alt and q.startswith(_prefix_alt))
    )
    _found_info[_set_name] = _matched

print(f"\nTotal issues: {all_issues.height}  ({time.perf_counter()-_t0:.2f}s)")
print(all_issues.select(["issue_type", "severity", "Q Name"]).head(15))










# ======================================================================
# SECTION: ## Step 11 -- Export
_step("Exporting report")
# ======================================================================
# --- CODE CELL 14 ---
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import difflib
import re

try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    _RICH_TEXT_AVAILABLE = True
except Exception:
    CellRichText = None
    TextBlock = None
    InlineFont = None
    _RICH_TEXT_AVAILABLE = False

FILL_HIGH   = PatternFill("solid", fgColor="F4CCCC")
FILL_MEDIUM = PatternFill("solid", fgColor="FCE5CD")
FILL_INFO   = PatternFill("solid", fgColor="CFE2F3")
FILL_PASS   = PatternFill("solid", fgColor="D9EAD3")
FILL_HEADER = PatternFill("solid", fgColor="274E13")
FILL_SECT   = PatternFill("solid", fgColor="E8F5E9")

FONT_TITLE  = Font(bold=True, size=13, color="274E13")
FONT_HDR    = Font(bold=True, color="FFFFFF", size=11)
FONT_SECT   = Font(bold=True, color="274E13", size=11)
FONT_NORM   = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
SEV_FILL = {"high": FILL_HIGH, "medium": FILL_MEDIUM, "info": FILL_INFO, "pass": FILL_PASS}

_ISSUE_LABELS = {
    "removed_question"         : "Question removed",
    "mandatory_to_optional"    : "Mandatory question moved to optional",
    "added_question"           : "Question added",
    "mandatory_column_mismatch": "Mandatory flag changed",
    "type_changed"             : "Question type changed",
    "label_mismatch"           : "Label text changed",
    "required_changed"         : "Required changed",
    "choice_filter_modified"   : "Choice filter changed",
    "appearance_changed"       : "Appearance changed",
    "calculation_changed"      : "Calculation changed",
    "constraint_modified"      : "Constraint changed",
    "hint_changed"             : "Hint text changed",
    "crop_harvest_violation"   : "Crop harvest rule violation",
    "choices_list_changed"     : "Choices list changed",
    "additional_information_replacement_change": "Additional information replacement change",
    "removed_choice"           : "Choice removed",
    "added_choice"             : "Choice added",
    "choice_label_mismatch"    : "Choice label changed",
    "cluster_ea_choice_changes_summary": "Cluster/EA choice changes summary",
    "placeholder_not_found"    : "Placeholder not in template/additional info",
    "placeholder_should_use_kobo_ref": "Placeholder looks like variable (use ${var})",
    "kobo_ref_loose_syntax"    : "Loose $var syntax (use ${var})",
    "kobo_ref_missing_variable": "${var} references missing variable",
    "duplicate_qname"          : "Duplicate question name",
    "duplicate_choice_name"    : "Duplicate choice name in list",
}

def _hdr(ws, row, vals):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = FILL_HEADER; cell.font = FONT_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 18

def _sect(ws, row, title, n):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = FILL_SECT; c.font = FONT_SECT
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20

def _style_row(ws, row, n, severity=""):
    fill = SEV_FILL.get(severity)
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_NORM; cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if fill: cell.fill = fill

def _autofit(ws, mn=12, mx=60):
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w+2, mn), mx)

ISSUE_ACTION_MAP = {
    "added_question": "Review added question and confirm it is intentional",
    "removed_question": "Restore missing question or document approved removal",
    "mandatory_to_optional": "Review mandatory status against reference",
    "mandatory_column_mismatch": "Review mandatory status against reference",
    "choices_list_changed": "Review list_name change and downstream skip/relevance",
    "label_mismatch": "Review question label mismatch",
    "constraint_modified": "Review constraint change and respondent impact",
    "required_changed": "Review required flag change",
    "choice_filter_modified": "Review choice_filter change and dependent list filtering",
    "appearance_changed": "Review appearance change",
    "calculation_changed": "Review calculation change",
    "type_changed": "Review question type change",
    "hint_changed": "Review hint text change",
    "removed_choice": "Review removed choice and downstream logic",
    "added_choice": "Review added choice and downstream logic",
    "choice_label_mismatch": "Review choice label mismatch",
    "cluster_ea_choice_changes_summary": "Informational only: review aggregate cluster/ea choice turnover counts",
    "broken_relevant_reference": "Fix relevant expression to valid referenced variables",
    "relevant_inexact_reference": "Review relevant expression equivalence with reference",
    "relevant_modified": "Review relevant expression change",
    "duplicate_qname": "Rename duplicate question names to unique values",
    "duplicate_choice_name": "Rename duplicate choice names in list",
    "placeholder_should_use_kobo_ref": "Use KoBo variable syntax ${var} instead of plain placeholder",
    "placeholder_not_found": "Add missing placeholder key or correct typo",
    "kobo_ref_loose_syntax": "Replace loose $var syntax with ${var}",
    "kobo_ref_missing_variable": "Fix ${var} reference to an existing survey variable",
    "missing_critical_question": "Add missing critical question",
    "advisory_question": "Review advisory question omission",
    "critical_mandatory_mismatch": "Align mandatory flag for critical question",
    "crop_harvest_violation": "Fix crop_harvest sequence/order",
}


def _action_for_issue_type(issue_type: str) -> str:
    key = str(issue_type or "").strip()
    if key.startswith("additional_information_replacement_change"):
        return "Review replacement-driven text change (Additional info / crop / AGOL)"
    return ISSUE_ACTION_MAP.get(key, "Review issue details")


COL_MAP = [
    ("issue_type",    "Issue type"),
    ("mandatory_cat", "Type"),
    ("set_name",      "Set"),
    ("Q Name",        "Q Name"),
    ("list_name",     "list_name"),
    ("option_name",   "name"),
    ("field",         "Field"),
    ("current",       "Current value"),
    ("reference",     "Reference / rule"),
    ("action",        "Action"),
    ("severity",      "Severity"),
    ("excel_row",     "Excel row"),
]

def _table(ws, start_row, df, apply_view=True, col_map=None):
    if "issue_type" in df.columns and "action" not in df.columns:
        df = df.with_columns(
            pl.col("issue_type")
            .map_elements(_action_for_issue_type, return_dtype=pl.Utf8)
            .alias("action")
        )
    cmap = col_map or COL_MAP
    cols = [(s, d) for s, d in cmap if s in df.columns]
    _hdr(ws, start_row, [d for _, d in cols])
    r = start_row + 1
    if df.height == 0:
        ws.cell(row=r, column=1, value="No issues in this category").font = Font(bold=True, color="274E13", size=10)
        return r + 2
    for rd in df.to_dicts():
        for c, (src, _) in enumerate(cols, 1):
            ws.cell(row=r, column=c, value=rd.get(src))
        _style_row(ws, r, len(cols), rd.get("severity", ""))
        r += 1
    if apply_view:
        ws.freeze_panes = f"A{start_row+1}"
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(cols))}{start_row}"
    return r + 1


def _tokenize_for_word_diff(text: str):
    # Tokenize into non-space chunks and attach following spaces to the same token.
    # This reduces rich-text run fragmentation and avoids visual space-loss artifacts.
    raw = re.findall(r"\s+|\S+", text)
    out = []
    for tok in raw:
        if tok.isspace() and out:
            out[-1] = out[-1] + tok
        else:
            out.append(tok)
    return out


def _coalesce_segments(parts):
    merged = []
    for txt, is_equal in parts:
        if not txt:
            continue
        if merged and merged[-1][1] == is_equal:
            merged[-1] = (merged[-1][0] + txt, is_equal)
        else:
            merged.append((txt, is_equal))
    return merged


def _build_diff_segments(current_text: str, reference_text: str):
    a = _tokenize_for_word_diff(current_text)
    b = _tokenize_for_word_diff(reference_text)
    sm = difflib.SequenceMatcher(a=a, b=b)
    current_parts, reference_parts = [], []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            current_parts.append(("".join(a[i1:i2]), True))
            reference_parts.append(("".join(b[j1:j2]), True))
        elif tag == "replace":
            current_parts.append(("".join(a[i1:i2]), False))
            reference_parts.append(("".join(b[j1:j2]), False))
        elif tag == "delete":
            current_parts.append(("".join(a[i1:i2]), False))
        elif tag == "insert":
            reference_parts.append(("".join(b[j1:j2]), False))
    return _coalesce_segments(current_parts), _coalesce_segments(reference_parts)


def _set_rich_text(cell, full_text: str, parts):
    if not _RICH_TEXT_AVAILABLE:
        cell.value = full_text
        return
    blocks = []
    for txt, is_equal in parts:
        if not txt:
            continue
        color = "FF000000" if is_equal else "FFFF0000"
        blocks.append(TextBlock(InlineFont(color=color), txt))
    if not blocks:
        cell.value = full_text
        return
    cell.value = CellRichText(*blocks)


def _apply_inline_diff_for_issue(ws, start_row, df, cols, issue_types):
    if df.height == 0:
        return
    cidx = {src: i for i, (src, _) in enumerate(cols, start=1)}
    issue_col = cidx.get("issue_type")
    cur_col = cidx.get("current")
    ref_col = cidx.get("reference")
    if not issue_col or not cur_col or not ref_col:
        return

    _allowed = set(issue_types or [])

    def _match_issue(issue_val: str) -> bool:
        for allowed in _allowed:
            if isinstance(allowed, str) and allowed.endswith("*"):
                if issue_val.startswith(allowed[:-1]):
                    return True
            elif issue_val == allowed:
                return True
        return False

    for offset, rd in enumerate(df.to_dicts(), start=1):
        _issue_val = str(rd.get("issue_type") or "")
        if not _match_issue(_issue_val):
            continue
        cur_txt = str(rd.get("current") or "")
        ref_txt = str(rd.get("reference") or "")
        cur_parts, ref_parts = _build_diff_segments(cur_txt, ref_txt)
        _set_rich_text(ws.cell(row=start_row + offset, column=cur_col), cur_txt, cur_parts)
        _set_rich_text(ws.cell(row=start_row + offset, column=ref_col), ref_txt, ref_parts)


CRITICAL_ISSUE_TYPES = {
    "missing_critical_question", "advisory_question",
    "critical_mandatory_mismatch", "crop_harvest_violation",
}
RELEVANT_ISSUE_TYPES  = {"broken_relevant_reference", "relevant_inexact_reference", "relevant_modified"}
STRUCTURE_ISSUE_TYPES = {
    "kobo_ref_loose_syntax", "kobo_ref_missing_variable",
    "duplicate_qname", "duplicate_choice_name", "type_changed",
}
REPLACEMENT_ISSUE_TYPES = {
    "placeholder_not_found", "placeholder_should_use_kobo_ref",
}
QUESTION_CHANGE_TYPES = {
    "mandatory_column_mismatch", "added_question", "removed_question", "mandatory_to_optional",
    "choices_list_changed", "label_mismatch", "constraint_modified",
    "required_changed", "choice_filter_modified", "appearance_changed", "calculation_changed",
    "hint_changed",
}


def _question_change_expr():
    return pl.col("issue_type").is_in(list(QUESTION_CHANGE_TYPES))


def _normalise_qtype_for_summary(df):
    """Fill missing/blank mandatory_cat for summary tables.

    Rule:
      - keep valid mandatory_cat values as-is
      - if blank and Q Name starts with o_ -> optional
      - otherwise -> non-mandatory
    """
    if df.height == 0:
        return df
    if "mandatory_cat" not in df.columns:
        return df.with_columns(pl.lit("non-mandatory").alias("mandatory_cat"))

    _valid = ["mandatory", "mandatory-panel", "non-mandatory", "optional"]
    _mand = pl.col("mandatory_cat").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_lowercase()
    _qnm = (
        pl.when(pl.col("Q Name").is_null()).then(pl.lit(""))
        .otherwise(pl.col("Q Name").cast(pl.Utf8))
        .str.strip_chars()
        .str.to_lowercase()
    )
    return df.with_columns(
        pl.when(_mand.is_in(_valid)).then(_mand)
        .when(_qnm.str.starts_with("o_")).then(pl.lit("optional"))
        .otherwise(pl.lit("non-mandatory"))
        .alias("mandatory_cat")
    )


def _grouped_table(ws, r, df):
    """Render (issue_type, mandatory_cat, severity) -> count grouped summary table."""
    _hdr(ws, r, ["Issue type", "Q type", "Count", "Severity"]); r += 1
    if df.height == 0:
        ws.cell(row=r, column=1, value="No issues").font = Font(bold=True, color="274E13", size=10)
        return r + 2

    df = _normalise_qtype_for_summary(df)

    grouped = (
        df
        .group_by(["issue_type", "mandatory_cat", "severity"])
        .agg(pl.len().alias("count"))
        .with_columns(
            pl.when(pl.col("severity") == "high").then(pl.lit(0))
            .when(pl.col("severity") == "medium").then(pl.lit(1))
            .otherwise(pl.lit(2))
            .alias("_s"),
            pl.when(pl.col("mandatory_cat") == "mandatory").then(pl.lit(0))
            .when(pl.col("mandatory_cat") == "mandatory-panel").then(pl.lit(1))
            .when(pl.col("mandatory_cat") == "non-mandatory").then(pl.lit(2))
            .when(pl.col("mandatory_cat") == "optional").then(pl.lit(3))
            .otherwise(pl.lit(4))
            .alias("_q")
        )
        .sort(["_s", "issue_type", "_q"])
        .drop(["_s", "_q"])
    )
    for rd in grouped.to_dicts():
        ws.cell(row=r, column=1, value=_ISSUE_LABELS.get(rd["issue_type"], rd["issue_type"]))
        ws.cell(row=r, column=2, value=rd["mandatory_cat"] or "")
        ws.cell(row=r, column=3, value=rd["count"])
        ws.cell(row=r, column=4, value=rd["severity"])
        _style_row(ws, r, 4, rd["severity"])
        r += 1
    return r + 1


def write_summary_sheet(wb, all_issues, rules=None, critical_issues=None,
                        count_issues=None, found_info=None, replacement_status=None):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", [38, 20, 42, 12]):
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = "KoBO Questionnaire Validation Report"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Comparison basis: {_reference_scope_label()}"
    ws["A2"].font = Font(size=10, color="274E13", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A3:D3")
    _curr_name = Path(run.get("questionnaire_path", "")).name
    _curr_iso = str(run.get("iso3", "") or "").upper()
    _curr_round = _extract_round_token_kobo(_curr_name) or "R?"
    ws["A3"] = f"Current questionnaire: {_curr_iso} {_curr_round} | {_curr_name}"
    ws["A3"].font = Font(size=10, color="1F4E78", italic=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A4:D4")
    ws["A4"] = f"Checked against: {_reference_descriptor_kobo()}"
    ws["A4"].font = Font(size=10, color="1F4E78", italic=True)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A5:D5")
    ws["A5"] = f"Language scope: {_language_scope_descriptor_kobo()}"
    ws["A5"].font = Font(size=10, color="1F4E78", italic=True)
    ws["A5"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A6:D6")
    _tpl_name = Path(run.get("template_path", "")).name if run.get("template_path") else "N/A"
    ws["A6"] = f"Template used for placeholder mapping: {_tpl_name}"
    ws["A6"].font = Font(size=10, color="1F4E78", italic=True)
    ws["A6"].alignment = Alignment(horizontal="left", vertical="center")

    r = 7
    _rules = rules or {}
    _crit  = (critical_issues if critical_issues is not None
              else pl.DataFrame(schema={"set_name": pl.Utf8, "issue_type": pl.Utf8, "Q Name": pl.Utf8}))
    _cnt   = (count_issues if count_issues is not None
              else pl.DataFrame(schema={"set_name": pl.Utf8, "issue_type": pl.Utf8}))
    _found = found_info or {}

    # -- Critical Sets Status -------------------------------------------------
    _sect(ws, r, "CRITICAL SETS STATUS", 4); r += 1
    _hdr(ws, r, ["Critical set", "Status", "Details", ""]); r += 1

    def _set_row(ws, r, name, passed, details):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value="PASS" if passed else "FAIL")
        ws.cell(row=r, column=3, value=details)
        _style_row(ws, r, 4, "pass" if passed else "high")
        ws.cell(row=r, column=2).font = Font(bold=True, size=10,
                                              color="274E13" if passed else "CC0000")

    _EXACT_FAIL_TYPES = ["missing_critical_question", "critical_mandatory_mismatch"]
    for set_name, set_rules in _rules.get("exact_sets", {}).items():
        required_qs = [r.get("q_name", "") for r in set_rules if r.get("required", True)]
        advisory_qs = [r.get("q_name", "") for r in set_rules if not r.get("required", True)]
        found_exact = set(_found.get(set_name, []))
        missing_required = [q for q in required_qs if q and q not in found_exact]

        set_fail = all_issues.filter(
            (pl.col("set_name") == set_name) & pl.col("issue_type").is_in(_EXACT_FAIL_TYPES)
        )
        set_warn = all_issues.filter(
            (pl.col("set_name") == set_name) & (pl.col("issue_type") == "advisory_question")
        )

        # Fallback to rule-based presence check so status stays correct
        # even if issue tagging is affected by execution order.
        passed = (set_fail.height == 0) and (len(missing_required) == 0)

        if passed and set_warn.height == 0:
            details = "All required questions present"
        elif passed:
            warn_qs = [q for q in advisory_qs if q and q not in found_exact]
            if not warn_qs:
                warn_qs = [q for q in set_warn["Q Name"].to_list() if q]
            details = (f"Required OK  (advisory missing: {', '.join(warn_qs)})"
                       if warn_qs else "All required questions present")
        else:
            missing = list(dict.fromkeys(missing_required + [q for q in set_fail["Q Name"].to_list() if q]))
            details = f"Missing: {', '.join(missing)}" if missing else "Issues found"
        _set_row(ws, r, set_name, passed, details); r += 1
    for set_name, rule in _rules.get("min_count_sets", {}).items():
        min_count = rule.get("min_count", 1)
        prefix    = rule.get("prefix", "")
        found_qs  = _found.get(set_name, [])
        failed    = _cnt.filter(pl.col("set_name") == set_name).height > 0
        if not failed:
            details = f">={min_count} {prefix}* confirmed  ({len(found_qs)} found)"
        else:
            details = f"{len(found_qs)} found -- need >={min_count} {prefix}* questions"
        _set_row(ws, r, set_name, not failed, details); r += 1

    # Crop harvest row
    if _rules.get("crop_harvest"):
        _harv_fail = all_issues.filter(pl.col("issue_type") == "crop_harvest_violation").height > 0
        _harv_det = ("Partial crop-harvest set -- see issues" if _harv_fail
                     else "Crop harvest rule satisfied")
        _set_row(ws, r, "CRP_HARV", not _harv_fail, _harv_det); r += 1

    if not _rules.get("exact_sets") and not _rules.get("min_count_sets"):
        ws.cell(row=r, column=1, value="No critical sets defined")
        _style_row(ws, r, 4, "info"); r += 1

    r += 1

    # -- Questionnaire Structure Checks --------------------------------------
    _sect(ws, r, "QUESTIONNAIRE STRUCTURE CHECKS", 4); r += 1
    _hdr(ws, r, ["Check", "Status", "Details", ""]); r += 1

    def _check_row(ws, r, name, passed, details):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value="PASS" if passed else "FAIL")
        ws.cell(row=r, column=3, value=details)
        _style_row(ws, r, 4, "pass" if passed else "high")
        ws.cell(row=r, column=2).font = Font(bold=True, size=10,
                                              color="274E13" if passed else "CC0000")

    _broken   = all_issues.filter(pl.col("issue_type").is_in(["broken_relevant_reference", "relevant_inexact_reference"])).height
    _modified = all_issues.filter(pl.col("issue_type") == "relevant_modified").height
    skip_ok = _broken == 0
    skip_det = (f"{_broken} broken reference(s); {_modified} modified" if not skip_ok
                else f"No broken references  ({_modified} modified)")
    _check_row(ws, r, "Skip logic references (relevant)", skip_ok, skip_det); r += 1

    _struct_df   = all_issues.filter(pl.col("issue_type").is_in(list(STRUCTURE_ISSUE_TYPES)))
    _struct_all  = _struct_df.height
    _dup_q       = _struct_df.filter(pl.col("issue_type") == "duplicate_qname").height
    _dup_choice  = _struct_df.filter(pl.col("issue_type") == "duplicate_choice_name").height
    _qtype       = _struct_df.filter(pl.col("issue_type") == "type_changed").height
    _other_struct = _struct_all - _dup_q - _dup_choice - _qtype
    _struct_ok   = _struct_all == 0
    _struct_det  = (
        f"{_struct_all} issue(s): Q type integrity={_qtype}; duplicate Q Name={_dup_q}; duplicate choice name={_dup_choice}; KoBo refs={_other_struct}"
        if not _struct_ok else
        "No issues in Q type integrity, duplicates, or KoBo references"
    )
    _check_row(ws, r, "Q type integrity, duplicates and KoBo references", _struct_ok, _struct_det); r += 1

    r += 1

    # -- Replacement Status ---------------------------------------------------
    _sect(ws, r, "REPLACEMENT STATUS  (validated questionnaire output)", 4); r += 1
    _hdr(ws, r, ["Check", "Status", "Details", ""]); r += 1

    _repl_rows = (replacement_status or {}).get("rows", [])
    if not _repl_rows:
        ws.cell(row=r, column=1, value="Validated questionnaire step not run yet")
        ws.cell(row=r, column=2, value="INFO")
        ws.cell(row=r, column=3, value="Run Step 12, then rerun export to include replacement checks.")
        _style_row(ws, r, 4, "info"); r += 1
    else:
        for rd in _repl_rows:
            sev = rd.get("severity", "info")
            st  = rd.get("status", "INFO")
            ws.cell(row=r, column=1, value=rd.get("check", ""))
            ws.cell(row=r, column=2, value=st)
            ws.cell(row=r, column=3, value=rd.get("details", ""))
            _style_row(ws, r, 4, sev)
            if st == "PASS":
                ws.cell(row=r, column=2).font = Font(bold=True, size=10, color="274E13")
            elif st == "WARN":
                ws.cell(row=r, column=2).font = Font(bold=True, size=10, color="B45F06")
            else:
                ws.cell(row=r, column=2).font = Font(bold=True, size=10, color="CC0000")
            r += 1

    r += 1

    # -- Question Changes -----------------------------------------------------
    _sect(ws, r, "QUESTION CHANGES", 4); r += 1
    r = _grouped_table(ws, r, all_issues.filter(_question_change_expr()))
    r += 1

    # -- Choice Changes -------------------------------------------------------
    _sect(ws, r, "CHOICE CHANGES  (questions present in both files)", 4); r += 1
    r = _grouped_table(ws, r,
                       all_issues.filter(pl.col("issue_type").is_in(
                           ["removed_choice", "added_choice", "choice_label_mismatch", "cluster_ea_choice_changes_summary"])))


def write_critical_sets_sheet(wb, all_issues, found_info=None):
    ws = wb.create_sheet("Critical Sets")
    ws.sheet_view.showGridLines = False

    df = all_issues.filter(
        pl.col("issue_type").is_in(list(CRITICAL_ISSUE_TYPES))
    ).sort(["set_name", "severity", "Q Name"])
    _sect(ws, 1, "CRITICAL SETS  Structural validation issues", 8)
    next_row = _table(ws, 2, df, apply_view=False)

    if found_info:
        next_row += 1
        _sect(ws, next_row, "QUESTIONS FOUND IN CURRENT QUESTIONNAIRE", 3)
        next_row += 1
        _hdr(ws, next_row, ["Set", "Count found", "Questions"])
        next_row += 1
        for set_name, questions in found_info.items():
            ws.cell(row=next_row, column=1, value=set_name)
            ws.cell(row=next_row, column=2, value=len(questions))
            ws.cell(row=next_row, column=3, value=", ".join(questions) if questions else "none found")
            _style_row(ws, next_row, 3, "info" if questions else "high")
            next_row += 1
    ws.freeze_panes = "A1"
    _autofit(ws)


def write_questionnaire_structure_sheet(wb, all_issues):
    ws = wb.create_sheet("Questionnaire Structure")
    ws.sheet_view.showGridLines = False
    row = 1

    # Box 1: legacy relevant checks
    rel_df = all_issues.filter(
        pl.col("issue_type").is_in(list(RELEVANT_ISSUE_TYPES))
    ).sort(["severity", "Q Name"])
    _sect(ws, row, "QUESTIONNAIRE STRUCTURE CHECKS  Skip logic references (relevant)", 8)
    _rel_start = row + 1
    row = _table(ws, _rel_start, rel_df, apply_view=False)
    _apply_inline_diff_for_issue(
        ws,
        _rel_start,
        rel_df,
        [(s, d) for s, d in COL_MAP if s in rel_df.columns],
        {"relevant_modified"},
    )

    # Box 2: Q Type integrity
    row += 1
    qtype_df = all_issues.filter(
        pl.col("issue_type") == "type_changed"
    ).sort(["severity", "Q Name", "field"])
    _sect(ws, row, "Q TYPE INTEGRITY ISSUES", 8)
    row = _table(ws, row + 1, qtype_df, apply_view=False)

    # Box 3: duplicate/token syntax checks
    row += 1
    struct_df = all_issues.filter(
        pl.col("issue_type").is_in([
            "duplicate_qname", "duplicate_choice_name",
            "kobo_ref_loose_syntax", "kobo_ref_missing_variable",
        ])
    ).sort(["severity", "Q Name", "field"])
    _sect(ws, row, "QUESTIONNAIRE STRUCTURE CHECKS  Duplicates and KoBO references", 8)
    _table(ws, row + 1, struct_df, apply_view=False)

    # Avoid frozen panes in multi-box sheets (this was causing navigation/display issues).
    ws.freeze_panes = "A1"
    _autofit(ws)


def write_replacement_issues_sheet(wb, all_issues):
    ws = wb.create_sheet("Replacement Issues")
    ws.sheet_view.showGridLines = False
    df = all_issues.filter(pl.col("issue_type").is_in(list(REPLACEMENT_ISSUE_TYPES))).sort(["severity", "Q Name", "field"])
    _sect(ws, 1, "REPLACEMENT ISSUES  Placeholder and Additional information checks", 8)
    next_row = _table(ws, 2, df, apply_view=False)

    rp_df = (
        all_issues
        .filter(pl.col("issue_type").cast(pl.Utf8).str.starts_with("additional_information_replacement_change ("))
        .sort(["severity", "issue_type", "Q Name"])
    )
    next_row += 1
    _sect(ws, next_row, "ADDITIONAL INFORMATION REPLACEMENT CHANGES  (previous_round informational)", 8)
    rp_start = next_row + 1
    _table(ws, rp_start, rp_df, apply_view=False)
    _apply_inline_diff_for_issue(
        ws,
        rp_start,
        rp_df,
        [(s, d) for s, d in COL_MAP if s in rp_df.columns],
        {"additional_information_replacement_change*"},
    )

    ws.freeze_panes = "A1"
    _autofit(ws)


def write_relevant_sheet(wb, all_issues):
    ws = wb.create_sheet("Relevant Changes")
    ws.sheet_view.showGridLines = False
    df = all_issues.filter(pl.col("issue_type").is_in(list(RELEVANT_ISSUE_TYPES)))
    _sect(ws, 1, "RELEVANT CHANGES  KoBO skip-logic (relevant column)", 8)
    _table(ws, 2, df)
    _autofit(ws)


def write_question_changes_sheet(wb, all_issues):
    ws = wb.create_sheet("Question Changes")
    ws.sheet_view.showGridLines = False
    df = all_issues.filter(_question_change_expr())
    col_map = [c for c in COL_MAP if c[0] not in {"set_name", "list_name", "option_name"}]
    _sect(ws, 1, "QUESTION CHANGES  Presence, mandatory, label, hint, required, choice_filter, appearance, calculation, constraint, choices list", 8)
    _table(ws, 2, df, col_map=col_map)
    _apply_inline_diff_for_issue(ws, 2, df, [(s, d) for s, d in col_map if s in df.columns], {"label_mismatch", "constraint_modified", "hint_changed", "choice_filter_modified"})
    _autofit(ws)


def write_option_changes_sheet(wb, all_issues):
    ws = wb.create_sheet("Choice Changes")
    ws.sheet_view.showGridLines = False
    df = all_issues.filter(pl.col("issue_type").is_in([
        "removed_choice", "added_choice", "choice_label_mismatch", "cluster_ea_choice_changes_summary",
    ]))
    col_map = [c for c in COL_MAP if c[0] != "set_name"]
    _sect(ws, 1, "CHOICE CHANGES  Answer choices for questions in both files", 8)
    _table(ws, 2, df, col_map=col_map)
    _apply_inline_diff_for_issue(ws, 2, df, [(s, d) for s, d in col_map if s in df.columns], {"choice_label_mismatch"})
    _autofit(ws)


def export_report(all_issues, result_file, found_info=None,
                  rules=None, critical_issues=None, count_issues=None,
                  replacement_status=None):
    wb = Workbook()
    wb.remove(wb.active)
    write_summary_sheet(wb, all_issues, rules=rules,
                        critical_issues=critical_issues, count_issues=count_issues,
                        found_info=found_info, replacement_status=replacement_status)
    write_critical_sets_sheet(wb, all_issues, found_info=found_info)
    write_questionnaire_structure_sheet(wb, all_issues)
    write_replacement_issues_sheet(wb, all_issues)
    write_question_changes_sheet(wb, all_issues)
    write_option_changes_sheet(wb, all_issues)
    wb.save(result_file)
    print(f"Report saved: {result_file}")
    print(f"Sheets: {wb.sheetnames}")


# -- run export ----------------------------------------------------------------
from pathlib import Path as _P
_q_path = _P(run["questionnaire_path"])
_out_dir = _P(run.get("output_dir") or _q_path.parent)
_out_dir.mkdir(parents=True, exist_ok=True)
_round_kobo = _extract_round_token_kobo(_q_path.name)
_rtag_kobo = (f"_{_round_kobo}" if _round_kobo else "")
_dtag_kobo = _time.strftime("%Y%m%d")
_lang_kobo = str(cfg.get("language", "")).lower() if 'cfg' in dir() else "xx"
_iso3_kobo = str(cfg.get("iso3", "")).upper() if 'cfg' in dir() else "XXX"
_report_file = str(_out_dir / f"report_kobo_{_lang_kobo}_{_iso3_kobo}{_rtag_kobo}_{_dtag_kobo}.xlsx")
export_report(all_issues, _report_file, found_info=_found_info,
              rules=rules, critical_issues=critical_issues, count_issues=count_issues,
              replacement_status=globals().get("_replacement_status"))













# ======================================================================
# SECTION: ## Step 12 -- Validated questionnaire output
_step("Producing validated questionnaire")
# ======================================================================
# --- CODE CELL 15 ---
import urllib.request as _urlreq
import json as _json
import re as _re
from shutil import copy2 as _copy2
from datetime import date as _date

# Special answer codes appended to each crop list
_CROP_SPECIALS = {
    "crop" : [("777", "No crop production"), ("888", "Don't know"), ("999", "Refused")],
    "crop2": [("666", "No other crop"),       ("888", "Don't know"), ("999", "Refused")],
    "crop3": [("666", "No other crop"),       ("888", "Don't know"), ("999", "Refused")],
}


def _read_crop_choices(src_path: str, language: str) -> tuple[dict[str, list[tuple]], dict]:
    """
    Read 'Crop list' sheet -> {list_name: [(code, label), ...]} for crop/crop2/crop3.
    Also returns crop-selection diagnostics for reporting (expected exactly 10 selected).
    """
    status = {
        "expected_selected": 10,
        "selected_count": 0,
        "candidate_count": 0,
        "status": "FAIL",
        "severity": "high",
        "message": "Crop list sheet not found",
    }

    wb = openpyxl.load_workbook(src_path, data_only=True, read_only=True)
    if "Crop list" not in wb.sheetnames:
        wb.close()
        return {}, status
    rows = list(wb["Crop list"].iter_rows(values_only=True))
    wb.close()

    # Layout: row 1 = title, row 2 = empty/sub-title, row 3 = header  (skiprows=2)
    if len(rows) < 3:
        status["message"] = "Crop list sheet has no crop rows"
        return {}, status

    headers  = [str(h).strip() if h is not None else "" for h in rows[2]]
    sel_idx  = next((i for i, h in enumerate(headers) if "Select top" in h), None)
    code_idx = next((i for i, h in enumerate(headers) if "Dataset code" in h), None)
    if language == "fr":
        lbl_idx = next((i for i, h in enumerate(headers) if "Label" in h and "FR" in h), None)
    else:
        lbl_idx = next((i for i, h in enumerate(headers) if "Label" in h and "EN" in h), None)

    if code_idx is None or lbl_idx is None:
        status["message"] = "Crop list headers missing Dataset code / language label"
        return {}, status

    selected, unselected = [], []
    for row in rows[3:]:
        code = row[code_idx] if code_idx < len(row) else None
        lbl  = row[lbl_idx]  if lbl_idx  < len(row) else None
        sel  = row[sel_idx]  if sel_idx is not None and sel_idx < len(row) else None
        if code is None or str(code).strip() in ("", "nan"):
            continue
        entry = (str(code).strip(), str(lbl or "").strip())
        (selected if sel is not None and str(sel).strip() else unselected).append(entry)

    selected.sort(key=lambda x: x[0])
    unselected.sort(key=lambda x: x[0])
    combined = selected + unselected

    selected_count = len(selected)
    candidate_count = len(combined)
    status["selected_count"] = selected_count
    status["candidate_count"] = candidate_count

    if selected_count == 10:
        status["status"] = "PASS"
        status["severity"] = "pass"
        status["message"] = f"10 crops selected (OK) out of {candidate_count} candidates"
    elif selected_count == 0:
        status["status"] = "FAIL"
        status["severity"] = "high"
        status["message"] = f"No crops selected (0/10) out of {candidate_count} candidates"
    elif selected_count < 10:
        status["status"] = "WARN"
        status["severity"] = "medium"
        status["message"] = f"Only {selected_count}/10 crops selected out of {candidate_count} candidates"
    else:
        status["status"] = "WARN"
        status["severity"] = "medium"
        status["message"] = f"{selected_count}/10 crops selected (more than expected) out of {candidate_count} candidates"

    return ({name: combined + specials for name, specials in _CROP_SPECIALS.items()}, status)


def _fetch_admin_choices(iso3: str) -> dict[str, list[tuple]]:
    """
    Fetch adm1 and adm2 from public FAO AGOL REST service (no credentials needed).
    Returns:
      admin1: [(adm1_pcode, adm1_name), ...]
      admin2: [(adm2_pcode, adm2_name, adm1_pcode), ...]  <- adm1_pcode for choice_filter (col 5)
    """
    _BASE = ("https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services"
             "/Administrative_Boundaries_Reference_(view_layer)/FeatureServer")

    def _get(layer, fields):
        url = (f"{_BASE}/{layer}/query"
               f"?where=adm0_ISO3+%3D+%27{iso3}%27"
               f"&outFields={fields}&returnGeometry=false&outSR=4326&f=json")
        with _urlreq.urlopen(url, timeout=30) as r:
            return _json.loads(r.read().decode())["features"]

    adm1 = sorted(
        [(f["attributes"]["adm1_pcode"], f["attributes"]["adm1_name"])
         for f in _get(1, "adm1_name,adm1_pcode")],
        key=lambda x: x[0],
    )
    adm2 = sorted(
        [(f["attributes"]["adm2_pcode"], f["attributes"]["adm2_name"], f["attributes"]["adm1_pcode"])
         for f in _get(0, "adm2_name,adm2_pcode,adm1_pcode")],
        key=lambda x: x[0],
    )
    print(f"  AGOL  adm1: {len(adm1)} rows   adm2: {len(adm2)} rows")
    return {"admin1": adm1, "admin2": adm2}


def _rebuild_choices_sheet(wb, country_rows: dict[str, list[tuple]]) -> None:
    """
    Replace all rows belonging to country_rows list_names with fresh data.
    Non-country rows and empty separator rows are preserved in their original order.
    New country blocks are appended at the end, each preceded by an empty separator row.
    admin2 rows also populate col 5 (adm1_pcode) for choice_filter cascading.
    """
    ws       = wb["choices"]
    all_rows = list(ws.iter_rows(values_only=True))
    n_cols   = max((len(r) for r in all_rows), default=5)
    skip     = set(country_rows.keys())

    # Keep header + non-country rows + empty separator rows (r[0] is None)
    kept = [all_rows[0]] + [
        r for r in all_rows[1:]
        if r[0] is None or r[0] not in skip
    ]

    # Build new country blocks, each preceded by an empty separator row
    new_rows = []
    for list_name, entries in country_rows.items():
        new_rows.append([None] * n_cols)   # blank separator between blocks
        for entry in entries:
            row    = [None] * n_cols
            row[0] = list_name
            row[1] = entry[0]   # code / pcode
            row[2] = entry[1]   # label / name
            if len(entry) > 2 and n_cols >= 5:
                row[4] = entry[2]   # adm1_pcode in col 5 (choice_filter)
            new_rows.append(row)

    ws.delete_rows(1, ws.max_row)
    for r in kept + new_rows:
        ws.append(list(r))


def _apply_replacements_to_wb(wb, replacement_pairs: dict) -> None:
    """Apply #placeholder# -> value to every string cell in survey and choices sheets."""
    if not replacement_pairs:
        return
    for sheet_name in ("survey", "choices"):
        if sheet_name not in wb.sheetnames:
            continue
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "#" in cell.value:
                    v = cell.value
                    for key, val in replacement_pairs.items():
                        if key in v:
                            v = v.replace(key, str(val))
                    cell.value = v


def _count_choices_list_rows(wb, list_name: str) -> int:
    """Count rows in choices sheet belonging to a given list_name."""
    if "choices" not in wb.sheetnames:
        return 0
    ws = wb["choices"]
    return sum(
        1
        for r in ws.iter_rows(min_row=2, values_only=True)
        if str((r[0] if r and len(r) > 0 else "") or "").strip() == list_name
    )




def _scan_unresolved_placeholders(wb, sheet_name: str, limit: int = 5) -> tuple[int, list[str]]:
    """Return (count, examples) of remaining #placeholder# tokens in a sheet."""
    if sheet_name not in wb.sheetnames:
        return 0, []
    patt = _re.compile(r"#[^#]+#")
    ws = wb[sheet_name]
    count = 0
    examples: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and patt.search(v):
                count += 1
                if len(examples) < limit:
                    examples.append(f"{cell.coordinate}: {v[:90]}")
    return count, examples
def produce_validated_questionnaire(
    cfg             : dict,
    reference_survey: pl.DataFrame,
    replacement_pairs: dict,
) -> tuple[str, dict]:
    """
    Build validated_questionnaire_kobo_<lang>_<iso3>_<date>.xlsx:
      1. Copy country questionnaire to output_dir
      2. Inject crop choices from 'Crop list' sheet
      3. Inject admin choices from AGOL
      4. Restore #placeholder# labels from template (then replace in step 5)
      5. Apply all #placeholder# replacements from Additional information sheet
    Returns (saved_file_path, replacement_status).
    """
    src  = cfg["questionnaire_path"]
    iso3 = cfg["iso3"]
    lang = cfg["language"]
    out  = Path(cfg.get("output_dir") or str(Path(src).parent))
    _round_kobo = _extract_round_token_kobo(Path(src).name)
    _rt_kobo = (f"_{_round_kobo}" if _round_kobo else "")
    dest = str(out / f"validated_questionnaire_kobo_{lang}_{iso3.upper()}{_rt_kobo}_{_date.today():%Y%m%d}.xlsx")

    replacement_status = {"rows": []}

    def _add_row(check: str, status: str, details: str, severity: str):
        replacement_status["rows"].append({
            "check": check,
            "status": status,
            "details": details,
            "severity": severity,
        })

    print("Building validated questionnaire ...")
    print(f"  Source : {Path(src).name}")
    print(f"  Output : {dest}")

    _copy2(src, dest)
    wb = openpyxl.load_workbook(dest)

    # -- 2. Crop choices --------------------------------------------------------
    crop_rows, crop_sel_status = _read_crop_choices(src, lang)
    print(f"  Crop selection: {crop_sel_status['message']}")
    if crop_sel_status["status"] == "PASS":
        _add_row("Crop selection (top 10)", "PASS", crop_sel_status["message"], "pass")
    elif crop_sel_status["status"] == "WARN":
        _add_row("Crop selection (top 10)", "WARN", crop_sel_status["message"], "medium")
    else:
        _add_row("Crop selection (top 10)", "FAIL", crop_sel_status["message"], "high")

    if crop_rows:
        print(f"  Crops  : {sum(len(v) for v in crop_rows.values())} entries  "
              f"({', '.join(crop_rows)})")

    # -- 3. Admin choices from AGOL ---------------------------------------------
    admin_rows: dict = {}
    agol_error = None
    try:
        admin_rows = _fetch_admin_choices(iso3)
    except Exception as e:
        agol_error = str(e)
        print(f"  Warning: AGOL fetch failed -- admin choices not updated  ({e})")

    adm1_n = len(admin_rows.get("admin1", []))
    adm2_n = len(admin_rows.get("admin2", []))
    if agol_error:
        _add_row("AGOL fetch", "FAIL", f"Fetch failed: {agol_error}", "high")
    elif adm1_n > 0 and adm2_n > 0:
        _add_row("AGOL fetch", "PASS", f"admin1={adm1_n}, admin2={adm2_n}", "pass")
    else:
        _add_row("AGOL fetch", "WARN", f"admin1={adm1_n}, admin2={adm2_n}", "medium")

    country_rows = {**crop_rows, **admin_rows}
    if country_rows:
        _rebuild_choices_sheet(wb, country_rows)

    # Verify choices replacement counts in workbook
    for list_name in ["crop", "crop2", "crop3"]:
        expected = len(crop_rows.get(list_name, []))
        actual = _count_choices_list_rows(wb, list_name)
        if expected > 0 and actual == expected:
            _add_row(f"Choices replaced: {list_name}", "PASS", f"{actual} rows written", "pass")
        elif expected == 0:
            _add_row(f"Choices replaced: {list_name}", "FAIL", "No rows prepared from Crop list", "high")
        else:
            _add_row(f"Choices replaced: {list_name}", "FAIL", f"Expected {expected}, wrote {actual}", "high")

    for list_name in ["admin1", "admin2"]:
        expected = len(admin_rows.get(list_name, []))
        actual = _count_choices_list_rows(wb, list_name)
        if expected > 0 and actual == expected:
            _add_row(f"Choices replaced: {list_name}", "PASS", f"{actual} rows written", "pass")
        elif agol_error:
            _add_row(f"Choices replaced: {list_name}", "FAIL", "Skipped because AGOL fetch failed", "high")
        elif expected == 0:
            _add_row(f"Choices replaced: {list_name}", "WARN", "No rows fetched from AGOL", "medium")
        else:
            _add_row(f"Choices replaced: {list_name}", "FAIL", f"Expected {expected}, wrote {actual}", "high")

    # -- 4. Restore template labels for #placeholder# questions -----------------
    #      Reset labels to canonical template form so step 5 substitutes correctly.
    if "survey" in wb.sheetnames and reference_survey.height > 0:
        ws_s      = wb["survey"]
        hdr       = [str(c.value or "").strip() for c in next(ws_s.iter_rows(min_row=1, max_row=1))]
        name_idx  = next((i for i, h in enumerate(hdr) if h == "name"), None)
        label_col = _find_label_col(hdr, lang)
        label_idx = hdr.index(label_col) if label_col and label_col in hdr else None

        if name_idx is not None and label_idx is not None:
            ref_ph = {
                r["Q Name"]: r["label"]
                for r in reference_survey
                    .filter(pl.col("label").str.contains(r"#[^#]+#"))
                    .select(["Q Name", "label"])
                    .iter_rows(named=True)
            }
            for row in ws_s.iter_rows(min_row=2):
                qn = str(row[name_idx].value or "").strip()
                if qn in ref_ph:
                    row[label_idx].value = ref_ph[qn]

    # -- 5. Apply #placeholder# -> actual values ---------------------------------
    _apply_replacements_to_wb(wb, replacement_pairs)

    # Explicit choices-level replacement integrity check
    _ch_unresolved, _ch_examples = _scan_unresolved_placeholders(wb, "choices")
    if _ch_unresolved == 0:
        _add_row(
            "Placeholder replacement in choices",
            "PASS",
            "No unresolved #placeholder# token remains in choices sheet",
            "pass",
        )
    else:
        _det = f"{_ch_unresolved} unresolved #placeholder# token(s) in choices"
        if _ch_examples:
            _det += " | examples: " + " ; ".join(_ch_examples)
        _add_row("Placeholder replacement in choices", "FAIL", _det, "high")

    wb.save(dest)
    print(f"  Saved  : {dest}")

    print("  Replacement checks:")
    for rr in replacement_status["rows"]:
        print(f"    - {rr['check']}: {rr['status']}  ({rr['details']})")

    return dest, replacement_status


# -- Run ------------------------------------------------------------------------
_validated_path, _replacement_status = produce_validated_questionnaire(cfg, reference_survey, replacement_pairs)

# Refresh report so Summary includes replacement status from Step 12
if all(k in globals() for k in ["export_report", "all_issues", "_report_file", "_found_info", "rules", "critical_issues", "count_issues"]):
    export_report(
        all_issues,
        _report_file,
        found_info=_found_info,
        rules=rules,
        critical_issues=critical_issues,
        count_issues=count_issues,
        replacement_status=_replacement_status,
    )
    print(f"  Report refreshed with replacement status: {_report_file}")
else:
    print("  Note: Replacement status prepared, but report context is not loaded in memory.")

_summary(
    all_issues  = all_issues if 'all_issues' in dir() else None,
    report_path = _report_file if '_report_file' in dir() else None,
    extra_paths = [("Validated", _validated_path)] if '_validated_path' in dir() and _validated_path else None,
)
