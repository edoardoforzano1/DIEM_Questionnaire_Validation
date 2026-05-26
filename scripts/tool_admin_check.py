from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from admin_tools_common import (
    detect_label_column,
    fetch_agol_admin_rows,
    merge_runtime_config,
    now_stamp,
    print_runtime_banner,
    resolve_source_workbook,
    safe_json,
)


def _norm(txt) -> str:
    return str(txt or "").strip().lower()


def _get_choices_context(wb, language: str):
    if "choices" not in wb.sheetnames:
        raise KeyError("Workbook has no 'choices' sheet.")
    ws = wb["choices"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    if not rows:
        raise ValueError("choices sheet is empty.")

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    list_col = idx.get("list_name")
    name_col = idx.get("name")
    if list_col is None or name_col is None:
        raise ValueError("choices sheet must contain 'list_name' and 'name' columns.")

    label_col = detect_label_column(headers, language)
    filter_col = None
    for c in ("my_filter_admin", "choice_filter", "filter"):
        if c in idx:
            filter_col = idx[c]
            break
    return ws, rows, headers, list_col, name_col, label_col, filter_col


def _extract_current_admin(rows, list_col, name_col, label_col, filter_col, include_admin3: bool):
    target_lists = {"admin1", "admin2"} | ({"admin3"} if include_admin3 else set())
    by_list: dict[str, list[dict]] = {k: [] for k in target_lists}
    duplicates: list[dict] = []
    blank_codes: list[dict] = []
    seen_codes: dict[str, set[str]] = {k: set() for k in target_lists}

    for excel_row_idx, row in enumerate(rows[1:], start=2):
        list_name = str(row[list_col] if list_col < len(row) else "").strip()
        if list_name not in target_lists:
            continue
        code = str(row[name_col] if name_col < len(row) else "").strip()
        label = str(row[label_col] if (label_col is not None and label_col < len(row)) else "").strip()
        parent = str(row[filter_col] if (filter_col is not None and filter_col < len(row)) else "").strip()

        if not code:
            blank_codes.append(
                {
                    "issue_type": "blank_code",
                    "list_name": list_name,
                    "code": "",
                    "current_label": label,
                    "agol_label": "",
                    "current_parent": parent,
                    "agol_parent": "",
                    "source_row": excel_row_idx,
                    "severity": "high",
                    "details": "Admin row has blank name/code in choices sheet",
                }
            )
            continue

        if code in seen_codes[list_name]:
            duplicates.append(
                {
                    "issue_type": "duplicate_code",
                    "list_name": list_name,
                    "code": code,
                    "current_label": label,
                    "agol_label": "",
                    "current_parent": parent,
                    "agol_parent": "",
                    "source_row": excel_row_idx,
                    "severity": "high",
                    "details": "Duplicate code in questionnaire admin list",
                }
            )
            continue

        seen_codes[list_name].add(code)
        by_list[list_name].append(
            {
                "list_name": list_name,
                "code": code,
                "label": label,
                "parent": parent,
                "excel_row": excel_row_idx,
            }
        )
    return by_list, duplicates + blank_codes


def _to_map(entries: list[dict]) -> dict[str, dict]:
    return {str(e.get("code") or "").strip(): e for e in entries if str(e.get("code") or "").strip()}


def _build_issues(current_rows: dict[str, list[dict]], agol_rows: dict[str, list[tuple[str, str, str]]], include_admin3: bool):
    target_lists = ["admin1", "admin2"] + (["admin3"] if include_admin3 else [])
    issues: list[dict] = []

    for list_name in target_lists:
        current_map = _to_map(current_rows.get(list_name, []))
        agol_map = {c: {"code": c, "label": l, "parent": p} for c, l, p in agol_rows.get(list_name, []) if c}

        for code, a in agol_map.items():
            if code not in current_map:
                issues.append(
                    {
                        "issue_type": "missing_in_questionnaire",
                        "list_name": list_name,
                        "code": code,
                        "current_label": "",
                        "agol_label": a["label"],
                        "current_parent": "",
                        "agol_parent": a["parent"],
                        "source_row": "",
                        "severity": "medium",
                        "details": "Code exists in AGOL but not in questionnaire",
                    }
                )

        for code, c in current_map.items():
            if code not in agol_map:
                issues.append(
                    {
                        "issue_type": "extra_in_questionnaire",
                        "list_name": list_name,
                        "code": code,
                        "current_label": c.get("label", ""),
                        "agol_label": "",
                        "current_parent": c.get("parent", ""),
                        "agol_parent": "",
                        "source_row": c.get("excel_row", ""),
                        "severity": "medium",
                        "details": "Code exists in questionnaire but not in AGOL",
                    }
                )
                continue

            a = agol_map[code]
            if _norm(c.get("label")) != _norm(a.get("label")):
                issues.append(
                    {
                        "issue_type": "label_mismatch",
                        "list_name": list_name,
                        "code": code,
                        "current_label": c.get("label", ""),
                        "agol_label": a.get("label", ""),
                        "current_parent": c.get("parent", ""),
                        "agol_parent": a.get("parent", ""),
                        "source_row": c.get("excel_row", ""),
                        "severity": "low",
                        "details": "Same code found, label text differs",
                    }
                )

            if list_name in {"admin2", "admin3"} and _norm(c.get("parent")) != _norm(a.get("parent")):
                issues.append(
                    {
                        "issue_type": "parent_mismatch",
                        "list_name": list_name,
                        "code": code,
                        "current_label": c.get("label", ""),
                        "agol_label": a.get("label", ""),
                        "current_parent": c.get("parent", ""),
                        "agol_parent": a.get("parent", ""),
                        "source_row": c.get("excel_row", ""),
                        "severity": "medium",
                        "details": "Same code found, parent filter/admin code differs",
                    }
                )
    return issues


def _write_check_report(out_xlsx: Path, issues: list[dict], summary_rows: list[dict], run_meta: dict) -> Path:
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_det = wb.create_sheet("Details")

    ws_sum.append(["field", "value"])
    for k, v in run_meta.items():
        ws_sum.append([k, safe_json(v)])
    ws_sum.append([])
    ws_sum.append(["list_name", "questionnaire_count", "agol_count", "missing", "extra", "label_mismatch", "parent_mismatch"])
    for r in summary_rows:
        ws_sum.append([
            r["list_name"],
            r["questionnaire_count"],
            r["agol_count"],
            r["missing"],
            r["extra"],
            r["label_mismatch"],
            r["parent_mismatch"],
        ])

    headers = [
        "issue_type",
        "severity",
        "list_name",
        "code",
        "current_label",
        "agol_label",
        "current_parent",
        "agol_parent",
        "source_row",
        "details",
    ]
    ws_det.append(headers)
    for row in issues:
        ws_det.append([row.get(h, "") for h in headers])

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return out_xlsx


def run(cfg: dict) -> int:
    print_runtime_banner("tool_admin_check", cfg)
    source = resolve_source_workbook(cfg)
    print(f"  source      : {source}")

    wb = openpyxl.load_workbook(source, data_only=True)
    _, rows, _, list_col, name_col, label_col, filter_col = _get_choices_context(wb, cfg.get("language", "en"))

    include_admin3 = bool(cfg.get("include_admin3"))
    current_rows, basic_issues = _extract_current_admin(
        rows, list_col, name_col, label_col, filter_col, include_admin3=include_admin3
    )

    agol_rows = fetch_agol_admin_rows(str(cfg.get("iso3") or ""), include_admin3=include_admin3)
    core_issues = _build_issues(current_rows, agol_rows, include_admin3=include_admin3)
    all_issues = basic_issues + core_issues

    target_lists = ["admin1", "admin2"] + (["admin3"] if include_admin3 else [])
    counts = defaultdict(lambda: defaultdict(int))
    for it in all_issues:
        counts[it["list_name"]][it["issue_type"]] += 1

    summary_rows = []
    for ln in target_lists:
        summary_rows.append(
            {
                "list_name": ln,
                "questionnaire_count": len(current_rows.get(ln, [])),
                "agol_count": len(agol_rows.get(ln, [])),
                "missing": counts[ln]["missing_in_questionnaire"],
                "extra": counts[ln]["extra_in_questionnaire"],
                "label_mismatch": counts[ln]["label_mismatch"],
                "parent_mismatch": counts[ln]["parent_mismatch"],
            }
        )

    out_dir = Path(str(cfg.get("output_dir")))
    out_xlsx = out_dir / f"admin_check_report_{Path(source).stem}_{now_stamp()}.xlsx"
    out_csv = out_dir / f"admin_check_report_{Path(source).stem}_{now_stamp()}.csv"
    _write_check_report(
        out_xlsx=out_xlsx,
        issues=all_issues,
        summary_rows=summary_rows,
        run_meta={
            "tool": "tool_admin_check",
            "source_file": str(source),
            "iso3": cfg.get("iso3"),
            "language": cfg.get("language"),
            "include_admin3": include_admin3,
            "issues_total": len(all_issues),
        },
    )

    out_dir.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        headers = [
            "issue_type",
            "severity",
            "list_name",
            "code",
            "current_label",
            "agol_label",
            "current_parent",
            "agol_parent",
            "source_row",
            "details",
        ]
        import csv

        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for row in all_issues:
            w.writerow({h: row.get(h, "") for h in headers})

    print(f"  issues      : {len(all_issues)}")
    for r in summary_rows:
        print(
            "  "
            + f"{r['list_name']}: q={r['questionnaire_count']} agol={r['agol_count']} "
            + f"missing={r['missing']} extra={r['extra']} "
            + f"label={r['label_mismatch']} parent={r['parent_mismatch']}"
        )
    print(f"  report_xlsx : {out_xlsx}")
    print(f"  report_csv  : {out_csv}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Standalone admin mismatch diagnosis against AGOL."
    )
    parser.add_argument("--config", default="", help="Optional path to admin_tools_config.yaml")
    parser.add_argument("--table", dest="config", help=argparse.SUPPRESS)
    parser.add_argument("--source-file", default="", help="Optional explicit source workbook path override")
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parents[1]
    cfg_path = Path(args.config).expanduser().resolve() if args.config else None
    cfg = merge_runtime_config(repo_root, explicit_cfg=cfg_path)
    if args.source_file:
        cfg["source_mode"] = "custom"
        cfg["source_file"] = args.source_file

    try:
        return run(cfg)
    except Exception as e:
        print(f"ERROR: {e}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
