"""
Generate documentation screenshots from real report Excel files.
Renders each sheet as a styled PNG preserving cell colors from the workbook.
Output goes to docs/assets/images/reports/ with the filenames expected by the docs.
"""
import os
import sys
import pathlib
import openpyxl
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch
import numpy as np

REPO_ROOT = pathlib.Path(__file__).parent.parent
OUTPUT_DIR = REPO_ROOT / "docs" / "assets" / "images" / "reports"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Map (tool, sheet_keyword) -> output filename stem
SHEET_MAP = {
    # GeoPoll
    ("geopoll", "summary"):             "geopoll-summary",
    ("geopoll", "critical"):            "geopoll-critical-sets",
    ("geopoll", "structure"):           "geopoll-structure",
    ("geopoll", "replacement"):         "geopoll-replacement-issues",
    ("geopoll", "question changes"):    "geopoll-question-changes",
    ("geopoll", "option changes"):      "geopoll-option-changes",
    # KoBo
    ("kobo", "summary"):                "kobo-summary",
    ("kobo", "critical"):               "kobo-critical-sets",
    ("kobo", "structure"):              "kobo-structure",
    ("kobo", "replacement"):            "kobo-replacement-issues",
    ("kobo", "question changes"):       "kobo-question-changes",
    ("kobo", "choice changes"):         "kobo-choice-changes",
    ("kobo", "validated"):              "kobo-validated-output",
}

# Fallback fill colors for severity keywords found in cell values
SEV_COLORS = {
    "high":   "#FFB3B3",
    "medium": "#FFD9A0",
    "info":   "#B3D9FF",
    "pass":   "#B3F0C0",
}

MAX_ROWS = 30
MAX_COLS = 12


def hex_to_rgb01(hex_color: str):
    h = hex_color.lstrip("#")
    if len(h) == 6:
        return tuple(int(h[i:i+2], 16) / 255 for i in (0, 2, 4))
    return (1.0, 1.0, 1.0)


def argb_to_hex(argb: str) -> str:
    """Convert openpyxl ARGB string (8-char) to #RRGGBB."""
    if argb and len(argb) == 8:
        return "#" + argb[2:]
    if argb and len(argb) == 6:
        return "#" + argb
    return "#FFFFFF"


def get_cell_bg(cell) -> str:
    try:
        fill = cell.fill
        if fill and fill.fill_type not in (None, "none"):
            fg = fill.fgColor
            if fg:
                if fg.type == "rgb" and fg.rgb and fg.rgb != "00000000":
                    return argb_to_hex(fg.rgb)
                if fg.type == "theme":
                    return "#F0F0F0"
    except Exception:
        pass
    return "#FFFFFF"


def get_cell_font_bold(cell) -> bool:
    try:
        return cell.font and cell.font.bold
    except Exception:
        return False


def render_sheet_to_png(ws, out_path: pathlib.Path, title: str):
    # Collect data
    rows_data = []
    col_widths = []

    all_rows = list(ws.iter_rows())
    # Find last meaningful row/col
    data_rows = all_rows[:MAX_ROWS]
    n_rows = len(data_rows)
    n_cols = min(ws.max_column or 1, MAX_COLS)

    for row in data_rows:
        row_cells = []
        for ci in range(n_cols):
            cell = row[ci] if ci < len(row) else None
            val = str(cell.value) if cell and cell.value is not None else ""
            # Truncate long text
            if len(val) > 55:
                val = val[:52] + "..."
            bg = get_cell_bg(cell) if cell else "#FFFFFF"
            bold = get_cell_font_bold(cell) if cell else False
            row_cells.append({"val": val, "bg": bg, "bold": bold})
        rows_data.append(row_cells)

    # Compute column widths based on content
    col_widths = []
    for ci in range(n_cols):
        max_len = 0
        for row in rows_data:
            if ci < len(row):
                max_len = max(max_len, len(row[ci]["val"]))
        col_widths.append(max(max_len * 0.11 + 0.2, 0.8))

    total_width = sum(col_widths)
    total_height = n_rows * 0.38 + 0.6

    fig_w = min(max(total_width, 10), 22)
    fig_h = min(max(total_height, 3), 14)
    scale = fig_w / total_width if total_width > 0 else 1.0
    col_widths_scaled = [w * scale for w in col_widths]

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(0, fig_w)
    ax.set_ylim(0, fig_h)
    ax.axis("off")

    row_h = (fig_h - 0.5) / max(n_rows, 1)
    y_start = fig_h - 0.45

    for ri, row in enumerate(rows_data):
        x = 0
        y = y_start - ri * row_h
        for ci, cell_data in enumerate(row):
            if ci >= len(col_widths_scaled):
                break
            cw = col_widths_scaled[ci]
            bg_color = hex_to_rgb01(cell_data["bg"])
            rect = plt.Rectangle((x, y - row_h), cw, row_h,
                                  facecolor=bg_color, edgecolor="#CCCCCC",
                                  linewidth=0.4)
            ax.add_patch(rect)

            text_color = "black"
            fontsize = 6.5 if n_cols > 8 else 7.5
            fw = "bold" if cell_data["bold"] else "normal"
            ax.text(x + 0.04, y - row_h / 2, cell_data["val"],
                    va="center", ha="left", fontsize=fontsize,
                    color=text_color, fontweight=fw,
                    clip_on=True,
                    transform=ax.transData)
            x += cw

    # Title bar
    ax.text(fig_w / 2, fig_h - 0.18, title,
            ha="center", va="center", fontsize=9, fontweight="bold",
            color="#173b6d")

    fig.patch.set_facecolor("#F7FBFE")
    plt.tight_layout(pad=0.1)
    plt.savefig(out_path, dpi=130, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"  Saved: {out_path.name}")


def match_sheet(tool: str, sheet_name: str) -> str | None:
    sl = sheet_name.lower()
    for (t, kw), stem in SHEET_MAP.items():
        if t == tool and kw in sl:
            return stem
    return None


def process_report(xlsx_path: pathlib.Path, tool: str):
    print(f"\nProcessing: {xlsx_path.name}")
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    except Exception as e:
        print(f"  ERROR loading: {e}")
        return

    for sheet_name in wb.sheetnames:
        stem = match_sheet(tool, sheet_name)
        if stem is None:
            continue
        out_path = OUTPUT_DIR / f"{stem}.png"
        ws = wb[sheet_name]
        render_sheet_to_png(ws, out_path, f"{sheet_name}")

    wb.close()


def process_validated_questionnaire(xlsx_path: pathlib.Path):
    """Render the survey sheet from a validated questionnaire file."""
    print(f"\nProcessing validated output: {xlsx_path.name}")
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    except Exception as e:
        print(f"  ERROR loading: {e}")
        return
    sheet_name = "survey" if "survey" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    out_path = OUTPUT_DIR / "kobo-validated-output.png"
    render_sheet_to_png(ws, out_path, f"Validated Questionnaire — {sheet_name} sheet")
    wb.close()


def main():
    batch = REPO_ROOT / "batch_output"

    # GeoPoll: pick the first available report
    geopoll_reports = sorted(batch.rglob("report_geopoll_*.xlsx"))
    if geopoll_reports:
        geopoll_reports.sort(key=lambda p: p.stat().st_size, reverse=True)
        process_report(geopoll_reports[0], "geopoll")
    else:
        print("No GeoPoll reports found.")

    # KoBo report
    kobo_reports = sorted(batch.rglob("report_kobo_en_*.xlsx"))
    if not kobo_reports:
        kobo_reports = sorted(batch.rglob("report_kobo_*.xlsx"))
    if kobo_reports:
        kobo_reports.sort(key=lambda p: p.stat().st_size, reverse=True)
        process_report(kobo_reports[0], "kobo")
    else:
        print("No KoBo reports found.")

    # KoBo validated questionnaire output
    validated = sorted(batch.rglob("validated_questionnaire_kobo_en_*.xlsx"))
    if not validated:
        validated = sorted(batch.rglob("validated_questionnaire_kobo_*.xlsx"))
    if validated:
        validated.sort(key=lambda p: p.stat().st_size, reverse=True)
        process_validated_questionnaire(validated[0])
    else:
        print("No validated questionnaire found.")

    print("\nDone. Files written to:", OUTPUT_DIR)


if __name__ == "__main__":
    main()
