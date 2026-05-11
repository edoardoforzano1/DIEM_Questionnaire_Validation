# ---------------------------------------------------------------------------
# Terminal output helpers (injected during conversion from notebook)
# ---------------------------------------------------------------------------
import sys as _sys, time as _time, pathlib, warnings as _warnings, atexit as _atexit

_warnings.filterwarnings("ignore")

_STEP_N     = 0
_STEP_TOTAL = 0
_T_START    = 0.0
_BANNER_W   = 66
_REAL_STDOUT = _sys.stdout


class _NullOutput:
    """Swallows intermediate notebook prints."""
    def write(self, *a, **kw): pass
    def flush(self, *a, **kw): pass
    def fileno(self): raise OSError("not a real file")


_atexit.register(lambda: setattr(_sys, "stdout", _REAL_STDOUT))


def _banner(tool: str, language: str, iso3: str, mode: str) -> None:
    global _T_START
    _T_START = _time.perf_counter()
    _sys.stdout = _NullOutput()
    print("", file=_REAL_STDOUT)
    print("=" * _BANNER_W, file=_REAL_STDOUT)
    print("  DIEM Questionnaire Validation v2", file=_REAL_STDOUT)
    print(f"  Tool: {tool.upper()}  |  Language: {language.upper()}  |  Country: {iso3.upper()}", file=_REAL_STDOUT)
    print(f"  Reference mode: {mode}", file=_REAL_STDOUT)
    print("=" * _BANNER_W, file=_REAL_STDOUT)


def _info(label: str, value: str) -> None:
    print(f"  {label:<12} {value}", file=_REAL_STDOUT)


def _step(label: str, detail: str = "") -> None:
    global _STEP_N
    _STEP_N += 1
    tag  = f"[{_STEP_N}/{_STEP_TOTAL}]"
    pad  = "." * max(1, 44 - len(label))
    tail = f"  {detail}" if detail else ""
    print(f"  {tag} {label} {pad}{tail}", file=_REAL_STDOUT, flush=True)


def _summary(all_issues=None, report_path=None, extra_paths=None) -> None:
    _sys.stdout = _REAL_STDOUT
    print()
    print("-" * _BANNER_W)
    print("  VALIDATION SUMMARY")
    print("-" * _BANNER_W)
    try:
        import polars as _pl
        if all_issues is not None and all_issues.height > 0:
            sev = {r["severity"]: r["n"] for r in
                   all_issues.group_by("severity").agg(_pl.len().alias("n")).to_dicts()}
            high   = sev.get("high",   0)
            medium = sev.get("medium", 0)
            info   = sev.get("info",   0)
            print(f"  {'HIGH':<10} {high:>4}   {'Must fix before deployment' if high else 'None'}")
            print(f"  {'MEDIUM':<10} {medium:>4}   {'Review suggested' if medium else 'None'}")
            print(f"  {'INFO':<10} {info:>4}   Informational")
            print(f"  {'PASS':<10} {'yes' if high == 0 and medium == 0 else 'no':>4}")
        else:
            print("  No issues found — questionnaire passed all checks.")
    except Exception as _e:
        print(f"  (summary unavailable: {_e})")
    print()
    if report_path:
        print(f"  Report  : {report_path}")
    if extra_paths:
        for _lbl, _p in extra_paths:
            print(f"  {_lbl:<8}: {_p}")
    print(f"  Time    : {_time.perf_counter() - _T_START:.1f}s")
    print("=" * _BANNER_W)
    print()


# ---------------------------------------------------------------------------

# ======================================================================
_STEP_TOTAL = 10
# SECTION: ## GeoPoll Questionnaire Validator
# ======================================================================
# --- CODE CELL 1 ---
import re
import shutil
import tempfile
import uuid
import yaml
import polars as pl
import polars.selectors as cs
import openpyxl
from pathlib import Path
from dataclasses import dataclass
from typing import Literal, Optional




# ======================================================================
# SECTION: ## Step 1  Configuration
_step("Loading configuration")
# ======================================================================
# --- CODE CELL 2 ---
ReferenceMode = Literal["latest_template", "previous_round"]
Enumerator    = Literal["geopoll", "kobo"]
Language      = Literal["EN", "FR", "ES", "AR", "PT"]

@dataclass
class ValidationConfig:
    template_repo          : Path
    working_dir            : Path
    questionnaire_file     : str
    reference_mode         : ReferenceMode
    enumerator             : Enumerator
    language               : Language
    iso3                   : str
    previous_round_file    : Optional[str]  = None
    output_subfolder       : str            = "output"
    output_dir             : Optional[Path] = None    # base output folder; output goes to output_dir/output_subfolder
    treat_blank_as_no      : bool           = True   # treat blank Mandatory cell as "no"
    critical_sets_file     : Optional[Path] = None    # path to critical_sets.yaml; None = auto-detect

import yaml
from datetime import datetime


def _load_effective_config(cfg_path: Path) -> tuple[dict, Path | None, Path]:
    base_cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
    if not isinstance(base_cfg, dict):
        raise TypeError(f"Config must be a mapping in {cfg_path}")

    profile_name = str(base_cfg.get("config_profile") or "").strip()
    profile_path: Path | None = None
    effective_cfg = dict(base_cfg)

    if profile_name:
        profiles_dir = Path(base_cfg.get("config_profiles_dir") or (Path(base_cfg.get("output_dir") or base_cfg.get("working_dir") or "C:/Temp/questionnaire_validation") / "config_profiles"))
        candidate = Path(profile_name)
        if not candidate.is_absolute():
            candidate = profiles_dir / candidate

        if candidate.suffix.lower() not in {".yaml", ".yml"}:
            if candidate.with_suffix(".yaml").exists():
                candidate = candidate.with_suffix(".yaml")
            elif candidate.with_suffix(".yml").exists():
                candidate = candidate.with_suffix(".yml")

        if not candidate.exists():
            raise FileNotFoundError(
                f"Config profile not found: {candidate}\n"
                f"Set 'config_profile' to an existing file or leave it empty."
            )

        profile_cfg = yaml.safe_load(candidate.read_text(encoding="utf-8")) or {}
        if not isinstance(profile_cfg, dict):
            raise TypeError(f"Config profile must be a mapping in {candidate}")

        effective_cfg.update(profile_cfg)
        profile_path = candidate.resolve()

    active_cfg_path = profile_path or cfg_path.resolve()
    return effective_cfg, profile_path, active_cfg_path


# -- Load centralized config ---------------------------------------------------
# Edit validation_config.yaml -- never touch notebook code.
_cfg_base_path = Path(__file__).parent.parent / "configuration" / "validation_config.yaml"
_y, _cfg_profile_path, _cfg_active_path = _load_effective_config(_cfg_base_path)

cfg = ValidationConfig(
    template_repo         = Path(_y["templates_dir"]),
    working_dir           = Path(_y["working_dir"]),
    questionnaire_file    = _y["questionnaire_file"],
    reference_mode        = _y["reference_mode"],
    enumerator            = "geopoll",
    language              = str(_y["language"]).upper(),
    iso3                  = _y["iso3"],
    output_dir            = Path(_y.get("output_dir") or _y.get("working_dir") or "C:/Temp"),
    output_subfolder      = "geopoll_output",
    previous_round_file   = _y.get("previous_round_file"),
    critical_sets_file    = Path(_y.get("critical_sets_file")) if _y.get("critical_sets_file") else (_cfg_base_path.resolve().parent / "critical_sets.yaml"),
)

_banner("geopoll", cfg.language, cfg.iso3, cfg.reference_mode)
_info("Config", str(_cfg_active_path))
if _cfg_profile_path:
    _info("Base cfg", str(_cfg_base_path.resolve()))
_info("File", cfg.questionnaire_file)
_info("Output", str(cfg.output_dir / cfg.output_subfolder))


def _reference_scope_label() -> str:
    mode = str(getattr(globals().get("cfg", None), "reference_mode", "") or "").strip().lower()
    if mode == "previous_round":
        return "previous round"
    return "latest template"


def _not_in_reference_text(suffix: str = "") -> str:
    base = _reference_scope_label()
    if suffix:
        return f"(not in {base} {suffix})"
    return f"(not in {base})"


# ======================================================================
# SECTION: ## Step 2  Reference file resolution
_step("Resolving reference file")
# ======================================================================
# --- CODE CELL 3 ---
# Maps language code (cfg.language) to the canonical column name in the survey sheet
LANGUAGE_COLUMN: dict[str, str] = {
    "EN": "English",
    "FR": "French",
    "ES": "Spanish",
    "AR": "Arabic",
    "PT": "Portuguese",
}

# Extra aliases used to resolve non-standard headers (e.g. fr_French, sw_Swahili)
LANGUAGE_ALIASES: dict[str, list[str]] = {
    "EN": ["english", "eng", "en"],
    "FR": ["french", "francais", "franï¿½ais", "fr"],
    "ES": ["spanish", "espanol", "espaï¿½ol", "es"],
    "AR": ["arabic", "arabe", "ar"],
    "PT": ["portuguese", "portugues", "portuguï¿½s", "pt"],
}


def _normalize_header_name(name: str) -> str:
    s = str(name or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "", s)


def _header_tokens(name: str) -> list[str]:
    s = str(name or "").strip().lower()
    return [t for t in re.split(r"[^a-z0-9]+", s) if t]


def _is_language_candidate_header(name: str) -> bool:
    """Reject obvious non-language columns to avoid false alias matches."""
    n = _normalize_header_name(name)
    blocked = {
        "qname", "suggestedqname", "qtype", "mandatory", "excelrow", "sourcefile",
        "length", "len", "codes", "additionalnotes",
        "defaultskippatternsconditional", "specifyskippatternvariablefrombluetext",
        "specifyskippatternvariable",
    }
    return n not in blocked


def resolve_language_column(columns: list[str], language: str) -> str | None:
    """
    Resolves the best matching text column for a language, supporting variants
    such as 'fr_French' while remaining deterministic.
    """
    lang = str(language or "EN").upper()
    preferred = LANGUAGE_COLUMN.get(lang, "English")
    if preferred in columns:
        return preferred

    preferred_norm = _normalize_header_name(preferred)
    for c in columns:
        if _normalize_header_name(c) == preferred_norm:
            return c

    aliases = [_normalize_header_name(a) for a in LANGUAGE_ALIASES.get(lang, [])]
    ranked: list[tuple[int, int, str]] = []
    for c in columns:
        if not _is_language_candidate_header(c):
            continue
        c_norm = _normalize_header_name(c)
        c_toks = set(_header_tokens(c))
        score = 0

        if preferred_norm and preferred_norm in c_norm:
            score += 6

        for a in aliases:
            if not a:
                continue
            if len(a) <= 2:
                # short language codes (fr/es/ar/pt/en): require token/prefix match,
                # never generic substring match (prevents 'Codes' -> 'Spanish').
                c_low = str(c or "").strip().lower()
                if a in c_toks:
                    score += 5
                if c_low.startswith(a + "_") or c_low.startswith(a + "-"):
                    score += 4
            else:
                if a in c_norm:
                    score += 3
                if a in c_toks:
                    score += 2

        if score > 0:
            ranked.append((score, -len(c), c))

    if not ranked:
        return None
    ranked.sort(reverse=True)
    return ranked[0][2]


def detect_language_from_filename(path: Path | str) -> str | None:
    """Returns EN/FR/ES/AR/PT when filename contains a language token like _ar_."""
    name = Path(path).name.lower()
    m = re.search(r"_(en|fr|es|ar|pt)(?=_|\.|$)", name)
    return m.group(1).upper() if m else None


def choose_target_language(questionnaire_path: Path, questions_df: pl.DataFrame, preferred: str = "EN") -> tuple[str, str]:
    """
    Chooses the analysis language for text/option comparisons.
    Priority:
      1) language token in questionnaire filename (if non-EN and present in columns)
      2) config language (if non-EN and present in columns)
      3) best available non-EN column with most non-empty cells
      4) EN fallback
    """
    cols = questions_df.columns

    file_lang = detect_language_from_filename(questionnaire_path)
    if file_lang and file_lang != "EN" and resolve_language_column(cols, file_lang):
        return file_lang, "filename"

    pref = str(preferred or "EN").upper()
    if pref != "EN" and resolve_language_column(cols, pref):
        return pref, "config"

    scored: list[tuple[int, str]] = []
    for lang in ["AR", "FR", "ES", "PT"]:
        col = resolve_language_column(cols, lang)
        if not col:
            continue
        n_non_empty = (
            questions_df
            .select(pl.col(col).cast(pl.Utf8).fill_null("").str.strip_chars().alias("_v"))
            .filter(pl.col("_v") != "")
            .height
        )
        if n_non_empty > 0:
            scored.append((n_non_empty, lang))

    if scored:
        scored.sort(reverse=True)
        return scored[0][1], "auto_non_en"

    return "EN", "fallback_en"


def extract_date_from_name(path: Path) -> int:
    """
    Pulls the 8-digit date out of a filename like:
      household_questionnaire_geopoll_EN_template_20250708_ISO3.xlsx
                                                    ^^^^^^^^
    Returns 0 if no date is found, so the file sorts last.
    """
    m = re.search(r"template_(\d{8})", path.name)
    return int(m.group(1)) if m else 0


def find_latest_template(template_repo: Path, enumerator: str, language: str) -> Path:
    """
    Scans the template repo folder and returns the most recent template
    matching the given enumerator (geopoll/kobo) and language (EN/FR/etc.).
    Sorts by date in filename first, then by file modification time as fallback.
    """
    matches = [
        p for p in template_repo.glob("*.xlsx")
        if enumerator.lower() in p.name.lower()
        and f"_{language.lower()}_" in p.name.lower()
        and "template" in p.name.lower()
    ]
    if not matches:
        raise FileNotFoundError(
            f"No template found for enumerator='{enumerator}', language='{language}' in:\n  {template_repo}"
        )
    matches.sort(key=lambda p: (extract_date_from_name(p), p.stat().st_mtime), reverse=True)
    return matches[0]


def resolve_reference_file(cfg: ValidationConfig) -> Path:
    """Returns the path to the reference file based on reference_mode in the config."""
    if cfg.reference_mode == "latest_template":
        return find_latest_template(cfg.template_repo, cfg.enumerator, cfg.language)
    if cfg.reference_mode == "previous_round":
        if not cfg.previous_round_file:
            raise ValueError("previous_round_file must be set when reference_mode='previous_round'")
        return cfg.working_dir / cfg.previous_round_file
    raise ValueError(f"Unknown reference_mode: '{cfg.reference_mode}'")


def load_critical_sets(cfg: ValidationConfig) -> dict:
    """
    Loads validation rules from critical_sets.yaml.
    Search order:
      1. cfg.critical_sets_file (if set)
      2. critical_sets.yaml next to validation_config.yaml
      3. critical_sets.yaml in the parent folder of validation_config.yaml
      4. critical_sets.yaml in the current working directory
      5. critical_sets.yaml in a scripts/ subfolder of the current directory

    Returns a dict with keys: exact_sets, min_count_sets, crop_harvest.
    If the file is not found, structural validation is disabled (empty rules returned).
    """
    _cfg_dir = Path(__file__).parent.parent / "configuration"
    candidates = [
        Path(cfg.critical_sets_file) if cfg.critical_sets_file else None,
        _cfg_dir / "critical_sets.yaml",
        _cfg_dir.parent / "critical_sets.yaml",
        Path.cwd() / "critical_sets.yaml",
        Path.cwd() / "scripts" / "critical_sets.yaml",
    ]
    for path in candidates:
        if path is not None and path.exists():
            with open(path, "r", encoding="utf-8") as f:
                data = yaml.safe_load(f)
            print(f"Rules loaded  : {path}")
            return {
                "exact_sets"    : data.get("exact_sets", {}),
                "min_count_sets": data.get("min_count_sets", {}),
                "crop_harvest"  : data.get("crop_harvest", {}),
            }
    print("  critical_sets.yaml not found  structural validation disabled.")
    print(f"   Searched: {[str(p) for p in candidates if p is not None]}")
    return {"exact_sets": {}, "min_count_sets": {}, "crop_harvest": {}}


def prepare_run(cfg: ValidationConfig) -> dict:
    """
    Creates output directories and resolves all file paths.
    Returns a dict with the three paths the rest of the notebook needs.
    Prints a summary so you can confirm the right files were picked up.
    """
    cfg.working_dir.mkdir(parents=True, exist_ok=True)
    output_dir = (cfg.output_dir if cfg.output_dir else cfg.working_dir) / cfg.output_subfolder
    output_dir.mkdir(parents=True, exist_ok=True)

    questionnaire_path = cfg.working_dir / cfg.questionnaire_file
    reference_path     = resolve_reference_file(cfg)

    if not questionnaire_path.exists():
        raise FileNotFoundError(f"Questionnaire file not found:\n  {questionnaire_path}")
    if not reference_path.exists():
        raise FileNotFoundError(f"Reference file not found:\n  {reference_path}")

    _qname = questionnaire_path.name
    _round_tag = ""
    for _pat in [
        r"(?i)(?:^|[_\-\s])R(\d{1,3})(?=$|[_\-\s\.])",
        r"(?i)(?:^|[_\-\s])ROUND[_\-\s]*(\d{1,3})(?=$|[_\-\s\.])",
        r"(?i)R(?:OUND)?[_\-\s]*([0-9]{1,3})(?=(?:\.[A-Za-z0-9]+)?$)",
    ]:
        _m = re.search(_pat, _qname)
        if _m:
            _round_tag = f"_R{_m.group(1)}"
            break
    _rtag = _round_tag
    _dtag = _time.strftime("%Y%m%d")
    report_file = output_dir / f"report_geopoll_{cfg.language.lower()}_{cfg.iso3.upper()}{_rtag}_{_dtag}.xlsx"
    validated_questionnaire_file = output_dir / f"validated_questionnaire_geopoll_{cfg.language.lower()}_{cfg.iso3.upper()}{_rtag}_{_dtag}.xlsx"

    run = {
        "questionnaire_path" : questionnaire_path,
        "reference_path"     : reference_path,
        "report_file"        : report_file,
        "validated_questionnaire_file": validated_questionnaire_file,
    }

    print("Questionnaire :", questionnaire_path)
    print("Reference     :", reference_path)
    print("Report output :", report_file)
    print("Validated out :", validated_questionnaire_file)
    return run


def _write_config_snapshot_geopoll(
    cfg: ValidationConfig,
    run: dict,
    cfg_base_path: Path,
    cfg_active_path: Path,
    effective_cfg: dict,
) -> Path:
    logs_root = cfg.output_dir if cfg.output_dir else run["report_file"].parent.parent
    logs_dir = logs_root / "configuration"
    logs_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = logs_dir / f"config_geopoll_{cfg.language}_{cfg.iso3}_{ts}.yaml"

    payload = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "tool": cfg.enumerator,
        "active_config_file": str(cfg_active_path),
        "base_config_file": str(cfg_base_path.resolve()),
        "config_profile": str(effective_cfg.get("config_profile") or ""),
        "resolved_config": {
            "templates_dir": str(cfg.template_repo),
            "working_dir": str(cfg.working_dir),
            "questionnaire_file": cfg.questionnaire_file,
            "reference_mode": cfg.reference_mode,
            "language": cfg.language,
            "iso3": cfg.iso3,
            "output_dir": str(cfg.output_dir) if cfg.output_dir else "",
            "output_subfolder": cfg.output_subfolder,
            "previous_round_file": cfg.previous_round_file or "",
            "critical_sets_file": str(cfg.critical_sets_file) if cfg.critical_sets_file else "",
        },
        "resolved_run_paths": {k: str(v) for k, v in run.items()},
    }

    with out_file.open("w", encoding="utf-8") as f:
        yaml.safe_dump(payload, f, sort_keys=False, allow_unicode=True)
    return out_file


run   = prepare_run(cfg)
rules = load_critical_sets(cfg)

try:
    _config_snapshot_file = _write_config_snapshot_geopoll(
        cfg=cfg,
        run=run,
        cfg_base_path=_cfg_base_path,
        cfg_active_path=_cfg_active_path,
        effective_cfg=_y,
    )
    print(f"Config snapshot: {_config_snapshot_file}")
except Exception as _cfg_log_err:
    print(f"Warning: could not write config snapshot: {_cfg_log_err}")

# Initial language from config (template/reference selection).
# Actual comparison language is resolved after loading survey columns.
target_lang = cfg.language.upper()
print(f"Configured language : {target_lang}")




# ======================================================================
# SECTION: ## Step 3  Reading and normalising the survey sheet
_step("Reading survey sheets")
# ======================================================================
# --- CODE CELL 4 ---
CORE_COLUMNS = [
    "Q Name",
    "Suggested Qname",
    "English",
    "French",
    "Spanish",
    "Arabic",
    "Portuguese",
    "Q Type",
    "Mandatory",
    "Randomize",
    "Conditional",
    "Programming Instructions",
    "Core questions only",
    "Skip Pattern",
    "Codes",
    "Default skip patterns & conditional ",   # note: trailing space is in the real header
    "Specify skip pattern variable (from blue text)",
    "Additional notes",
    "excel_row",
    "source_file",
]


COLUMN_ALIASES: dict[str, list[str]] = {
    "Q Name": [
        "Q Name", "QName", "Qname", "Q name", "Q_Name", "Q-Name",
    ],
    "Q Type": [
        "Q Type", "QType", "Q type", "Qtype", "Question Type", "Question type",
    ],
    "Mandatory": [
        "Mandatory", "mandatory", "Mandatory ", "Required", "required",
    ],
    "Skip Pattern": [
        "Skip Pattern", "Skip pattern", "Skip Pattern ",
    ],
    "Randomize": [
        "Randomize", "Randomise", "Randomize ",
    ],
    "Conditional": [
        "Conditional", "Conditional ",
    ],
    "Programming Instructions": [
        "Programming Instructions", "Programming instructions", "Programming Instruction", "Programming Instructions ",
    ],
    "Core questions only": [
        "Core questions only", "Core question only", "Core questions only ",
    ],
    "Codes": [
        "Codes", "Codes ", "Code", "code",
    ],
    "Default skip patterns & conditional ": [
        "Default skip patterns & conditional", "Default skip patterns & conditional ",
        "Default skip patterns and conditional",
    ],
    "Specify skip pattern variable (from blue text)": [
        "Specify skip pattern variable (from blue text)",
        "Specify skip pattern variable",
    ],
}


def _resolve_column_by_aliases(columns: list[str], aliases: list[str]) -> str | None:
    alias_norm = {_normalize_header_name(a) for a in aliases}
    for c in columns:
        if _normalize_header_name(c) in alias_norm:
            return c
    return None


def _canonical_column_from_header(name: str) -> str | None:
    n = _normalize_header_name(name)
    for canonical, aliases in COLUMN_ALIASES.items():
        alias_norm = {_normalize_header_name(canonical)} | {_normalize_header_name(a) for a in aliases}
        if n in alias_norm:
            return canonical
    return None


def _as_text_cell(value):
    """
    Preserve cell content while forcing a stable string/None shape.
    This avoids Polars schema crashes on mixed typed columns (e.g. '#REF!' strings
    appearing in otherwise numeric-looking columns).
    """
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _load_workbook_readonly(path: Path, data_only: bool = True, read_only: bool = True):
    """
    Best-effort workbook loader:
    1) normal open
    2) if locked (PermissionError), try a temp copy and open the copy
    """
    p = Path(path)
    try:
        return openpyxl.load_workbook(p, data_only=data_only, read_only=read_only)
    except PermissionError as e:
        tmp_dir = Path(tempfile.gettempdir()) / "diem_qval_wb_cache"
        tmp_dir.mkdir(parents=True, exist_ok=True)
        tmp_path = tmp_dir / f"{uuid.uuid4().hex}_{p.name}"
        try:
            shutil.copy2(p, tmp_path)
            wb = openpyxl.load_workbook(tmp_path, data_only=data_only, read_only=read_only)
            setattr(wb, "_temp_copy_path", str(tmp_path))
            print(f"  Warning: workbook locked, reading temp copy: {p.name}")
            return wb
        except Exception as copy_err:
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass
            raise PermissionError(
                f"Cannot open workbook '{p}'. It may be locked by Excel/OneDrive sync. "
                "Close the workbook (or disable Office co-authoring sync on this file) and retry."
            ) from copy_err


def _close_workbook(wb) -> None:
    if wb is None:
        return
    tmp_copy = getattr(wb, "_temp_copy_path", "")
    try:
        wb.close()
    finally:
        if tmp_copy:
            try:
                tmp_path = Path(tmp_copy)
                if tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass


def read_survey_sheet(path: Path, sheet_name: str = "survey", header_row: int = 3, _wb=None) -> pl.DataFrame:
    """
    Reads the survey sheet from an xlsx file into a Polars DataFrame.

    header_row=3 because the GeoPoll template has two title rows above the
    real column headers.

    Two columns are added automatically:
      excel_row    the actual row number in the Excel file (for tracing mismatches)
      source_file  the filename (tells you which file a row came from)

    Pass _wb to reuse an already-opened openpyxl workbook (avoids re-opening
    the same file multiple times).
    """
    owns_wb = _wb is None
    wb = _wb or _load_workbook_readonly(path, data_only=True, read_only=True)
    try:
        ws = next(
            (wb[n] for n in wb.sheetnames if n.strip().lower() == sheet_name.lower()),
            None
        )
        if ws is None:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

        row_iter = ws.iter_rows(values_only=True)
        for _ in range(header_row - 1):
            next(row_iter, None)   # skip title rows safely

        raw_headers = next(row_iter, None)
        if raw_headers is None:
            raise ValueError(f"Header row {header_row} is missing in sheet '{sheet_name}' for file {path.name}")
        headers = [
            str(h).replace("\n", " ").strip() if h is not None else f"unnamed_{i}"
            for i, h in enumerate(raw_headers, 1)
        ]

        rows      = []
        excel_row = header_row + 1

        for values in row_iter:
            if all(v is None for v in values):
                excel_row += 1
                continue
            row_dict = {}
            for i in range(len(headers)):
                raw_val = values[i] if i < len(values) else None
                row_dict[headers[i]] = _as_text_cell(raw_val)
            row_dict["excel_row"]   = excel_row
            row_dict["source_file"] = Path(path).name
            rows.append(row_dict)
            excel_row += 1

        if rows:
            df = pl.DataFrame(rows, infer_schema_length=None)
        else:
            schema = {h: pl.Utf8 for h in headers}
            schema["excel_row"] = pl.Int64
            schema["source_file"] = pl.Utf8
            df = pl.DataFrame(schema=schema)

        # Canonicalize known alias columns before any strict presence checks.
        mapped_cols: list[tuple[str, str]] = []
        for canonical, aliases in COLUMN_ALIASES.items():
            if canonical in df.columns:
                continue
            matched = _resolve_column_by_aliases(df.columns, [canonical] + aliases)
            if matched and matched != canonical:
                df = df.with_columns(pl.col(matched).alias(canonical))
                mapped_cols.append((matched, canonical))
        if mapped_cols:
            print(f"  Header aliases mapped in {path.name}: {mapped_cols}")

        if "Q Name" not in df.columns:
            raise KeyError(f"'Q Name' column not found after reading the sheet. Available columns: {df.columns}")

        return (
            df.with_columns(pl.col("Q Name").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Q Name"))
            .filter(pl.col("Q Name") != "")
        )
    finally:
        if owns_wb:
            _close_workbook(wb)

def build_questions_df(df: pl.DataFrame) -> pl.DataFrame:
    """
    Selects and cleans the core validation columns.
    Warns (does not crash) if an expected column is missing  silent drops
    would make mismatches invisible.
    """
    # Canonicalise language columns first (e.g. fr_French -> French)
    # so downstream logic can rely on stable names.
    mapped: list[tuple[str, str]] = []
    for lang_code, canonical in LANGUAGE_COLUMN.items():
        if canonical in df.columns:
            continue
        matched = resolve_language_column(df.columns, lang_code)
        if matched and matched != canonical:
            df = df.with_columns(pl.col(matched).alias(canonical))
            mapped.append((matched, canonical))
    if mapped:
        print(f"  Language columns mapped: {mapped}")

    # Canonicalise known non-language column variants (especially skip columns).
    mapped_cols: list[tuple[str, str]] = []
    for canonical, aliases in COLUMN_ALIASES.items():
        if canonical in df.columns:
            continue
        matched = _resolve_column_by_aliases(df.columns, aliases)
        if matched and matched != canonical:
            df = df.with_columns(pl.col(matched).alias(canonical))
            mapped_cols.append((matched, canonical))
    if mapped_cols:
        print(f"  Columns mapped: {mapped_cols}")

    # Mandatory may be absent in some operational files; keep pipeline running
    # by synthesizing a blank column so risk logic still applies deterministically.
    if "Mandatory" not in df.columns:
        df = df.with_columns(pl.lit("").alias("Mandatory"))
        print("  Column synthesized: Mandatory (blank default)")

    missing = [
        c for c in CORE_COLUMNS
        if c not in df.columns and c not in ("excel_row", "source_file")
    ]
    if missing:
        print(f"  Expected columns not found (will be skipped): {missing}")

    available = [c for c in CORE_COLUMNS if c in df.columns]
    out = df.select(available)

    out = out.with_columns(
        cs.exclude("excel_row").cast(pl.Utf8).fill_null("").str.strip_chars()
    )
    return out





# ======================================================================
# SECTION: ## Step 4  Exploding answer options
_step("Parsing answer options")
# ======================================================================
# --- CODE CELL 5 ---
def _normalize_q_type(value: str) -> str:
    txt = str(value or "").strip().lower()
    txt = re.sub(r"[-_/]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt)
    return txt


def is_option_bearing_qtype(value: str) -> bool:
    """
    Robust Q Type classifier for option-bearing questions.
    Handles spacing/punctuation variants (e.g. Open Ended vs Open-Ended).
    """
    t = _normalize_q_type(value)
    if not t:
        return False
    if t == "single choice":
        return True
    if "select all that apply" in t:
        return True
    if "single choice" in t and "open ended" in t:
        return True
    return False

# Matches blocks like:
#   1) Label text
#   2) Next label
#
# (?m)     ^ matches the start of each line
# DOTALL   . also matches newlines, so multi-line labels are captured in full
OPTION_PATTERN = re.compile(r"(?m)^\s*(\d+)\)\s*(.*?)(?=^\s*\d+\)|\Z)", re.DOTALL)


def parse_options(text: str) -> list[tuple[int, str]]:
    """
    Parses numbered options from a question text cell.
    Returns a list of (option_code, option_label) tuples.

    Labels are whitespace-collapsed but otherwise kept raw 
    

    Example:
        input:  "What is your gender?\\n\\n1)Male\\n2)Female\\n3)DON'T want to answer"
        output: [(1, "Male"), (2, "Female"), (3, "DON'T want to answer")]
    """
    if not text or not isinstance(text, str):
        return []
    results = []
    for m in OPTION_PATTERN.finditer(text):
        code  = int(m.group(1))
        label = " ".join(m.group(2).split())   # collapse newlines/extra spaces
        if label:
            results.append((code, label))
    return results


def extract_question_stem(text: str) -> str:
    """
    Returns only the question/instruction part, removing numbered option blocks.
    Used for Question Changes so option text is handled only in Option Changes.
    """
    if not text or not isinstance(text, str):
        return ""
    m = OPTION_PATTERN.search(text)
    stem = text[:m.start()] if m else text
    return stem.strip()


def explode_options(questions_df: pl.DataFrame, text_col: str = "English") -> pl.DataFrame:
    """
    Turns the questions DataFrame (one row per question) into an options
    DataFrame (one row per answer option).

    Output columns:
        Q Name, Q Type, option_code (Int64), option_label (Utf8),
        excel_row, source_file
    """
    EMPTY_SCHEMA = {
        "Q Name"      : pl.Utf8,
        "Q Type"      : pl.Utf8,
        "option_code" : pl.Int64,
        "option_label": pl.Utf8,
        "excel_row"   : pl.Int64,
        "source_file" : pl.Utf8,
    }

    if text_col not in questions_df.columns:
        print(f"  Column '{text_col}' not found  returning empty options frame.")
        return pl.DataFrame(schema=EMPTY_SCHEMA)

    df = (
        questions_df.filter(
            pl.col("Q Type").map_elements(is_option_bearing_qtype, return_dtype=pl.Boolean)
        )
        if "Q Type" in questions_df.columns
        else questions_df
    )

    carry_cols = [c for c in ["Q Name", "Q Type", "excel_row", "source_file"] if c in df.columns]
    rows = []

    for row in df.select(carry_cols + [text_col]).iter_rows(named=True):
        text = row.get(text_col) or ""
        for code, label in parse_options(text):
            out_row = {c: row.get(c) for c in carry_cols}
            out_row["option_code"]  = code
            out_row["option_label"] = label
            rows.append(out_row)

    if not rows:
        return pl.DataFrame(schema=EMPTY_SCHEMA)

    return pl.DataFrame(rows).with_columns(
        pl.col("option_code").cast(pl.Int64),
        pl.col("option_label").cast(pl.Utf8),
    )


def parse_codes_cell(text: str) -> list[tuple[int, str]]:
    """
    Parses numbered entries from the survey 'Codes' column, e.g.:
      1)shock_none
      2)shock_flood
    Returns (code_num, code_token) rows.
    """
    if not text or not isinstance(text, str):
        return []
    out = []
    for m in OPTION_PATTERN.finditer(text):
        num = int(m.group(1))
        token = " ".join(m.group(2).split())
        if token:
            out.append((num, token))
    return out


def explode_codes(questions_df: pl.DataFrame, codes_col: str = "Codes") -> pl.DataFrame:
    """
    Turns the survey 'Codes' column into one row per numbered code token.
    Output columns: Q Name, code_num, code_token, excel_row, source_file
    """
    EMPTY_SCHEMA = {
        "Q Name": pl.Utf8,
        "code_num": pl.Int64,
        "code_token": pl.Utf8,
        "excel_row": pl.Int64,
        "source_file": pl.Utf8,
    }
    if codes_col not in questions_df.columns:
        return pl.DataFrame(schema=EMPTY_SCHEMA)

    carry_cols = [c for c in ["Q Name", "excel_row", "source_file"] if c in questions_df.columns]
    rows = []
    for row in questions_df.select(carry_cols + [codes_col]).iter_rows(named=True):
        for num, tok in parse_codes_cell(str(row.get(codes_col) or "")):
            rd = {c: row.get(c) for c in carry_cols}
            rd["code_num"] = num
            rd["code_token"] = tok
            rows.append(rd)
    if not rows:
        return pl.DataFrame(schema=EMPTY_SCHEMA)
    return pl.DataFrame(rows).with_columns([
        pl.col("code_num").cast(pl.Int64),
        pl.col("code_token").cast(pl.Utf8),
    ])



_step("Normalising placeholders")
# --- CODE CELL 6 ---
PLACEHOLDER_PATTERN = re.compile(r"\$[^$\r\n]+\$")

RESTORE_TEXT_COLUMNS = [
    "English",
    "French",
    "Spanish",
    "Arabic",
    "Portuguese",
    "Skip Pattern",
    "Default skip patterns & conditional ",
    "Default skip patterns & conditional",
    "Specify skip pattern variable (from blue text)",
    "Additional notes",
]

SEASON_PHASE_EXPECTED_VALUES = {
    "not yet in season",
    "land preparation",
    "planting",
    "early growing",
    "growing",
    "maturing",
    "Harvesting",
    "Recently finished"
}

CROP_MARKER_HINTS = {
    "$TEN MOST COMMON CROP$",
    "$TEN MOST COMMON CROPS$",
    "$TEN MOST COMMON CROP FR$",
    "$TEN MOST COMMON CROPS FR$",
    "$CROPS SOLD$",
    "$CROPS SOLD FR$",
    "$CROP CODES$",
    "$CROP CODES FR$",
    "$CROP SOLD CODES$",
    "$CROP SOLD CODES FR$",
}

CROP_LABEL_MARKER_PATTERN = r"\$TEN MOST COMMON CROP(?:S)?(?:\s+(?:EN|FR|ES|AR|PT))?\$"
CROP_SOLD_LABEL_MARKER_PATTERN = r"\$CROPS SOLD(?:\s+(?:EN|FR|ES|AR|PT))?\$"
CROP_CODE_MARKER_PATTERN = r"\$CROP CODES(?:\s+(?:EN|FR|ES|AR|PT))?\$"
CROP_SOLD_CODE_MARKER_PATTERN = r"\$CROP SOLD CODES(?:\s+(?:EN|FR|ES|AR|PT))?\$"


def _normalize_placeholder_token(token: str) -> str:
    s = str(token or "").strip()
    if s.startswith("$") and s.endswith("$") and len(s) >= 2:
        s = s[1:-1]
    s = re.sub(r"\s+", " ", s).strip().upper()
    s = re.sub(r" (EN|FR|ES|AR|PT)$", "", s)
    return s


def _is_crop_placeholder_token(token: str) -> bool:
    norm = _normalize_placeholder_token(token)
    if norm.startswith("TEN MOST COMMON CROP"):
        return True
    return norm in {"CROPS SOLD", "CROP CODES", "CROP SOLD CODES"}


def _contains_crop_marker_text(text: str) -> bool:
    for tok in re.findall(r"\$[^$\r\n]+\$", str(text or "")):
        if _is_crop_placeholder_token(tok):
            return True
    return False


def _clean_header(value) -> str:
    return str(value or "").replace("\n", " ").strip()


def _find_column(columns: list[str], *candidates: str) -> Optional[str]:
    lookup = {_clean_header(c).lower(): c for c in columns}
    for candidate in candidates:
        match = lookup.get(_clean_header(candidate).lower())
        if match:
            return match
    return None


def read_generic_sheet(path: Path, sheet_name: str, header_row: int, _wb=None) -> pl.DataFrame:
    """
    Reads any workbook sheet into Polars using the provided 1-based header row.
    Used for auxiliary sheets such as Crop list and Additional information.
    """
    owns_wb = _wb is None
    wb = _wb or _load_workbook_readonly(path, data_only=True, read_only=True)
    try:
        ws = next((wb[n] for n in wb.sheetnames if n.strip().lower() == sheet_name.lower()), None)
        if ws is None:
            raise KeyError(f"Sheet '{sheet_name}' not found in {path.name}. Available sheets: {wb.sheetnames}")

        row_iter = ws.iter_rows(values_only=True)
        for _ in range(header_row - 1):
            next(row_iter, None)

        raw_headers = next(row_iter, None)
        if raw_headers is None:
            raise ValueError(f"Header row {header_row} is missing in sheet '{sheet_name}' for file {path.name}")
        headers = [
            _clean_header(h) if h is not None else f"unnamed_{i}"
            for i, h in enumerate(raw_headers, 1)
        ]

        rows = []
        for values in row_iter:
            if all(v is None for v in values):
                continue
            row = {}
            for i in range(len(headers)):
                raw_val = values[i] if i < len(values) else None
                row[headers[i]] = _as_text_cell(raw_val)
            rows.append(row)

        return pl.DataFrame(rows, infer_schema_length=None) if rows else pl.DataFrame(schema={h: pl.Utf8 for h in headers})
    finally:
        if owns_wb:
            _close_workbook(wb)

def read_additional_information(path: Path, _wb=None) -> pl.DataFrame:
    return read_generic_sheet(path, sheet_name="Additional information", header_row=2, _wb=_wb)


def read_crop_list(path: Path, _wb=None) -> pl.DataFrame:
    return read_generic_sheet(path, sheet_name="Crop list", header_row=3, _wb=_wb)

# Returns True if the value contains at least one $...$ placeholder.
def has_placeholder(value) -> bool:
    return bool(PLACEHOLDER_PATTERN.search(str(value or "")))


def build_placeholder_restore_map(
    reference_questions: pl.DataFrame,
    candidate_columns: Optional[list[str]] = None,
) -> dict[str, dict[str, str]]:
    """
    Returns a per-question map of columns whose template/reference cell still
    contains a $...$ placeholder. Those exact reference cells are what we use
    to restore the current questionnaire before comparison.
    """
    columns = [c for c in (candidate_columns or RESTORE_TEXT_COLUMNS) if c in reference_questions.columns]
    restore_map: dict[str, dict[str, str]] = {}

    for row in reference_questions.select(["Q Name"] + columns).iter_rows(named=True):
        qname = row["Q Name"]
        hits = {
            col: str(row[col] or "")
            for col in columns
            if has_placeholder(row.get(col))
        }
        if hits:
            restore_map[qname] = hits
    return restore_map


def restore_placeholder_cells(
    current_questions: pl.DataFrame,
    reference_questions: pl.DataFrame,
    candidate_columns: Optional[list[str]] = None,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    """
    Restores current questionnaire cells to the template/reference version,
    but only for cells where the reference still contains a $...$ placeholder.
    Returns (restored_questions, restore_log).
    """
    restore_map = build_placeholder_restore_map(reference_questions, candidate_columns)
    if not restore_map:
        empty_log = pl.DataFrame(schema={"Q Name": pl.Utf8, "field": pl.Utf8, "restored_value": pl.Utf8})
        return current_questions.clone(), empty_log

    restore_log_rows = []
    for qname, hits in restore_map.items():
        for field, restored_value in hits.items():
            restore_log_rows.append({
                "Q Name": qname,
                "field": field,
                "restored_value": restored_value,
            })

    if not restore_log_rows:
        empty_log = pl.DataFrame(schema={"Q Name": pl.Utf8, "field": pl.Utf8, "restored_value": pl.Utf8})
        return current_questions.clone(), empty_log

    restore_log = pl.DataFrame(restore_log_rows).select(["Q Name", "field", "restored_value"])
    restore_wide = restore_log.pivot(
        values="restored_value",
        index="Q Name",
        columns="field",
        aggregate_function="first",
    )

    joined = current_questions.join(restore_wide, on="Q Name", how="left", suffix="_restore")

    fields = sorted(set(restore_log["field"].to_list()))
    for field in fields:
        restore_col = f"{field}_restore"
        if field in current_questions.columns and restore_col in joined.columns:
            joined = (
                joined
                .with_columns(pl.coalesce([pl.col(restore_col), pl.col(field)]).alias(field))
                .drop(restore_col)
            )

    leftover_restore_cols = [c for c in joined.columns if c.endswith("_restore")]
    if leftover_restore_cols:
        joined = joined.drop(leftover_restore_cols)

    restored = joined.select(current_questions.columns)
    return restored, restore_log

def _normalize_placeholder_key(original_value: str) -> str:
    """
    Normalize an Additional information 'Original' key to a placeholder token.

    Accepts both strict "$...$" keys and plain labels like "currency".
    Plain labels are canonicalized to "$label$".
    """
    original = str(original_value or "").strip()
    if not original:
        return ""
    if has_placeholder(original):
        m = PLACEHOLDER_PATTERN.search(original)
        return m.group(0).strip() if m else ""
    return f"${original}$"


def _placeholder_lookup_aliases(token: str) -> list[str]:
    """Build normalized lookup aliases for placeholder tokens (e.g. $season FR$ -> $season$)."""
    raw = str(token or "").strip()
    if not raw:
        return []

    aliases = []
    if raw.startswith("$") and raw.endswith("$") and len(raw) >= 2:
        inner = re.sub(r"\s+", " ", raw[1:-1]).strip()
        if inner:
            aliases.append(f"${inner}$")
            base = re.sub(r"\s+(EN|FR|ES|AR|PT)$", "", inner, flags=re.IGNORECASE).strip()
            if base and base != inner:
                aliases.append(f"${base}$")
    else:
        aliases.append(raw)

    out = []
    seen = set()
    for a in aliases:
        k = a.lower()
        if k and k not in seen:
            seen.add(k)
            out.append(k)
    return out


def _build_placeholder_index(replacements: dict[str, str]) -> dict[str, str]:
    idx = {}
    for k, v in (replacements or {}).items():
        if not v:
            continue
        for alias in _placeholder_lookup_aliases(k):
            idx.setdefault(alias, v)
    return idx


def _lookup_placeholder_replacement(token: str, replacements: dict[str, str], idx: dict[str, str] | None = None) -> str | None:
    mapping = idx if idx is not None else _build_placeholder_index(replacements)
    for alias in _placeholder_lookup_aliases(token):
        val = mapping.get(alias)
        if val is not None and str(val) != "":
            return val
    return None


def _language_code_for_text_column(column_name: str) -> str | None:
    c = str(column_name or "")
    c_norm = _normalize_header_name(c)
    for lang in ["EN", "FR", "ES", "AR", "PT"]:
        canonical = LANGUAGE_COLUMN.get(lang, "")
        if canonical and c_norm == _normalize_header_name(canonical):
            return lang
        aliases = LANGUAGE_ALIASES.get(lang, [])
        toks = set(_header_tokens(c))
        for a in aliases:
            a_norm = _normalize_header_name(a)
            if not a_norm:
                continue
            if len(a_norm) <= 2:
                c_low = c.lower().strip()
                if a.lower() in toks or c_low.startswith(a.lower() + "_") or c_low.startswith(a.lower() + "-"):
                    return lang
            elif a_norm in c_norm:
                return lang
    return None


def _placeholder_token_explicit_lang(token: str) -> str | None:
    raw = str(token or "").strip()
    if not (raw.startswith("$") and raw.endswith("$") and len(raw) >= 2):
        return None
    inner = re.sub(r"\s+", " ", raw[1:-1]).strip()
    m = re.search(r"\s+(EN|FR|ES|AR|PT)$", inner, flags=re.IGNORECASE)
    return m.group(1).upper() if m else None


def analyze_additional_info_substitutions(path: Path, language: str, _wb=None) -> dict:
    """Return diagnostics about Additional information substitution readiness."""
    df = read_additional_information(path, _wb=_wb)
    out = {
        "has_sheet": True,
        "has_required_columns": True,
        "replacement_col": "",
        "original_col": "",
        "rows_with_any_data": 0,
        "rows_with_placeholder_style_original": 0,
        "rows_with_non_placeholder_original": 0,
        "loaded_keys": set(),
        "blank_value_keys": set(),
    }

    if df.height == 0:
        out["has_sheet"] = False
        return out

    replacement_col = _find_column(
        df.columns,
        f"Replacement ({language})",
        "Replacement",
        "Replacement (EN)",
    )
    original_col = _find_column(df.columns, "Original")
    out["replacement_col"] = replacement_col or ""
    out["original_col"] = original_col or ""
    if not replacement_col or not original_col:
        out["has_required_columns"] = False
        return out

    for row in df.iter_rows(named=True):
        original = str(row.get(original_col) or "").strip()
        replacement = str(row.get(replacement_col) or "").strip()
        if not original and not replacement:
            continue

        out["rows_with_any_data"] += 1
        if has_placeholder(original):
            out["rows_with_placeholder_style_original"] += 1
        elif original:
            out["rows_with_non_placeholder_original"] += 1

        key = _normalize_placeholder_key(original)
        if not key:
            continue
        if replacement:
            out["loaded_keys"].add(key)
        else:
            out["blank_value_keys"].add(key)

    return out


def build_additional_info_replacements(path: Path, language: str, _wb=None) -> dict[str, str]:
    """Backward-compatible single-language map from Additional information."""
    per_lang = build_additional_info_replacements_by_language(path, _wb=_wb)
    lang = str(language or "EN").upper()
    if per_lang.get(lang):
        return per_lang[lang]
    if per_lang.get("EN"):
        return per_lang["EN"]
    for m in per_lang.values():
        if m:
            return m
    return {}


def build_additional_info_replacements_by_language(path: Path, _wb=None) -> dict[str, dict[str, str]]:
    """Build per-language substitution maps from Additional information."""
    df = read_additional_information(path, _wb=_wb)
    original_col = _find_column(df.columns, "Original")
    if not original_col or df.height == 0:
        return {lang: {} for lang in ["EN", "FR", "ES", "AR", "PT"]}

    generic_replacement_col = _find_column(df.columns, "Replacement")
    en_replacement_col = _find_column(df.columns, "Replacement (EN)", "Replacement EN")

    replacement_col_by_lang = {}
    for lang in ["EN", "FR", "ES", "AR", "PT"]:
        replacement_col_by_lang[lang] = _find_column(
            df.columns,
            f"Replacement ({lang})",
            f"Replacement {lang}",
        )

    out = {lang: {} for lang in ["EN", "FR", "ES", "AR", "PT"]}
    season_phase_by_lang = {lang: "" for lang in out.keys()}

    for row in df.iter_rows(named=True):
        key = _normalize_placeholder_key(str(row.get(original_col) or "").strip())
        if not key:
            continue

        for lang in out.keys():
            replacement = ""
            col = replacement_col_by_lang.get(lang)
            if col:
                replacement = str(row.get(col) or "").strip()
            if not replacement and generic_replacement_col:
                replacement = str(row.get(generic_replacement_col) or "").strip()
            if not replacement and lang != "EN" and en_replacement_col:
                replacement = str(row.get(en_replacement_col) or "").strip()

            if replacement:
                out[lang][key] = replacement
                if key.lower() in {"$season phase$", "$season phase $"}:
                    season_phase_by_lang[lang] = replacement.lower()

    for lang, mapping in out.items():
        if "$expected or nothing$" in mapping:
            phase = season_phase_by_lang.get(lang, "")
            mapping["$expected or nothing$"] = "expected" if phase in SEASON_PHASE_EXPECTED_VALUES else ""

    return out


def apply_text_replacements(
    value,
    replacements: dict[str, str],
    replacements_by_language: Optional[dict[str, dict[str, str]]] = None,
) -> str:
    text = str(value or "")
    has_lang_maps = any((replacements_by_language or {}).values())
    if not replacements and not has_lang_maps:
        return text

    default_map = replacements or {}
    default_idx = _build_placeholder_index(default_map)
    lang_idx_cache: dict[str, dict[str, str]] = {}

    def _lang_map(lang: str) -> dict[str, str]:
        if not replacements_by_language:
            return {}
        return replacements_by_language.get(lang, {}) or {}

    for token in sorted(set(re.findall(r"\$[^$\r\n]+\$", text)), key=len, reverse=True):
        replacement = None
        token_lang = _placeholder_token_explicit_lang(token)

        # If token explicitly carries a language suffix (e.g. $season FR$),
        # prioritize that language replacement column.
        if token_lang:
            token_lang_map = _lang_map(token_lang)
            if token_lang_map:
                if token_lang not in lang_idx_cache:
                    lang_idx_cache[token_lang] = _build_placeholder_index(token_lang_map)
                replacement = _lookup_placeholder_replacement(token, token_lang_map, lang_idx_cache[token_lang])

        # Fallback to the active column/default language map.
        if replacement is None and default_map:
            replacement = _lookup_placeholder_replacement(token, default_map, default_idx)

        # Final fallback: EN map if available (useful when EN is selected and
        # only English replacements are filled).
        if replacement is None:
            en_map = _lang_map("EN")
            if en_map:
                if "EN" not in lang_idx_cache:
                    lang_idx_cache["EN"] = _build_placeholder_index(en_map)
                replacement = _lookup_placeholder_replacement(token, en_map, lang_idx_cache["EN"])

        if replacement is not None:
            text = text.replace(token, replacement)
    return text


def apply_placeholder_conversions(
    questions_df: pl.DataFrame,
    replacements: dict[str, str],
    candidate_columns: Optional[list[str]] = None,
    replacements_by_language: Optional[dict[str, dict[str, str]]] = None,
) -> pl.DataFrame:
    """
    Forward-applies placeholder replacements to text-bearing columns.
    Language columns use language-specific replacement maps when available.
    """
    columns = [c for c in (candidate_columns or RESTORE_TEXT_COLUMNS) if c in questions_df.columns]
    if not columns:
        return questions_df.clone()

    default_replacements = replacements or {}
    has_lang_maps = any((replacements_by_language or {}).values())
    out = questions_df.clone()

    for col in columns:
        lang = _language_code_for_text_column(col)
        col_replacements = default_replacements
        if replacements_by_language and lang and replacements_by_language.get(lang):
            col_replacements = replacements_by_language[lang]

        if not col_replacements and not has_lang_maps:
            continue

        out = out.with_columns(
            pl.col(col)
            .cast(pl.Utf8)
            .fill_null("")
            .map_elements(
                lambda v: apply_text_replacements(v, col_replacements, replacements_by_language=replacements_by_language),
                return_dtype=pl.Utf8,
            )
            .alias(col)
        )

    return out


def extract_crop_metadata(crop_df: pl.DataFrame, language: str) -> dict:
    """
    Pulls the key columns from the Crop list sheet in a language-aware way.
    Returned labels include the full crop pool, not only the selected rows.
    """
    label_col = _find_column(crop_df.columns, f"Label ({language})", "Label (EN)")
    dataset_col = _find_column(crop_df.columns, "Dataset code")
    select_col = _find_column(crop_df.columns, "Select top 10 crops", "Select top 10 crops ")

    if not label_col:
        return {"label_col": None, "dataset_col": dataset_col, "select_col": select_col, "labels": []}

    labels = [
        str(v).strip()
        for v in crop_df.get_column(label_col).to_list()
        if str(v or "").strip()
    ]
    return {
        "label_col": label_col,
        "dataset_col": dataset_col,
        "select_col": select_col,
        "labels": labels,
    }


def find_crop_placeholder_questions(reference_questions: pl.DataFrame) -> list[str]:
    """
    Identifies questions whose reference/template text carries a crop marker.
    Those are the questions that need template-style restoration for comparison
    and deployed-style rebuilding for validated output.
    """
    text_columns = [c for c in RESTORE_TEXT_COLUMNS if c in reference_questions.columns]
    hits = []
    for row in reference_questions.select(["Q Name"] + text_columns).iter_rows(named=True):
        values = "\n".join(str(row.get(col) or "") for col in text_columns)
        if _contains_crop_marker_text(values):
            hits.append(row["Q Name"])
    return sorted(set(hits))


def _is_selected_crop_value(value) -> bool:
    return str(value or "").strip().lower() not in {"", "0", "0.0", "none", "nan"}


def _build_sorted_crop_entries(crop_df: pl.DataFrame, language: str) -> list[dict]:
    meta = extract_crop_metadata(crop_df, language)
    label_col = meta.get("label_col")
    dataset_col = meta.get("dataset_col")
    select_col = meta.get("select_col")
    if not label_col or label_col not in crop_df.columns:
        return []

    needed = [label_col]
    if dataset_col and dataset_col in crop_df.columns:
        needed.append(dataset_col)
    if select_col and select_col in crop_df.columns:
        needed.append(select_col)

    entries = []
    for idx, row in enumerate(crop_df.select(needed).iter_rows(named=True), start=1):
        label = str(row.get(label_col) or "").strip()
        code = str(row.get(dataset_col) or "").strip() if dataset_col and dataset_col in crop_df.columns else ""
        selected = _is_selected_crop_value(row.get(select_col)) if select_col and select_col in crop_df.columns else False
        if not label and not code:
            continue
        entries.append({"row_idx": idx, "label": label, "code": code, "selected": selected})

    entries.sort(key=lambda e: (0 if e["selected"] else 1, e["row_idx"]))
    return entries


def build_crop_option_block(crop_df: pl.DataFrame, language: str, include_no_crop: bool = True) -> str:
    """Builds a numbered crop label block: selected crops first, then others."""
    entries = _build_sorted_crop_entries(crop_df, language)
    labels = [e["label"] for e in entries if e["label"]]
    if not labels:
        return ""

    lines = [f"{idx}){label}" for idx, label in enumerate(labels, start=1)]
    if include_no_crop:
        lines.extend(["91)No crop production", "92)DON'T KNOW", "93)REFUSED"])
    else:
        lines.extend(["92)DON'T KNOW", "93)REFUSED"])
    return "\n".join(lines)


def build_crop_code_block(crop_df: pl.DataFrame, language: str, include_no_crop: bool = True) -> str:
    """Builds a numbered crop dataset-code block: selected crops first, then others."""
    entries = _build_sorted_crop_entries(crop_df, language)
    codes = [e["code"] for e in entries if e["code"]]
    if not codes:
        return ""

    lines = [f"{idx}){code}" for idx, code in enumerate(codes, start=1)]
    if include_no_crop:
        lines.extend(["91)No crop production", "92)DON'T KNOW", "93)REFUSED"])
    else:
        lines.extend(["92)DON'T KNOW", "93)REFUSED"])
    return "\n".join(lines)


def rebuild_crop_questions_for_deployed_form(
    questions_df: pl.DataFrame,
    template_questions: pl.DataFrame,
    crop_df: pl.DataFrame,
    language: str,
    candidate_columns: Optional[list[str]] = None,
) -> pl.DataFrame:
    """
    Replaces crop placeholders using the current questionnaire Crop list.
    Handles label and code markers, singular/plural and language-suffixed forms.
    """
    crop_qnames = set(find_crop_placeholder_questions(template_questions))
    columns = [c for c in (candidate_columns or RESTORE_TEXT_COLUMNS) if c in questions_df.columns]
    if not crop_qnames or not columns:
        return questions_df.clone()

    main_block = build_crop_option_block(crop_df, language, include_no_crop=True)
    sold_block = build_crop_option_block(crop_df, language, include_no_crop=False)
    main_code_block = build_crop_code_block(crop_df, language, include_no_crop=True)
    sold_code_block = build_crop_code_block(crop_df, language, include_no_crop=False)
    if not main_block:
        return questions_df.clone()

    crop_qnames_list = sorted(crop_qnames)
    qname_lc = pl.col("Q Name").cast(pl.Utf8).str.to_lowercase()
    is_crop_q = pl.col("Q Name").is_in(crop_qnames_list)
    is_sales_q = qname_lc.str.contains("sales")

    exprs = []
    for col in columns:
        base = pl.col(col).cast(pl.Utf8).fill_null("")

        has_marker = (
            base.str.contains(CROP_LABEL_MARKER_PATTERN)
            | base.str.contains(CROP_SOLD_LABEL_MARKER_PATTERN)
            | base.str.contains(CROP_CODE_MARKER_PATTERN)
            | base.str.contains(CROP_SOLD_CODE_MARKER_PATTERN)
        )

        replaced_main = (
            base
            .str.replace_all(CROP_LABEL_MARKER_PATTERN, main_block)
            .str.replace_all(CROP_SOLD_LABEL_MARKER_PATTERN, sold_block)
            .str.replace_all(CROP_CODE_MARKER_PATTERN, main_code_block)
            .str.replace_all(CROP_SOLD_CODE_MARKER_PATTERN, sold_code_block)
        )

        replaced_sold = (
            base
            .str.replace_all(CROP_LABEL_MARKER_PATTERN, sold_block)
            .str.replace_all(CROP_SOLD_LABEL_MARKER_PATTERN, sold_block)
            .str.replace_all(CROP_CODE_MARKER_PATTERN, sold_code_block)
            .str.replace_all(CROP_SOLD_CODE_MARKER_PATTERN, sold_code_block)
        )

        exprs.append(
            pl.when(is_crop_q & has_marker)
            .then(
                pl.when(is_sales_q).then(replaced_sold).otherwise(replaced_main)
            )
            .otherwise(base)
            .alias(col)
        )

    return questions_df.with_columns(exprs)






# --- CODE CELL 7 ---
current_wb = _load_workbook_readonly(run["questionnaire_path"], data_only=True, read_only=True)
current_raw = read_survey_sheet(run["questionnaire_path"], _wb=current_wb)
reference_raw = read_survey_sheet(run["reference_path"])

current_questions_raw = build_questions_df(current_raw)
reference_questions_raw = build_questions_df(reference_raw)

current_crop_list = read_crop_list(run["questionnaire_path"], _wb=current_wb)
try:
    reference_crop_list = read_crop_list(run["reference_path"])
except Exception:
    reference_crop_list = pl.DataFrame()
current_text_replacements_by_language = build_additional_info_replacements_by_language(run["questionnaire_path"], _wb=current_wb)
current_text_replacements = current_text_replacements_by_language.get(cfg.language, {}) or current_text_replacements_by_language.get("EN", {})
_close_workbook(current_wb)
restore_log                = pl.DataFrame(schema={"Q Name": pl.Utf8, "field": pl.Utf8, "restored_value": pl.Utf8})
template_questions_for_restore = reference_questions_raw

# Comparison restore policy:
# - Always restore standard text placeholder columns.
# - For template/reference comparison modes, also restore 'Codes' placeholders
#   so dynamic crop/domain expansions do not appear as false code additions.
compare_restore_columns = list(RESTORE_TEXT_COLUMNS)
if cfg.reference_mode != "previous_round":
    compare_restore_columns.append("Codes")

if cfg.reference_mode == "previous_round":
    template_path = find_latest_template(cfg.template_repo, cfg.enumerator, cfg.language)
    template_raw  = read_survey_sheet(template_path)
    template_questions_for_restore = build_questions_df(template_raw)

    restored_current_questions, restore_log = restore_placeholder_cells(
        current_questions_raw,
        template_questions_for_restore,
        candidate_columns=compare_restore_columns,
    )
    comparison_current_questions = apply_placeholder_conversions(
        restored_current_questions,
        current_text_replacements,
        replacements_by_language=current_text_replacements_by_language,
    )
    comparison_current_questions = rebuild_crop_questions_for_deployed_form(
        comparison_current_questions,
        template_questions_for_restore,
        current_crop_list,
        cfg.language,
    )
    comparison_reference_questions = reference_questions_raw
    validated_output_questions = comparison_current_questions
else:
    comparison_current_questions, restore_log = restore_placeholder_cells(
        current_questions_raw,
        reference_questions_raw,
        candidate_columns=compare_restore_columns,
    )
    comparison_reference_questions = reference_questions_raw
    validated_output_questions = apply_placeholder_conversions(
        current_questions_raw,
        current_text_replacements,
        replacements_by_language=current_text_replacements_by_language,
    )
    validated_output_questions = rebuild_crop_questions_for_deployed_form(
        validated_output_questions,
        reference_questions_raw,
        current_crop_list,
        cfg.language,
    )

current_questions   = comparison_current_questions
reference_questions = comparison_reference_questions

target_lang, target_lang_source = choose_target_language(
    run["questionnaire_path"],
    current_questions_raw,
    preferred=cfg.language,
)
print(f"Comparison language : {target_lang} ({target_lang_source})")

current_en_col   = resolve_language_column(current_questions.columns, "EN")
reference_en_col = resolve_language_column(reference_questions.columns, "EN")
if not current_en_col or not reference_en_col:
    raise KeyError("English column not found in one of the files; option checks require English baseline.")

current_tgt_col   = resolve_language_column(current_questions.columns, target_lang)
reference_tgt_col = resolve_language_column(reference_questions.columns, target_lang)

if target_lang == "EN":
    current_tgt_col   = current_en_col
    reference_tgt_col = reference_en_col
else:
    if not current_tgt_col:
        print(f"  Target language column unresolved in current file for {target_lang}; using English in current file.")
        current_tgt_col = current_en_col
    if not reference_tgt_col:
        print(f"  Target language column unresolved in reference file for {target_lang}; target-language reference values will be blank in report.")

print(f"  Option columns -> EN: current='{current_en_col}', reference='{reference_en_col}'")
print(f"  Option columns -> {target_lang}: current='{current_tgt_col}', reference='{reference_tgt_col}'")

# English options are the stable baseline for presence (added/removed option codes)
current_options_en   = explode_options(current_questions, text_col=current_en_col)
reference_options_en = explode_options(reference_questions, text_col=reference_en_col)

# Target-language options are used for label mismatch checks.
# If target columns resolve to EN, reuse EN exploded frames to avoid duplicate work.
if current_tgt_col == current_en_col and reference_tgt_col == reference_en_col:
    current_options_tgt   = current_options_en.clone()
    reference_options_tgt = reference_options_en.clone()
else:
    current_options_tgt   = explode_options(current_questions, text_col=current_tgt_col)
    if reference_tgt_col:
        reference_options_tgt = explode_options(reference_questions, text_col=reference_tgt_col)
    else:
        reference_options_tgt = pl.DataFrame(schema=current_options_tgt.schema)

current_codes_col = "Codes" if "Codes" in current_questions.columns else _resolve_column_by_aliases(current_questions.columns, COLUMN_ALIASES.get("Codes", ["Codes"]))
reference_codes_col = "Codes" if "Codes" in reference_questions.columns else _resolve_column_by_aliases(reference_questions.columns, COLUMN_ALIASES.get("Codes", ["Codes"]))
current_codes = explode_codes(current_questions, codes_col=current_codes_col) if current_codes_col else pl.DataFrame(schema={"Q Name": pl.Utf8, "code_num": pl.Int64, "code_token": pl.Utf8, "excel_row": pl.Int64, "source_file": pl.Utf8})
reference_codes = explode_codes(reference_questions, codes_col=reference_codes_col) if reference_codes_col else pl.DataFrame(schema={"Q Name": pl.Utf8, "code_num": pl.Int64, "code_token": pl.Utf8, "excel_row": pl.Int64, "source_file": pl.Utf8})

print("--- comparison current questionnaire ---")
print(f"  questions : {current_questions.shape}")
print(f"  options EN      : {current_options_en.shape}")
print(f"  options {target_lang:<2}      : {current_options_tgt.shape}")
print(f"  codes rows      : {current_codes.shape}")
print(f"  restored placeholder cells : {restore_log.height}")
print(f"  validated output questions : {validated_output_questions.shape}")

print("\n--- comparison reference file ---")
print(f"  questions : {reference_questions.shape}")
print(f"  options EN      : {reference_options_en.shape}")
print(f"  options {target_lang:<2}      : {reference_options_tgt.shape}")
print(f"  codes rows      : {reference_codes.shape}")

print("\n--- sample: resp_gender options (comparison current, target language) ---")
print(current_options_tgt.filter(pl.col("Q Name") == "resp_gender"))





# ======================================================================
# SECTION: ## Step 6  Normalisation helpers
_step("Normalisation helpers")
# ======================================================================
# --- CODE CELL 8 ---
def normalize_text_expr(col_name: str) -> pl.Expr:
    """
    Polars expression that lowercases a string column, removes punctuation,
    and collapses whitespace. Apply as an alias so the original column is preserved.

    Usage:
        df.with_columns(normalize_text_expr("option_label").alias("option_label_norm"))
    """
    return (
        pl.col(col_name)
        .cast(pl.Utf8)
        .fill_null("")
        .str.to_lowercase()
        .str.replace_all(r"[^\w\s]", "")   # remove punctuation
        .str.replace_all(r"\s+", " ")       # collapse multiple spaces
        .str.strip_chars()
    )


def normalize_code_token_expr(col_name: str) -> pl.Expr:
    """Normalizes code tokens from the 'Codes' column for stable comparisons."""
    return (
        pl.col(col_name)
        .cast(pl.Utf8)
        .fill_null("")
        .str.to_lowercase()
        .str.replace_all(r"[^a-z0-9_]+", "")
        .str.strip_chars()
    )


def normalize_mandatory_expr(col_name: str) -> pl.Expr:
    """
    Polars expression that maps any common yes/no variant to canonical "yes"/"no"/""
    so that "Yes", "YES", "y", "true", "1" are all treated as equivalent.

    Using when/then/otherwise keeps this inside the Polars execution engine
    (faster and parallelisable) vs a Python lambda with map_elements.
    """
    cleaned = (
        pl.col(col_name)
        .cast(pl.Utf8)
        .fill_null("")
        .str.to_lowercase()
        .str.strip_chars()
    )
    return (
        pl.when(cleaned.is_in(["yes", "y", "true", "1"])).then(pl.lit("yes"))
        .when(cleaned.is_in(["no",  "n", "false", "0"])).then(pl.lit("no"))
        .otherwise(pl.lit(""))
    )



# ======================================================================
# SECTION: ## Step 7  Comparison and validation functions
_step("Comparison functions")
# ======================================================================
# --- CODE CELL 9 ---
SKIP_PATTERN_COLS = [
    "Skip Pattern",
    "Default skip patterns & conditional ",   # trailing space is intentional
    "Specify skip pattern variable (from blue text)",
]


def _normalize_skip_text(value) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text).strip()


def _parse_skip_option_codes(
    skip_text: str,
    max_range_span: int = 200,
    max_total_codes: int = 500,
) -> set[int]:
    """
    Extracts numeric option codes from the left side of skip rules.
    Examples handled: "1=...", "1-4 = ...", "1,3,5=...", "2 to 4 = ...".

    Safety guards:
      - do not fully expand very wide ranges,
      - cap total extracted codes to avoid pathological blowups.
    """
    if not skip_text:
        return set()

    codes = set()
    parts = [p.strip() for p in re.split(r"[\r\n;]+", str(skip_text)) if p.strip()]
    for part in parts:
        if len(codes) >= max_total_codes:
            break

        left = part.split("=", 1)[0] if "=" in part else part

        for a, b in re.findall(r"\b(\d+)\s*(?:-|to)\s*(\d+)\b", left, flags=re.IGNORECASE):
            start, end = int(a), int(b)
            lo, hi = sorted((start, end))
            span = hi - lo + 1

            if span <= max_range_span and (len(codes) + span) <= max_total_codes:
                codes.update(range(lo, hi + 1))
            else:
                # Keep endpoints so we still validate something without exploding runtime.
                codes.add(lo)
                codes.add(hi)
                if len(codes) >= max_total_codes:
                    break

        if len(codes) >= max_total_codes:
            break

        left_no_ranges = re.sub(r"\b\d+\s*(?:-|to)\s*\d+\b", " ", left, flags=re.IGNORECASE)
        for n in re.findall(r"\b\d+\b", left_no_ranges):
            codes.add(int(n))
            if len(codes) >= max_total_codes:
                break

    return codes

def _extract_referenced_qnames(skip_text: str, source_q: str, ref_qnames: set[str], curr_qnames: set[str]) -> set[str]:
    """
    Extract target question names from skip rules using a strict heuristic:
      - parse only RHS of '=' (or line/segment after '='),
      - take the first token as candidate target,
      - ignore common routing labels (ineligible, end, quota, etc.),
      - keep candidates that are known qnames OR look like qnames (underscore).
    """
    text = str(skip_text or "")
    if len(text) > 4000:
        text = text[:4000]

    stopwords = {
        "ineligible", "eligible", "end", "poll", "quota", "reached",
        "terminate", "terminated", "screenout", "screen_out",
        "refusal", "refused", "closeout", "close", "callback",
        "yes", "no", "na", "n_a", "none",
    }

    out = set()
    segments = [s.strip() for s in re.split(r"[\r\n;]+", text) if s.strip()]
    for seg in segments:
        if "=" not in seg:
            continue
        rhs = seg.split("=", 1)[1].strip()
        if not rhs:
            continue

        m = re.match(r"^\s*([A-Za-z_]\w*)", rhs)
        if not m:
            continue

        cand = m.group(1)
        cand_l = cand.lower()
        if cand == source_q or cand_l in stopwords:
            continue

        if cand in ref_qnames or cand in curr_qnames or ("_" in cand):
            out.add(cand)

    return out


# â”€â”€â”€ helpers for Default-column consistency check â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

_QNAME_TOKEN_RE = re.compile(r'\b([a-z][a-z0-9_]{2,})\b')
_GO_TO_RE       = re.compile(r'\bgo\s+to\s+([^\n;]+)', re.IGNORECASE)
_NEXT_Q_RE      = re.compile(r'\bnext\s+question\b', re.IGNORECASE)
_OPTIONAL_RE    = re.compile(r'\boptional\s+question\b', re.IGNORECASE)


def _extract_qname_tokens(text: str, known_qnames: set[str]) -> list[str]:
    """Return word tokens from text that are known Q Names."""
    return [t for t in _QNAME_TOKEN_RE.findall(str(text or "")) if t in known_qnames]


def _extract_qname_like(text: str, known_qnames: set[str]) -> list[str]:
    """
    Return word tokens that are either known Q Names OR look like Q Names
    (contain an underscore).  This lets us detect references to Q Names
    that were renamed / removed from the current questionnaire.
    """
    return [t for t in _QNAME_TOKEN_RE.findall(str(text or ""))
            if t in known_qnames or "_" in t]


def _extract_codes_from_numeric_spec(
    spec: str,
    max_range_span: int = 200,
    max_total_codes: int = 500,
) -> set[int]:
    """
    Safely parse numeric code specs like "1-4, 7, 9 to 11".
    Large ranges are truncated to endpoints to prevent memory blowups while still
    preserving strong validation signal.
    """
    out: set[int] = set()
    if not spec:
        return out

    for a, b in re.findall(r"\b(\d+)\s*(?:-|to)\s*(\d+)\b", spec, re.IGNORECASE):
        lo, hi = sorted((int(a), int(b)))
        span = hi - lo + 1
        if span <= max_range_span and (len(out) + span) <= max_total_codes:
            out.update(range(lo, hi + 1))
        else:
            out.add(lo)
            out.add(hi)
        if len(out) >= max_total_codes:
            return out

    clean = re.sub(r"\b\d+\s*(?:-|to)\s*\d+\b", " ", spec, flags=re.IGNORECASE)
    for n in re.findall(r"\b\d+\b", clean):
        out.add(int(n))
        if len(out) >= max_total_codes:
            break
    return out


def _extract_condition_codes(stmt: str) -> set[int]:
    """
    Extract option codes from a full Default-column rule statement by finding
    the range between '=' and 'go to', e.g.:
      "if hh_gender = 1-4, go to hh_education"  -> {1, 2, 3, 4}
      "For any response, go to X"               -> {} (any response = no range)
    Receives the FULL statement (not sliced) so that 'go to' can anchor the match.
    """
    match = re.search(
        r'=\s*([\d][^;=\n]*?)\s*,?\s*go\s+to',
        stmt, re.IGNORECASE,
    )
    if not match:
        return set()
    rng = match.group(1).strip()
    return _extract_codes_from_numeric_spec(rng)

def _extract_skip_codes_for_target(skip_text: str, target: str) -> set[int]:
    """
    Extract option codes from Skip-Pattern lines that send flow to *target*.
    Handles both compact and verbose styles:
      "1-3 = hh_education"
      "if hh_gender = 1-3, go to hh_education"
      "1 = q_a 2-4 = q_b"
      "if q = 1, go to a if q = 2-4, go to b"
    """
    def _codes_from_spec(spec: str) -> set[int]:
        return _extract_codes_from_numeric_spec(spec)

    compact_re = re.compile(
        r'(\d+(?:\s*(?:-|to)\s*\d+)?(?:\s*,\s*\d+(?:\s*(?:-|to)\s*\d+)?)*)\s*=\s*(\[[^\]]+\]|[A-Za-z_]\w*)',
        re.IGNORECASE,
    )
    verbose_re = re.compile(
        r'if\s+[A-Za-z_]\w*\s*=\s*([^;]+?)\s*,?\s*go\s+to\s*(.+?)(?=(?:\s+if\s+[A-Za-z_]\w*\s*=)|$)',
        re.IGNORECASE,
    )

    codes: set[int] = set()
    for part in re.split(r'[\r\n;]+', str(skip_text or "")):
        if target not in part:
            continue

        # Compact style supports multiple mappings in one segment.
        for m in compact_re.finditer(part):
            code_spec = m.group(1)
            target_text = m.group(2)
            target_tokens = _extract_qname_like(target_text, {target})
            if target in target_tokens or target_text.strip('[] ') == target:
                codes.update(_codes_from_spec(code_spec))

        # Verbose style supports multiple "if ... go to ..." clauses in one segment.
        for vm in verbose_re.finditer(part):
            code_spec = vm.group(1)
            target_text = vm.group(2).strip().strip('.')
            target_tokens = _extract_qname_like(target_text, {target})
            if target in target_tokens or target_text.strip('[] ') == target:
                codes.update(_codes_from_spec(code_spec))

    return codes

def _parse_default_skip_rules(default_text: str, known_qnames: set[str]) -> list[dict]:
    """
    Parse "Default skip patterns & conditional" into structured rules.
    Each rule: {targets, is_flexible, is_next_question, raw}

    Handles both compact and verbose styles, including multiple clauses in one line.
    """
    text = _normalize_skip_text(str(default_text or ""))
    if not text:
        return []

    rules = []

    # Verbose style, including chained clauses in one segment.
    verbose_re = re.compile(
        r'if\s+[A-Za-z_]\w*\s*=\s*([^;]+?)\s*,?\s*go\s+to\s*(.+?)(?=(?:\s+if\s+[A-Za-z_]\w*\s*=)|$)',
        re.IGNORECASE,
    )
    for vm in verbose_re.finditer(text):
        code_spec = vm.group(1)
        target_text = vm.group(2).strip().strip('.')
        is_next_question = bool(_NEXT_Q_RE.search(target_text))
        is_flexible = bool(_OPTIONAL_RE.search(target_text))
        targets = _extract_qname_like(target_text, known_qnames)
        option_codes = _extract_condition_codes(f"x = {code_spec}, go to y")
        if targets or is_next_question:
            rules.append({
                "targets": targets,
                "is_flexible": is_flexible,
                "is_next_question": is_next_question,
                "option_codes": option_codes,
                "raw": vm.group(0),
            })

    # Compact form fallback (no "go to"): "1 = q1 2-4 = q2"
    # Keep full RHS clause so patterns like "OR optional question" are preserved.
    if not rules:
        compact_re = re.compile(
            r'(\d+(?:\s*(?:-|to)\s*\d+)?(?:\s*,\s*\d+(?:\s*(?:-|to)\s*\d+)?)*)\s*=\s*(.+?)(?=(?:\s+\d+(?:\s*(?:-|to)\s*\d+)?(?:\s*,\s*\d+(?:\s*(?:-|to)\s*\d+)?)*)\s*=|$)',
            re.IGNORECASE,
        )
        for m in compact_re.finditer(text):
            code_spec = m.group(1)
            target_text = m.group(2).strip().strip('.')
            targets = _extract_qname_like(target_text, known_qnames)
            if not targets:
                continue
            rules.append({
                "targets": targets,
                "is_flexible": bool(_OPTIONAL_RE.search(target_text)),
                "is_next_question": bool(_NEXT_Q_RE.search(target_text)),
                "option_codes": _extract_condition_codes(f"x = {code_spec}, go to y"),
                "raw": m.group(0),
            })

    # Minimal fallback for "For any response, go to target" when no numeric condition exists.
    if not rules:
        any_re = re.compile(r'for\s+any\s+response\s*,?\s*go\s+to\s*(.+)$', re.IGNORECASE)
        m = any_re.search(text)
        if m:
            target_text = m.group(1).strip().strip('.')
            targets = _extract_qname_like(target_text, known_qnames)
            if targets:
                rules.append({
                    "targets": targets,
                    "is_flexible": bool(_OPTIONAL_RE.search(target_text)),
                    "is_next_question": bool(_NEXT_Q_RE.search(target_text)),
                    "option_codes": set(),
                    "raw": m.group(0),
                })

    return rules

def _semantic_skip_signature(skip_text: str, known_qnames: set[str]) -> tuple[set[tuple], bool]:
    """
    Build a semantic signature for skip logic text so phrasing differences
    can still compare as equal.

    Signature item:
      (target, sorted_codes_tuple, is_flexible, is_next_question)
    """
    text = _normalize_skip_text(skip_text)
    if not text:
        return set(), False

    rules = _parse_default_skip_rules(text, known_qnames)
    if not rules:
        return set(), False

    sig: set[tuple] = set()
    for rule in rules:
        codes = tuple(sorted(int(c) for c in (rule.get("option_codes") or set())))
        is_flexible = bool(rule.get("is_flexible"))
        is_next_question = bool(rule.get("is_next_question"))
        targets = sorted(set(rule.get("targets") or []))

        if not targets and is_next_question:
            sig.add(("__NEXT__", codes, is_flexible, is_next_question))
            continue

        for t in targets:
            sig.add((t, codes, is_flexible, is_next_question))

    return sig, True


def _skip_signatures_equivalent(sig_a: set[tuple], sig_b: set[tuple]) -> bool:
    """
    Semantic equivalence for skip signatures with permissive wildcard handling:
    - same routing keys (target/flex/next) are required
    - for a given key, empty code tuple () means "any response" and is treated as wildcard
    """
    def _group(sig: set[tuple]) -> dict[tuple, set[tuple]]:
        out: dict[tuple, set[tuple]] = {}
        for item in sig:
            # item = (target, codes_tuple, is_flexible, is_next_question)
            target, codes, is_flexible, is_next = item
            key = (target, is_flexible, is_next)
            out.setdefault(key, set()).add(tuple(codes))
        return out

    ga = _group(sig_a)
    gb = _group(sig_b)
    if set(ga.keys()) != set(gb.keys()):
        return False

    for key in ga.keys():
        ca = ga[key]
        cb = gb[key]
        if () in ca or () in cb:
            # one side says "any response" for this target route
            continue
        if ca != cb:
            return False
    return True

def _check_skip_consistency(
    skip_text      : str,
    rules          : list[dict],
    q_mandatory_map: dict[str, str],
    curr_qnames    : set[str],
    known_qnames   : set[str],
    source_q       : str,
    next_qname     : str | None,
) -> list[str]:
    """
    Returns a list of issue descriptions (empty = consistent).

    Rules:
    - Each required target Q Name must appear in skip_text.
    - If the rule is flexible ("OR optional question"), a non-mandatory Q Name
      in skip_text is also acceptable.
    - [next question] is treated as equivalent to the concrete next Q Name.
    - Any explicit RHS target that does not exist in current questionnaire
      is treated as a broken reference.

    Range-check policy:
    - Compare ranges only when a target has one unambiguous numeric mapping.
    - Skip range comparison for ambiguous authoring forms (e.g. "otherwise").
    """
    skip_text_s = str(skip_text or "")
    skip_present = set(_extract_qname_like(skip_text_s, curr_qnames))
    skip_mentions_next = bool(_NEXT_Q_RE.search(skip_text_s))
    issues = []

    # Broken target references must stay high severity:
    # RHS targets that do not exist in the current questionnaire.
    skip_targets = _extract_referenced_qnames(
        skip_text_s,
        source_q=source_q,
        ref_qnames=known_qnames,
        curr_qnames=curr_qnames,
    )
    invalid_targets = sorted(t for t in skip_targets if t not in curr_qnames)
    if invalid_targets:
        issues.append(
            f"target question(s) not found in current questionnaire: {invalid_targets}"
        )

    expected_codes_by_target: dict[str, set[int]] = {}
    range_skip_targets: set[str] = set()

    for rule in rules:
        targets = list(rule.get("targets") or [])
        is_flexible = bool(rule.get("is_flexible"))
        is_next_question = bool(rule.get("is_next_question"))
        expected_codes = set(rule.get("option_codes") or set())
        raw_l = str(rule.get("raw") or "").lower()

        # Resolve [next question] to concrete target when possible.
        if is_next_question and not targets and next_qname:
            targets = [next_qname]

        if not targets:
            continue

        found_targets = [t for t in targets if t in skip_present]

        # Equivalence: skip says [next question], rule says concrete next qname.
        if not found_targets and skip_mentions_next and next_qname and next_qname in targets:
            found_targets = [next_qname]

        # Equivalence: rule says [next question], skip says concrete next qname.
        if not found_targets and is_next_question and next_qname and next_qname in skip_present:
            found_targets = [next_qname]

        if not found_targets:
            if is_flexible:
                non_mand = [
                    q for q in skip_present
                    if q_mandatory_map.get(q) not in ("mandatory", "mandatory-panel")
                ]
                if not non_mand:
                    mandatory_targets = sorted(
                        q for q in skip_present
                        if q_mandatory_map.get(q) in ("mandatory", "mandatory-panel")
                    )
                    if mandatory_targets:
                        issues.append(
                            "invalid qname category: "
                            f"expected optional/non-mandatory target; got mandatory target(s) {mandatory_targets}"
                        )
                        continue
                    issues.append(
                        f"expected {targets} (or non-mandatory alternative) not found in skip pattern"
                    )
            else:
                if not skip_text_s.strip():
                    issues.append(
                        f"skip pattern is empty but default/spec says go to {targets}"
                    )
                else:
                    issues.append(
                        f"expected {targets} in skip pattern; got: {skip_text_s[:80]}"
                    )

        ambiguous = (len(targets) != 1) or ("otherwise" in raw_l)
        if ambiguous or not expected_codes:
            for t in targets:
                range_skip_targets.add(t)
        else:
            t = targets[0]
            expected_codes_by_target.setdefault(t, set()).update(expected_codes)

    # Range verification per target (single comparison per target).
    for t, expected in expected_codes_by_target.items():
        if t in range_skip_targets:
            continue

        # next-question equivalence may hide explicit target token in skip text.
        if t not in skip_present and not (skip_mentions_next and next_qname == t):
            continue

        actual_codes = _extract_skip_codes_for_target(skip_text_s, t)
        if actual_codes and actual_codes != expected:
            issues.append(
                f"option range mismatch for target '{t}': "
                f"default/spec says {sorted(expected)}, "
                f"skip pattern has {sorted(actual_codes)}"
            )

    return issues

def validate_skip_patterns(
    current_questions    : pl.DataFrame,
    reference_questions  : pl.DataFrame,
    current_options_en   : pl.DataFrame | None = None,
    reference_options_en : pl.DataFrame | None = None,
    q_mandatory_map      : dict[str, str] | None = None,
) -> pl.DataFrame:
    """
    Three-layer skip-pattern validation:

    1. Template consistency check (mostly medium): current effective skip rule
       should align with template effective rule. Broken references remain high.

    Effective rule selection per side:
       - Current (analyzed): use "Specify skip pattern variable (from blue text)"
         when filled; otherwise fall back to "Skip Pattern", then "Default".
       - Reference/template: use template "Specify" when filled; otherwise
         template "Skip Pattern", then template "Default".

    2. Option code validity (high): numeric codes referenced in Skip Pattern
       must actually exist in the question's answer options.

    3. Option code validity (high): numeric codes referenced in Skip Pattern
       must actually exist in the question's answer options.
    """
    EMPTY_SCHEMA = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field"     : pl.Utf8, "current" : pl.Utf8, "reference": pl.Utf8,
        "severity"  : pl.Utf8, "excel_row": pl.Int64,
    }

    DEF_COL  = "Default skip patterns & conditional "
    SPEC_COL = "Specify skip pattern variable (from blue text)"
    SKIP_COL = "Skip Pattern"

    has_def_cur  = DEF_COL  in current_questions.columns
    has_spec_cur = SPEC_COL in current_questions.columns
    has_skip_cur = SKIP_COL in current_questions.columns

    has_def_ref  = DEF_COL  in reference_questions.columns
    has_spec_ref = SPEC_COL in reference_questions.columns
    has_skip_ref = SKIP_COL in reference_questions.columns

    if not has_def_cur and not has_skip_cur and not has_spec_cur:
        return pl.DataFrame(schema=EMPTY_SCHEMA)

    curr_qnames = set(current_questions["Q Name"].to_list())
    ref_qnames  = set(reference_questions["Q Name"].to_list())
    known_qnames = curr_qnames | ref_qnames
    q_mandatory_map     = q_mandatory_map or {}

    # Option code lookup for validity check (layer 3)
    curr_opt_map: dict[str, set[int]] = {}
    if current_options_en is not None and current_options_en.height > 0:
        for row in current_options_en.select(["Q Name", "option_code"]).iter_rows(named=True):
            q, c = row.get("Q Name"), row.get("option_code")
            if q and c is not None:
                curr_opt_map.setdefault(q, set()).add(int(c))

    # Minimal row maps from current + reference questionnaires
    cur_cols = ["Q Name"] + [c for c in [SKIP_COL, DEF_COL, SPEC_COL, "excel_row"]
                             if c in current_questions.columns]
    ref_cols = ["Q Name"] + [c for c in [SKIP_COL, DEF_COL, SPEC_COL]
                             if c in reference_questions.columns]
    cur_rows = {r["Q Name"]: r
                for r in current_questions.select(cur_cols).iter_rows(named=True)}
    ref_rows = {r["Q Name"]: r
                for r in reference_questions.select(ref_cols).iter_rows(named=True)}

    # Q Name ordering for "next question" resolution
    qname_list = current_questions["Q Name"].to_list()
    qname_next = {q: qname_list[i + 1]
                  for i, q in enumerate(qname_list) if i + 1 < len(qname_list)}

    issues = []

    for qname, cr in cur_rows.items():
        excel_row = cr.get("excel_row")
        skip_text = _normalize_skip_text(cr.get(SKIP_COL) or "")
        def_text  = _normalize_skip_text(cr.get(DEF_COL)  or "")
        spec_text = _normalize_skip_text(cr.get(SPEC_COL) or "")

        rr = ref_rows.get(qname, {})
        ref_skip_text = _normalize_skip_text(rr.get(SKIP_COL) or "") if has_skip_ref else ""
        ref_def_text  = _normalize_skip_text(rr.get(DEF_COL)  or "") if has_def_ref else ""
        ref_spec_text = _normalize_skip_text(rr.get(SPEC_COL) or "") if has_spec_ref else ""

        # Current effective rule: Specify > Skip Pattern > Default
        current_effective = spec_text if spec_text else (skip_text if skip_text else def_text)

        # Reference/template effective rule: Specify > Skip Pattern > Default
        reference_effective = ref_spec_text if ref_spec_text else (ref_skip_text if ref_skip_text else ref_def_text)

        # Current-file emptiness check:
        # Flag when default exists but both Specify and Skip Pattern are empty.
        if def_text and (not spec_text) and (not skip_text):
            issues.append({
                "issue_type": "skip_pattern_empty",
                "set_name": "",
                "Q Name": qname,
                "field": SKIP_COL,
                "current": "Specify + Skip Pattern are empty",
                "reference": f"default present: {def_text[:200]}",
                "severity": "high",
                "excel_row": excel_row,
            })

        # Current-file semantic warning:
        # If Specify is blank and both Skip Pattern + Default are present, flag only when
        # Skip Pattern is semantically inconsistent with the current Default rule.
        if (not spec_text) and skip_text and def_text:
            _def_rules = _parse_default_skip_rules(def_text, known_qnames)
            _default_consistency_issues = []
            if _def_rules:
                _default_consistency_issues = _check_skip_consistency(
                    skip_text,
                    _def_rules,
                    q_mandatory_map,
                    curr_qnames,
                    known_qnames,
                    qname,
                    qname_next.get(qname),
                )
            if _default_consistency_issues:
                issues.append({
                    "issue_type": "default_skip_modified",
                    "set_name": "",
                    "Q Name": qname,
                    "field": SKIP_COL,
                    "current": skip_text[:220],
                    "reference": f"default: {def_text[:200]}",
                    "severity": "info",
                    "excel_row": excel_row,
                })

        # Layer 1: Current effective rule must align with template effective rule.
        if reference_effective and current_effective:
            rules = _parse_default_skip_rules(reference_effective, known_qnames)
            for desc in _check_skip_consistency(
                current_effective,
                rules,
                q_mandatory_map,
                curr_qnames,
                known_qnames,
                qname,
                qname_next.get(qname),
            ):
                is_range = desc.startswith("option range mismatch")
                is_broken_target = desc.startswith("target question(s) not found in current questionnaire")
                is_invalid_category = desc.startswith("invalid qname category")
                is_empty = desc.startswith("skip pattern is empty")
                severity = "high" if (is_broken_target or is_invalid_category or is_empty) else "info"
                if is_broken_target:
                    issue_type = "skipPattern_invalid_qname"
                elif is_invalid_category:
                    issue_type = "skipPattern_invalid_qnameCategory"
                elif is_range:
                    issue_type = "skipPattern_range_mismatch"
                elif is_empty:
                    issue_type = "skip_pattern_empty"
                else:
                    issue_type = "skipPattern_changes"
                issues.append({
                    "issue_type": issue_type,
                    "set_name": "",
                    "Q Name": qname, "field": SKIP_COL,
                    "current": current_effective[:220], "reference": reference_effective[:220],
                    "severity": severity,
                    "excel_row": excel_row,
                })

        # Layer 2: no additional override-only notes are emitted to avoid noise.

        # Layer 3: Option codes in effective current skip rule must exist in options
        if current_effective:
            mentioned = _parse_skip_option_codes(current_effective)
            available = curr_opt_map.get(qname, set())
            if mentioned and available:
                invalid = sorted(c for c in mentioned if c not in available)
                if invalid:
                    issues.append({
                        "issue_type": "skipPattern_range_invalid", "set_name": "",
                        "Q Name": qname, "field": SKIP_COL,
                        "current": f"invalid option code(s): {invalid}",
                        "reference": f"valid codes: {sorted(available)}",
                        "severity": "high", "excel_row": excel_row,
                    })

        # Speculative drift warnings are intentionally not emitted.
        # Option changes are already surfaced in Option Changes and should be reviewed there.

    if not issues:
        return pl.DataFrame(schema=EMPTY_SCHEMA)

    return (
        pl.DataFrame(issues)
        .unique(subset=["issue_type", "Q Name", "field", "current", "reference"], keep="first")
    )






# ======================================================================
# SECTION: ## Step 8  Issue unifiers
_step("Issue unifiers")
# ======================================================================
# --- CODE CELL 10 ---
def make_presence_issues(
    added: pl.DataFrame,
    removed: pl.DataFrame,
    reference_questions: pl.DataFrame | None = None,
) -> pl.DataFrame:
    """
    Converts added/removed question lists into the common issues schema.

    Severity rules:
      - Optional questions (o_ prefix): 'info'   tracked but not a blocker
      - Non-optional added            : 'medium'  unexpected addition, review needed
      - Non-optional removed          : 'high'    mandatory question missing
    """
    added_opt_lookup = (
        added
        .with_columns([
            pl.col("Q Name").cast(pl.Utf8).str.strip_chars().alias("_q"),
            pl.when(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.starts_with("o_"))
            .then(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.slice(2))
            .otherwise(pl.col("Q Name").cast(pl.Utf8).str.strip_chars())
            .str.to_lowercase()
            .alias("_core"),
        ])
        .filter(pl.col("_q").str.starts_with("o_"))
        .select(["_core", pl.col("Q Name").alias("_optional_qname")])
        .unique(subset=["_core"], keep="first")
    )

    if reference_questions is not None and "Mandatory" in reference_questions.columns:
        removed_aug = (
            removed.join(
                reference_questions.select(["Q Name", "Mandatory"]),
                on="Q Name", how="left"
            )
            .with_columns(
                pl.when(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.starts_with("o_"))
                .then(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.slice(2))
                .otherwise(pl.col("Q Name").cast(pl.Utf8).str.strip_chars())
                .str.to_lowercase()
                .alias("_core")
            )
        )
        severity_expr = (
            pl.when(normalize_mandatory_expr("Mandatory") == "yes")
            .then(pl.lit("high"))
            .otherwise(pl.lit("info"))
            .alias("severity")
        )
        moved_to_optional = (
            removed_aug
            .join(added_opt_lookup, on="_core", how="inner")
            .filter(normalize_mandatory_expr("Mandatory") == "yes")
        )
    else:
        removed_aug = removed.with_columns([
            pl.lit("").alias("Mandatory"),
            pl.when(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.starts_with("o_"))
            .then(pl.col("Q Name").cast(pl.Utf8).str.strip_chars().str.slice(2))
            .otherwise(pl.col("Q Name").cast(pl.Utf8).str.strip_chars())
            .str.to_lowercase()
            .alias("_core"),
        ])
        severity_expr = (
            pl.when(pl.col("is_optional")).then(pl.lit("info")).otherwise(pl.lit("high"))
            .alias("severity")
        )
        moved_to_optional = pl.DataFrame(schema={
            "Q Name": pl.Utf8, "is_optional": pl.Boolean, "Mandatory": pl.Utf8,
            "_core": pl.Utf8, "_optional_qname": pl.Utf8,
        })

    moved_issues = (
        moved_to_optional.with_columns([
            pl.lit("mandatory_to_optional").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("Q Name").alias("field"),
            pl.col("_optional_qname").alias("current"),
            pl.lit("present (mandatory in reference)").alias("reference"),
            pl.lit("high").alias("severity"),
            pl.lit(None).cast(pl.Int64).alias("excel_row"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )

    added_base = (
        added.join(moved_to_optional.select(pl.col("_optional_qname").alias("Q Name")), on="Q Name", how="anti")
        if moved_to_optional.height > 0 else added
    )
    added_issues = added_base.with_columns([
        pl.lit("added_question").alias("issue_type"),
        pl.lit("").alias("set_name"),
        pl.lit("Q Name").alias("field"),
        pl.lit("present").alias("current"),
        pl.lit("missing_in_reference").alias("reference"),
        pl.lit("info").alias("severity"),
        pl.lit(None).cast(pl.Int64).alias("excel_row"),
    ]).select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])

    removed_base = (
        removed_aug.join(moved_to_optional.select(["Q Name"]), on="Q Name", how="anti")
        if moved_to_optional.height > 0 else removed_aug
    )
    removed_issues = removed_base.with_columns([
        pl.lit("removed_question").alias("issue_type"),
        pl.lit("").alias("set_name"),
        pl.lit("Q Name").alias("field"),
        pl.lit("missing_in_current").alias("current"),
        pl.lit("present").alias("reference"),
        severity_expr,
        pl.lit(None).cast(pl.Int64).alias("excel_row"),
    ]).select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])

    return pl.concat([added_issues, removed_issues, moved_issues], how="vertical")


def make_mandatory_issues(mandatory_diff: pl.DataFrame) -> pl.DataFrame:
    """Converts mandatory mismatches into the common issues schema."""
    return (
        mandatory_diff
        .with_columns([
            pl.lit("").alias("set_name"),
            pl.col("Mandatory").alias("current"),
            pl.col("Mandatory_ref").alias("reference"),
            pl.lit("high").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )


def make_option_issues(option_diff: pl.DataFrame) -> pl.DataFrame:
    """Converts option label mismatches into the common issues schema."""
    return (
        option_diff
        .with_columns([
            pl.lit("").alias("set_name"),
            pl.col("option_label").alias("current"),
            pl.col("option_label_ref").alias("reference"),
            pl.lit("medium").alias("severity"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )


# Note: validate_critical_sets, validate_prefix_counts, validate_crop_harvest,
# and validate_skip_patterns already return the common schema directly.


def make_option_presence_issues(added_options: pl.DataFrame, removed_options: pl.DataFrame) -> pl.DataFrame:
    """Converts added/removed options into the common issues schema."""
    added_issues = added_options.with_columns([
        pl.lit("").alias("set_name"),
        pl.col("option_label").alias("current"),
        pl.lit(_not_in_reference_text()).alias("reference"),
        pl.lit("medium").alias("severity"),
        pl.lit(None).cast(pl.Int64).alias("excel_row"),
    ]).select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])

    removed_issues = removed_options.with_columns([
        pl.lit("").alias("set_name"),
        pl.lit("(removed)").alias("current"),
        pl.col("option_label").alias("reference"),
        pl.lit("high").alias("severity"),
        pl.lit(None).cast(pl.Int64).alias("excel_row"),
    ]).select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])

    return pl.concat([added_issues, removed_issues], how="vertical")


def build_option_changes_view(
    en_option_diff: pl.DataFrame,
    en_added_options: pl.DataFrame,
    en_removed_options: pl.DataFrame,
    en_code_renumber: pl.DataFrame,
    tgt_option_diff: pl.DataFrame,
    tgt_added_options: pl.DataFrame,
    tgt_removed_options: pl.DataFrame,
    tgt_code_renumber: pl.DataFrame,
    target_lang: str,
) -> pl.DataFrame:
    """
    Builds Option Changes rows with independent EN and target-language checks.

    Layout policy:
    - Current value / Reference / rule are always EN baseline.
    - For non-EN runs, target language values are added in separate columns.
    """
    is_en = str(target_lang).upper() == "EN"

    def _scope_view(option_diff, added_opts, removed_opts, renumber_opts, scope):
        mismatch = option_diff.with_columns([
            pl.col("option_label").alias("current"),
            pl.col("option_label_ref").alias("reference"),
            pl.col("excel_row").cast(pl.Int64).alias("excel_row"),
        ]).select(["issue_type", "Q Name", "field", "current", "reference", "severity", "excel_row"])

        added = added_opts.with_columns([
            pl.col("option_label").alias("current"),
            pl.lit(_not_in_reference_text()).alias("reference"),
            pl.lit(None).cast(pl.Int64).alias("excel_row"),
        ]).select(["issue_type", "Q Name", "field", "current", "reference", "severity", "excel_row"])

        removed = removed_opts.with_columns([
            pl.lit("(removed)").alias("current"),
            pl.col("option_label").alias("reference"),
            pl.lit(None).cast(pl.Int64).alias("excel_row"),
        ]).select(["issue_type", "Q Name", "field", "current", "reference", "severity", "excel_row"])

        renumber = renumber_opts.with_columns([
            pl.concat_str([
                pl.col("new_code").cast(pl.Utf8),
                pl.lit(") "),
                pl.col("option_label").cast(pl.Utf8),
            ]).alias("current"),
            pl.concat_str([
                pl.col("old_code").cast(pl.Utf8),
                pl.lit(") "),
                pl.col("option_label_ref").cast(pl.Utf8),
            ]).alias("reference"),
            pl.col("excel_row").cast(pl.Int64).alias("excel_row"),
        ]).select(["issue_type", "Q Name", "field", "current", "reference", "severity", "excel_row"])

        base = pl.concat([mismatch, added, removed, renumber], how="vertical")
        return base.rename({
            "current": f"current_{scope.lower()}",
            "reference": f"reference_{scope.lower()}",
            "severity": f"severity_{scope.lower()}",
            "excel_row": f"excel_row_{scope.lower()}",
        })

    en_view = _scope_view(en_option_diff, en_added_options, en_removed_options, en_code_renumber, "EN")
    if is_en:
        return (
            en_view
            .with_columns([
                pl.lit("").alias("set_name"),
                pl.col("current_en").alias("current"),
                pl.col("reference_en").alias("reference"),
                pl.col("severity_en").alias("severity"),
                pl.col("excel_row_en").alias("excel_row"),
            ])
            .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
        )

    tgt_view = _scope_view(tgt_option_diff, tgt_added_options, tgt_removed_options, tgt_code_renumber, "TGT")
    keys = ["issue_type", "Q Name", "field"]
    out = en_view.join(tgt_view, on=keys, how="full", suffix="_t")

    return (
        out
        .with_columns([
            pl.coalesce([pl.col("issue_type"), pl.col("issue_type_t")]).alias("issue_type"),
            pl.coalesce([pl.col("Q Name"), pl.col("Q Name_t")]).alias("Q Name"),
            pl.coalesce([pl.col("field"), pl.col("field_t")]).alias("field"),
            pl.lit("").alias("set_name"),
            pl.coalesce([pl.col("current_en"), pl.lit("")]).alias("current"),
            pl.coalesce([pl.col("reference_en"), pl.lit("")]).alias("reference"),
            pl.coalesce([pl.col("current_tgt"), pl.lit("")]).alias("current_lang"),
            pl.coalesce([pl.col("reference_tgt"), pl.lit("")]).alias("reference_lang"),
            pl.when(pl.col("severity_tgt") == "high").then(pl.lit("high"))
            .when(pl.col("severity_en") == "high").then(pl.lit("high"))
            .when(pl.col("severity_tgt") == "medium").then(pl.lit("medium"))
            .when(pl.col("severity_en") == "medium").then(pl.lit("medium"))
            .otherwise(pl.lit("info")).alias("severity"),
            pl.coalesce([pl.col("excel_row_tgt"), pl.col("excel_row_en")]).alias("excel_row"),
        ])
        .select([
            "issue_type", "set_name", "Q Name", "field", "current", "reference",
            "current_lang", "reference_lang", "severity", "excel_row",
        ])
    )




# ======================================================================
# SECTION: ## Step 9  Run the full pipeline
_step("Running pipeline")
# ======================================================================
# --- CODE CELL 11 ---
def compare_question_presence(
    current_questions : pl.DataFrame,
    reference_questions: pl.DataFrame,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    """
    Returns (added, removed). Each DataFrame includes `is_optional`.
    """
    key_col = "Q Name"

    added = (
        current_questions.select([key_col])
        .join(reference_questions.select([key_col]), on=key_col, how="anti")
        .with_columns(pl.col(key_col).str.starts_with("o_").alias("is_optional"))
    )

    removed = (
        reference_questions.select([key_col])
        .join(current_questions.select([key_col]), on=key_col, how="anti")
        .with_columns(pl.col(key_col).str.starts_with("o_").alias("is_optional"))
    )

    return added, removed


def compare_mandatory(
    current_questions : pl.DataFrame,
    reference_questions: pl.DataFrame,
    treat_blank_as_no: bool = True,
) -> pl.DataFrame:
    """
    Compares Mandatory column values by Q Name.
    """
    if "Mandatory" not in current_questions.columns or "Mandatory" not in reference_questions.columns:
        return pl.DataFrame(schema={
            "issue_type": pl.Utf8, "Q Name": pl.Utf8, "field": pl.Utf8,
            "Mandatory": pl.Utf8, "Mandatory_ref": pl.Utf8, "excel_row": pl.Int64,
        })

    curr_non_blank = (
        current_questions
        .select(pl.col("Mandatory").cast(pl.Utf8).fill_null("").str.strip_chars().alias("_m"))
        .filter(pl.col("_m") != "")
        .height
    )
    ref_non_blank = (
        reference_questions
        .select(pl.col("Mandatory").cast(pl.Utf8).fill_null("").str.strip_chars().alias("_m"))
        .filter(pl.col("_m") != "")
        .height
    )

    # If one side has no usable mandatory values at all, row-by-row mismatch output
    # becomes noisy and low-signal. Emit one structural issue instead.
    if curr_non_blank == 0 and ref_non_blank == 0:
        return pl.DataFrame(schema={
            "issue_type": pl.Utf8, "Q Name": pl.Utf8, "field": pl.Utf8,
            "Mandatory": pl.Utf8, "Mandatory_ref": pl.Utf8, "excel_row": pl.Int64,
        })
    if curr_non_blank == 0 or ref_non_blank == 0:
        return pl.DataFrame([
            {
                "issue_type": "mandatory_source_missing",
                "Q Name": "",
                "field": "Mandatory",
                "Mandatory": f"non_blank_values={curr_non_blank}",
                "Mandatory_ref": f"non_blank_values={ref_non_blank}",
                "excel_row": None,
            }
        ], schema={
            "issue_type": pl.Utf8, "Q Name": pl.Utf8, "field": pl.Utf8,
            "Mandatory": pl.Utf8, "Mandatory_ref": pl.Utf8, "excel_row": pl.Int64,
        })

    joined = (
        current_questions.select(["Q Name", "Mandatory", "excel_row"])
        .join(reference_questions.select(["Q Name", "Mandatory"]), on="Q Name", how="inner", suffix="_ref")
        .with_columns([
            pl.col("Mandatory").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Mandatory"),
            pl.col("Mandatory_ref").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Mandatory_ref"),
        ])
    )

    if treat_blank_as_no:
        joined = joined.with_columns([
            pl.when(pl.col("Mandatory") == "").then(pl.lit("No")).otherwise(pl.col("Mandatory")).alias("Mandatory"),
            pl.when(pl.col("Mandatory_ref") == "").then(pl.lit("No")).otherwise(pl.col("Mandatory_ref")).alias("Mandatory_ref"),
        ])

    return (
        joined
        .filter(normalize_text_expr("Mandatory") != normalize_text_expr("Mandatory_ref"))
        .with_columns([
            pl.lit("mandatory_column_mismatch").alias("issue_type"),
            pl.lit("Mandatory").alias("field"),
        ])
        .select(["issue_type", "Q Name", "field", "Mandatory", "Mandatory_ref", "excel_row"])
    )


def _mandatory_cat_expr(qname_col: str, mandatory_col: str) -> pl.Expr:
    qn = pl.col(qname_col).cast(pl.Utf8).fill_null("").str.strip_chars().str.to_lowercase()
    md = pl.col(mandatory_col).cast(pl.Utf8).fill_null("").str.strip_chars().str.to_lowercase()
    return (
        pl.when(qn.str.starts_with("o_")).then(pl.lit("optional"))
        .when(md.str.contains("panel")).then(pl.lit("mandatory-panel"))
        .when(md.is_in(["yes", "y", "true", "1"])).then(pl.lit("mandatory"))
        .otherwise(pl.lit("non-mandatory"))
    )


def _normalize_qtype_value(value: str) -> str:
    s = str(value or "").strip().lower()
    if not s:
        return ""
    s = re.sub(r"[\\/_-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Canonical families for stable comparisons.
    if "select all that apply" in s:
        return "select all that apply"
    if "single choice" in s:
        return "single choice"
    return s


def _qtype_mode(value: str) -> str:
    t = _normalize_qtype_value(value)
    if not t:
        return ""
    if "select all that apply" in t:
        return "multi_select"
    if "single choice" in t:
        return "single_select"
    if is_option_bearing_qtype(t):
        return "option_bearing"
    return "non_option"


def _qtype_change_severity(
    curr_mode: str,
    ref_mode: str,
    curr_option_count: int,
    ref_option_count: int,
) -> str:
    """
    Severity policy:
    - HIGH for incompatible/invalid transitions (including non-mandatory questions)
    - MEDIUM for valid type changes
    """
    cm = str(curr_mode or "")
    rm = str(ref_mode or "")
    modes = {cm, rm}

    # Incompatible family change: single <-> multi.
    if "single_select" in modes and "multi_select" in modes:
        return "high"

    # Incompatible option-bearing <-> non-option mode.
    if "non_option" in modes and (("single_select" in modes) or ("multi_select" in modes) or ("option_bearing" in modes)):
        return "high"

    # Invalid option-count correctness checks:
    # option-bearing/select questions should normally have at least 2 options.
    curr_option_like = cm in {"single_select", "multi_select", "option_bearing"}
    ref_option_like = rm in {"single_select", "multi_select", "option_bearing"}
    if curr_option_like and int(curr_option_count or 0) <= 1:
        return "high"
    if ref_option_like and int(ref_option_count or 0) <= 1:
        return "high"
    if (cm == "non_option") and int(curr_option_count or 0) > 0:
        return "high"
    if (rm == "non_option") and int(ref_option_count or 0) > 0:
        return "high"

    # Changed but compatible/valid.
    return "medium"


def compare_qtype_changes(
    current_questions: pl.DataFrame,
    reference_questions: pl.DataFrame,
    current_options_en: pl.DataFrame | None = None,
    reference_options_en: pl.DataFrame | None = None,
) -> pl.DataFrame:
    """
    Risk-based Q Type comparison:
    - normalises wording to avoid cosmetic false positives
    - escalates severity for mandatory questions and option-bearing mode shifts
    """
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if "Q Type" not in current_questions.columns or "Q Type" not in reference_questions.columns:
        return pl.DataFrame(schema=EMPTY)

    # Option counts per Q Name (English options baseline).
    if current_options_en is not None and current_options_en.height > 0:
        _cur_opt_counts = (
            current_options_en
            .group_by("Q Name")
            .agg(pl.len().alias("_curr_opt_n"))
        )
    else:
        _cur_opt_counts = pl.DataFrame(schema={"Q Name": pl.Utf8, "_curr_opt_n": pl.Int64})

    if reference_options_en is not None and reference_options_en.height > 0:
        _ref_opt_counts = (
            reference_options_en
            .group_by("Q Name")
            .agg(pl.len().alias("_ref_opt_n"))
        )
    else:
        _ref_opt_counts = pl.DataFrame(schema={"Q Name": pl.Utf8, "_ref_opt_n": pl.Int64})

    joined = (
        current_questions
        .select(["Q Name", "Q Type", "Mandatory", "excel_row"])
        .join(
            reference_questions.select(["Q Name", "Q Type", "Mandatory"]),
            on="Q Name", how="inner", suffix="_ref",
        )
        .with_columns([
            pl.col("Q Type").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Q Type"),
            pl.col("Q Type_ref").cast(pl.Utf8).fill_null("").str.strip_chars().alias("Q Type_ref"),
            pl.coalesce([pl.col("Mandatory"), pl.col("Mandatory_ref")]).alias("_mand_base"),
        ])
        .with_columns([
            pl.col("Q Type").map_elements(_normalize_qtype_value, return_dtype=pl.Utf8).alias("_type_norm"),
            pl.col("Q Type_ref").map_elements(_normalize_qtype_value, return_dtype=pl.Utf8).alias("_type_norm_ref"),
            pl.col("Q Type").map_elements(_qtype_mode, return_dtype=pl.Utf8).alias("_type_mode"),
            pl.col("Q Type_ref").map_elements(_qtype_mode, return_dtype=pl.Utf8).alias("_type_mode_ref"),
        ])
        .join(_cur_opt_counts, on="Q Name", how="left")
        .join(_ref_opt_counts, on="Q Name", how="left")
        .with_columns([
            pl.col("_curr_opt_n").fill_null(0).cast(pl.Int64),
            pl.col("_ref_opt_n").fill_null(0).cast(pl.Int64),
        ])
        .filter(pl.col("_type_norm") != pl.col("_type_norm_ref"))
        .with_columns(
            pl.struct(["_type_mode", "_type_mode_ref", "_curr_opt_n", "_ref_opt_n"])
            .map_elements(
                lambda r: _qtype_change_severity(
                    r.get("_type_mode"),
                    r.get("_type_mode_ref"),
                    r.get("_curr_opt_n"),
                    r.get("_ref_opt_n"),
                ),
                return_dtype=pl.Utf8,
            )
            .alias("severity")
        )
    )
    if joined.height == 0:
        return pl.DataFrame(schema=EMPTY)

    return (
        joined
        .with_columns([
            pl.lit("qtype_changed").alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit("Q Type").alias("field"),
            pl.col("Q Type").alias("current"),
            pl.col("Q Type_ref").alias("reference"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )


def compare_legacy_text_field_changes(
    current_questions: pl.DataFrame,
    reference_questions: pl.DataFrame,
    column_name: str,
    issue_type: str,
    default_severity: str = "medium",
    high_when_mandatory: bool = False,
) -> pl.DataFrame:
    """Legacy column parity checks with normalization to avoid cosmetic false positives."""
    EMPTY = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if column_name not in current_questions.columns or column_name not in reference_questions.columns:
        return pl.DataFrame(schema=EMPTY)

    joined = (
        current_questions
        .select(["Q Name", column_name, "Mandatory", "excel_row"])
        .join(
            reference_questions.select(["Q Name", column_name, "Mandatory"]),
            on="Q Name", how="inner", suffix="_ref",
        )
        .with_columns([
            pl.col(column_name).cast(pl.Utf8).fill_null("").str.strip_chars().alias("_curr"),
            pl.col(f"{column_name}_ref").cast(pl.Utf8).fill_null("").str.strip_chars().alias("_ref"),
            pl.coalesce([pl.col("Mandatory"), pl.col("Mandatory_ref")]).alias("_mand_base"),
        ])
        .with_columns([
            normalize_text_expr("_curr").alias("_curr_norm"),
            normalize_text_expr("_ref").alias("_ref_norm"),
            _mandatory_cat_expr("Q Name", "_mand_base").alias("_mandatory_cat"),
        ])
        .filter((pl.col("_curr_norm") != pl.col("_ref_norm")) & ((pl.col("_curr_norm") != "") | (pl.col("_ref_norm") != "")))
    )
    if joined.height == 0:
        return pl.DataFrame(schema=EMPTY)

    if high_when_mandatory:
        joined = joined.with_columns(
            pl.when(pl.col("_mandatory_cat").is_in(["mandatory", "mandatory-panel"]))
            .then(pl.lit("high"))
            .otherwise(pl.lit(default_severity))
            .alias("severity")
        )
    else:
        joined = joined.with_columns(pl.lit(default_severity).alias("severity"))

    return (
        joined
        .with_columns([
            pl.lit(issue_type).alias("issue_type"),
            pl.lit("").alias("set_name"),
            pl.lit(column_name).alias("field"),
            pl.col("_curr").alias("current"),
            pl.col("_ref").alias("reference"),
        ])
        .select(["issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"])
    )


def compare_question_labels_single(
    current_questions : pl.DataFrame,
    reference_questions: pl.DataFrame,
    current_text_col: str,
    reference_text_col: str,
    lang_scope: str,
) -> pl.DataFrame:
    """
    Question stem (text label) changes for matching Q Name in one language scope.
    Numbered option blocks are removed so option labels are handled only in Option Changes.
    """
    if current_text_col not in current_questions.columns or reference_text_col not in reference_questions.columns:
        return pl.DataFrame(schema={
            "issue_type": pl.Utf8,
            "Q Name": pl.Utf8,
            "field": pl.Utf8,
            "current": pl.Utf8,
            "reference": pl.Utf8,
            "severity": pl.Utf8,
            "excel_row": pl.Int64,
            "lang_scope": pl.Utf8,
        })

    return (
        current_questions
        .select([
            "Q Name",
            "excel_row",
            pl.col(current_text_col).alias("_curr_text"),
        ])
        .join(
            reference_questions.select([
                "Q Name",
                pl.col(reference_text_col).alias("_ref_text"),
            ]),
            on="Q Name",
            how="inner",
        )
        .with_columns([
            pl.col("_curr_text")
            .map_elements(extract_question_stem, return_dtype=pl.Utf8)
            .fill_null("")
            .str.strip_chars()
            .alias("current"),
            pl.col("_ref_text")
            .map_elements(extract_question_stem, return_dtype=pl.Utf8)
            .fill_null("")
            .str.strip_chars()
            .alias("reference"),
        ])
        .with_columns([
            normalize_text_expr("current").alias("_current_norm"),
            normalize_text_expr("reference").alias("_reference_norm"),
        ])
        .filter(pl.col("_current_norm") != pl.col("_reference_norm"))
        .with_columns([
            pl.lit("question_label_mismatch").alias("issue_type"),
            pl.lit("label").alias("field"),
            pl.lit("medium").alias("severity"),
            pl.lit(lang_scope).alias("lang_scope"),
            pl.col("excel_row").cast(pl.Int64).alias("excel_row"),
        ])
        .select(["issue_type", "Q Name", "field", "current", "reference", "severity", "excel_row", "lang_scope"])
    )


def build_question_changes_view(
    core_question_issues: pl.DataFrame,
    en_label_diff: pl.DataFrame,
    tgt_label_diff: pl.DataFrame,
    target_lang: str,
) -> pl.DataFrame:
    """
    Builds Question Changes rows with independent EN and target-language label checks.

    Layout policy:
    - Current value / Reference / rule are always EN baseline.
    - For non-EN runs, target language values are added in separate columns.
    """
    is_en = str(target_lang).upper() == "EN"

    if is_en:
        label_view = (
            en_label_diff
            .with_columns([
                pl.lit("").alias("set_name"),
                pl.lit("EN").alias("lang_scope"),
                pl.col("current").cast(pl.Utf8).fill_null("").alias("current"),
                pl.col("reference").cast(pl.Utf8).fill_null("").alias("reference"),
                pl.col("severity").cast(pl.Utf8).fill_null("info").alias("severity"),
                pl.col("excel_row").cast(pl.Int64).alias("excel_row"),
            ])
            .select(["issue_type", "set_name", "Q Name", "field", "lang_scope", "current", "reference", "severity", "excel_row"])
        )

        core_view = (
            core_question_issues
            .with_columns([
                pl.col("set_name").cast(pl.Utf8).fill_null("").alias("set_name"),
                pl.col("field").cast(pl.Utf8).fill_null("").alias("field"),
                pl.lit("N/A").alias("lang_scope"),
                pl.col("current").cast(pl.Utf8).fill_null("").alias("current"),
                pl.col("reference").cast(pl.Utf8).fill_null("").alias("reference"),
                pl.col("severity").cast(pl.Utf8).fill_null("info").alias("severity"),
                pl.col("excel_row").cast(pl.Int64),
            ])
            .select(["issue_type", "set_name", "Q Name", "field", "lang_scope", "current", "reference", "severity", "excel_row"])
        )
        return pl.concat([core_view, label_view], how="vertical")

    def _scope_view(label_diff: pl.DataFrame, scope: str) -> pl.DataFrame:
        return (
            label_diff
            .rename({
                "current": f"current_{scope.lower()}",
                "reference": f"reference_{scope.lower()}",
                "severity": f"severity_{scope.lower()}",
                "excel_row": f"excel_row_{scope.lower()}",
            })
            .select([
                "issue_type", "Q Name", "field",
                f"current_{scope.lower()}", f"reference_{scope.lower()}",
                f"severity_{scope.lower()}", f"excel_row_{scope.lower()}",
            ])
        )

    en_view = _scope_view(en_label_diff, "EN")
    tgt_view = _scope_view(tgt_label_diff, "TGT")
    keys = ["issue_type", "Q Name", "field"]
    out = en_view.join(tgt_view, on=keys, how="full", suffix="_t")
    label_view = (
        out
        .with_columns([
            pl.coalesce([pl.col("issue_type"), pl.col("issue_type_t")]).alias("issue_type"),
            pl.coalesce([pl.col("Q Name"), pl.col("Q Name_t")]).alias("Q Name"),
            pl.coalesce([pl.col("field"), pl.col("field_t")]).alias("field"),
            pl.lit("").alias("set_name"),
            # EN baseline in primary columns
            pl.coalesce([pl.col("current_en"), pl.lit("")]).alias("current"),
            pl.coalesce([pl.col("reference_en"), pl.lit("")]).alias("reference"),
            # Target language shown in dedicated columns
            pl.coalesce([pl.col("current_tgt"), pl.lit("")]).alias("current_lang"),
            pl.coalesce([pl.col("reference_tgt"), pl.lit("")]).alias("reference_lang"),
            pl.when(pl.col("severity_tgt") == "high").then(pl.lit("high"))
            .when(pl.col("severity_en") == "high").then(pl.lit("high"))
            .when(pl.col("severity_tgt") == "medium").then(pl.lit("medium"))
            .when(pl.col("severity_en") == "medium").then(pl.lit("medium"))
            .otherwise(pl.lit("info")).alias("severity"),
            pl.coalesce([pl.col("excel_row_tgt"), pl.col("excel_row_en")]).alias("excel_row"),
            pl.when(
                ((pl.col("current_en").cast(pl.Utf8).fill_null("").str.strip_chars() != "")
                 | (pl.col("reference_en").cast(pl.Utf8).fill_null("").str.strip_chars() != ""))
                &
                ((pl.col("current_tgt").cast(pl.Utf8).fill_null("").str.strip_chars() != "")
                 | (pl.col("reference_tgt").cast(pl.Utf8).fill_null("").str.strip_chars() != ""))
            ).then(pl.lit(f"EN+{target_lang}"))
            .when(
                (pl.col("current_en").cast(pl.Utf8).fill_null("").str.strip_chars() != "")
                | (pl.col("reference_en").cast(pl.Utf8).fill_null("").str.strip_chars() != "")
            ).then(pl.lit("EN"))
            .when(
                (pl.col("current_tgt").cast(pl.Utf8).fill_null("").str.strip_chars() != "")
                | (pl.col("reference_tgt").cast(pl.Utf8).fill_null("").str.strip_chars() != "")
            ).then(pl.lit(target_lang))
            .otherwise(pl.lit(target_lang))
            .alias("lang_scope"),
        ])
        .select([
            "issue_type", "set_name", "Q Name", "field", "lang_scope", "current", "reference",
            "current_lang", "reference_lang", "severity", "excel_row",
        ])
    )

    core_view = (
        core_question_issues
        .with_columns([
            pl.col("set_name").cast(pl.Utf8).fill_null("").alias("set_name"),
            pl.col("field").cast(pl.Utf8).fill_null("").alias("field"),
            pl.lit("N/A").alias("lang_scope"),
            pl.col("current").cast(pl.Utf8).fill_null("").alias("current"),
            pl.col("reference").cast(pl.Utf8).fill_null("").alias("reference"),
            pl.lit("").alias("current_lang"),
            pl.lit("").alias("reference_lang"),
            pl.col("severity").cast(pl.Utf8).fill_null("info").alias("severity"),
            pl.col("excel_row").cast(pl.Int64),
        ])
        .select([
            "issue_type", "set_name", "Q Name", "field", "lang_scope", "current", "reference",
            "current_lang", "reference_lang", "severity", "excel_row",
        ])
    )

    return pl.concat([core_view, label_view], how="vertical")


def compare_option_labels_single(
    current_options : pl.DataFrame,
    reference_options: pl.DataFrame,
    lang_scope: str,
) -> pl.DataFrame:
    """
    Option label changes for matching Q Name + option_code.
    """
    return (
        current_options
        .join(reference_options, on=["Q Name", "option_code"], how="inner", suffix="_ref")
        .with_columns([
            normalize_text_expr("option_label").alias("option_label_norm"),
            normalize_text_expr("option_label_ref").alias("option_label_ref_norm"),
        ])
        .filter(pl.col("option_label_norm") != pl.col("option_label_ref_norm"))
        .with_columns([
            pl.lit("option_label_mismatch").alias("issue_type"),
            pl.concat_str([pl.lit("option_"), pl.col("option_code").cast(pl.Utf8)]).alias("field"),
            pl.lit("medium").alias("severity"),
            pl.lit(lang_scope).alias("lang_scope"),
        ])
        .select(["issue_type", "Q Name", "field", "option_label", "option_label_ref", "severity", "excel_row", "lang_scope"])
    )


def compare_option_presence_single(
    current_options : pl.DataFrame,
    reference_options: pl.DataFrame,
    lang_scope: str,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    """
    Added/removed options for one language scope.
    """
    key_cols = ["Q Name", "option_code"]

    removed_options = (
        reference_options.select(key_cols + ["option_label"])
        .join(current_options.select(key_cols), on=key_cols, how="anti")
        .with_columns([
            pl.lit("removed_option").alias("issue_type"),
            pl.concat_str([pl.lit("option_"), pl.col("option_code").cast(pl.Utf8)]).alias("field"),
            pl.lit("high").alias("severity"),
            pl.lit(lang_scope).alias("lang_scope"),
        ])
        .select(["issue_type", "Q Name", "field", "option_label", "severity", "lang_scope"])
    )

    added_options = (
        current_options.select(key_cols + ["option_label"])
        .join(reference_options.select(key_cols), on=key_cols, how="anti")
        .with_columns([
            pl.lit("added_option").alias("issue_type"),
            pl.concat_str([pl.lit("option_"), pl.col("option_code").cast(pl.Utf8)]).alias("field"),
            pl.lit("medium").alias("severity"),
            pl.lit(lang_scope).alias("lang_scope"),
        ])
        .select(["issue_type", "Q Name", "field", "option_label", "severity", "lang_scope"])
    )

    return added_options, removed_options


def compare_option_code_renumber_single(
    current_options: pl.DataFrame,
    reference_options: pl.DataFrame,
    lang_scope: str,
) -> pl.DataFrame:
    """
    Detect explicit option-position renumbering for the same normalized label
    within the same question: e.g., position 12 becomes 15.

    We only emit when label->code mapping is 1:1 on both sides to avoid
    ambiguous many-to-many matches.
    """
    cur = (
        current_options
        .with_columns(normalize_text_expr("option_label").alias("label_norm"))
        .filter(pl.col("label_norm") != "")
        .group_by(["Q Name", "label_norm"])
        .agg([
            pl.col("option_code").n_unique().alias("cur_n_codes"),
            pl.col("option_code").first().alias("new_code"),
            pl.col("option_label").first().alias("option_label"),
            pl.col("excel_row").min().alias("excel_row"),
        ])
    )

    ref = (
        reference_options
        .with_columns(normalize_text_expr("option_label").alias("label_norm"))
        .filter(pl.col("label_norm") != "")
        .group_by(["Q Name", "label_norm"])
        .agg([
            pl.col("option_code").n_unique().alias("ref_n_codes"),
            pl.col("option_code").first().alias("old_code"),
            pl.col("option_label").first().alias("option_label_ref"),
        ])
    )

    return (
        cur.join(ref, on=["Q Name", "label_norm"], how="inner")
        .filter(
            (pl.col("cur_n_codes") == 1)
            & (pl.col("ref_n_codes") == 1)
            & (pl.col("new_code") != pl.col("old_code"))
        )
        .with_columns([
            pl.lit("option_position_renumbered_same_label").alias("issue_type"),
            pl.concat_str([
                pl.lit("position_"),
                pl.col("old_code").cast(pl.Utf8),
                pl.lit("_to_"),
                pl.col("new_code").cast(pl.Utf8),
            ]).alias("field"),
            pl.lit("high").alias("severity"),
            pl.lit(lang_scope).alias("lang_scope"),
        ])
        .select([
            "issue_type", "Q Name", "field",
            "old_code", "new_code",
            "option_label", "option_label_ref",
            "severity", "excel_row", "lang_scope",
        ])
    )


def compare_codes_presence_single(
    current_codes: pl.DataFrame,
    reference_codes: pl.DataFrame,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    """Added/removed numeric code entries in the survey 'Codes' column."""
    key_cols = ["Q Name", "code_num"]

    removed_codes = (
        reference_codes.select(key_cols + ["code_token"])
        .join(current_codes.select(key_cols), on=key_cols, how="anti")
        .with_columns([
            pl.lit("codes_col_removed").alias("issue_type"),
            pl.concat_str([pl.lit("codes_"), pl.col("code_num").cast(pl.Utf8)]).alias("field"),
            pl.lit("high").alias("severity"),
        ])
        .select(["issue_type", "Q Name", "field", "code_num", "code_token", "severity"])
    )

    added_codes = (
        current_codes.select(key_cols + ["code_token", "excel_row"])
        .join(reference_codes.select(key_cols), on=key_cols, how="anti")
        .with_columns([
            pl.lit("codes_col_added").alias("issue_type"),
            pl.concat_str([pl.lit("codes_"), pl.col("code_num").cast(pl.Utf8)]).alias("field"),
            pl.lit("medium").alias("severity"),
        ])
        .select(["issue_type", "Q Name", "field", "code_num", "code_token", "severity", "excel_row"])
    )

    return added_codes, removed_codes


def compare_codes_token_mismatch_single(
    current_codes: pl.DataFrame,
    reference_codes: pl.DataFrame,
    renumber_codes: pl.DataFrame | None = None,
) -> pl.DataFrame:
    """
    Same code number exists in both, but token text differs in 'Codes' column.

    Precedence rule:
    - If a code position is already explained by a same-token renumber event
      (old_code_num->new_code_num) for the same Q Name, suppress token-mismatch
      rows on those positions to avoid double-reporting reorder as rename.
    """
    mismatch = (
        current_codes
        .join(reference_codes, on=["Q Name", "code_num"], how="inner", suffix="_ref")
        .with_columns([
            normalize_code_token_expr("code_token").alias("code_token_norm"),
            normalize_code_token_expr("code_token_ref").alias("code_token_ref_norm"),
        ])
        .filter(pl.col("code_token_norm") != pl.col("code_token_ref_norm"))
        .with_columns([
            pl.lit("codes_col_token_mismatch").alias("issue_type"),
            pl.concat_str([pl.lit("codes_"), pl.col("code_num").cast(pl.Utf8)]).alias("field"),
            pl.lit("high").alias("severity"),
        ])
        .select([
            "issue_type", "Q Name", "field", "code_num",
            "code_token", "code_token_ref", "severity", "excel_row",
        ])
    )

    if renumber_codes is None or renumber_codes.height == 0:
        return mismatch

    renumber_positions = (
        pl.concat([
            renumber_codes.select([
                pl.col("Q Name"),
                pl.col("old_code_num").cast(pl.Int64).alias("code_num"),
            ]),
            renumber_codes.select([
                pl.col("Q Name"),
                pl.col("new_code_num").cast(pl.Int64).alias("code_num"),
            ]),
        ], how="vertical")
        .unique()
    )

    return mismatch.join(renumber_positions, on=["Q Name", "code_num"], how="anti")


def compare_codes_renumber_single(
    current_codes: pl.DataFrame,
    reference_codes: pl.DataFrame,
) -> pl.DataFrame:
    """
    Detect same normalized code token moved to a different number in 'Codes'.
    Emits only 1:1 token->number mappings to avoid ambiguous many-to-many cases.
    """
    cur = (
        current_codes
        .with_columns(normalize_code_token_expr("code_token").alias("token_norm"))
        .filter(pl.col("token_norm") != "")
        .group_by(["Q Name", "token_norm"])
        .agg([
            pl.col("code_num").n_unique().alias("cur_n_nums"),
            pl.col("code_num").first().alias("new_code_num"),
            pl.col("code_token").first().alias("code_token"),
            pl.col("excel_row").min().alias("excel_row"),
        ])
    )

    ref = (
        reference_codes
        .with_columns(normalize_code_token_expr("code_token").alias("token_norm"))
        .filter(pl.col("token_norm") != "")
        .group_by(["Q Name", "token_norm"])
        .agg([
            pl.col("code_num").n_unique().alias("ref_n_nums"),
            pl.col("code_num").first().alias("old_code_num"),
            pl.col("code_token").first().alias("code_token_ref"),
        ])
    )

    return (
        cur.join(ref, on=["Q Name", "token_norm"], how="inner")
        .filter(
            (pl.col("cur_n_nums") == 1)
            & (pl.col("ref_n_nums") == 1)
            & (pl.col("new_code_num") != pl.col("old_code_num"))
        )
        .with_columns([
            pl.lit("codes_col_renumbered_same_token").alias("issue_type"),
            pl.concat_str([
                pl.lit("codes_"),
                pl.col("old_code_num").cast(pl.Utf8),
                pl.lit("_to_"),
                pl.col("new_code_num").cast(pl.Utf8),
            ]).alias("field"),
            pl.lit("high").alias("severity"),
        ])
        .select([
            "issue_type", "Q Name", "field",
            "old_code_num", "new_code_num",
            "code_token", "code_token_ref",
            "severity", "excel_row",
        ])
    )


def build_codes_changes_view(
    token_mismatch: pl.DataFrame,
    added_codes: pl.DataFrame,
    removed_codes: pl.DataFrame,
    renumber_codes: pl.DataFrame,
    target_lang: str,
) -> pl.DataFrame:
    """Converts Codes-column checks to the common Option Changes view schema."""
    rows = []

    if token_mismatch.height > 0:
        rows.append(
            token_mismatch.with_columns([
                pl.col("code_token").alias("current"),
                pl.col("code_token_ref").alias("reference"),
            ]).select(["issue_type", "Q Name", "field", "current", "reference", "severity", "excel_row"])
        )

    if added_codes.height > 0:
        rows.append(
            added_codes.with_columns([
                pl.col("code_token").alias("current"),
                pl.lit(_not_in_reference_text("Codes")).alias("reference"),
            ]).select(["issue_type", "Q Name", "field", "current", "reference", "severity", "excel_row"])
        )

    if removed_codes.height > 0:
        rows.append(
            removed_codes.with_columns([
                pl.lit("(removed from Codes)").alias("current"),
                pl.col("code_token").alias("reference"),
                pl.lit(None).cast(pl.Int64).alias("excel_row"),
            ]).select(["issue_type", "Q Name", "field", "current", "reference", "severity", "excel_row"])
        )

    if renumber_codes.height > 0:
        rows.append(
            renumber_codes.with_columns([
                pl.concat_str([
                    pl.col("new_code_num").cast(pl.Utf8), pl.lit(") "), pl.col("code_token")
                ]).alias("current"),
                pl.concat_str([
                    pl.col("old_code_num").cast(pl.Utf8), pl.lit(") "), pl.col("code_token_ref")
                ]).alias("reference"),
            ]).select(["issue_type", "Q Name", "field", "current", "reference", "severity", "excel_row"])
        )

    if rows:
        base = pl.concat(rows, how="vertical")
    else:
        base = pl.DataFrame(schema={
            "issue_type": pl.Utf8, "Q Name": pl.Utf8, "field": pl.Utf8,
            "current": pl.Utf8, "reference": pl.Utf8,
            "severity": pl.Utf8, "excel_row": pl.Int64,
        })

    if str(target_lang).upper() == "EN":
        return base.with_columns(pl.lit("").alias("set_name")).select([
            "issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"
        ])

    return base.with_columns([
        pl.lit("").alias("set_name"),
        pl.lit("").alias("current_lang"),
        pl.lit("").alias("reference_lang"),
    ]).select([
        "issue_type", "set_name", "Q Name", "field", "current", "reference",
        "current_lang", "reference_lang", "severity", "excel_row",
    ])





# --- CODE CELL 12 ---
def validate_critical_sets(
    questions_df: pl.DataFrame,
    exact_sets  : dict,
) -> pl.DataFrame:
    """
    Checks that each critical question group is fully present with the expected
    Mandatory value. Rules come from critical_sets.yaml -> exact_sets.

    required=true  -> missing = issue_type 'missing_critical_question', severity HIGH
    required=false -> missing = issue_type 'advisory_question',          severity MEDIUM
    """
    EMPTY_SCHEMA = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }

    if not exact_sets:
        return pl.DataFrame(schema=EMPTY_SCHEMA)

    if "Mandatory_norm" not in questions_df.columns:
        questions_df = questions_df.with_columns(
            normalize_mandatory_expr("Mandatory").alias("Mandatory_norm")
        )

    present_qnames = set(questions_df["Q Name"].to_list())

    def _wealth_optional_variant(q_name: str) -> str:
        q = str(q_name or "").strip()
        if q.startswith("hh_wealth_"):
            return "o_hh_wealth_" + q[len("hh_wealth_"):]
        return ""

    issues = []

    for set_name, rules in exact_sets.items():
        required_names   = [r["q_name"] for r in rules if r.get("required", True)]

        for rule in rules:
            q_name   = rule["q_name"]
            expected = rule.get("expected_mandatory", "")
            required = rule.get("required", True)
            alt_q_name = _wealth_optional_variant(q_name)
            matched_q_name = q_name if q_name in present_qnames else (alt_q_name if alt_q_name in present_qnames else "")

            if not matched_q_name:
                issues.append({
                    "issue_type": "missing_critical_question" if required else "advisory_question",
                    "set_name"  : set_name, "Q Name": q_name,
                    "field"     : "Q Name", "current": "",
                    "reference" : (f"present (or optional variant: {alt_q_name})" if alt_q_name else "present"),
                    "severity"  : "high" if required else "medium",
                    "excel_row" : None,
                })
                continue

            if not expected:
                continue

            # If a wealth optional variant is accepted as substitute, treat it as pass
            # and avoid forcing mandatory parity on the alternate form.
            if matched_q_name != q_name:
                continue

            row = (
                questions_df
                .filter(pl.col("Q Name") == matched_q_name)
                .select(["Q Name", "Mandatory", "Mandatory_norm", "excel_row"])
                .to_dicts()
            )
            if not row:
                continue
            row = row[0]
            if row["Mandatory_norm"] != expected:
                issues.append({
                    "issue_type": "critical_mandatory_mismatch",
                    "set_name"  : set_name, "Q Name": q_name,
                    "field"     : "Mandatory",
                    "current"   : row["Mandatory"], "reference": expected,
                    "severity"  : "high", "excel_row": row.get("excel_row"),
                })

    if not issues:
        return pl.DataFrame(schema=EMPTY_SCHEMA)
    return pl.DataFrame(issues)


def validate_prefix_counts(
    current_questions: pl.DataFrame,
    min_count_sets   : dict,
) -> pl.DataFrame:
    """
    Checks that questions with a given prefix appear at least min_count times.
    Covers CS groups (cs_stress_*, cs_crisis_*, cs_emergency_*) and HDDS count.
    Rules come from critical_sets.yaml -> min_count_sets.
    """
    EMPTY_SCHEMA = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if not min_count_sets:
        return pl.DataFrame(schema=EMPTY_SCHEMA)

    all_qnames = current_questions["Q Name"].to_list()
    issues = []
    for set_name, rule in min_count_sets.items():
        prefix    = rule.get("prefix", "")
        min_count = rule.get("min_count", 1)
        desc      = rule.get("description", f"At least {min_count} '{prefix}*' questions required")
        prefix_alt = "o_hh_wealth_" if str(prefix) == "hh_wealth_" or str(set_name).upper() == "WEALTH" else ""
        matched = sorted(
            q for q in all_qnames
            if (prefix and q.startswith(prefix)) or (prefix_alt and q.startswith(prefix_alt))
        )
        if len(matched) < min_count:
            found_str = f"{len(matched)} found" + (f": {', '.join(matched)}" if matched else " (none)")
            issues.append({
                "issue_type": "missing_critical_question", "set_name": set_name, "Q Name": "",
                "field": "count", "current": found_str, "reference": desc,
                "severity": "high", "excel_row": None,
            })

    if not issues:
        return pl.DataFrame(schema=EMPTY_SCHEMA)
    return pl.DataFrame(issues)


def validate_crop_harvest(
    current_questions: pl.DataFrame,
    crop_rules       : dict,
) -> pl.DataFrame:
    """
    Questionnaire must contain EITHER only the minimal set OR all questions in
    the full set. Rules come from critical_sets.yaml -> crop_harvest.
    """
    EMPTY_SCHEMA = {
        "issue_type": pl.Utf8, "set_name": pl.Utf8, "Q Name": pl.Utf8,
        "field": pl.Utf8, "current": pl.Utf8, "reference": pl.Utf8,
        "severity": pl.Utf8, "excel_row": pl.Int64,
    }
    if not crop_rules:
        return pl.DataFrame(schema=EMPTY_SCHEMA)

    minimal = set(crop_rules.get("minimal", []))
    full    = set(crop_rules.get("full", []))
    if not minimal and not full:
        return pl.DataFrame(schema=EMPTY_SCHEMA)

    all_qnames     = set(current_questions["Q Name"].to_list())
    relevant_found = {q for q in all_qnames if q in full or q in minimal}

    if relevant_found == minimal or full.issubset(all_qnames):
        return pl.DataFrame(schema=EMPTY_SCHEMA)

    issues = [{
        "issue_type": "crop_harvest_violation", "set_name": "CRP_HARV", "Q Name": "",
        "field": "Q Name",
        "current"  : f"found: {', '.join(sorted(relevant_found)) or 'none'}",
        "reference": f"either only [{', '.join(sorted(minimal))}] OR all of [{', '.join(sorted(full))}]",
        "severity" : "high", "excel_row": None,
    }]
    return pl.DataFrame(issues)




# --- CODE CELL 13 ---
#  Run all comparisons
import time
_t0 = time.perf_counter()
print("[timing] run block started", flush=True)
_t_prev = _t0
def _tick(label: str):
    global _t_prev
    _now = time.perf_counter()
    print(f"[timing] {label:<34} {_now - _t_prev:8.3f}s", flush=True)
    _t_prev = _now

added, removed = compare_question_presence(current_questions, reference_questions)
_tick("compare_question_presence")

mandatory_diff = compare_mandatory(
    current_questions,
    reference_questions,
    treat_blank_as_no=cfg.treat_blank_as_no,
)
_tick("compare_mandatory")

qtype_issues = compare_qtype_changes(
    current_questions,
    reference_questions,
    current_options_en=current_options_en,
    reference_options_en=reference_options_en,
)
randomize_issues = compare_legacy_text_field_changes(
    current_questions, reference_questions,
    column_name="Randomize", issue_type="randomize_changed",
    default_severity="info", high_when_mandatory=False,
)
conditional_issues = compare_legacy_text_field_changes(
    current_questions, reference_questions,
    column_name="Conditional", issue_type="conditional_changed",
    default_severity="info", high_when_mandatory=False,
)
programming_issues = compare_legacy_text_field_changes(
    current_questions, reference_questions,
    column_name="Programming Instructions", issue_type="programming_instructions_changed",
    default_severity="info", high_when_mandatory=False,
)
core_only_issues = compare_legacy_text_field_changes(
    current_questions, reference_questions,
    column_name="Core questions only", issue_type="core_questions_only_changed",
    default_severity="info", high_when_mandatory=False,
)
print(
    f"QType: {qtype_issues.height}  Randomize: {randomize_issues.height}  "
    f"Conditional: {conditional_issues.height}  Prog.Instr: {programming_issues.height}  "
    f"Core-only: {core_only_issues.height}"
)
_tick("qtype + legacy control checks")

# Label diffs use comparison questions (after placeholder restore) to avoid
# false mismatches caused only by $...$ placeholder substitution.
cmp_current_en_col   = resolve_language_column(current_questions.columns, "EN")
cmp_reference_en_col = resolve_language_column(reference_questions.columns, "EN")
if not cmp_current_en_col or not cmp_reference_en_col:
    raise KeyError("English column not found in comparison survey frames; question label checks require English baseline.")
cmp_current_tgt_col   = resolve_language_column(current_questions.columns, target_lang)
cmp_reference_tgt_col = resolve_language_column(reference_questions.columns, target_lang)
if target_lang == "EN":
    cmp_current_tgt_col = cmp_current_en_col
    cmp_reference_tgt_col = cmp_reference_en_col
else:
    if not cmp_current_tgt_col:
        cmp_current_tgt_col = cmp_current_en_col

en_question_label_diff = compare_question_labels_single(
    current_questions,
    reference_questions,
    current_text_col=cmp_current_en_col,
    reference_text_col=cmp_reference_en_col,
    lang_scope="EN",
)
_tick("compare question labels EN")

if target_lang == "EN":
    tgt_question_label_diff = en_question_label_diff.clone()
elif not cmp_reference_tgt_col:
    _ql_keys = en_question_label_diff.select(["Q Name", "field", "excel_row"]).unique()
    _ql_cur = current_questions.select(["Q Name", cmp_current_tgt_col]).rename({cmp_current_tgt_col: "_curr_tgt"})
    tgt_question_label_diff = (
        _ql_keys
        .join(_ql_cur, on="Q Name", how="left")
        .with_columns([
            pl.lit("question_label_mismatch").alias("issue_type"),
            pl.col("_curr_tgt").cast(pl.Utf8).fill_null("").alias("current"),
            pl.lit("").alias("reference"),
            pl.lit("medium").alias("severity"),
            pl.lit(target_lang).alias("lang_scope"),
        ])
        .select(["issue_type", "Q Name", "field", "current", "reference", "severity", "excel_row", "lang_scope"])
    )
else:
    tgt_question_label_diff = compare_question_labels_single(
        current_questions,
        reference_questions,
        current_text_col=cmp_current_tgt_col,
        reference_text_col=cmp_reference_tgt_col,
        lang_scope=target_lang,
    )
_tick("compare question labels target")

en_option_diff = compare_option_labels_single(current_options_en, reference_options_en, "EN")
en_added_opts, en_removed_opts = compare_option_presence_single(current_options_en, reference_options_en, "EN")
en_code_renumber = compare_option_code_renumber_single(current_options_en, reference_options_en, "EN")
_tick("compare options EN")

if target_lang == "EN":
    tgt_option_diff = en_option_diff.clone()
    tgt_added_opts = en_added_opts.clone()
    tgt_removed_opts = en_removed_opts.clone()
    tgt_code_renumber = en_code_renumber.clone()
elif not reference_tgt_col:
    _opt_cur_lookup = (
        current_options_tgt
        .with_columns(pl.concat_str([pl.lit("option_"), pl.col("option_code").cast(pl.Utf8)]).alias("field"))
        .select(["Q Name", "field", "option_label", "excel_row"])
        .unique(subset=["Q Name", "field"], keep="first")
    )
    tgt_option_diff = (
        en_option_diff
        .select(["issue_type", "Q Name", "field", "severity", "excel_row"])
        .join(_opt_cur_lookup, on=["Q Name", "field"], how="left")
        .with_columns([
            pl.col("option_label").cast(pl.Utf8).fill_null("").alias("option_label"),
            pl.lit("").alias("option_label_ref"),
            pl.lit(target_lang).alias("lang_scope"),
        ])
        .select(["issue_type", "Q Name", "field", "option_label", "option_label_ref", "severity", "excel_row", "lang_scope"])
    )
    tgt_added_opts = (
        en_added_opts
        .select(["issue_type", "Q Name", "field", "severity"])
        .join(_opt_cur_lookup.select(["Q Name", "field", "option_label"]), on=["Q Name", "field"], how="left")
        .with_columns([
            pl.col("option_label").cast(pl.Utf8).fill_null("").alias("option_label"),
            pl.lit(target_lang).alias("lang_scope"),
        ])
        .select(["issue_type", "Q Name", "field", "option_label", "severity", "lang_scope"])
    )
    tgt_removed_opts = (
        en_removed_opts
        .select(["issue_type", "Q Name", "field", "severity"])
        .with_columns([
            pl.lit("").alias("option_label"),
            pl.lit(target_lang).alias("lang_scope"),
        ])
        .select(["issue_type", "Q Name", "field", "option_label", "severity", "lang_scope"])
    )
    tgt_code_renumber = (
        en_code_renumber
        .join(_opt_cur_lookup.select(["Q Name", "field", "option_label"]), on=["Q Name", "field"], how="left")
        .with_columns([
            pl.col("option_label").cast(pl.Utf8).fill_null("").alias("option_label"),
            pl.lit("").alias("option_label_ref"),
            pl.lit(target_lang).alias("lang_scope"),
        ])
        .select([
            "issue_type", "Q Name", "field",
            "old_code", "new_code", "option_label", "option_label_ref",
            "severity", "excel_row", "lang_scope",
        ])
    )
else:
    tgt_option_diff = compare_option_labels_single(current_options_tgt, reference_options_tgt, target_lang)
    tgt_added_opts, tgt_removed_opts = compare_option_presence_single(current_options_tgt, reference_options_tgt, target_lang)
    tgt_code_renumber = compare_option_code_renumber_single(current_options_tgt, reference_options_tgt, target_lang)
_tick("compare options target")

codes_renumber = compare_codes_renumber_single(current_codes, reference_codes)
codes_token_diff = compare_codes_token_mismatch_single(current_codes, reference_codes, renumber_codes=codes_renumber)
codes_added, codes_removed = compare_codes_presence_single(current_codes, reference_codes)
_tick("compare Codes column")

critical_issues = validate_critical_sets(current_questions, rules["exact_sets"])
count_issues = validate_prefix_counts(current_questions, rules["min_count_sets"])
harvest_issues = validate_crop_harvest(current_questions, rules["crop_harvest"])

_tick("critical/count/harvest checks")

# Build mandatory map needed by validate_skip_patterns (layer 1 flexibility check)
_q_mand_map: dict[str, str] = {}
for _df in [reference_questions, current_questions]:
    for _r in _df.select(["Q Name", "Mandatory"]).iter_rows(named=True):
        _q = _r["Q Name"]
        _m = str(_r.get("Mandatory") or "").strip().lower()
        if str(_q).startswith("o_"):
            _q_mand_map[_q] = "optional"
        elif "panel" in _m:
            _q_mand_map[_q] = "mandatory-panel"
        elif _m in ("yes", "y", "true", "1"):
            _q_mand_map[_q] = "mandatory"
        else:
            _q_mand_map[_q] = "non-mandatory"
_tick("mandatory map")

skip_issues = validate_skip_patterns(
    current_questions,
    reference_questions,
    current_options_en=current_options_en,
    reference_options_en=reference_options_en,
    q_mandatory_map=_q_mand_map,
)
_tick("validate_skip_patterns")

# Reduce noisy duplication: when a high-severity broken skip issue exists for a Q,
# suppress informational "change" rows for the same Q.
if skip_issues.height > 0:
    _high_skip_types = {
        "skipPattern_invalid_qname",
        "skipPattern_invalid_qnameCategory",
        "skipPattern_range_invalid",
        "skip_pattern_empty",
    }
    _info_skip_types = {
        "skipPattern_changes",
        "default_skip_modified",
        "skipPattern_range_mismatch",
    }
    _high_q = set(
        skip_issues
        .filter((pl.col("severity") == "high") & pl.col("issue_type").is_in(list(_high_skip_types)))
        .get_column("Q Name")
        .drop_nulls()
        .to_list()
    )
    if _high_q:
        skip_issues = skip_issues.filter(
            ~(
                pl.col("Q Name").is_in(list(_high_q))
                & pl.col("issue_type").is_in(list(_info_skip_types))
                & (pl.col("severity") == "info")
            )
        )


#  Convert to common schema
presence_issues = make_presence_issues(added, removed, reference_questions=reference_questions)
mandatory_issues = make_mandatory_issues(mandatory_diff)
core_question_issues = pl.concat(
    [
        presence_issues,
        mandatory_issues,
        qtype_issues,
        randomize_issues,
        conditional_issues,
        programming_issues,
        core_only_issues,
    ],
    how="vertical",
)
question_changes_view = build_question_changes_view(
    core_question_issues,
    en_question_label_diff,
    tgt_question_label_diff,
    target_lang,
)
question_issues = question_changes_view.select([
    "issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"
])
option_changes_view = build_option_changes_view(
    en_option_diff, en_added_opts, en_removed_opts, en_code_renumber,
    tgt_option_diff, tgt_added_opts, tgt_removed_opts, tgt_code_renumber,
    target_lang,
)
codes_changes_view = build_codes_changes_view(
    codes_token_diff,
    codes_added,
    codes_removed,
    codes_renumber,
    target_lang,
)
option_changes_view = pl.concat([option_changes_view, codes_changes_view], how="vertical")
option_issues = option_changes_view.select([
    "issue_type", "set_name", "Q Name", "field", "current", "reference", "severity", "excel_row"
])
_tick("build question/option issue frames")

#  Filter option issues: only report options for questions in BOTH files
_excluded_qnames = set(added["Q Name"].to_list()) | set(removed["Q Name"].to_list())
if _excluded_qnames:
    _excl = list(_excluded_qnames)
    option_issues = option_issues.filter(~pl.col("Q Name").is_in(_excl))
    option_changes_view = option_changes_view.filter(~pl.col("Q Name").is_in(_excl))
_tick("filter option issues")

#  Filter crop-placeholder option issues (expected round-to-round crop list drift)
_crop_placeholder_qnames = set(find_crop_placeholder_questions(template_questions_for_restore))
_crop_placeholder_qnames &= set(current_questions["Q Name"].to_list())
_crop_placeholder_qnames &= set(reference_questions["Q Name"].to_list())
if _crop_placeholder_qnames:
    _crop_q = list(_crop_placeholder_qnames)
    option_issues = option_issues.filter(~pl.col("Q Name").is_in(_crop_q))
    option_changes_view = option_changes_view.filter(~pl.col("Q Name").is_in(_crop_q))
_tick("filter crop replacement option issues")

#  Stack all issues
all_issues = pl.concat(
    [question_issues, option_issues,
     critical_issues, count_issues, harvest_issues, skip_issues],
    how="vertical",
)

#  Sort: high -> medium -> info
all_issues = (
    all_issues
    .with_columns(
        pl.when(pl.col("severity") == "high").then(pl.lit(0))
        .when(pl.col("severity") == "medium").then(pl.lit(1))
        .otherwise(pl.lit(2))
        .alias("_sort_order")
    )
    .sort(["_sort_order", "issue_type", "Q Name"])
    .drop("_sort_order")
)
_tick("concat + severity sort")

#  CS downgrade: if count minimum is met, removed CS questions -> info
_passing_prefixes = []
for _set_name, _rule in rules.get("min_count_sets", {}).items():
    if count_issues.filter(pl.col("set_name") == _set_name).height == 0:
        _prefix = _rule.get("prefix", "")
        if _prefix:
            _passing_prefixes.append(_prefix)

if _passing_prefixes:
    _starts_exprs = [pl.col("Q Name").cast(pl.Utf8).str.starts_with(p) for p in _passing_prefixes]
    _q_matches_prefix = pl.any_horizontal(_starts_exprs) if len(_starts_exprs) > 1 else _starts_exprs[0]
    all_issues = all_issues.with_columns(
        pl.when(
            (pl.col("issue_type") == "removed_question") & _q_matches_prefix
        ).then(pl.lit("info"))
        .otherwise(pl.col("severity"))
        .alias("severity")
    )
    question_changes_view = question_changes_view.with_columns(
        pl.when(
            (pl.col("issue_type") == "removed_question") & _q_matches_prefix
        ).then(pl.lit("info"))
        .otherwise(pl.col("severity"))
        .alias("severity")
    )
_tick("cs downgrade")

#  Add mandatory_cat via Polars lookup (current values override reference values)
_ref_m = reference_questions.select(["Q Name", "Mandatory"]).with_columns(pl.lit(0).alias("_prio"))
_cur_m = current_questions.select(["Q Name", "Mandatory"]).with_columns(pl.lit(1).alias("_prio"))
_mand_lookup = (
    pl.concat([_ref_m, _cur_m], how="vertical")
    .with_columns([
        pl.col("Q Name").cast(pl.Utf8).alias("Q Name"),
        pl.col("Mandatory").cast(pl.Utf8).fill_null("").str.strip_chars().str.to_lowercase().alias("_m"),
    ])
    .with_columns([
        pl.when(pl.col("Q Name").str.starts_with("o_")).then(pl.lit("optional"))
        .when(pl.col("_m").str.contains("panel")).then(pl.lit("mandatory-panel"))
        .when(pl.col("_m").is_in(["yes", "y", "true", "1"])).then(pl.lit("mandatory"))
        .otherwise(pl.lit("non-mandatory"))
        .alias("mandatory_cat")
    ])
    .sort(["Q Name", "_prio"])
    .group_by("Q Name")
    .agg(pl.col("mandatory_cat").last().alias("mandatory_cat"))
)

all_issues = (
    all_issues
    .join(_mand_lookup, on="Q Name", how="left")
    .with_columns(pl.col("mandatory_cat").fill_null(""))
)

option_changes_view = (
    option_changes_view
    .join(_mand_lookup, on="Q Name", how="left")
    .with_columns(pl.col("mandatory_cat").fill_null(""))
)

question_changes_view = (
    question_changes_view
    .join(_mand_lookup, on="Q Name", how="left")
    .with_columns(pl.col("mandatory_cat").fill_null(""))
)
_tick("mandatory_cat lookup join")

question_changes_view = (
    question_changes_view
    .with_columns(
        pl.when(pl.col("severity") == "high").then(pl.lit(0))
        .when(pl.col("severity") == "medium").then(pl.lit(1))
        .otherwise(pl.lit(2))
        .alias("_s")
    )
    .sort(["_s", "Q Name", "field"])
    .drop("_s")
)
_tick("sort question_changes_view")

option_changes_view = (
    option_changes_view
    .with_columns(
        pl.when(pl.col("severity") == "high").then(pl.lit(0))
        .when(pl.col("severity") == "medium").then(pl.lit(1))
        .otherwise(pl.lit(2))
        .alias("_s")
    )
    .sort(["_s", "Q Name", "field"])
    .drop("_s")
)
_tick("sort option_changes_view")

#  Build found_info for the Critical Sets sheet
_found_info = {}
_cur_qnames = current_questions["Q Name"].to_list()

# Exact sets (for example HDDS)
for _set_name, _set_rules in rules.get("exact_sets", {}).items():
    _expected = [str(r.get("q_name", "")).strip() for r in _set_rules]
    _matched = []
    for q in _expected:
        if not q:
            continue
        if q in _cur_qnames:
            _matched.append(q)
            continue
        if q.startswith("hh_wealth_"):
            _alt = "o_hh_wealth_" + q[len("hh_wealth_"):]
            if _alt in _cur_qnames:
                _matched.append(f"{q} (matched via {_alt})")
    _found_info[_set_name] = _matched

# Prefix-based sets (for example WEALTH / CS groups)
for _set_name, _rule in rules.get("min_count_sets", {}).items():
    _prefix = str(_rule.get("prefix", "")).strip()
    _prefix_alt = "o_hh_wealth_" if _prefix == "hh_wealth_" or _set_name.upper() == "WEALTH" else ""
    _matched = sorted(
        q for q in _cur_qnames
        if (_prefix and q.startswith(_prefix)) or (_prefix_alt and q.startswith(_prefix_alt))
    )
    _found_info[_set_name] = _matched

# Crop harvest set info
_crop_cfg = rules.get("crop_harvest", {}) or {}
_crop_candidates = set(_crop_cfg.get("minimal", [])) | set(_crop_cfg.get("full", []))
if _crop_candidates:
    _found_info["crop_harvest"] = sorted(q for q in _cur_qnames if q in _crop_candidates)

_tick("build found_info")

#  Summary
print(f"Questions added       : {added.height}")
print(f"Questions removed     : {removed.height}")
print(f"Mandatory mismatches  : {mandatory_diff.height}")
print(f"Question label changes: {question_issues.filter(pl.col('issue_type') == 'question_label_mismatch').height}")
print(f"Option label changes  : {option_issues.filter(pl.col('issue_type') == 'option_label_mismatch').height}")
print(f"Option positions changed: {option_issues.filter(pl.col('issue_type') == 'option_position_renumbered_same_label').height}")
print(f"Options removed       : {option_issues.filter(pl.col('issue_type') == 'removed_option').height}")
print(f"Options added         : {option_issues.filter(pl.col('issue_type') == 'added_option').height}")
print(f"Codes token mismatch  : {option_issues.filter(pl.col('issue_type') == 'codes_col_token_mismatch').height}")
print(f"Codes removed         : {option_issues.filter(pl.col('issue_type') == 'codes_col_removed').height}")
print(f"Codes added           : {option_issues.filter(pl.col('issue_type') == 'codes_col_added').height}")
print(f"Codes renumbered      : {option_issues.filter(pl.col('issue_type') == 'codes_col_renumbered_same_token').height}")
print(f"Critical set issues   : {critical_issues.height}")
print(f"Count rule violations : {count_issues.height}")
print(f"Crop harvest issues   : {harvest_issues.height}")
print(f"Skip pattern issues   : {skip_issues.height}")
print(f"")
print(f"Total issues          : {all_issues.height}")
print(f"[timing] TOTAL run block                    {time.perf_counter() - _t0:8.3f}s", flush=True)

if all_issues.height > 0:
    print("\nFirst 20 issues:")
    print(all_issues.head(20))
else:
    print("\nNo issues found.")





# ======================================================================
# SECTION: ## Step 10  Export to Excel
_step("Exporting report")
# ======================================================================
# --- CODE CELL 14 ---
from shutil import copyfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import difflib

#  Colours 
FILL_HIGH    = PatternFill("solid", fgColor="F4CCCC")
FILL_MEDIUM  = PatternFill("solid", fgColor="FCE5CD")
FILL_INFO    = PatternFill("solid", fgColor="CFE2F3")
FILL_PASS    = PatternFill("solid", fgColor="D9EAD3")
FILL_HEADER  = PatternFill("solid", fgColor="274E13")
FILL_SECTION = PatternFill("solid", fgColor="E8F5E9")

FONT_TITLE   = Font(bold=True, size=13, color="274E13")
FONT_HEADER  = Font(bold=True, color="FFFFFF", size=11)
FONT_SECTION = Font(bold=True, color="274E13", size=11)
FONT_NORMAL  = Font(size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

SEVERITY_FILL = {"high": FILL_HIGH, "medium": FILL_MEDIUM, "info": FILL_INFO}
STATUS_FILL   = {"PASS": FILL_PASS, "FAIL": FILL_HIGH, "WARNING": FILL_MEDIUM, "WARN": FILL_MEDIUM}

CRITICAL_SET_ISSUE_TYPES = {
    "missing_critical_question", "advisory_question",
    "critical_mandatory_mismatch", "crop_harvest_violation",
}

SKIP_PATTERN_ISSUE_TYPES = {
    "skipPattern_changes",
    "skipPattern_range_mismatch",
    "default_skip_modified",
    "skipPattern_range_invalid",
    "skipPattern_invalid_qname",
    "skipPattern_invalid_qnameCategory",
    "skip_pattern_empty",
}
QTYPE_ISSUE_TYPES = {"qtype_changed"}

CORE_QUESTION_CHANGE_ISSUE_TYPES = {
    "mandatory_source_missing",
    "mandatory_column_mismatch",
    "mandatory_to_optional",
    "removed_question",
    "added_question",
    "question_label_mismatch",
}

ADDITIONAL_QUESTION_FIELD_ISSUE_TYPES = {
    "conditional_changed",
    "randomize_changed",
    "programming_instructions_changed",
    "core_questions_only_changed",
}

REPLACEMENT_ISSUE_TYPES = {
    "replacement_additional_info_missing",
    "replacement_crop_selection_mismatch",
    "replacement_crop_round_delta",
    "replacement_unresolved_placeholder",
    "replacement_missing_key",
    "replacement_malformed_placeholder",
}

#  Shared helpers 
def _header_row(ws, row, values):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = FILL_HEADER; c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 18

def _section_header(ws, row, title, n_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=title)
    c.fill = FILL_SECTION; c.font = FONT_SECTION
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20

def _data_row(ws, row, n_cols, severity="", fill=None):
    use_fill = fill or SEVERITY_FILL.get(severity)
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = FONT_NORMAL; c.border = THIN_BORDER
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if use_fill:
            c.fill = use_fill

def _autofit(ws, mn=12, mx=55):
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(w + 2, mn), mx)

ISSUE_ACTION_MAP = {
    "added_question": "Review added question and confirm it is intentional",
    "removed_question": "Restore missing question or document approved removal",
    "mandatory_to_optional": "Review mandatory status against reference",
    "mandatory_column_mismatch": "Review mandatory status against reference",
    "mandatory_source_missing": "Mandatory values are unavailable in one file; verify source column completeness first",
    "qtype_changed": "Review Q Type change and questionnaire behavior impact",
    "question_label_mismatch": "Review question text mismatch",
    "randomize_changed": "Review randomization behavior change",
    "conditional_changed": "Review conditional-routing text change",
    "programming_instructions_changed": "Review programming instruction change",
    "core_questions_only_changed": "Review core-questions-only flag change",
    "added_option": "Review added option and downstream logic",
    "removed_option": "Review removed option and downstream logic",
    "option_label_mismatch": "Review option label mismatch",
    "option_position_renumbered_same_label": "Verify option reordering and skip/code alignment",
    "codes_col_token_mismatch": "Review code token mismatch in Codes column",
    "codes_col_added": "Review added code token and routing",
    "codes_col_removed": "Review removed code token and routing",
    "codes_col_renumbered_same_token": "Verify code renumbering and skip references",
    "skipPattern_changes": "Skip pattern changed versus reference (routing remains valid)",
    "skipPattern_range_mismatch": "Skip pattern range differs from reference but remains valid",
    "default_skip_modified": "Specify is blank and current Skip Pattern is inconsistent with current Default rule",
    "skipPattern_range_invalid": "Fix invalid option codes/range referenced in Skip Pattern",
    "skipPattern_invalid_qname": "Fix Skip Pattern target question name (target does not exist)",
    "skipPattern_invalid_qnameCategory": "Fix Skip Pattern target category (must route to optional/non-mandatory target)",
    "skip_pattern_empty": "Default skip rule exists but both Specify and Skip Pattern are empty",
    "missing_critical_question": "Add missing critical question",
    "advisory_question": "Review advisory question omission",
    "critical_mandatory_mismatch": "Align mandatory flag for critical question",
    "crop_harvest_violation": "Fix crop_harvest sequence/order",
    "duplicate_qname": "Rename duplicate Q Name(s) to unique identifiers",
    "replacement_additional_info_missing": "Populate Additional information sheet with replacement keys/values",
    "replacement_crop_selection_mismatch": "Review top-10 crop list selection and replacement outputs",
    "replacement_crop_round_delta": "Review crop-list round differences documented in Questionnaire Structure",
    "replacement_unresolved_placeholder": "Resolve remaining placeholder token(s) in output",
    "replacement_missing_key": "Add missing replacement key or fix placeholder spelling",
    "replacement_malformed_placeholder": "Fix malformed placeholder format (use $key$)",
}


def _action_for_issue_type(issue_type: str) -> str:
    key = str(issue_type or "").strip()
    return ISSUE_ACTION_MAP.get(key, "Review issue details")


def _optional_counterparts_for_missing_skip_targets(source_q: str, current_text: str) -> list[str]:
    """
    For skipPattern_invalid_qname rows, detect whether missing target(s) have
    optional counterparts (o_<target>) in the current questionnaire.
    """
    try:
        _curr_df = globals().get("current_questions", None)
        _ref_df = globals().get("reference_questions", None)
        curr_qnames = set(_curr_df.get_column("Q Name").drop_nulls().to_list()) if _curr_df is not None else set()
        ref_qnames = set(_ref_df.get_column("Q Name").drop_nulls().to_list()) if _ref_df is not None else set()
        if not curr_qnames:
            return []

        invalid_targets = sorted(
            t for t in _extract_referenced_qnames(
                str(current_text or ""),
                source_q=str(source_q or ""),
                ref_qnames=ref_qnames,
                curr_qnames=curr_qnames,
            )
            if t not in curr_qnames
        )
        hits = []
        for t in invalid_targets:
            if str(t).startswith("o_"):
                continue
            cand = f"o_{t}"
            if cand in curr_qnames:
                hits.append(cand)
        return sorted(set(hits))
    except Exception:
        return []


def _action_for_issue_row(issue_type: str, q_name: str = "", current_text: str = "") -> str:
    key = str(issue_type or "").strip()
    if key == "skipPattern_invalid_qname":
        opt_hits = _optional_counterparts_for_missing_skip_targets(q_name, current_text)
        if opt_hits:
            return (
                "Fix Skip Pattern target question name (target does not exist). "
                f"Optional counterpart found ({', '.join(opt_hits)}): verify compatibility before replacing."
            )
    return _action_for_issue_type(key)


def _issues_col_map():
    lang_code = str(globals().get("target_lang", "")).upper().strip()
    lang_label = lang_code if (lang_code and lang_code != "EN") else "Language"
    return [
        ("issue_type",    "Issue type"),
        ("mandatory_cat", "Type"),
        ("set_name",      "Set"),
        ("Q Name",        "Q Name"),
        ("field",         "Field"),
        ("current",       "Current value"),
        ("reference",     "Reference / rule"),
        ("lang_scope",    "Language scope"),
        ("current_lang",    f"Current value ({lang_label})"),
        ("reference_lang",  f"Reference / rule ({lang_label})"),
        ("current_en",    "Current value (EN)"),
        ("reference_en",  "Reference / rule (EN)"),
        ("action",        "Action"),
        ("severity",      "Severity"),
        ("excel_row",     "Excel row"),
    ]


def _issues_table(ws, start_row, df):
    """Renders a header + data table. Only columns present in df are shown."""
    if "issue_type" in df.columns:
        _q_expr = pl.col("Q Name") if "Q Name" in df.columns else pl.lit("")
        _cur_expr = pl.col("current") if "current" in df.columns else pl.lit("")
        _action_expr = (
            pl.struct([
                pl.col("issue_type").alias("_issue"),
                _q_expr.alias("_q"),
                _cur_expr.alias("_cur"),
            ])
            .map_elements(
                lambda r: _action_for_issue_row(r.get("_issue"), r.get("_q"), r.get("_cur")),
                return_dtype=pl.Utf8,
            )
            .alias("_action_default")
        )
        if "action" not in df.columns:
            df = df.with_columns(_action_expr).rename({"_action_default": "action"})
        else:
            df = df.with_columns([
                _action_expr,
                pl.coalesce([pl.col("action"), pl.col("_action_default")]).alias("action"),
            ]).drop("_action_default")
    cols = [(s, d) for s, d in _issues_col_map() if s in df.columns]
    _header_row(ws, start_row, [d for _, d in cols])
    r = start_row + 1
    if df.height == 0:
        c = ws.cell(row=r, column=1, value="No issues in this category")
        c.font = Font(bold=True, color="274E13", size=10)
        return r + 2
    for row_data in df.to_dicts():
        for col, (src, _) in enumerate(cols, 1):
            ws.cell(row=r, column=col, value=row_data.get(src))
        _data_row(ws, r, len(cols), row_data.get("severity", ""))
        r += 1
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(cols))}{start_row}"
    return r + 1


try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    _RICH_TEXT_AVAILABLE = True
except Exception:
    CellRichText = None
    TextBlock = None
    InlineFont = None
    _RICH_TEXT_AVAILABLE = False


def _tokenize_for_word_diff(text: str):
    raw = re.findall(r"\s+|\S+", text)
    out = []
    for tok in raw:
        if tok.isspace() and out:
            out[-1] = out[-1] + tok
        else:
            out.append(tok)
    return out


def _coalesce_segments(parts):
    merged = []
    for txt, is_equal in parts:
        if not txt:
            continue
        if merged and merged[-1][1] == is_equal:
            merged[-1] = (merged[-1][0] + txt, is_equal)
        else:
            merged.append((txt, is_equal))
    return merged


def _build_diff_segments(current_text: str, reference_text: str):
    a = _tokenize_for_word_diff(current_text)
    b = _tokenize_for_word_diff(reference_text)
    sm = difflib.SequenceMatcher(a=a, b=b)
    current_parts, reference_parts = [], []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            current_parts.append(("".join(a[i1:i2]), True))
            reference_parts.append(("".join(b[j1:j2]), True))
        elif tag == "replace":
            current_parts.append(("".join(a[i1:i2]), False))
            reference_parts.append(("".join(b[j1:j2]), False))
        elif tag == "delete":
            current_parts.append(("".join(a[i1:i2]), False))
        elif tag == "insert":
            reference_parts.append(("".join(b[j1:j2]), False))
    return _coalesce_segments(current_parts), _coalesce_segments(reference_parts)


def _set_rich_text(cell, full_text: str, parts):
    if not _RICH_TEXT_AVAILABLE:
        cell.value = full_text
        return
    blocks = []
    for txt, is_equal in parts:
        if not txt:
            continue
        color = "FF000000" if is_equal else "FFFF0000"
        blocks.append(TextBlock(InlineFont(color=color), txt))
    if not blocks:
        cell.value = full_text
        return
    cell.value = CellRichText(*blocks)


def _apply_inline_diff_for_issue(ws, start_row, df, issue_types):
    if df.height == 0:
        return

    cols = [(s, d) for s, d in _issues_col_map() if s in df.columns]
    cidx = {src: i for i, (src, _) in enumerate(cols, start=1)}

    _allowed = set(issue_types or [])

    def _match_issue(issue_val: str) -> bool:
        for allowed in _allowed:
            if isinstance(allowed, str) and allowed.endswith("*"):
                if issue_val.startswith(allowed[:-1]):
                    return True
            elif issue_val == allowed:
                return True
        return False

    compare_pairs = [("current", "reference")]
    if "current_lang" in cidx and "reference_lang" in cidx:
        compare_pairs.append(("current_lang", "reference_lang"))
    if "current_en" in cidx and "reference_en" in cidx:
        compare_pairs.append(("current_en", "reference_en"))

    for offset, rd in enumerate(df.to_dicts(), start=1):
        issue_val = str(rd.get("issue_type") or "")
        if not _match_issue(issue_val):
            continue

        for cur_key, ref_key in compare_pairs:
            cur_txt = str(rd.get(cur_key) or "")
            ref_txt = str(rd.get(ref_key) or "")
            if cur_txt == ref_txt:
                continue
            cur_parts, ref_parts = _build_diff_segments(cur_txt, ref_txt)
            _set_rich_text(ws.cell(row=start_row + offset, column=cidx[cur_key]), cur_txt, cur_parts)
            _set_rich_text(ws.cell(row=start_row + offset, column=cidx[ref_key]), ref_txt, ref_parts)


#  Per-set status builder 

def _set_status(all_issues, rules):
    results = []
    for set_name in rules.get("exact_sets", {}).keys():
        si   = all_issues.filter(pl.col("set_name") == set_name)
        high = si.filter(pl.col("severity") == "high").height
        med  = si.filter(pl.col("severity") == "medium").height
        if high > 0:
            status  = "FAIL"
            details = "; ".join(f"{r['Q Name'] or r['current']}" for r in si.filter(pl.col("severity") == "high").to_dicts())
        elif med > 0:
            status  = "WARNING"
            details = "; ".join(f"{r['Q Name'] or r['current']}" for r in si.to_dicts())
        else:
            status, details = "PASS", "All required questions present and correct"
        results.append({"set": set_name, "status": status, "details": details})
    for set_name, rule in rules.get("min_count_sets", {}).items():
        si = all_issues.filter(pl.col("set_name") == set_name)
        if si.filter(pl.col("severity").is_in(["high", "medium"])).height > 0:
            status, details = "FAIL", si["current"].to_list()[0]
        else:
            status  = "PASS"
            details = f">={rule.get('min_count',0)} {rule.get('prefix','')}* questions confirmed"
        results.append({"set": set_name, "status": status, "details": details})
    if rules.get("crop_harvest"):
        crp = all_issues.filter(pl.col("set_name") == "CRP_HARV")
        if crp.height > 0:
            status, details = "FAIL", crp["current"].to_list()[0]
        else:
            status, details = "PASS", "Crop harvest rule respected"
        results.append({"set": "CRP_HARV", "status": status, "details": details})

    return results


def _cat_counts(df):
    """Returns counts by mandatory_cat for the four standard categories."""
    cats = ["mandatory", "mandatory-panel", "non-mandatory", "optional"]
    if "mandatory_cat" not in df.columns or df.height == 0:
        return {c: 0 for c in cats}
    return {cat: df.filter(pl.col("mandatory_cat") == cat).height for cat in cats}


def _extract_round_token(name: str) -> str:
    txt = str(name or "")
    patterns = [
        r"(?i)(?:^|[_\-\s])R(\d{1,3})(?=$|[_\-\s\.])",
        r"(?i)(?:^|[_\-\s])ROUND[_\-\s]*(\d{1,3})(?=$|[_\-\s\.])",
        # Fallback for suffix forms like "...copyr10.xlsx"
        r"(?i)R(?:OUND)?[_\-\s]*([0-9]{1,3})(?=(?:\.[A-Za-z0-9]+)?$)",
    ]
    for pat in patterns:
        m = re.search(pat, txt)
        if m:
            return f"R{m.group(1)}"
    return ""


def _extract_iso3_token(name: str) -> str:
    txt = str(name or "").upper()
    cfg_iso = str(getattr(globals().get("cfg", None), "iso3", "") or "").upper().strip()
    if cfg_iso and re.search(rf"(?:^|[_\-\s]){re.escape(cfg_iso)}(?:$|[_\-\s])", txt):
        return cfg_iso

    # Handle common country-name tokens used in operational filenames.
    if "DRC" in txt or "CONGO" in txt:
        return "COD"
    if "YEMEN" in txt:
        return "YEM"
    if "MYANMAR" in txt:
        return "MMR"

    # Fallback: pick the first 3-letter token that is not a known boilerplate tag.
    banned = {
        "FAO", "WFP", "CATI", "VHF", "VFG", "VFF", "VFA", "ISO", "HHQ",
        "QNR", "EN", "AR", "FR", "ES", "PT",
    }
    tokens = re.findall(r"\b([A-Z]{3})\b", txt)
    for tok in tokens:
        if tok == "DRC":
            return "COD"
        if tok not in banned:
            return tok
    return ""


def _reference_descriptor() -> str:
    """
    Human-readable reference descriptor for Summary header.
    Example:
      previous_round | COD R8 | GeoPoll_FAO_DRC_Round_8_CATI_V4_copy.xlsx
      latest_template | EN template | household_questionnaire_geopoll_EN_template_20250708_ISO3.xlsx
    """
    _run = globals().get("run", {}) or {}
    _cfg = globals().get("cfg", None)
    mode = _reference_scope_label()
    ref_name = Path(_run.get("reference_path", "")).name if _run.get("reference_path") else ""
    lang = str(getattr(_cfg, "language", "") or "").upper()

    if mode == "previous round":
        iso = _extract_iso3_token(ref_name) or str(getattr(_cfg, "iso3", "") or "").upper()
        rnd = _extract_round_token(ref_name) or "R?"
        return f"{mode} | {iso} {rnd} | {ref_name}"
    return f"{mode} | {lang} template | {ref_name}"


#  Sheet 1: Summary 
def write_summary_sheet(wb, all_issues, rules, replacement_status=None):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    # A=category/set  B=total/status  C=mandatory/details  D=m-panel  E=non-mand.  F=optional  G=severity
    for col, w in zip("ABCDEFG", [32, 8, 13, 10, 13, 10, 14]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:G1")
    ws["A1"] = "GeoPoll Questionnaire Validation Report"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Comparison basis: {_reference_scope_label()}"
    ws["A2"].font = Font(size=10, color="274E13", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A3:G3")
    _curr_name = Path((globals().get("run", {}) or {}).get("questionnaire_path", "")).name
    _curr_iso = str(getattr(globals().get("cfg", None), "iso3", "") or "").upper()
    _curr_round = _extract_round_token(_curr_name) or "R?"
    ws["A3"] = f"Current questionnaire: {_curr_iso} {_curr_round} | {_curr_name}"
    ws["A3"].font = Font(size=10, color="1F4E78", italic=True)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A4:G4")
    ws["A4"] = f"Checked against: {_reference_descriptor()}"
    ws["A4"].font = Font(size=10, color="1F4E78", italic=True)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A5:G5")
    _cfg_lang = str(getattr(globals().get("cfg", None), "language", "") or "").upper()
    _effective_lang = str(globals().get("target_lang", "") or "").upper()
    _effective_src = str(globals().get("target_lang_source", "") or "").strip()
    ws["A5"] = f"Language scope: EN baseline + {_effective_lang or _cfg_lang} (configured={_cfg_lang or 'N/A'}; source={_effective_src or 'config'})"
    ws["A5"].font = Font(size=10, color="1F4E78", italic=True)
    ws["A5"].alignment = Alignment(horizontal="left", vertical="center")

    r = 6
    #  Critical sets 
    _section_header(ws, r, "CRITICAL SETS STATUS", 7); r += 1
    _header_row(ws, r, ["Critical set", "Status", "Details", "", "", "", ""]); r += 1
    for row_data in _set_status(all_issues, rules):
        fill = STATUS_FILL.get(row_data["status"])
        ws.cell(row=r, column=1, value=row_data["set"])
        ws.cell(row=r, column=2, value=row_data["status"])
        ws.cell(row=r, column=3, value=row_data["details"])
        _data_row(ws, r, 7, fill=fill); r += 1

    r += 1
    #  Structure checks (aligned with KoBo summary order)
    _section_header(ws, r, "QUESTIONNAIRE STRUCTURE CHECKS", 7); r += 1
    _header_row(ws, r, ["Check", "Status", "Details", "", "", "", ""]); r += 1

    _skip_df = all_issues.filter(pl.col("issue_type").is_in(list(SKIP_PATTERN_ISSUE_TYPES)))
    _skip_high = _skip_df.filter(pl.col("severity") == "high").height
    _skip_med = _skip_df.filter(pl.col("severity") == "medium").height
    _skip_info = _skip_df.filter(pl.col("severity") == "info").height

    ws.cell(row=r, column=1, value="Skip logic references")
    ws.cell(row=r, column=2, value="FAIL" if _skip_high > 0 else "PASS")
    ws.cell(row=r, column=3, value=(f"High severity issues: {_skip_high}" if _skip_high > 0 else "No high severity issues"))
    _data_row(ws, r, 7, fill=STATUS_FILL.get("FAIL" if _skip_high > 0 else "PASS")); r += 1

    _skip_warn_total = _skip_med + _skip_info
    ws.cell(row=r, column=1, value="Skip logic references")
    ws.cell(row=r, column=2, value="WARN" if _skip_warn_total > 0 else "PASS")
    ws.cell(row=r, column=3, value=(f"Medium severity issues: {_skip_med}; Info: {_skip_info}" if _skip_warn_total > 0 else "No medium/info issues"))
    _data_row(ws, r, 7, fill=STATUS_FILL.get("WARN" if _skip_warn_total > 0 else "PASS")); r += 1

    _dup_n = all_issues.filter(pl.col("issue_type") == "duplicate_qname").height
    _dup_st = "FAIL" if _dup_n > 0 else "PASS"
    ws.cell(row=r, column=1, value="Duplicate Q Names")
    ws.cell(row=r, column=2, value=_dup_st)
    ws.cell(row=r, column=3, value=f"{_dup_n} duplicate(s) found" if _dup_n else "No duplicates")
    _data_row(ws, r, 7, fill=STATUS_FILL.get(_dup_st)); r += 1

    _qt_df = all_issues.filter(pl.col("issue_type").is_in(list(QTYPE_ISSUE_TYPES)))
    _qt_high = _qt_df.filter(pl.col("severity") == "high").height
    _qt_med = _qt_df.filter(pl.col("severity") == "medium").height
    _qt_info = _qt_df.filter(pl.col("severity") == "info").height
    _qt_status = "FAIL" if _qt_high > 0 else ("WARN" if (_qt_med + _qt_info) > 0 else "PASS")
    ws.cell(row=r, column=1, value="Q Type integrity")
    ws.cell(row=r, column=2, value=_qt_status)
    ws.cell(row=r, column=3, value=f"High: {_qt_high}; Medium: {_qt_med}; Info: {_qt_info}")
    _data_row(ws, r, 7, fill=STATUS_FILL.get(_qt_status)); r += 1

    r += 1
    #  Replacement status
    _section_header(ws, r, "REPLACEMENT STATUS (validated questionnaire output)", 7); r += 1
    _header_row(ws, r, ["Check", "Status", "Details", "", "", "", ""]); r += 1

    _repl_rows = (replacement_status or {}).get("rows", [])
    if not _repl_rows:
        ws.cell(row=r, column=1, value="Replacement step not run")
        ws.cell(row=r, column=2, value="INFO")
        ws.cell(row=r, column=3, value="No replacement diagnostics available for this run.")
        _data_row(ws, r, 7, "info"); r += 1
    else:
        for rd in _repl_rows:
            sev = rd.get("severity", "info")
            st = str(rd.get("status", "INFO") or "INFO").upper()
            ws.cell(row=r, column=1, value=rd.get("check", ""))
            ws.cell(row=r, column=2, value=st)
            ws.cell(row=r, column=3, value=rd.get("details", ""))
            row_fill = STATUS_FILL.get(st) or SEVERITY_FILL.get(sev)
            _data_row(ws, r, 7, fill=row_fill)
            if st == "PASS":
                ws.cell(row=r, column=2).font = Font(bold=True, size=10, color="274E13")
            elif st in {"WARN", "WARNING"}:
                ws.cell(row=r, column=2).font = Font(bold=True, size=10, color="B45F06")
            elif st == "FAIL":
                ws.cell(row=r, column=2).font = Font(bold=True, size=10, color="CC0000")
            r += 1

    r += 1
    #  Core question changes (high/medium focus)
    _section_header(ws, r, "QUESTION CHANGES (CORE)", 7); r += 1
    _header_row(ws, r, ["Category", "Total", "Mandatory", "M-Panel", "Non-mand.", "Optional", "Severity"]); r += 1
    for label, itype, sev_filter, disp_sev in [
        ("Mandatory source unavailable", "mandatory_source_missing", None, "high"),
        ("Mandatory flag changes",       "mandatory_column_mismatch", None, "high"),
        ("Mandatory to optional",        "mandatory_to_optional", None, "high"),
        ("Questions removed (high)",     "removed_question", "high", "high"),
        ("Questions removed (info)",     "removed_question", "info", "info"),
        ("Questions added",              "added_question", None, "info"),
        ("Question labels changed",      "question_label_mismatch", None, "medium"),
    ]:
        q = all_issues.filter(pl.col("issue_type") == itype)
        if sev_filter:
            q = q.filter(pl.col("severity") == sev_filter)
        counts = _cat_counts(q)
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=q.height)
        ws.cell(row=r, column=3, value=counts["mandatory"])
        ws.cell(row=r, column=4, value=counts["mandatory-panel"])
        ws.cell(row=r, column=5, value=counts["non-mandatory"])
        ws.cell(row=r, column=6, value=counts["optional"])
        if q.height == 0:
            ws.cell(row=r, column=7, value="none")
            _data_row(ws, r, 7, fill=FILL_PASS)
        else:
            ws.cell(row=r, column=7, value=disp_sev)
            _data_row(ws, r, 7, disp_sev)
        r += 1

    r += 1
    #  Additional control-field diffs
    _section_header(ws, r, "QUESTION CHANGES (ADDITIONAL FIELDS)", 7); r += 1
    _header_row(ws, r, ["Category", "Total", "Mandatory", "M-Panel", "Non-mand.", "Optional", "Severity"]); r += 1
    for label, itype in [
        ("Conditional changed", "conditional_changed"),
        ("Randomize changed", "randomize_changed"),
        ("Programming instructions changed", "programming_instructions_changed"),
        ("Core questions-only changed", "core_questions_only_changed"),
    ]:
        q = all_issues.filter(pl.col("issue_type") == itype)
        counts = _cat_counts(q)
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=q.height)
        ws.cell(row=r, column=3, value=counts["mandatory"])
        ws.cell(row=r, column=4, value=counts["mandatory-panel"])
        ws.cell(row=r, column=5, value=counts["non-mandatory"])
        ws.cell(row=r, column=6, value=counts["optional"])
        if q.height == 0:
            ws.cell(row=r, column=7, value="none")
            _data_row(ws, r, 7, fill=FILL_PASS)
        else:
            ws.cell(row=r, column=7, value="info")
            _data_row(ws, r, 7, "info")
        r += 1

    r += 1
    #  Option changes (7 cols with mandatory breakdown)
    _section_header(ws, r, "OPTION CHANGES (questions present in both files)", 7); r += 1
    _header_row(ws, r, ["Category", "Total", "Mandatory", "M-Panel", "Non-mand.", "Optional", "Severity"]); r += 1
    for label, itype, disp_sev in [
        ("Options removed",                 "removed_option",                     "high"),
        ("Options added",                   "added_option",                       "medium"),
        ("Option labels changed",           "option_label_mismatch",              "medium"),
        ("Option positions changed",   "option_position_renumbered_same_label",  "high"),
        ("Codes token mismatch",            "codes_col_token_mismatch",           "high"),
        ("Codes removed",                   "codes_col_removed",                  "high"),
        ("Codes added",                     "codes_col_added",                    "medium"),
        ("Codes renumbered (same token)",   "codes_col_renumbered_same_token",    "high"),
    ]:
        q = all_issues.filter(pl.col("issue_type") == itype)
        counts = _cat_counts(q)
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=q.height)
        ws.cell(row=r, column=3, value=counts["mandatory"])
        ws.cell(row=r, column=4, value=counts["mandatory-panel"])
        ws.cell(row=r, column=5, value=counts["non-mandatory"])
        ws.cell(row=r, column=6, value=counts["optional"])
        if q.height == 0:
            ws.cell(row=r, column=7, value="none")
            _data_row(ws, r, 7, fill=FILL_PASS)
        else:
            ws.cell(row=r, column=7, value=disp_sev)
            _data_row(ws, r, 7, disp_sev)
        r += 1


#  Sheet: Structure/SkipPattern 
def write_structure_skip_sheet(wb, all_issues: pl.DataFrame):
    ws = wb.create_sheet("Questionnaire Structure")
    ws.sheet_view.showGridLines = False

    def _sorted(df: pl.DataFrame) -> pl.DataFrame:
        if df.height == 0:
            return df
        return (
            df.with_columns(
                pl.when(pl.col("severity") == "high").then(pl.lit(0))
                .when(pl.col("severity") == "medium").then(pl.lit(1))
                .otherwise(pl.lit(2))
                .alias("_s")
            )
            .sort(["_s", "issue_type", "Q Name"])
            .drop("_s")
        )

    skip_df = _sorted(all_issues.filter(pl.col("issue_type").is_in(list(SKIP_PATTERN_ISSUE_TYPES))))
    qtype_df = _sorted(all_issues.filter(pl.col("issue_type").is_in(list(QTYPE_ISSUE_TYPES))))
    dup_df = _sorted(all_issues.filter(pl.col("issue_type") == "duplicate_qname"))

    _section_header(ws, 1, "QUESTIONNAIRE STRUCTURE  Skip pattern issues", 8)
    next_row = _issues_table(ws, 2, skip_df)

    _section_header(ws, next_row, "Q TYPE INTEGRITY ISSUES", 8)
    next_row = _issues_table(ws, next_row + 1, qtype_df)

    _section_header(ws, next_row, "DUPLICATE Q NAME ISSUES", 8)
    _issues_table(ws, next_row + 1, dup_df)

    ws.freeze_panes = "A3"
    _autofit(ws)


def write_replacement_issues_sheet(wb, all_issues: pl.DataFrame):
    ws = wb.create_sheet("Replacement Issues")
    ws.sheet_view.showGridLines = False

    df = all_issues.filter(
        pl.col("issue_type").is_in(list(REPLACEMENT_ISSUE_TYPES))
    )
    if df.height > 0:
        df = (
            df.with_columns(
                pl.when(pl.col("severity") == "high").then(pl.lit(0))
                .when(pl.col("severity") == "medium").then(pl.lit(1))
                .otherwise(pl.lit(2))
                .alias("_s")
            )
            .sort(["_s", "issue_type", "Q Name"])
            .drop("_s")
        )

    _section_header(ws, 1, "REPLACEMENT ISSUES  Placeholder/additional-info diagnostics", 8)
    _issues_table(ws, 2, df)
    ws.freeze_panes = "A3"
    _autofit(ws)


#  Sheet: Critical Sets 
def write_critical_sets_sheet(wb, all_issues, found_info=None):
    ws = wb.create_sheet("Critical Sets")
    ws.sheet_view.showGridLines = False

    df = all_issues.filter(
        pl.col("issue_type").is_in(list(CRITICAL_SET_ISSUE_TYPES))
    ).sort(["set_name", "severity", "Q Name"])

    _section_header(ws, 1, "CRITICAL SETS  Structural validation issues", 8)
    next_row = _issues_table(ws, 2, df)

    if found_info:
        next_row += 1
        _section_header(ws, next_row, "QUESTIONS FOUND IN CURRENT QUESTIONNAIRE", 3)
        next_row += 1
        _header_row(ws, next_row, ["Set", "Count found", "Questions"])
        next_row += 1
        for set_name, questions in found_info.items():
            ws.cell(row=next_row, column=1, value=set_name)
            ws.cell(row=next_row, column=2, value=len(questions))
            ws.cell(row=next_row, column=3, value=", ".join(questions) if questions else "none found")
            _data_row(ws, next_row, 3, "info" if questions else "high")
            next_row += 1

    _autofit(ws)


#  Sheet 3: Question Changes 
def write_question_changes_sheet(wb, question_changes_view):
    """
    Question-level changes split into:
    1) core changes (presence/mandatory/label)
    2) additional informational control-field changes.
    """
    ws = wb.create_sheet("Question Changes")
    ws.sheet_view.showGridLines = False

    df = (
        question_changes_view
        .with_columns(
            pl.when(pl.col("severity") == "high").then(pl.lit(0))
            .when(pl.col("severity") == "medium").then(pl.lit(1))
            .otherwise(pl.lit(2))
            .alias("_s")
        )
        .sort(["_s", "issue_type", "Q Name"])
        .drop("_s")
    )
    if "set_name" in df.columns:
        df = df.drop("set_name")

    core_df = df.filter(pl.col("issue_type").is_in(list(CORE_QUESTION_CHANGE_ISSUE_TYPES)))
    additional_df = df.filter(pl.col("issue_type").is_in(list(ADDITIONAL_QUESTION_FIELD_ISSUE_TYPES)))
    additional_df = additional_df.with_columns(pl.lit("info").alias("severity"))

    _section_header(ws, 1, "QUESTION CHANGES (CORE)  Presence, mandatory, labels", 10)
    next_row = _issues_table(ws, 2, core_df)
    _apply_inline_diff_for_issue(
        ws,
        2,
        core_df,
        {"question_label_mismatch"},
    )

    _section_header(ws, next_row, "QUESTION CHANGES (ADDITIONAL FIELDS)", 10)
    _issues_table(ws, next_row + 1, additional_df)
    _apply_inline_diff_for_issue(
        ws,
        next_row + 1,
        additional_df,
        {"conditional_changed", "programming_instructions_changed", "core_questions_only_changed"},
    )
    # Keep navigation stable even with multiple tables in this sheet.
    # Without resetting here, the second table can push freeze panes far down.
    ws.freeze_panes = "A3"
    _cols = [(s, d) for s, d in _issues_col_map() if s in df.columns]
    if _cols:
        _n = len(_cols)
        ws.auto_filter.ref = f"A2:{get_column_letter(_n)}2"
    _autofit(ws)


#  Sheet 4: Option Changes 
def write_option_changes_sheet(wb, option_changes_view):
    """
    Option-level changes for questions present in both files. The 'Field'
    column shows option_1 / option_2 /  so you know which answer choice is
    affected. The 'Type' column shows the mandatory category of the parent question.
    """
    ws = wb.create_sheet("Option Changes")
    ws.sheet_view.showGridLines = False

    df = option_changes_view
    if "set_name" in df.columns:
        df = df.drop("set_name")

    code_issue_types = {
        "codes_col_token_mismatch",
        "codes_col_removed",
        "codes_col_added",
        "codes_col_renumbered_same_token",
    }
    option_df = df.filter(~pl.col("issue_type").is_in(list(code_issue_types)))
    codes_df = df.filter(pl.col("issue_type").is_in(list(code_issue_types)))

    _section_header(ws, 1, "OPTION CHANGES  Answer options for questions in both files", max(8, len(df.columns)))
    option_table_start = 2
    next_row = _issues_table(ws, option_table_start, option_df)
    _apply_inline_diff_for_issue(ws, option_table_start, option_df, {"option_label_mismatch", "option_position_renumbered_same_label"})

    next_row += 1
    _section_header(ws, next_row, "CODES COLUMN CHECKS  Number/token consistency in 'Codes'", max(8, len(df.columns)))
    codes_table_start = next_row + 1
    _issues_table(ws, codes_table_start, codes_df)
    _apply_inline_diff_for_issue(ws, codes_table_start, codes_df, {"codes_col_token_mismatch", "codes_col_renumbered_same_token"})

    # Keep sheet navigation usable: second table must not move freeze panes down.
    ws.freeze_panes = "A3"
    _cols = [(s, d) for s, d in _issues_col_map() if s in df.columns]
    if _cols:
        _n = len(_cols)
        _hdr = option_table_start if option_df.height > 0 else codes_table_start
        ws.auto_filter.ref = f"A{_hdr}:{get_column_letter(_n)}{_hdr}"
    _autofit(ws)


#  Main export 
def export_validation_report(all_issues, question_changes_view, option_changes_view, result_file, rules, found_info=None, replacement_status=None):
    wb = Workbook()
    wb.remove(wb.active)
    write_summary_sheet(wb, all_issues, rules, replacement_status=replacement_status)
    write_critical_sets_sheet(wb, all_issues, found_info=found_info)
    write_structure_skip_sheet(wb, all_issues)
    write_replacement_issues_sheet(wb, all_issues)
    write_question_changes_sheet(wb, question_changes_view)
    write_option_changes_sheet(wb, option_changes_view)
    wb.save(result_file)
    print(f"Report saved to: {result_file}")
    print(f"Sheets: {wb.sheetnames}")


def _resolve_survey_sheet(workbook):
    ws = next((workbook[n] for n in workbook.sheetnames if n.strip().lower() == "survey"), None)
    if ws is None:
        raise KeyError(f"Survey sheet not found. Available sheets: {workbook.sheetnames}")
    return ws


def _sheet_header_map(ws, header_row: int = 3) -> dict[str, int]:
    mapping = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is None:
            continue
        header = str(value).strip()
        mapping[header] = col
        canonical = _canonical_column_from_header(header)
        if canonical and canonical not in mapping:
            mapping[canonical] = col
    return mapping


def _count_selected_crops(crop_df: pl.DataFrame, language: str) -> tuple[int, int, str]:
    meta = extract_crop_metadata(crop_df, language)
    total_candidates = len(meta.get("labels", []))
    select_col = meta.get("select_col")
    if not select_col or select_col not in crop_df.columns:
        return 0, total_candidates, "missing_select_column"
    selected_values = crop_df.get_column(select_col).to_list()
    selected_count = sum(
        1 for v in selected_values
        if str(v or "").strip() not in {"", "0", "0.0", "none", "nan"}
    )
    return selected_count, total_candidates, "ok"


def _placeholder_scan_columns(questions_df: pl.DataFrame, candidate_columns: Optional[list[str]] = None) -> list[str]:
    cols = [c for c in (candidate_columns or RESTORE_TEXT_COLUMNS) if c in questions_df.columns]
    if "Codes" in questions_df.columns and "Codes" not in cols:
        cols.append("Codes")
    return cols


def _extract_placeholder_tokens(text: str) -> list[str]:
    return re.findall(r"\$[^$\r\n]+\$", str(text or ""))


def _has_broken_placeholder_token(text: str) -> bool:
    txt = str(text or "")
    if "$" not in txt:
        return False
    stripped = re.sub(r"\$[^$\r\n]+\$", "", txt)
    return "$" in stripped


def _count_unresolved_placeholders(questions_df: pl.DataFrame, candidate_columns: Optional[list[str]] = None) -> int:
    cols = _placeholder_scan_columns(questions_df, candidate_columns)
    if not cols:
        return 0
    unresolved = 0
    for row in questions_df.select(cols).iter_rows(named=True):
        for col in cols:
            txt = str(row.get(col) or "")
            if _extract_placeholder_tokens(txt):
                unresolved += 1
    return unresolved


def _selected_crop_labels(crop_df: pl.DataFrame, language: str, limit: int = 15) -> list[str]:
    meta = extract_crop_metadata(crop_df, language)
    label_col = meta.get("label_col")
    select_col = meta.get("select_col")
    if not label_col or not select_col or label_col not in crop_df.columns or select_col not in crop_df.columns:
        return []
    labels = []
    for row in crop_df.select([label_col, select_col]).iter_rows(named=True):
        sel = str(row.get(select_col) or "").strip().lower()
        if sel not in {"", "0", "0.0", "none", "nan"}:
            lbl = str(row.get(label_col) or "").strip()
            if lbl:
                labels.append(lbl)
    if len(labels) > limit:
        return labels[:limit] + [f"... (+{len(labels) - limit} more)"]
    return labels


def _selected_crop_entries(crop_df: pl.DataFrame, language: str) -> list[dict]:
    if crop_df is None or crop_df.height == 0:
        return []
    meta = extract_crop_metadata(crop_df, language)
    label_col = meta.get("label_col")
    dataset_col = meta.get("dataset_col")
    select_col = meta.get("select_col")
    if not label_col or not select_col or label_col not in crop_df.columns or select_col not in crop_df.columns:
        return []

    cols = [label_col, select_col]
    if dataset_col and dataset_col in crop_df.columns:
        cols.append(dataset_col)

    out = []
    for row in crop_df.select(cols).iter_rows(named=True):
        if not _is_selected_crop_value(row.get(select_col)):
            continue
        label = str(row.get(label_col) or "").strip()
        code = str(row.get(dataset_col) or "").strip() if dataset_col and dataset_col in crop_df.columns else ""
        if not label and not code:
            continue
        key = code.lower() if code else label.lower()
        display = f"{code} - {label}" if code and label else (label or code)
        out.append({"key": key, "display": display})
    return out


def _join_short(items: list[str], limit: int = 12) -> str:
    if not items:
        return "none"
    if len(items) <= limit:
        return ", ".join(items)
    return ", ".join(items[:limit]) + f", ... (+{len(items) - limit} more)"


def _build_crop_round_delta_issue_rows(
    current_crop_df: pl.DataFrame,
    reference_crop_df: pl.DataFrame,
    language: str,
    crop_qnames: Optional[list[str]] = None,
) -> pl.DataFrame:
    schema = {
        "issue_type": pl.Utf8,
        "set_name": pl.Utf8,
        "Q Name": pl.Utf8,
        "field": pl.Utf8,
        "current": pl.Utf8,
        "reference": pl.Utf8,
        "severity": pl.Utf8,
        "excel_row": pl.Int64,
    }
    if reference_crop_df is None or reference_crop_df.height == 0:
        return pl.DataFrame(schema=schema)

    cur_entries = _selected_crop_entries(current_crop_df, language)
    ref_entries = _selected_crop_entries(reference_crop_df, language)
    if not cur_entries and not ref_entries:
        return pl.DataFrame(schema=schema)

    cur_map = {e["key"]: e["display"] for e in cur_entries}
    ref_map = {e["key"]: e["display"] for e in ref_entries}
    cur_keys = set(cur_map.keys())
    ref_keys = set(ref_map.keys())

    added = sorted(cur_map[k] for k in (cur_keys - ref_keys))
    removed = sorted(ref_map[k] for k in (ref_keys - cur_keys))
    qname_txt = ", ".join(crop_qnames) if crop_qnames else "crop placeholder question(s)"

    row = {
        "issue_type": "replacement_crop_round_delta",
        "set_name": "REPLACEMENT",
        "Q Name": qname_txt,
        "field": "Crop list (selected top 10)",
        "current": (
            f"Current selected ({len(cur_entries)}): {_join_short(sorted(cur_map.values()))} | "
            f"Added vs reference ({len(added)}): {_join_short(added)}"
        ),
        "reference": (
            f"Reference selected ({len(ref_entries)}): {_join_short(sorted(ref_map.values()))} | "
            f"Removed vs reference ({len(removed)}): {_join_short(removed)}"
        ),
        "severity": "info",
        "excel_row": None,
    }
    return pl.DataFrame([row], schema=schema)


def _build_replacement_issue_rows(
    validated_questions: pl.DataFrame,
    replacements: dict[str, str],
    crop_df: pl.DataFrame,
    language: str,
    reference_questions: pl.DataFrame | None = None,
    additional_info_diag: dict | None = None,
    replacements_by_language: dict[str, dict[str, str]] | None = None,
) -> pl.DataFrame:
    schema = {
        "issue_type": pl.Utf8,
        "set_name": pl.Utf8,
        "Q Name": pl.Utf8,
        "field": pl.Utf8,
        "current": pl.Utf8,
        "reference": pl.Utf8,
        "severity": pl.Utf8,
        "excel_row": pl.Int64,
    }
    rows = []

    selected_count, total_candidates, crop_mode = _count_selected_crops(crop_df, language)
    if crop_mode == "missing_select_column":
        rows.append({
            "issue_type": "replacement_crop_selection_mismatch",
            "set_name": "REPLACEMENT",
            "Q Name": "",
            "field": "Crop list",
            "current": f"selection column missing ({total_candidates} candidate crops)",
            "reference": "Crop list should mark exactly 10 selected crops",
            "severity": "medium",
            "excel_row": None,
        })
    elif selected_count != 10:
        selected_labels = _selected_crop_labels(crop_df, language)
        rows.append({
            "issue_type": "replacement_crop_selection_mismatch",
            "set_name": "REPLACEMENT",
            "Q Name": "",
            "field": "Crop list",
            "current": f"{selected_count}/10 selected | {', '.join(selected_labels) if selected_labels else 'no selected labels'}",
            "reference": "Expected exactly 10 selected crops for $TEN MOST COMMON CROPS$ / $CROP CODES$ substitution",
            "severity": "medium",
            "excel_row": None,
        })

    if not replacements and not any((replacements_by_language or {}).values()):
        detail = "No replacement keys loaded"
        if additional_info_diag:
            detail = (
                f"No replacement keys loaded | rows={additional_info_diag.get('rows_with_any_data', 0)} "
                f"placeholder-style={additional_info_diag.get('rows_with_placeholder_style_original', 0)} "
                f"plain-style={additional_info_diag.get('rows_with_non_placeholder_original', 0)}"
            )
        rows.append({
            "issue_type": "replacement_additional_info_missing",
            "set_name": "REPLACEMENT",
            "Q Name": "",
            "field": "Additional information",
            "current": detail,
            "reference": "Additional information must define all $...$ placeholders used in survey",
            "severity": "high",
            "excel_row": None,
        })

    def _effective_replacements_for_col(col_name: str) -> dict[str, str]:
        lang = _language_code_for_text_column(col_name)
        if replacements_by_language and lang and replacements_by_language.get(lang):
            return replacements_by_language[lang]
        return replacements or {}

    seen = set()
    missing_key_locs = set()

    if reference_questions is not None and reference_questions.height > 0:
        ref_cols = _placeholder_scan_columns(reference_questions)
        for row in reference_questions.select(["Q Name", "excel_row"] + ref_cols).iter_rows(named=True):
            qname = str(row.get("Q Name") or "")
            excel_row = row.get("excel_row")
            for col in ref_cols:
                txt = str(row.get(col) or "")
                if not txt:
                    continue
                col_repl = _effective_replacements_for_col(col)
                col_idx = _build_placeholder_index(col_repl)
                for token in _extract_placeholder_tokens(txt):
                    if _is_crop_placeholder_token(token):
                        continue
                    if _lookup_placeholder_replacement(token, col_repl, col_idx) is not None:
                        continue
                    tok_norm = _normalize_placeholder_token(token)
                    key = ("replacement_missing_key", qname, col, token)
                    if key in seen:
                        continue
                    seen.add(key)
                    missing_key_locs.add((qname, col, tok_norm))
                    sev = "high"
                    rows.append({
                        "issue_type": "replacement_missing_key",
                        "set_name": "REPLACEMENT",
                        "Q Name": qname,
                        "field": col,
                        "current": token,
                        "reference": "Placeholder key missing/blank in Additional information",
                        "severity": sev,
                        "excel_row": excel_row,
                    })

    cols = _placeholder_scan_columns(validated_questions)
    for row in validated_questions.select(["Q Name", "excel_row"] + cols).iter_rows(named=True):
        qname = str(row.get("Q Name") or "")
        excel_row = row.get("excel_row")
        for col in cols:
            txt = str(row.get(col) or "")
            if not txt:
                continue

            if _has_broken_placeholder_token(txt):
                key = ("replacement_malformed_placeholder", qname, col, txt)
                if key not in seen:
                    seen.add(key)
                    rows.append({
                        "issue_type": "replacement_malformed_placeholder",
                        "set_name": "REPLACEMENT",
                        "Q Name": qname,
                        "field": col,
                        "current": txt,
                        "reference": "Placeholder token should be balanced and closed (example: $season$)",
                        "severity": "medium",
                        "excel_row": excel_row,
                    })

            col_repl = _effective_replacements_for_col(col)
            col_idx = _build_placeholder_index(col_repl)
            for token in _extract_placeholder_tokens(txt):
                tok_norm = _normalize_placeholder_token(token)
                if _is_crop_placeholder_token(token):
                    issue_type = "replacement_unresolved_placeholder"
                    severity = "high"
                    reference = "Crop placeholders must be fully expanded in validated output"
                elif _lookup_placeholder_replacement(token, col_repl, col_idx) is None:
                    issue_type = "replacement_missing_key"
                    severity = "high"
                    reference = "Placeholder key missing from Additional information"
                    missing_key_locs.add((qname, col, tok_norm))
                else:
                    if (qname, col, tok_norm) in missing_key_locs:
                        continue
                    issue_type = "replacement_unresolved_placeholder"
                    severity = "medium"
                    reference = "Placeholder replacement exists but token remained unresolved"

                key = (issue_type, qname, col, token)
                if key in seen:
                    continue
                seen.add(key)
                rows.append({
                    "issue_type": issue_type,
                    "set_name": "REPLACEMENT",
                    "Q Name": qname,
                    "field": col,
                    "current": token,
                    "reference": reference,
                    "severity": severity,
                    "excel_row": excel_row,
                })

    if not rows:
        return pl.DataFrame(schema=schema)
    return pl.DataFrame(rows, schema=schema)


def _fetch_admin_reference(admin_level: str, iso3: str) -> tuple[list[dict], str]:
    import urllib.request
    import json

    iso3 = str(iso3 or "").upper().strip()
    if not iso3:
        return [], "missing_iso3"

    urls = {
        "Admin 1": "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/1/query?where=adm0_ISO3%20%3D%20'{iso3}'&outFields=adm1_name,adm1_name_local,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json",
        "Admin 2": "https://services5.arcgis.com/sjP4Ugu5s0dZWLjd/arcgis/rest/services/Administrative_Boundaries_Reference_(view_layer)/FeatureServer/0/query?where=adm0_ISO3%20%3D%20'{iso3}'&outFields=adm2_name,adm2_name_local,adm2_pcode,adm1_name,adm1_pcode,adm0_name,adm0_ISO3&returnGeometry=false&outSR=4326&f=json",
    }
    rename = {
        "Admin 1": {
            "adm1_name": "adm1_name",
            "adm1_name_local": "adm1_name_local",
            "adm1_pcode": "adm1_pcode",
            "adm0_name": "adm0_name",
            "adm0_ISO3": "adm0_ISO3",
        },
        "Admin 2": {
            "adm2_name": "adm2_name",
            "adm2_name_local": "adm2_name_local",
            "adm2_pcode": "adm2_pcode",
            "adm1_name": "adm1_name",
            "adm1_pcode": "adm1_pcode",
            "adm0_name": "adm0_name",
            "adm0_ISO3": "adm0_ISO3",
        },
    }

    if admin_level not in urls:
        return [], f"unsupported_admin_level:{admin_level}"

    url = urls[admin_level].format(iso3=iso3)
    try:
        with urllib.request.urlopen(url, timeout=25) as resp:
            data = json.loads(resp.read().decode())
    except Exception as e:
        return [], str(e)

    features = data.get("features", []) or []
    mapping = rename[admin_level]
    rows = []
    for feat in features:
        attrs = feat.get("attributes", {}) or {}
        rows.append({dst: attrs.get(src) for src, dst in mapping.items()})
    return rows, ""


def _write_table_sheet(ws, rows: list[dict]):
    if not rows:
        ws.cell(row=1, column=1, value="No rows")
        return 0
    headers = list(rows[0].keys())
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    r = 2
    for rd in rows:
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=rd.get(h))
        r += 1
    return len(rows)


def _apply_admin_sheets(output_questionnaire_path: Path, iso3: str) -> dict:
    rows = []
    fetched = {}
    errors = {}
    for lvl in ["Admin 1", "Admin 2"]:
        data_rows, err = _fetch_admin_reference(lvl, iso3)
        if err:
            errors[lvl] = err
        else:
            fetched[lvl] = data_rows

    if fetched:
        details = ", ".join([
            f"admin1={len(fetched.get('Admin 1', []))}",
            f"admin2={len(fetched.get('Admin 2', []))}",
        ])
        rows.append({"check": "AGOL fetch", "status": "PASS" if not errors else "WARN", "details": details, "severity": "info" if not errors else "medium"})
    else:
        err_txt = "; ".join(f"{k}: {v}" for k, v in errors.items()) or "AGOL returned no data"
        rows.append({"check": "AGOL fetch", "status": "FAIL", "details": err_txt, "severity": "high"})
        return {"rows": rows}

    try:
        wb = openpyxl.load_workbook(output_questionnaire_path)
        for lvl, data_rows in fetched.items():
            sheet_name = f"{lvl} info"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            n = _write_table_sheet(ws, data_rows)
            rows.append({
                "check": f"Admin sheet added: {lvl.lower().replace(' ', '')}",
                "status": "PASS",
                "details": f"{n} rows written",
                "severity": "info",
            })
        wb.save(output_questionnaire_path)
    except Exception as e:
        rows.append({"check": "Admin sheet write", "status": "FAIL", "details": str(e), "severity": "high"})

    return {"rows": rows}


def _build_replacement_status_rows(
    validated_questions: pl.DataFrame,
    crop_df: pl.DataFrame,
    replacements: dict[str, str],
    language: str,
    iso3: str,
    output_questionnaire_path: Optional[Path],
    write_status: dict,
    additional_info_diag: dict | None = None,
    replacement_issues: pl.DataFrame | None = None,
) -> dict:
    rows = []

    selected_count, total_candidates, crop_mode = _count_selected_crops(crop_df, language)
    if crop_mode == "missing_select_column":
        rows.append({
            "check": "Crop selection (top 10)",
            "status": "WARN",
            "details": f"selection column not found; {total_candidates} candidates detected",
            "severity": "medium",
        })
    else:
        crop_status = "PASS" if selected_count == 10 else "WARN"
        crop_sev = "info" if crop_status == "PASS" else "medium"
        rows.append({
            "check": "Crop selection (top 10)",
            "status": crop_status,
            "details": f"{selected_count}/10 selected ({total_candidates} candidates)",
            "severity": crop_sev,
        })

    repl_df = replacement_issues if replacement_issues is not None else pl.DataFrame(schema={"issue_type": pl.Utf8, "severity": pl.Utf8})
    miss_high = repl_df.filter((pl.col("issue_type") == "replacement_missing_key") & (pl.col("severity") == "high")).height if repl_df.height > 0 else 0
    miss_med = repl_df.filter((pl.col("issue_type") == "replacement_missing_key") & (pl.col("severity") == "medium")).height if repl_df.height > 0 else 0
    unresolved_high = repl_df.filter((pl.col("issue_type") == "replacement_unresolved_placeholder") & (pl.col("severity") == "high")).height if repl_df.height > 0 else 0
    unresolved_med = repl_df.filter((pl.col("issue_type") == "replacement_unresolved_placeholder") & (pl.col("severity") == "medium")).height if repl_df.height > 0 else 0
    malformed_n = repl_df.filter(pl.col("issue_type") == "replacement_malformed_placeholder").height if repl_df.height > 0 else 0

    key_load_ok = bool(replacements)
    key_load_details = f"{len(replacements)} placeholder key(s) loaded from Additional information"
    if not key_load_ok:
        if additional_info_diag:
            key_load_details = (
                f"No replacement keys loaded | rows={additional_info_diag.get('rows_with_any_data', 0)} "
                f"placeholder-style={additional_info_diag.get('rows_with_placeholder_style_original', 0)} "
                f"plain-style={additional_info_diag.get('rows_with_non_placeholder_original', 0)}"
            )
        else:
            key_load_details = "No replacement keys loaded from Additional information"

    fail_parts = []
    if miss_high > 0:
        fail_parts.append(f"missing keys (high)={miss_high}")
    if unresolved_high > 0:
        fail_parts.append(f"unresolved placeholders (high)={unresolved_high}")

    warn_parts = []
    if miss_med > 0:
        warn_parts.append(f"missing keys (medium)={miss_med}")
    if unresolved_med > 0:
        warn_parts.append(f"unresolved placeholders (medium)={unresolved_med}")
    if malformed_n > 0:
        warn_parts.append(f"malformed placeholders={malformed_n}")

    check_statuses = []
    check_statuses.append("PASS" if key_load_ok else "FAIL")
    check_statuses.append("FAIL" if fail_parts else "PASS")
    check_statuses.append("WARN" if warn_parts else "PASS")

    fail_n = sum(1 for s in check_statuses if s == "FAIL")
    warn_n = sum(1 for s in check_statuses if s == "WARN")
    pass_n = sum(1 for s in check_statuses if s == "PASS")

    if fail_n > 0:
        summary_status = "FAIL"
        summary_sev = "high"
    elif warn_n > 0:
        summary_status = "WARN"
        summary_sev = "medium"
    else:
        summary_status = "PASS"
        summary_sev = "info"

    detail_parts = [f"{summary_status} (pass={pass_n}, warn={warn_n}, fail={fail_n})", key_load_details]
    if fail_parts:
        detail_parts.append("; ".join(fail_parts))
    if warn_parts:
        detail_parts.append("; ".join(warn_parts))

    rows.append({
        "check": "Additional info replacement",
        "status": summary_status,
        "details": " | ".join([p for p in detail_parts if p]),
        "severity": summary_sev,
    })

    if write_status.get("status") == "OK" and output_questionnaire_path is not None:
        rows.extend(_apply_admin_sheets(output_questionnaire_path, iso3).get("rows", []))
    else:
        rows.append({
            "check": "AGOL fetch",
            "status": "WARN",
            "details": "Skipped because validated questionnaire could not be written",
            "severity": "medium",
        })

    return {"rows": rows}


def write_validated_questionnaire(
    source_questionnaire_path: Path,
    output_questionnaire_path: Path,
    questions_df: pl.DataFrame,
) -> None:
    """
    Creates a validated questionnaire workbook by copying the original file and
    writing the converted survey values row-by-row using Q Name as key.
    """
    copyfile(source_questionnaire_path, output_questionnaire_path)

    wb = openpyxl.load_workbook(output_questionnaire_path)
    ws = _resolve_survey_sheet(wb)
    header_map = _sheet_header_map(ws, header_row=3)

    q_col = header_map.get("Q Name")
    if not q_col:
        raise KeyError("Column 'Q Name' not found in survey header row.")

    qname_to_row = {}
    for row_idx in range(4, ws.max_row + 1):
        qname = ws.cell(row=row_idx, column=q_col).value
        key = str(qname).strip() if qname is not None else ""
        if key:
            qname_to_row[key] = row_idx

    writable_cols = [
        col for col in questions_df.columns
        if col not in {"Q Name", "excel_row", "source_file"} and col in header_map
    ]

    for row in questions_df.select(["Q Name"] + writable_cols).to_dicts():
        qname = str(row.get("Q Name") or "").strip()
        target_row = qname_to_row.get(qname)
        if not target_row:
            continue
        for col_name in writable_cols:
            cell = ws.cell(row=target_row, column=header_map[col_name])
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = row.get(col_name)

    wb.save(output_questionnaire_path)
    print(f"Validated questionnaire saved to: {output_questionnaire_path}")
    return {"status": "OK", "path": str(output_questionnaire_path)}


# Duplicate Q Name detection
_dup_df = (
    current_questions
    .group_by("Q Name")
    .agg(pl.len().alias("n"))
    .filter(pl.col("n") > 1)
)
if _dup_df.height > 0:
    _dup_issues = _dup_df.select([
        pl.lit("duplicate_qname").alias("issue_type"),
        pl.lit("").alias("set_name"),
        pl.col("Q Name"),
        pl.lit("Q Name").alias("field"),
        pl.col("n").cast(pl.Utf8).alias("current"),
        pl.lit("unique").alias("reference"),
        pl.lit("high").alias("severity"),
        pl.lit(None).cast(pl.Int64).alias("excel_row"),
    ])
    all_issues = pl.concat([all_issues, _dup_issues], how="diagonal")
    if "mandatory_cat" in all_issues.columns:
        all_issues = all_issues.with_columns(pl.col("mandatory_cat").fill_null(""))
    print(f"Duplicate Q Names: {_dup_df.height} name(s) duplicated")
else:
    print("Duplicate Q Names: none found")

try:
    _sub_status = write_validated_questionnaire(
        source_questionnaire_path = run["questionnaire_path"],
        output_questionnaire_path = run["validated_questionnaire_file"],
        questions_df              = validated_output_questions,
    )
except Exception as _e:
    _sub_status = {"status": "ERROR", "path": "", "error": str(_e)}
    print(f"Warning: validated questionnaire failed: {_e}")

_additional_info_diag = analyze_additional_info_substitutions(run["questionnaire_path"], cfg.language)

_replacement_issues = _build_replacement_issue_rows(
    validated_questions = validated_output_questions,
    replacements = current_text_replacements,
    crop_df = current_crop_list,
    language = cfg.language,
    reference_questions = reference_questions,
    additional_info_diag = _additional_info_diag,
    replacements_by_language = current_text_replacements_by_language,
)
if _replacement_issues.height > 0:
    all_issues = pl.concat([all_issues, _replacement_issues], how="diagonal")
    if "mandatory_cat" in all_issues.columns:
        all_issues = all_issues.with_columns(pl.col("mandatory_cat").fill_null(""))
_crop_qnames_for_delta = sorted(set(find_crop_placeholder_questions(template_questions_for_restore)))
_crop_round_delta_issues = _build_crop_round_delta_issue_rows(
    current_crop_df = current_crop_list,
    reference_crop_df = reference_crop_list,
    language = cfg.language,
    crop_qnames = _crop_qnames_for_delta,
)
if _crop_round_delta_issues.height > 0:
    all_issues = pl.concat([all_issues, _crop_round_delta_issues], how="diagonal")
    if "mandatory_cat" in all_issues.columns:
        all_issues = all_issues.with_columns(pl.col("mandatory_cat").fill_null(""))
_replacement_status = _build_replacement_status_rows(
    validated_questions      = validated_output_questions,
    crop_df                  = current_crop_list,
    replacements             = current_text_replacements,
    language                 = cfg.language,
    iso3                     = cfg.iso3,
    output_questionnaire_path= run["validated_questionnaire_file"],
    write_status             = _sub_status,
    additional_info_diag     = _additional_info_diag,
    replacement_issues       = _replacement_issues,
)

print(f"Replacement diagnostics: {_replacement_issues.height} issue(s)")

export_validation_report(
    all_issues           = all_issues,
    question_changes_view= question_changes_view,
    option_changes_view  = option_changes_view,
    result_file          = run["report_file"],
    rules                = rules,
    found_info           = _found_info,
    replacement_status   = _replacement_status,
)

_summary(
    all_issues  = all_issues if 'all_issues' in dir() else None,
    report_path = run.get('report_file') if isinstance(run, dict) else None,
    extra_paths = [("Validated", run.get("validated_questionnaire_file"))] if isinstance(run, dict) and run.get("validated_questionnaire_file") else None,
)
