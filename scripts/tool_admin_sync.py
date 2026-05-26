from __future__ import annotations

import argparse
from pathlib import Path
from shutil import copy2

import openpyxl

from admin_tools_common import (
    as_bool,
    detect_label_column,
    fetch_agol_admin_rows,
    merge_runtime_config,
    now_stamp,
    print_runtime_banner,
    resolve_source_workbook,
)


def _norm(v) -> str:
    return str(v or "").strip().lower()


def _style_copy(style_obj):
    try:
        return style_obj.copy()
    except Exception:
        return style_obj


def _read_choices_rows(ws):
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column))
    if not rows:
        return [], [], {}
    n_cols = ws.max_column or 1
    headers = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    def row_values(cells):
        vals = [c.value for c in cells]
        if len(vals) < n_cols:
            vals.extend([None] * (n_cols - len(vals)))
        return vals[:n_cols]

    def row_styles(cells):
        out = []
        for i in range(n_cols):
            if i < len(cells):
                out.append(_style_copy(cells[i]._style))
            else:
                out.append(None)
        return out

    all_rows = [(row_values(c), row_styles(c)) for c in rows]
    return all_rows, headers, idx


def _extract_existing_admin_entries(all_rows, idx, include_admin3: bool, language: str):
    list_col = idx.get("list_name")
    name_col = idx.get("name")
    label_col = detect_label_column(list(idx.keys()), language)
    parent_col = None
    for c in ("my_filter_admin", "choice_filter", "filter"):
        if c in idx:
            parent_col = idx[c]
            break
    sample_col = idx.get("my_filter_sample")

    if list_col is None or name_col is None:
        raise ValueError("choices sheet must contain 'list_name' and 'name' columns.")

    target_lists = ["admin1", "admin2"] + (["admin3"] if include_admin3 else [])
    per_list = {k: [] for k in target_lists}
    preserve_sample = {}

    for vals, _ in all_rows[1:]:
        list_name = str(vals[list_col] if list_col < len(vals) else "").strip()
        if list_name not in per_list:
            continue
        code = str(vals[name_col] if name_col < len(vals) else "").strip()
        if not code:
            continue
        label = str(vals[label_col] if (label_col is not None and label_col < len(vals)) else "").strip()
        parent = str(vals[parent_col] if (parent_col is not None and parent_col < len(vals)) else "").strip()
        per_list[list_name].append(
            {
                "list_name": list_name,
                "code": code,
                "label": label,
                "parent": parent,
                "base_vals": list(vals),
            }
        )
        if sample_col is not None and sample_col < len(vals):
            sample_val = vals[sample_col]
            if sample_val is not None and str(sample_val).strip() != "":
                preserve_sample[(list_name, code)] = sample_val
    return per_list, preserve_sample


def _build_subset_entries(current_per_list, agol_rows, include_admin3: bool):
    target_lists = ["admin1", "admin2"] + (["admin3"] if include_admin3 else [])
    entries_out = {}
    stats = {"used_from_agol": 0, "missing_in_agol_kept_previous": 0, "dropped_duplicates": 0}

    for list_name in target_lists:
        agol_map = {c: {"code": c, "label": l, "parent": p} for c, l, p in agol_rows.get(list_name, []) if c}
        seen = set()
        built = []
        for row in current_per_list.get(list_name, []):
            code = str(row.get("code") or "").strip()
            if not code:
                continue
            if code in seen:
                stats["dropped_duplicates"] += 1
                continue
            seen.add(code)

            if code in agol_map:
                a = agol_map[code]
                built.append(
                    {
                        "list_name": list_name,
                        "code": code,
                        "label": a.get("label", ""),
                        "parent": a.get("parent", ""),
                        "base_vals": row.get("base_vals"),
                    }
                )
                stats["used_from_agol"] += 1
            else:
                built.append(row)
                stats["missing_in_agol_kept_previous"] += 1
        entries_out[list_name] = built
    return entries_out, stats


def _build_full_entries(agol_rows, include_admin3: bool):
    target_lists = ["admin1", "admin2"] + (["admin3"] if include_admin3 else [])
    entries_out = {}
    for list_name in target_lists:
        built = []
        for code, label, parent in agol_rows.get(list_name, []):
            if not code:
                continue
            built.append(
                {
                    "list_name": list_name,
                    "code": code,
                    "label": label,
                    "parent": parent,
                    "base_vals": None,
                }
            )
        entries_out[list_name] = built
    return entries_out


def _rewrite_admin_choices(ws, entries_per_list, language: str, preserve_sample: dict, preserve_sample_column: bool):
    all_rows, headers, idx = _read_choices_rows(ws)
    if not all_rows:
        return
    n_cols = len(all_rows[0][0])
    list_col = idx.get("list_name", 0)
    name_col = idx.get("name", 1)
    label_col = detect_label_column(headers, language)

    parent_col = None
    for c in ("my_filter_admin", "choice_filter", "filter"):
        if c in idx:
            parent_col = idx[c]
            break
    if parent_col is None and n_cols >= 5:
        parent_col = 4
    sample_col = idx.get("my_filter_sample")

    skip = set(entries_per_list.keys())
    kept_rows = []
    style_by_list = {}
    blank_style = None
    generic_style = all_rows[1][1] if len(all_rows) > 1 else all_rows[0][1]

    def _is_blank(vals):
        return all(v is None or str(v).strip() == "" for v in vals)

    for vals, styles in all_rows[1:]:
        list_txt = str(vals[list_col] if list_col < len(vals) and vals[list_col] is not None else "").strip()
        if list_txt in skip:
            style_by_list.setdefault(list_txt, styles)
            if blank_style is None and _is_blank(vals):
                blank_style = styles
            continue
        if blank_style is None and _is_blank(vals):
            blank_style = styles
        kept_rows.append((vals, styles))
    if blank_style is None:
        blank_style = generic_style

    new_rows = []
    for list_name, entries in entries_per_list.items():
        new_rows.append(([None] * n_cols, blank_style))
        row_style = style_by_list.get(list_name, generic_style)
        for e in entries:
            base_vals = e.get("base_vals")
            if isinstance(base_vals, list) and len(base_vals) == n_cols:
                row = list(base_vals)
            else:
                row = [None] * n_cols

            row[list_col] = list_name
            row[name_col] = e.get("code", "")
            if label_col is not None and label_col < n_cols:
                row[label_col] = e.get("label", "")
            if parent_col is not None and parent_col < n_cols:
                row[parent_col] = e.get("parent", "")
            if preserve_sample_column and sample_col is not None and sample_col < n_cols:
                key = (list_name, str(e.get("code") or "").strip())
                if key in preserve_sample:
                    row[sample_col] = preserve_sample[key]
            new_rows.append((row, row_style))

    payload = [all_rows[0]] + kept_rows + new_rows
    ws.delete_rows(1, ws.max_row)
    for r_idx, (vals, styles) in enumerate(payload, start=1):
        for c_idx in range(1, n_cols + 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=vals[c_idx - 1] if c_idx - 1 < len(vals) else None)
            if c_idx - 1 < len(styles) and styles[c_idx - 1] is not None:
                cell._style = _style_copy(styles[c_idx - 1])


def run(cfg: dict) -> int:
    print_runtime_banner("tool_admin_sync", cfg)
    source = resolve_source_workbook(cfg)
    print(f"  source      : {source}")
    sync_mode = str(cfg.get("sync_mode") or "subset_from_agol").strip().lower()
    include_admin3 = as_bool(cfg.get("include_admin3"), False)
    preserve_sample = as_bool(cfg.get("preserve_my_filter_sample"), True)
    dry_run = as_bool(cfg.get("dry_run"), False)

    out_dir = Path(str(cfg.get("output_dir")))
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{source.stem}_adminsync_{sync_mode}_{now_stamp()}.xlsx"

    if sync_mode == "keep_previous":
        copy2(source, out_file)
        print("  mode        : keep_previous (no admin edits)")
        print(f"  output      : {out_file}")
        return 0

    if dry_run:
        print("  dry_run     : true (no file written)")
        return 0

    copy2(source, out_file)
    wb = openpyxl.load_workbook(out_file)
    if "choices" not in wb.sheetnames:
        raise KeyError("Workbook has no 'choices' sheet.")
    ws = wb["choices"]

    all_rows, _, idx = _read_choices_rows(ws)
    current_per_list, sample_map = _extract_existing_admin_entries(
        all_rows=all_rows,
        idx=idx,
        include_admin3=include_admin3,
        language=str(cfg.get("language") or "en"),
    )
    agol_rows = fetch_agol_admin_rows(str(cfg.get("iso3") or ""), include_admin3=include_admin3)

    if sync_mode == "subset_from_agol":
        entries_per_list, stats = _build_subset_entries(current_per_list, agol_rows, include_admin3=include_admin3)
        print(
            "  subset stats: "
            + f"used_from_agol={stats['used_from_agol']} "
            + f"missing_in_agol_kept_previous={stats['missing_in_agol_kept_previous']} "
            + f"dropped_duplicates={stats['dropped_duplicates']}"
        )
    elif sync_mode == "full_from_agol":
        entries_per_list = _build_full_entries(agol_rows, include_admin3=include_admin3)
    else:
        raise ValueError(
            "sync_mode must be one of: keep_previous, subset_from_agol, full_from_agol"
        )

    _rewrite_admin_choices(
        ws=ws,
        entries_per_list=entries_per_list,
        language=str(cfg.get("language") or "en"),
        preserve_sample=sample_map,
        preserve_sample_column=preserve_sample,
    )
    wb.save(out_file)

    print(f"  mode        : {sync_mode}")
    print(f"  output      : {out_file}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Standalone admin sync tool for validated questionnaire outputs."
    )
    parser.add_argument("--config", default="", help="Optional path to admin_tools_config.yaml")
    parser.add_argument("--table", dest="config", help=argparse.SUPPRESS)
    parser.add_argument("--source-file", default="", help="Optional explicit source workbook path override")
    parser.add_argument("--sync-mode", default="", help="Optional mode override")
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parents[1]
    cfg_path = Path(args.config).expanduser().resolve() if args.config else None
    cfg = merge_runtime_config(repo_root, explicit_cfg=cfg_path)
    if args.source_file:
        cfg["source_mode"] = "custom"
        cfg["source_file"] = args.source_file
    if args.sync_mode:
        cfg["sync_mode"] = args.sync_mode.strip().lower()

    try:
        return run(cfg)
    except Exception as e:
        print(f"ERROR: {e}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
