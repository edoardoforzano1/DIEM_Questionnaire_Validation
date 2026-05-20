from openpyxl import load_workbook
p=r'c:/git/DIEM_Questionnaire_Validation/tmp_report_out5/geopoll_output/report_geopoll_fr_COD_R11_20260520.xlsx'
wb=load_workbook(p)
ws=wb['Option Changes']
h=[ws.cell(2,c).value for c in range(1,20)]
idx={str(v):c for c,v in enumerate(h,1) if v}
issue=idx['Issue type']; q=idx['Q Name']; sev=idx['Severity']; typ=idx.get('Type'); det=idx['Detail']
for r in range(3, ws.max_row+1):
    it=ws.cell(r,issue).value
    if str(it)=='codes_option_count_mismatch':
        print(r, ws.cell(r,q).value, ws.cell(r,typ).value if typ else '', ws.cell(r,sev).value, ws.cell(r,det).value)
