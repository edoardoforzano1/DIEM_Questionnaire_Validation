"""
Smoke test for DIEM Questionnaire Validation.

Runs the full validation pipeline against the currently configured questionnaire
and verifies that the outputs have the expected structure.  Does not check
specific issue counts — only that the pipeline completed and produced valid files.

Usage:
    python smoke_test.py
    smoke_test           (via the .bat launcher)
"""
from __future__ import annotations

import subprocess
import sys
import time
from pathlib import Path

import openpyxl
import yaml

REPO = Path(__file__).resolve().parent
CFG_PATH = REPO / "configuration" / "validation_config.yaml"

# Minimum sheets each tool's report must contain
REQUIRED_SHEETS = {
    "kobo":    {"Summary", "Question Changes", "Choice Changes", "Questionnaire Structure", "Replacement Issues"},
    "geopoll": {"Summary", "Question Changes", "Option Changes",  "Questionnaire Structure", "Replacement Issues"},
}

# Columns that must appear in the Question Changes sheet header row
ISSUE_COLUMNS_REQUIRED = {"issue_type", "severity"}

# ---------------------------------------------------------------------------

_passes: list[bool] = []


def _check(label: str, ok: bool, detail: str = "") -> bool:
    tag = "PASS" if ok else "FAIL"
    suffix = f"  ({detail})" if detail and not ok else ""
    print(f"  {tag}  {label}{suffix}")
    _passes.append(ok)
    return ok


def _section(title: str) -> None:
    print(f"\n[{title}]")


# ---------------------------------------------------------------------------

def _load_config() -> dict | None:
    _section("Config files")
    if not _check("validation_config.yaml exists", CFG_PATH.exists(), str(CFG_PATH)):
        return None
    try:
        cfg = yaml.safe_load(CFG_PATH.read_text(encoding="utf-8")) or {}
    except Exception as exc:
        _check("validation_config.yaml parses", False, str(exc))
        return None
    _check("validation_config.yaml parses", True)

    # Profile overlay
    profile = str(cfg.get("config_profile") or "").strip()
    if profile:
        profiles_dir = Path(
            cfg.get("config_profiles_dir")
            or (Path(cfg.get("output_dir") or cfg.get("working_dir") or "C:/Temp") / "config_profiles")
        )
        candidate = Path(profile) if Path(profile).is_absolute() else profiles_dir / profile
        for suffix in ("", ".yaml", ".yml"):
            p = candidate if suffix == "" else candidate.with_suffix(suffix)
            if p.exists():
                override = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
                cfg.update(override)
                _check(f"config_profile '{profile}' loaded", True)
                break
        else:
            _check(f"config_profile '{profile}' found", False, str(candidate))

    tool = str(cfg.get("tool") or "").lower()
    _check("tool is kobo or geopoll", tool in {"kobo", "geopoll"}, f"got {tool!r}")
    _check("iso3 is set", bool(str(cfg.get("iso3") or "").strip()))
    _check("language is set", bool(str(cfg.get("language") or "").strip()))
    return cfg


def _check_inputs(cfg: dict) -> None:
    _section("Input files")
    working_dir = Path(str(cfg.get("working_dir") or "C:/Temp"))
    q_file = cfg.get("questionnaire_file") or cfg.get("questionnaire_path") or ""
    q_path = Path(q_file) if Path(q_file).is_absolute() else working_dir / q_file
    _check("questionnaire file exists", q_path.exists(), str(q_path))

    reference_mode = str(cfg.get("reference_mode") or "latest_template")
    if reference_mode == "latest_template":
        tpl_dir = Path(str(cfg.get("templates_dir") or ""))
        tpl_ok = _check("templates_dir exists", tpl_dir.exists(), str(tpl_dir))
        if tpl_ok:
            tpl_files = list(tpl_dir.glob("*.xlsx"))
            _check("templates_dir contains .xlsx files", bool(tpl_files), f"{len(tpl_files)} found")
    elif reference_mode == "previous_round":
        prev_file = str(cfg.get("previous_round_file") or "")
        prev_path = Path(prev_file) if Path(prev_file).is_absolute() else working_dir / prev_file
        _check("previous_round_file exists", prev_path.exists(), str(prev_path))


def _run_validator() -> bool:
    _section("Validation run")
    t0 = time.perf_counter()
    result = subprocess.run(
        [sys.executable, str(REPO / "validate.py")],
        cwd=str(REPO),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    elapsed = time.perf_counter() - t0
    ok = _check(
        f"validate.py exited cleanly ({elapsed:.1f}s)",
        result.returncode == 0,
        f"exit code {result.returncode}",
    )
    # Print last lines of stdout regardless
    if result.stdout:
        for line in result.stdout.strip().splitlines()[-12:]:
            print(f"    {line}")
    if result.stderr and not ok:
        print("  --- stderr (last 8 lines) ---")
        for line in result.stderr.strip().splitlines()[-8:]:
            print(f"    {line}")
    return ok


def _check_report(cfg: dict) -> None:
    _section("Report file")
    tool = str(cfg.get("tool") or "kobo").lower()
    language = str(cfg.get("language") or "").lower()
    iso3 = str(cfg.get("iso3") or "").upper()
    output_base = Path(str(cfg.get("output_dir") or "C:/Temp/questionnaire_validation"))
    out_dir = output_base / f"{tool}_output"
    today = time.strftime("%Y%m%d")

    candidates = sorted(
        out_dir.glob(f"report_{tool}_{language}_{iso3}*{today}.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not _check("report file produced today", bool(candidates), f"searched {out_dir}"):
        return
    report_path = candidates[0]
    print(f"    {report_path.name}")

    try:
        wb = openpyxl.load_workbook(report_path, data_only=True, read_only=True)
    except Exception as exc:
        _check("report opens without error", False, str(exc))
        return
    _check("report opens without error", True)

    for sheet in sorted(REQUIRED_SHEETS.get(tool, set())):
        _check(f"sheet '{sheet}' present", sheet in wb.sheetnames)

    # Summary must have content
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        rows = list(ws.iter_rows(min_row=1, max_row=6, values_only=True))
        has_content = any(any(c is not None for c in r) for r in rows)
        _check("Summary sheet has content", has_content)

    # Question Changes — look for the column header row (may not be row 1 due to section titles)
    if "Question Changes" in wb.sheetnames:
        ws = wb["Question Changes"]
        found_headers: set[str] = set()
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell is not None:
                    found_headers.add(str(cell).strip().lower())
        present = ISSUE_COLUMNS_REQUIRED & found_headers
        _check(
            "Question Changes has issue columns",
            len(present) == len(ISSUE_COLUMNS_REQUIRED),
            f"found {sorted(present)}, missing {sorted(ISSUE_COLUMNS_REQUIRED - present)}",
        )

    wb.close()


def _check_validated_questionnaire(cfg: dict) -> None:
    tool = str(cfg.get("tool") or "").lower()
    if tool != "kobo":
        return  # GeoPoll also produces a validated output but path handling differs
    _section("Validated questionnaire (KoBo)")
    language = str(cfg.get("language") or "").lower()
    iso3 = str(cfg.get("iso3") or "").upper()
    output_base = Path(str(cfg.get("output_dir") or "C:/Temp/questionnaire_validation"))
    out_dir = output_base / "kobo_output"
    today = time.strftime("%Y%m%d")

    candidates = sorted(
        out_dir.glob(f"validated_questionnaire_kobo_{language}_{iso3}*{today}.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not _check("validated questionnaire produced today", bool(candidates), f"searched {out_dir}"):
        return
    val_path = candidates[0]
    print(f"    {val_path.name}")

    try:
        vwb = openpyxl.load_workbook(val_path, data_only=True, read_only=True)
    except Exception as exc:
        _check("validated questionnaire opens without error", False, str(exc))
        return
    _check("validated questionnaire opens without error", True)
    _check("'survey' sheet present", "survey" in vwb.sheetnames)
    _check("'choices' sheet present", "choices" in vwb.sheetnames)
    vwb.close()


def _check_admin_config() -> None:
    _section("Admin tools config")
    admin_cfg_path = REPO / "configuration" / "admin_tools_config.yaml"
    if not admin_cfg_path.exists():
        print("  SKIP  admin_tools_config.yaml not present")
        return
    try:
        admin_cfg = yaml.safe_load(admin_cfg_path.read_text(encoding="utf-8")) or {}
    except Exception as exc:
        _check("admin_tools_config.yaml parses", False, str(exc))
        return
    _check("admin_tools_config.yaml parses", True)
    _check("iso3 configured", bool(str(admin_cfg.get("iso3") or "").strip()))
    source_mode = str((admin_cfg.get("source") or {}).get("mode") or "standard").lower()
    _check("source.mode is valid", source_mode in {"standard", "custom"}, f"got {source_mode!r}")
    sync_mode = str((admin_cfg.get("sync") or {}).get("mode") or "subset_from_agol").lower()
    _check("sync.mode is valid", sync_mode in {"subset_from_agol", "full_from_agol", "keep_previous"}, f"got {sync_mode!r}")


# ---------------------------------------------------------------------------

def main() -> int:
    print("\n" + "=" * 50)
    print("  DIEM Questionnaire Validation — Smoke Test")
    print("=" * 50)

    cfg = _load_config()
    if cfg is None:
        print("\nAborted — cannot read config.\n")
        return 1

    _check_inputs(cfg)
    run_ok = _run_validator()
    if run_ok:
        _check_report(cfg)
        _check_validated_questionnaire(cfg)
    else:
        print("\n  Skipping output checks — validator did not complete cleanly.")

    _check_admin_config()

    n_pass = sum(_passes)
    n_fail = len(_passes) - n_pass
    print(f"\n{'=' * 50}")
    print(f"  {n_pass} passed  |  {n_fail} failed  |  {len(_passes)} total")
    print("=" * 50 + "\n")
    return 0 if n_fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
