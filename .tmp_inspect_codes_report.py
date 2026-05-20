from collections import Counter
from openpyxl import load_workbook
wb=load_workbook(r'c:/git/DIEM_Questionnaire_Validation/tmp_report_out3/geopoll_output/report_geopoll_fr_COD_R11_20260520.xlsx')
ws=wb['Option Changes']
hdr=[ws.cell(2,c).value for c in range(1,20)]
idx={str(v):c for c,v in enumerate(hdr, start=1) if v}
issue=idx['Issue type']; qcol=idx['Q Name']; dcol=idx['Detail']; ccol=idx['Current value']; rcol=idx['Reference / rule']
rows=[]
for r in range(3, ws.max_row+1):
    it=ws.cell(r,issue).value
    if it:
        rows.append((r,str(it),ws.cell(r,qcol).value,ws.cell(r,dcol).value,ws.cell(r,ccol).value,ws.cell(r,rcol).value))
ctr=Counter([x[1] for x in rows])
print('counts',dict(ctr))
for t in ['codes_token_mismatch','codes_added','codes_removed','codes_position_drift']:
    print(t,ctr.get(t,0))
for rec in rows:
    if rec[1].startswith('codes_'):
        print('sample',rec[0],rec[1],rec[2],rec[3],rec[4],rec[5])
