import openpyxl
from pathlib import Path
src = Path(r'C:/Temp/DRC_R11 HH questionnaire_Geopoll_TEST.xlsx')
out = Path(r'c:/git/DIEM_Questionnaire_Validation/tmp_curr/DRC_R11_HH_questionnaire_Geopoll_TEST_CODESREMOVED.xlsx')
out.parent.mkdir(parents=True, exist_ok=True)
wb = openpyxl.load_workbook(src)
ws = wb['survey']
hdr=[ws.cell(3,c).value for c in range(1,80)]
idx={str(v).strip():c for c,v in enumerate(hdr,start=1) if v is not None}
qcol=idx['Q Name']; ccol=idx['Codes']
for r in range(4, ws.max_row+1):
    if str(ws.cell(r,qcol).value or '').strip() == 'assistance_provider':
        ws.cell(r,ccol).value = 'NA'
        print('edited row', r)
        break
wb.save(out)
print('written', out)
