import openpyxl
from collections import Counter
p = r"C:\git\DIEM_Questionnaire_Validation\batch_output\tmp_check\kobo_output\report_kobo_fr_TCD_R10_20260519.xlsx"
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb['Choice Changes']
rows=[]
for r in range(3, ws.max_row+1):
    q = ws.cell(r,3).value
    if q in ('need_','shock'):
        rows.append((r, ws.cell(r,1).value, ws.cell(r,6).value, ws.cell(r,5).value))
print('ROWS', len(rows))
for rr in rows:
    print(rr)
print('COUNTS', Counter([x[1] for x in rows]))
wb.close()

wb2 = openpyxl.load_workbook(p, data_only=False, rich_text=True)
ws2 = wb2['Choice Changes']
for r in range(3, ws2.max_row+1):
    if ws2.cell(r,1).value == 'choice_changes_general' and ws2.cell(r,3).value == 'need_':
        v = ws2.cell(r,6).value
        print('DETAIL_RICHTEXT_TYPE', type(v).__name__)
        print('DETAIL_TEXT', str(v)[:260])
        break
wb2.close()
