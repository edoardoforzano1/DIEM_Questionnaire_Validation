import openpyxl, re
from pathlib import Path
pat = re.compile(r'(?m)^\s*(\d+)\)\s*(.*?)(?=^\s*\d+\)|\Z)', re.DOTALL)
for label, path in [('current', Path(r'C:/Temp/DRC_R11 HH questionnaire_Geopoll_TEST.xlsx'))]:
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws=wb['survey']
    hdr=[ws.cell(3,c).value for c in range(1,80)]
    idx={str(v).strip():c for c,v in enumerate(hdr, start=1) if v is not None}
    qcol=idx.get('Q Name'); ccol=idx.get('Codes')
    nonna=0; parsed=0; byq=[]
    for r in range(4, ws.max_row+1):
        q=ws.cell(r,qcol).value if qcol else None
        t=ws.cell(r,ccol).value if ccol else None
        s=str(t or '').strip()
        if s and s.upper()!='NA':
            nonna+=1
            m=list(pat.finditer(s))
            if m:
                parsed += len(m)
                byq.append((q,r,len(m),s[:80].replace('\n',' | ')))
    print(label,'nonNA codes cells',nonna,'parsed rows',parsed,'questions with parsed',len(byq))
    for rec in byq[:8]:
        print(' ',rec)
