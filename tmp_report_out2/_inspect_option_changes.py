from collections import Counter
from openpyxl import load_workbook
p = r"tmp_report_out2/geopoll_output/report_geopoll_fr_COD_R11_20260520.xlsx"
wb = load_workbook(p)
ws = wb['Option Changes']
header = [ws.cell(row=2, column=i).value for i in range(1, 20)]
idx = {str(v): i + 1 for i, v in enumerate(header) if v}
print('HEADER', header[:12])
issue_col = idx.get('Issue type'); detail_col = idx.get('Detail'); q_col = idx.get('Q Name')
cur_col = idx.get('Current value'); ref_col = idx.get('Reference / rule')
lang_scope_col = idx.get('Language scope')
cur_l_col = idx.get('Current value (FR)'); ref_l_col = idx.get('Reference / rule (FR)')
rows=[]
for r in range(3, ws.max_row+1):
    issue=ws.cell(r, issue_col).value if issue_col else None
    if issue:
        rows.append({
            'r':r,
            'issue':issue,
            'q':ws.cell(r,q_col).value if q_col else '',
            'detail':ws.cell(r,detail_col).value if detail_col else '',
            'scope':ws.cell(r,lang_scope_col).value if lang_scope_col else '',
            'cur':ws.cell(r,cur_col).value if cur_col else '',
            'ref':ws.cell(r,ref_col).value if ref_col else '',
            'cur_l':ws.cell(r,cur_l_col).value if cur_l_col else '',
            'ref_l':ws.cell(r,ref_l_col).value if ref_l_col else '',
        })
c=Counter(x['issue'] for x in rows)
print('COUNTS',dict(c))
# rows where EN side looks same but target likely changed
for x in rows:
    if x['issue'] in ('added_option','removed_option','option_changes (added/removed)','option_label_mismatch','option_position_drift'):
        if (x['scope'] and 'FR' in str(x['scope']) and 'EN' not in str(x['scope'])) or (not x['cur'] and x['cur_l']):
            print('TARGET_ONLY',x['r'],x['issue'],x['q'],'scope=',x['scope'])
            print('  EN cur/ref empty?',not bool(x['cur']),not bool(x['ref']))
            print('  FR cur/ref empty?',not bool(x['cur_l']),not bool(x['ref_l']))
            break
# show one added row and confirm option list is label-only (no duplicated 1|1)
for x in rows:
    if x['issue'] in ('added_option','option_changes (added/removed)'):
        print('SAMPLE',x['r'],x['issue'],x['q'],'scope=',x['scope'])
        print('DETAIL',x['detail'])
        print('CUR_HEAD',str(x['cur']).split('\n')[:6])
        print('REF_HEAD',str(x['ref']).split('\n')[:6])
        print('CUR_FR_HEAD',str(x['cur_l']).split('\n')[:6])
        print('REF_FR_HEAD',str(x['ref_l']).split('\n')[:6])
        break
# verify code rows have empty FR columns
for x in rows:
    if str(x['issue']).startswith('codes_'):
        print('CODE_ROW',x['r'],x['issue'],'scope=',x['scope'],'fr_cols_empty=',(not bool(x['cur_l']) and not bool(x['ref_l'])))
        break
