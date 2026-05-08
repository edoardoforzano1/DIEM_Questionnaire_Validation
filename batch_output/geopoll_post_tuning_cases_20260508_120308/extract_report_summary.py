import json
from pathlib import Path
from collections import Counter
import openpyxl

root = Path(r"""C:\\git\\DIEM_Questionnaire_Validation\\batch_output\\geopoll_post_tuning_cases_20260508_120308""")
rows = json.loads((root / "run_results.json").read_text(encoding="utf-8-sig"))
out = []
for r in rows:
    report = Path(r["report"])
    wb = openpyxl.load_workbook(report, data_only=True, read_only=True)
    s = wb["Summary"]
    q = wb["Question Changes"]
    struct = wb["Questionnaire Structure"]

    section_titles = []
    issue_types = Counter()
    severities = Counter()
    lang_scope_values = Counter()

    header_map = None
    for row in q.iter_rows(values_only=True):
        vals = ["" if v is None else str(v).strip() for v in row]
        first = vals[0] if vals else ""
        if first.startswith("QUESTION CHANGES") or first.startswith("LESS IMPORTANT FIELD CHANGES"):
            section_titles.append(first)
        if "Issue type" in vals and "Severity" in vals:
            header_map = {v:i for i,v in enumerate(vals) if v}
            continue
        if header_map and any(vals):
            it_idx = header_map.get("Issue type")
            sev_idx = header_map.get("Severity")
            ls_idx = header_map.get("Language scope")
            if it_idx is not None and it_idx < len(vals):
                it = vals[it_idx]
                if it and it != "No issues in this category":
                    issue_types[it] += 1
            if sev_idx is not None and sev_idx < len(vals):
                sv = vals[sev_idx].lower()
                if sv in {"high","medium","info"}:
                    severities[sv] += 1
            if ls_idx is not None and ls_idx < len(vals):
                lv = vals[ls_idx]
                if lv:
                    lang_scope_values[lv] += 1

    drift_mentions = 0
    for row in struct.iter_rows(values_only=True):
        for v in row:
            if v is None:
                continue
            if "potential_skip_pattern_option_drift" in str(v):
                drift_mentions += 1

    out.append({
        "case": r["case"],
        "summary_A2": s["A2"].value,
        "summary_A3": s["A3"].value,
        "summary_A4": s["A4"].value,
        "question_sections": section_titles,
        "q_change_severity_counts": dict(severities),
        "q_change_lang_scope_top": dict(lang_scope_values),
        "has_drift_issue_mentions": drift_mentions > 0,
        "q_issue_types_count": len(issue_types),
    })

print(json.dumps(out, ensure_ascii=False, indent=2))
