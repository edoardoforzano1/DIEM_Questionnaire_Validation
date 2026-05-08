import json
from pathlib import Path
import openpyxl

root = Path(r"""C:\git\DIEM_Questionnaire_Validation\batch_output\geopoll_post_tuning_verify_20260508_120947""")
rows = json.loads((root / "run_results.json").read_text(encoding="utf-8-sig"))
for r in rows:
    wb = openpyxl.load_workbook(Path(r["report"]), data_only=True, read_only=True)
    s = wb["Summary"]
    print("---", r["case"], "---")
    print("A2:", s["A2"].value)
    print("A3:", s["A3"].value)
    print("A4:", s["A4"].value)
    print("A5:", s["A5"].value)
    q = wb["Question Changes"]
    found = False
    for row in q.iter_rows(values_only=True):
        vals=["" if v is None else str(v) for v in row]
        if any("LESS IMPORTANT FIELD CHANGES" in v for v in vals):
            found=True
            break
    print("Has less-important section:", found)
    struct = wb["Questionnaire Structure"]
    drift=False
    for row in struct.iter_rows(values_only=True):
        for v in row:
            if v and "potential_skip_pattern_option_drift" in str(v):
                drift=True
    print("Has drift issue:", drift)
