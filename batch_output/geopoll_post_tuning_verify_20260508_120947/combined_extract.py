import json, re
from pathlib import Path
from collections import Counter
import openpyxl

GROOT = Path(r"""C:\\git\\DIEM_Questionnaire_Validation\\batch_output\\geopoll_post_tuning_verify_20260508_120947""")
KROOT = Path(r"""C:\\git\\DIEM_Questionnaire_Validation\\batch_output\\kobo_post_check_20260508_121221""")

sev_re = re.compile(r"^\s*(HIGH|MEDIUM|INFO)\s+(\d+)\b")

def parse_log_counts(path: Path):
    out = {"HIGH":0,"MEDIUM":0,"INFO":0}
    for line in path.read_text(encoding='utf-8', errors='ignore').splitlines():
        m=sev_re.match(line)
        if m:
            out[m.group(1)] = int(m.group(2))
    return out

geo_rows = json.loads((GROOT / 'run_results.json').read_text(encoding='utf-8-sig'))
geo_out=[]
for r in geo_rows:
    report=Path(r['report'])
    wb=openpyxl.load_workbook(report,data_only=True,read_only=True)
    s=wb['Summary']
    q=wb['Question Changes']

    section = 'core'
    header=None
    counts=Counter()
    less_info_non_info=0
    for row in q.iter_rows(values_only=True):
        vals=["" if v is None else str(v).strip() for v in row]
        first=vals[0] if vals else ''
        if first.startswith('LESS IMPORTANT FIELD CHANGES'):
            section='less'
            header=None
            continue
        if 'Issue type' in vals and 'Severity' in vals:
            header={v:i for i,v in enumerate(vals) if v}
            continue
        if header and any(vals):
            it=vals[header.get('Issue type',-1)] if header.get('Issue type',-1)>=0 else ''
            sev=vals[header.get('Severity',-1)].lower() if header.get('Severity',-1)>=0 else ''
            if it and it!='No issues in this category':
                counts[f"{section}:{it}"] += 1
                if section=='less' and sev!='info':
                    less_info_non_info += 1

    drift=False
    struct=wb['Questionnaire Structure']
    for row in struct.iter_rows(values_only=True):
        for v in row:
            if v and 'potential_skip_pattern_option_drift' in str(v):
                drift=True

    geo_out.append({
        'case': r['case'],
        'log_counts': parse_log_counts(Path(r['log'])),
        'summary_A3': s['A3'].value,
        'summary_A4': s['A4'].value,
        'summary_A5': s['A5'].value,
        'less_section_non_info_rows': less_info_non_info,
        'less_conditional_count': counts.get('less:conditional_changed',0),
        'less_randomize_count': counts.get('less:randomize_changed',0),
        'less_programming_count': counts.get('less:programming_instructions_changed',0),
        'less_core_only_count': counts.get('less:core_questions_only_changed',0),
        'has_drift_issue': drift,
    })

kobo_rows = json.loads((KROOT / 'run_results.json').read_text(encoding='utf-8-sig'))
kobo_out=[]
for r in kobo_rows:
    kobo_out.append({'case': r['case'], 'log_counts': parse_log_counts(Path(r['log']))})

print(json.dumps({'geopoll_root': str(GROOT), 'kobo_root': str(KROOT), 'geopoll': geo_out, 'kobo': kobo_out}, ensure_ascii=False, indent=2))
