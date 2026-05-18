import openpyxl, collections, json
p = r"C:\Temp\questionnaire_validation\kobo_output\report_kobo_ar_PSE_R2_20260518.xlsx"
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb['Choice Changes']
hdr = None
hrow = None
for r in range(1, min(ws.max_row, 80) + 1):
    vals = [ws.cell(r, c).value for c in range(1, 40)]
    if any((str(v).strip() == 'Issue type') for v in vals if v is not None):
        hdr = vals
        hrow = r
        break
if hdr is None:
    raise SystemExit('Header row not found')
cols = {str(v).strip(): i + 1 for i, v in enumerate(hdr) if v is not None}
rows = []
for r in range(hrow + 1, ws.max_row + 1):
    it_col = cols.get('Issue type')
    if not it_col:
        continue
    it = ws.cell(r, it_col).value
    if it is None:
        continue
    it_s = str(it).strip()
    if not it_s or it_s.lower().startswith('no issues'):
        continue
    d = {k: ws.cell(r, c).value for k, c in cols.items()}
    rows.append(d)
print('rows', len(rows))
ct = collections.Counter(str(r.get('Issue type')) for r in rows)
print('issue_counts', json.dumps(ct, ensure_ascii=False))
root = {'removed_option_causes_index_drift', 'option_name_dictionary_replaced'}
by_key = collections.defaultdict(set)
for r in rows:
    q = str(r.get('Q Name') or '')
    l = str(r.get('list_name') or '')
    by_key[(q, l)].add(str(r.get('Issue type') or ''))
keys_root = {k for k, v in by_key.items() if (v & root)}
lm_on_root = sum(1 for r in rows if str(r.get('Issue type')) == 'choice_label_mismatch' and (str(r.get('Q Name') or ''), str(r.get('list_name') or '')) in keys_root)
lm_total = ct.get('choice_label_mismatch', 0)
print('label_mismatch_total', lm_total)
print('label_mismatch_on_root_keys', lm_on_root)
print('label_mismatch_off_root_keys', lm_total - lm_on_root)
add_total = ct.get('added_choice', 0)
add_high = sum(1 for r in rows if str(r.get('Issue type')) == 'added_choice' and str(r.get('Severity') or '').lower() == 'high')
add_medium = sum(1 for r in rows if str(r.get('Issue type')) == 'added_choice' and str(r.get('Severity') or '').lower() == 'medium')
print('added_choice_total', add_total, 'high', add_high, 'medium', add_medium)
print('added_choice_qnames', sorted(set(str(r.get('Q Name') or '') for r in rows if str(r.get('Issue type')) == 'added_choice')))
print('root_keys_count', len(keys_root))
for q, l in sorted(keys_root):
    issues = sorted(by_key[(q, l)])
    print('ROOT_KEY', q, '|', l, '|', ','.join(issues))
