import openpyxl
p = r"C:\git\DIEM_Questionnaire_Validation\batch_output\tmp_check\kobo_output\report_kobo_fr_TCD_R10_20260519.xlsx"
wb = openpyxl.load_workbook(p, data_only=False, rich_text=True)
ws = wb['Choice Changes']
for r in range(3, ws.max_row+1):
    if ws.cell(r,1).value == 'choice_changes_general' and ws.cell(r,3).value == 'need_':
        v = ws.cell(r,6).value
        try:
            for i, blk in enumerate(v):
                b = getattr(getattr(blk,'font',None),'b',None)
                txt = getattr(blk,'text','')
                if b:
                    print('BOLD_BLOCK', i, txt[:80])
        except Exception as e:
            print('ERR',e)
        break
wb.close()
